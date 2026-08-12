"""Workflow for applying diagonal borders to blank cells in PCE Testsheet and PCE VI worksheets."""

from __future__ import annotations

import os
import re
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import range_boundaries

from src.postprocessing.converters import _is_pce_testsheet_sheet, _is_pce_vi_sheet

# Define diagonal border style
diagonal = Side(style="thin")

# List of ranges to process
TESTSHEET_RANGES_TO_PROCESS = ["A3:Y55", "A56:Q65"]
VI_SHEET_RANGES_TO_PROCESS = [
    "C6:H6", "K6:N6", "C7:H7", "K7:N7", "C12:H12", "C13:F13", "O13",
    "C15:H15", "C16:F16", "J16:K16", "O16"
]


@dataclass(frozen=True)
class DiagonalSummary:
    target_path: Path
    processed_count: int


def _range_bounds(range_string: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    return min_col, min_row, max_col, max_row


def _ranges_overlap(left: str, right: str) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = _range_bounds(left)
    right_min_col, right_min_row, right_max_col, right_max_row = _range_bounds(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def _iter_range_cells(ws: object, range_string: str) -> list[object]:
    cells = ws[range_string]
    if isinstance(cells, tuple):
        if cells and isinstance(cells[0], tuple):
            return [cell for row in cells for cell in row]
        return [cells]
    return [cells]


def _is_merged_cell(ws: object, cell: object) -> bool:
    return any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges)


def _resolve_cell(ws: object, ref_str: str, wb: object | None) -> tuple[object | None, object | None]:
    if "!" in ref_str:
        sheet_part, coord_part = ref_str.split("!", 1)
        sheet_name = sheet_part.strip("'\"")
        if wb and sheet_name in wb.sheetnames:
            target_ws = wb[sheet_name]
        else:
            target_ws = ws
    else:
        target_ws = ws
        coord_part = ref_str

    coord_part = coord_part.strip().upper().replace("$", "")
    if not re.match(r"^[A-Z]+[0-9]+$", coord_part):
        return None, None

    try:
        return target_ws, target_ws[coord_part]
    except Exception:
        return None, None


def _is_blank_value(ws: object, cell: object, wb: object | None = None, visited: set | None = None) -> bool:
    if visited is None:
        visited = set()
    key = (ws.title, cell.coordinate)
    if key in visited:
        return True
    visited.add(key)

    val = cell.value
    if val is None:
        return True
    if isinstance(val, str):
        if not val.startswith("="):
            return val.strip() == ""
        return _evaluate_formula_blank(ws, val, wb, visited)
    return str(val).strip() == ""


def _evaluate_formula_blank(ws: object, formula_str: str, wb: object | None, visited: set) -> bool:
    s = formula_str.strip()
    if s.startswith("="):
        s = s[1:].strip()

    if s.upper().startswith("IF("):
        inner = s[3:-1] if s.endswith(")") else s[3:]
        args = []
        current = []
        paren_level = 0
        in_quote = False
        for char in inner:
            if char == '"':
                in_quote = not in_quote
                current.append(char)
            elif char == "(" and not in_quote:
                paren_level += 1
                current.append(char)
            elif char == ")" and not in_quote:
                paren_level -= 1
                current.append(char)
            elif char == "," and paren_level == 0 and not in_quote:
                args.append("".join(current).strip())
                current = []
            else:
                current.append(char)
        if current:
            args.append("".join(current).strip())

        if len(args) >= 3:
            cond, v_true, v_false = args[0], args[1], args[2]
            cond_res = _eval_condition(ws, cond, wb, visited.copy())
            target_expr = v_true if cond_res else v_false
            return _check_expr_blank(ws, target_expr, wb, visited)
        elif len(args) == 2:
            cond, v_true = args[0], args[1]
            cond_res = _eval_condition(ws, cond, wb, visited.copy())
            target_expr = v_true if cond_res else '""'
            return _check_expr_blank(ws, target_expr, wb, visited)

    return _check_expr_blank(ws, s, wb, visited)


def _eval_condition(ws: object, cond_str: str, wb: object | None, visited: set) -> bool:
    cond_str = cond_str.strip()
    m = re.match(r"^ISBLANK\((.+)\)$", cond_str, re.I)
    if m:
        ref = m.group(1).strip()
        t_ws, t_cell = _resolve_cell(ws, ref, wb)
        if t_cell:
            return _is_blank_value(t_ws, t_cell, wb, visited)
        return True

    if cond_str.upper().startswith("AND(") and cond_str.endswith(")"):
        inner = cond_str[4:-1]
        parts = [p.strip() for p in inner.split(",")]
        return all(_eval_condition(ws, p, wb, visited) for p in parts)

    m_eq = re.split(r"(=|<>)", cond_str)
    if len(m_eq) == 3:
        left, op, right = m_eq[0].strip(), m_eq[1].strip(), m_eq[2].strip()
        t_ws, t_cell = _resolve_cell(ws, left, wb)
        if t_cell:
            is_blank = _is_blank_value(t_ws, t_cell, wb, visited)
            if op == "=":
                if right in ('""', "''"):
                    return is_blank
                elif right == "0":
                    return is_blank or str(t_cell.value).strip() == "0"
            elif op == "<>":
                if right in ('""', "''"):
                    return not is_blank
                elif right == "0":
                    return not (is_blank or str(t_cell.value).strip() == "0")
    return False


def _check_expr_blank(ws: object, expr_str: str, wb: object | None, visited: set) -> bool:
    expr_str = expr_str.strip()
    if expr_str in ('""', "''", ""):
        return True
    if expr_str.startswith('"') and expr_str.endswith('"'):
        return expr_str[1:-1].strip() == ""
    if re.match(r"^-?\d+(\.\d+)?$", expr_str):
        return False
    t_ws, t_cell = _resolve_cell(ws, expr_str, wb)
    if t_cell:
        return _is_blank_value(t_ws, t_cell, wb, visited)
    return False


def _apply_diagonal_border(cell: object) -> None:
    current_border = cell.border
    cell.border = Border(
        left=current_border.left,
        right=current_border.right,
        top=current_border.top,
        bottom=current_border.bottom,
        diagonal=diagonal,
        diagonalDown=False,
        diagonalUp=True,
        vertical=current_border.vertical,
        horizontal=current_border.horizontal,
        outline=current_border.outline,
    )


def process_range(ws: object, range_string: str, wb: object | None = None) -> None:
    for merged_range in ws.merged_cells.ranges:
        if _ranges_overlap(merged_range.coord, range_string):
            cells = ws[merged_range.coord]
            if isinstance(cells, tuple) and cells and isinstance(cells[0], tuple):
                top_left_cell = cells[0][0]
            elif isinstance(cells, tuple) and cells:
                top_left_cell = cells[0]
            else:
                top_left_cell = cells
            if _is_blank_value(ws, top_left_cell, wb):
                if isinstance(cells, tuple):
                    if cells and isinstance(cells[0], tuple):
                        for row in cells:
                            for cell in row:
                                _apply_diagonal_border(cell)
                    else:
                        for cell in cells:
                            _apply_diagonal_border(cell)
                else:
                    _apply_diagonal_border(cells)

    for cell in _iter_range_cells(ws, range_string):
        if _is_merged_cell(ws, cell):
            continue
        if _is_blank_value(ws, cell, wb):
            _apply_diagonal_border(cell)


def _repair_openpyxl_zip(input_path: Path, output_path: Path) -> None:
    """Repair openpyxl saved workbook by fixing table relationship target paths and restoring missing drawing parts."""
    if not output_path.exists():
        return

    input_drawings: dict[str, bytes] = {}
    if input_path.exists():
        try:
            with zipfile.ZipFile(input_path, "r") as z_in:
                for name in z_in.namelist():
                    if name.startswith("xl/drawings/"):
                        input_drawings[name] = z_in.read(name)
        except Exception:
            pass

    temp_repaired = output_path.with_suffix(".repaired.tmp")
    try:
        with zipfile.ZipFile(output_path, "r") as z_out, zipfile.ZipFile(temp_repaired, "w") as z_rep:
            for item in z_out.infolist():
                content = z_out.read(item.filename)
                if item.filename.startswith("xl/worksheets/_rels/"):
                    content_str = content.decode("utf-8", errors="ignore")
                    fixed_str = re.sub(r'Target="/xl/tables/', 'Target="../tables/', content_str)
                    fixed_str = re.sub(r'Target="/xl/drawings/', 'Target="../drawings/', fixed_str)
                    content = fixed_str.encode("utf-8")
                z_rep.writestr(item, content)

            out_names = set(z_out.namelist())
            for d_name, d_bytes in input_drawings.items():
                if d_name not in out_names:
                    z_rep.writestr(d_name, d_bytes)

        shutil.move(temp_repaired, output_path)
    except Exception:
        if temp_repaired.exists():
            temp_repaired.unlink(missing_ok=True)


def process_workbook(file_path: str | Path) -> str:
    """Process a single workbook applying diagonal borders to blank cells."""
    path = Path(file_path).expanduser().resolve()
    wb = load_workbook(str(path), data_only=False)

    for ws in wb.worksheets:
        if _is_pce_testsheet_sheet(ws.title):
            for rng in TESTSHEET_RANGES_TO_PROCESS:
                process_range(ws, rng, wb)
        elif _is_pce_vi_sheet(ws.title):
            for rng in VI_SHEET_RANGES_TO_PROCESS:
                process_range(ws, rng, wb)

    for ws in wb.worksheets:
        ws._tables.clear()

    if path.parent.name == "processed_testsheet":
        new_file_path = path
        wb.save(str(new_file_path))
        print(f"Processed {path.name} -> saved in-place as {new_file_path}")
    else:
        output_dir = path.parent / "processed_testsheet"
        output_dir.mkdir(parents=True, exist_ok=True)
        new_file_path = output_dir / path.name
        wb.save(str(new_file_path))
        print(f"Processed {path.name} -> saved as {new_file_path}")

    wb.close()
    _repair_openpyxl_zip(path, new_file_path)
    return str(new_file_path)


def process_diagonal_target(target_path: str | Path) -> DiagonalSummary:
    """Process all Excel workbooks in target directory or a single workbook."""
    path = Path(target_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Target not found: {path}")

    processed_count = 0
    if path.is_file():
        if path.suffix.lower() == ".xlsx" and not path.name.startswith("~$"):
            process_workbook(path)
            processed_count = 1
    else:
        for filename in sorted(os.listdir(path)):
            if (
                filename.endswith(".xlsx")
                and not filename.startswith("processed_")
                and not filename.startswith("~$")
            ):
                file_path = path / filename
                if file_path.is_file():
                    process_workbook(file_path)
                    processed_count += 1

    return DiagonalSummary(target_path=path, processed_count=processed_count)


def run_diagonal() -> DiagonalSummary:
    """Interactive entrypoint for applying diagonal borders to blank testsheet cells."""
    target_folder = input("Enter the path to the target folder or Excel file: ").strip().strip('"')
    summary = process_diagonal_target(target_folder)
    print(f"Completed diagonal border processing for {summary.processed_count} file(s).")
    return summary
