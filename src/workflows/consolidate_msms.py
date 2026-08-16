"""Consolidate MSMS Workflow (PYTHON/MSMS/*.xls -> DATA MSMS.xlsx).

Ingests scattered .xls HTML files from PYTHON/MSMS/ into master DATA MSMS.xlsx,
deduplicates work orders, normalizes FL ERMS, and moves processed files to PYTHON/MSMS/COMPLETED/.

Resilience Policy: best-effort - Collects errors per file, continues processing remaining files.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
import shutil
from typing import Any, Sequence

import openpyxl
import pandas as pd

from src.core.normalizers import normalize_fl_erms
from src.project.environment import ProjectEnvironment
from src.workflows.models import (
    ConsolidateMsmsRequest,
    ConsolidateMsmsResult,
)

logger = logging.getLogger(__name__)


class ConsolidateMsmsPreflightGuard:
    """Pre-flight resource guard stage for Consolidate MSMS workflow."""

    def validate(self, environment: ProjectEnvironment) -> None:
        """Validate environmental preconditions before reading data."""
        data_msms_path = environment.storage.get_data_msms_path()
        if not data_msms_path.exists():
            raise FileNotFoundError(f"DATA MSMS workbook not found: {data_msms_path}")

        python_msms_dir = environment.storage.get_python_msms_dir()
        if not python_msms_dir.exists() or not python_msms_dir.is_dir():
            raise FileNotFoundError(f"PYTHON/MSMS directory not found: {python_msms_dir}")

        xls_files = environment.storage.list_msms_xls_files()
        if not xls_files:
            raise FileNotFoundError(f"No .xls files found in {python_msms_dir}")


@dataclass(frozen=True)
class ConsolidateMsmsRow:
    """Immutable representation of a normalized DATA MSMS row to append."""

    wo: str
    location: str
    description: str
    substation_name_erms: str | None = None
    fl_erms: str | None = None
    cycle_date: str | None = None
    substation_number: Any = None


@dataclass(frozen=True)
class ConsolidateMsmsPlan:
    """Transformation execution plan for Consolidate MSMS workflow."""

    target_data_msms: Path
    completed_dir: Path
    rows_to_append: tuple[ConsolidateMsmsRow, ...]
    files_to_move: tuple[Path, ...]
    duplicates_skipped: int = 0
    files_processed: int = 0
    errors: tuple[str, ...] = ()


class ConsolidateMsmsExtractor:
    """Pure I/O reading stage for Consolidate MSMS workflow."""

    def get_existing_wos(self, data_msms_path: Path) -> set[str]:
        """Read existing Work Order strings from DATA MSMS.xlsx Column A."""
        wb = openpyxl.load_workbook(data_msms_path, read_only=True)
        ws = wb.active
        existing_wos: set[str] = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0] if row else None
            if val is not None:
                val_str = str(val).strip()
                if val_str and val_str.lower() not in ("none", "nan"):
                    existing_wos.add(val_str)
        wb.close()
        return existing_wos

    def read_xls_file(self, xls_path: Path) -> tuple[list[dict[str, str]], str | None]:
        """Read HTML table from a Maximo .xls export file.

        Returns tuple of (extracted_rows, error_message).
        """
        if not xls_path.exists():
            return [], f"File not found: '{xls_path}'"

        try:
            tables = pd.read_html(str(xls_path), flavor="lxml")
        except ValueError as e:
            logger.warning(f"No tables found in '{xls_path.name}': {e}")
            return [], f"Failed to read HTML table in '{xls_path.name}': {e}"
        except Exception as e:
            return [], f"Failed to read HTML table in '{xls_path.name}': {e}"

        if not tables or tables[0].empty:
            logger.warning(f"No table data found in '{xls_path.name}'")
            return [], None

        df = tables[0]
        extracted: list[dict[str, str]] = []

        for _, row in df.iterrows():
            wo_raw = row.iloc[0] if len(row) > 0 else None
            if pd.isna(wo_raw):
                continue
            wo_str = str(wo_raw).strip()
            if not wo_str:
                continue

            status_raw = row.iloc[1] if len(row) > 1 else ""
            status_str = str(status_raw).strip() if not pd.isna(status_raw) else ""

            loc_raw = row.iloc[2] if len(row) > 2 else ""
            loc_str = str(loc_raw).strip() if not pd.isna(loc_raw) else ""
            if loc_str.upper() in ("LOCATION", "NAN", "NONE"):
                loc_str = ""

            desc_raw = row.iloc[3] if len(row) > 3 else ""
            desc_str = str(desc_raw).strip() if not pd.isna(desc_raw) else ""
            if desc_str.upper() in ("DESCRIPTION", "NAN", "NONE"):
                desc_str = ""

            extracted.append({
                "wo": wo_str,
                "status": status_str,
                "location": loc_str,
                "description": desc_str,
            })

        return extracted, None


class ConsolidateMsmsFilter:
    """Pure predicate logic stage for Consolidate MSMS workflow."""

    def filter_rows(
        self,
        extracted_files: Sequence[tuple[Path, Sequence[dict[str, str]]]],
        existing_wos: set[str],
    ) -> tuple[list[tuple[Path, list[dict[str, str]]]], int]:
        """Deduplicate rows across files and skip already existing Work Orders.

        Returns tuple of (filtered_files, duplicates_skipped_count).
        """
        seen_wos = set(existing_wos)
        duplicates_skipped = 0
        filtered_files: list[tuple[Path, list[dict[str, str]]]] = []

        for path, rows in extracted_files:
            valid_rows: list[dict[str, str]] = []
            for row in rows:
                wo = row["wo"]
                if wo.upper() in ("WONUM", "WORK ORDER", "WO", "NAN", "NONE"):
                    continue

                if wo in seen_wos:
                    duplicates_skipped += 1
                    continue

                seen_wos.add(wo)
                valid_rows.append(row)

            filtered_files.append((path, valid_rows))

        return filtered_files, duplicates_skipped


class ConsolidateMsmsTransformer:
    """Pure transformation plan construction stage for Consolidate MSMS workflow."""

    def build_plan(
        self,
        target_data_msms: Path,
        completed_dir: Path,
        filtered_files: Sequence[tuple[Path, Sequence[dict[str, str]]]],
        duplicates_skipped: int,
        files_processed: int,
        errors: Sequence[str],
    ) -> ConsolidateMsmsPlan:
        """Construct immutable ConsolidateMsmsPlan from filtered records."""
        rows_to_append: list[ConsolidateMsmsRow] = []
        files_to_move: list[Path] = []

        for path, rows in filtered_files:
            files_to_move.append(path)
            for row in rows:
                fl_erms = normalize_fl_erms(row["location"])
                rows_to_append.append(
                    ConsolidateMsmsRow(
                        wo=row["wo"],
                        location=row["location"],
                        description=row["description"],
                        substation_name_erms=None,
                        fl_erms=fl_erms,
                        cycle_date=None,
                        substation_number=None,
                    )
                )

        return ConsolidateMsmsPlan(
            target_data_msms=target_data_msms,
            completed_dir=completed_dir,
            rows_to_append=tuple(rows_to_append),
            files_to_move=tuple(files_to_move),
            duplicates_skipped=duplicates_skipped,
            files_processed=files_processed,
            errors=tuple(errors),
        )


class ConsolidateMsmsLoader:
    """Pure write I/O stage for Consolidate MSMS workflow."""

    def load(self, plan: ConsolidateMsmsPlan) -> None:
        """Append rows to DATA MSMS.xlsx contiguously, compact blank gaps, and move processed .xls files."""
        if plan.target_data_msms.exists():
            wb = openpyxl.load_workbook(plan.target_data_msms)
            ws = wb.active

            self._write_contiguous_and_compact(ws, plan.rows_to_append)

            wb.save(plan.target_data_msms)
            wb.close()

        if plan.files_to_move:
            plan.completed_dir.mkdir(parents=True, exist_ok=True)
            for file_path in plan.files_to_move:
                if file_path.exists():
                    dst = plan.completed_dir / file_path.name
                    if dst.exists():
                        dst.unlink()
                    shutil.move(str(file_path), str(dst))

    @staticmethod
    def _is_non_empty_row(row_vals: Sequence[Any]) -> bool:
        """Check if any cell in row values contains non-empty, non-sentinel data."""
        for val in row_vals:
            if val is not None:
                val_str = str(val).strip()
                if val_str and val_str.lower() not in ("none", "nan"):
                    return True
        return False

    def _write_contiguous_and_compact(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        rows_to_append: Sequence[ConsolidateMsmsRow],
    ) -> None:
        """Scan existing non-empty rows, compact them, append new rows contiguously, and truncate trailing orphan rows."""
        orig_max_row = ws.max_row or 1
        num_cols = max(ws.max_column or 7, 7)

        # Scan existing non-empty rows (skipping empty rows)
        existing_records: list[list[Any]] = []
        for r_idx in range(2, orig_max_row + 1):
            row_vals = [ws.cell(r_idx, c).value for c in range(1, num_cols + 1)]
            if self._is_non_empty_row(row_vals):
                existing_records.append(row_vals)

        new_records: list[list[Any]] = [
            [
                row.wo,
                row.location,
                row.description,
                row.substation_name_erms,
                row.fl_erms,
                row.cycle_date,
                row.substation_number,
            ]
            for row in rows_to_append
        ]

        all_records = existing_records + new_records

        # Write contiguously starting from row 2
        for r_idx, r_data in enumerate(all_records, start=2):
            for c_idx in range(1, num_cols + 1):
                val = r_data[c_idx - 1] if c_idx - 1 < len(r_data) else None
                ws.cell(row=r_idx, column=c_idx).value = val

        last_row = 1 + len(all_records)
        if orig_max_row > last_row:
            ws.delete_rows(last_row + 1, orig_max_row - last_row)



class ConsolidateMsmsAuditor:
    """Verification & Audit stage for Consolidate MSMS workflow."""

    def audit(self, plan: ConsolidateMsmsPlan) -> ConsolidateMsmsResult:
        """Verify output integrity and construct telemetry result."""
        self._verify_output(plan.target_data_msms)
        return ConsolidateMsmsResult(
            files_processed=plan.files_processed,
            rows_appended=len(plan.rows_to_append),
            duplicates_skipped=plan.duplicates_skipped,
            errors=plan.errors,
            files_moved=plan.files_to_move,
        )

    def _verify_output(self, target_data_msms: Path) -> None:
        if not target_data_msms.exists():
            raise RuntimeError(f"DATA MSMS workbook does not exist: {target_data_msms}")
        if target_data_msms.stat().st_size == 0:
            raise RuntimeError(f"DATA MSMS workbook is empty (0 bytes): {target_data_msms}")


class ConsolidateMsmsWorkflow:
    """6-stage ETL workflow for consolidating scattered Maximo .xls files into DATA MSMS.xlsx.

    Resilience Policy: best-effort per file.
    """

    def __init__(
        self,
        preflight_guard: ConsolidateMsmsPreflightGuard | None = None,
        extractor: ConsolidateMsmsExtractor | None = None,
        filter_stage: ConsolidateMsmsFilter | None = None,
        transformer: ConsolidateMsmsTransformer | None = None,
        loader: ConsolidateMsmsLoader | None = None,
        auditor: ConsolidateMsmsAuditor | None = None,
    ) -> None:
        self.preflight_guard = preflight_guard or ConsolidateMsmsPreflightGuard()
        self.extractor = extractor or ConsolidateMsmsExtractor()
        self.filter_stage = filter_stage or ConsolidateMsmsFilter()
        self.transformer = transformer or ConsolidateMsmsTransformer()
        self.loader = loader or ConsolidateMsmsLoader()
        self.auditor = auditor or ConsolidateMsmsAuditor()

    def execute(
        self,
        environment: ProjectEnvironment,
        request: ConsolidateMsmsRequest | None = None,
    ) -> ConsolidateMsmsResult:
        """Execute Consolidate MSMS workflow."""
        self.preflight_guard.validate(environment)

        req = request or ConsolidateMsmsRequest()
        if req.progress_sink:
            req.progress_sink("Starting Consolidate MSMS workflow...")

        data_msms_path = environment.storage.get_data_msms_path()
        completed_dir = environment.storage.get_python_msms_completed_dir()
        xls_files = environment.storage.list_msms_xls_files()

        existing_wos = self.extractor.get_existing_wos(data_msms_path)

        extracted_files: list[tuple[Path, list[dict[str, str]]]] = []
        errors: list[str] = []
        files_processed = 0

        for i, xls_path in enumerate(xls_files):
            if req.progress_sink:
                req.progress_sink(f"Reading file {i+1}/{len(xls_files)}: {xls_path.name}...")

            rows, err = self.extractor.read_xls_file(xls_path)
            if err:
                errors.append(err)
                continue

            extracted_files.append((xls_path, rows))
            files_processed += 1

        filtered_files, duplicates_skipped = self.filter_stage.filter_rows(
            extracted_files=extracted_files,
            existing_wos=existing_wos,
        )

        plan = self.transformer.build_plan(
            target_data_msms=data_msms_path,
            completed_dir=completed_dir,
            filtered_files=filtered_files,
            duplicates_skipped=duplicates_skipped,
            files_processed=files_processed,
            errors=errors,
        )

        if req.progress_sink:
            req.progress_sink(f"Appending {len(plan.rows_to_append)} rows to {data_msms_path.name}...")

        self.loader.load(plan)

        result = self.auditor.audit(plan)

        if req.progress_sink:
            req.progress_sink(
                f"Consolidate MSMS complete: {result.files_processed} files processed, "
                f"{result.rows_appended} rows appended, {result.duplicates_skipped} duplicates skipped."
            )

        return result
