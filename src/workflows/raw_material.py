"""Automated Raw Material Creation & Sorting Workflow for Pahang CLI."""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence
import zipfile
import openpyxl

from src.core.normalizers import format_month_folder, normalize_date_str
from src.project.environment import ProjectEnvironment
from src.project.models import CameraConfig
from src.testsheet.extractor import TestsheetExtractor
from src.testsheet.models import PhotoRange, SubstationTestsheetPackage, TestsheetData
from src.testsheet.repository import SubstationTestsheetRepository
from src.workflows.models import RawMaterialRequest, RawMaterialResult


@dataclass(frozen=True)
class AutomatedRawMaterialSummary:
    """Execution summary for Raw Material Creation & Sorting workflow."""

    substations_count: int = 0
    ir_copied_count: int = 0
    dg_copied_count: int = 0
    us_tev_extracted_count: int = 0
    warnings: tuple[str, ...] = ()
    errors: tuple[str, ...] = ()



@dataclass(frozen=True)
class ExtractedPackageData:
    """Extracted raw data and photo files for a package."""

    package: SubstationTestsheetPackage
    testsheet_data: TestsheetData | None
    ir_photos: tuple[Path, ...]
    dg_photos: tuple[Path, ...]
    substation_folder_name: str
    us_tev_archives: tuple[Path, ...] = ()


@dataclass(frozen=True)
class CopyInstruction:
    """Specification for copying a single photo to a destination path."""

    source_path: Path
    dest_path: Path
    tech_name: str
    substation_folder_name: str
    photo_number: int


@dataclass(frozen=True)
class ExtractInstruction:
    """Specification for unzipping a US+TEV archive to a destination directory."""

    source_archive: Path
    dest_dir: Path
    substation_folder_name: str


@dataclass(frozen=True)
class TransformationPlan:
    """Destination path resolution and copy plan for a package."""

    directories_to_create: tuple[Path, ...]
    copy_instructions: tuple[CopyInstruction, ...]
    substation_folder_name: str
    ir_count: int
    dg_count: int
    extract_instructions: tuple[ExtractInstruction, ...] = ()
    us_tev_count: int = 0


class RawMaterialPreflightGuard:
    """Pure I/O Check pre-flight validation stage."""

    def validate(self, environment: ProjectEnvironment, request: RawMaterialRequest) -> None:
        """Validate input directory, workspace directories, and TOTAL PE.xlsx preconditions."""
        input_dir = request.target_dir or request.output_path or environment.storage.get_testsheet_dir()
        self._validate_directories(environment, input_dir)
        total_pe_path = environment.storage.get_total_pe_path()
        self._validate_total_pe_workbook(total_pe_path)

    def _validate_directories(self, environment: ProjectEnvironment, input_dir: Path) -> None:
        if not input_dir.exists():
            raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

        raw_material_dir = environment.storage.get_raw_material_dir()
        if not raw_material_dir.exists():
            raise FileNotFoundError(f"Pre-flight failed: RAW MATERIAL directory missing at {raw_material_dir}")

        python_dir = environment.storage.get_python_dir()
        if not python_dir.exists():
            raise FileNotFoundError(f"Pre-flight failed: PYTHON directory missing at {python_dir}")

    def _validate_total_pe_workbook(self, total_pe_path: Path) -> None:
        if not total_pe_path.exists():
            raise RuntimeError(
                f"TOTAL PE.xlsx pre-check failed: File missing at {total_pe_path}. "
                "Please run 'Populate TOTAL PE' workflow first."
            )

        try:
            wb = openpyxl.load_workbook(total_pe_path, read_only=True)
            if "DataCycle1" not in wb.sheetnames:
                wb.close()
                raise RuntimeError(
                    "TOTAL PE.xlsx pre-check failed: 'DataCycle1' sheet missing in TOTAL PE.xlsx."
                )
            ws = wb["DataCycle1"]
            has_data = ws.max_row >= 2
            wb.close()
            if not has_data:
                raise RuntimeError(
                    "TOTAL PE.xlsx pre-check failed: 'DataCycle1' sheet is empty. "
                    "Please run 'Populate TOTAL PE' workflow first."
                )
        except Exception as e:
            if isinstance(e, RuntimeError):
                raise
            raise RuntimeError(f"TOTAL PE.xlsx pre-check failed while reading: {e}")


class RawMaterialExtractor:
    """Pure I/O reading stage for Raw Material workflow."""

    def __init__(
        self,
        repository: SubstationTestsheetRepository | None = None,
        extractor: TestsheetExtractor | None = None,
    ) -> None:
        self.repository = repository or SubstationTestsheetRepository()
        self.extractor = extractor or TestsheetExtractor()

    def discover_packages(
        self, environment: ProjectEnvironment, request: RawMaterialRequest
    ) -> list[SubstationTestsheetPackage]:
        """Discover testsheet packages and verify UNSORTED RAW DATA directories exist."""
        input_path = request.target_dir or request.output_path or environment.storage.get_testsheet_dir()
        packages = self.repository.discover_packages(input_path)

        if not packages:
            raise FileNotFoundError(
                f"Input directory verification failed: No testsheets or packages found in {input_path}"
            )

        self._verify_unsorted_raw_data_dirs(packages)
        return packages

    def _verify_unsorted_raw_data_dirs(self, packages: Sequence[SubstationTestsheetPackage]) -> None:
        for pkg in packages:
            if not pkg.unsorted_raw_data_dir.exists():
                raise RuntimeError(
                    f"Input directory verification failed: 'UNSORTED RAW DATA' directory missing in {pkg.testsheet_path.parent}. "
                    "Field crew must provide UNSORTED RAW DATA/ folder containing raw photos."
                )

    def load_pe_alignment_data(self, total_pe_path: Path) -> set[tuple[str, str]]:
        """Read existing (PE/Station, Date) tuples from TOTAL PE.xlsx."""
        wb_pe = openpyxl.load_workbook(total_pe_path, data_only=True)
        try:
            if "DataCycle1" not in wb_pe.sheetnames:
                raise RuntimeError("TOTAL PE.xlsx missing 'DataCycle1' sheet.")
            ws_pe = wb_pe["DataCycle1"]
            existing_tuples: set[tuple[str, str]] = set()
            for r_idx in range(2, ws_pe.max_row + 1):
                existing_tuples.update(self._extract_tuples_from_row(ws_pe, r_idx))
            return existing_tuples
        finally:
            wb_pe.close()

    def _extract_tuples_from_row(self, ws_pe: openpyxl.worksheet.worksheet.Worksheet, r_idx: int) -> set[tuple[str, str]]:
        substation_val_raw = str(ws_pe.cell(r_idx, 1).value or "").strip()
        sub_val = str(ws_pe.cell(r_idx, 3).value or "").strip().upper()
        dt_val = str(ws_pe.cell(r_idx, 4).value or "").strip()
        norm_dt = normalize_date_str(dt_val)

        row_tuples: set[tuple[str, str]] = set()
        for d in (dt_val, norm_dt):
            if not d:
                continue
            if substation_val_raw:
                row_tuples.add((substation_val_raw, d))
                try:
                    num_int = int(substation_val_raw)
                    row_tuples.add((str(num_int), d))
                    row_tuples.add((f"{num_int:03d}", d))
                except ValueError:
                    pass
            if sub_val:
                row_tuples.add((sub_val, d))
        return row_tuples

    def extract_package_data(
        self, pkg: SubstationTestsheetPackage, warnings: list[str]
    ) -> ExtractedPackageData:
        """Read testsheet data and scan photo files for a package."""
        data = self._load_testsheet_data(pkg, warnings)
        substation_folder_name = f"{pkg.substation_number:03d}"
        unsorted_dir = pkg.unsorted_raw_data_dir

        if not unsorted_dir.exists():
            warnings.append(f"PE {substation_folder_name}: Unsorted raw data directory missing at {unsorted_dir}")
            return ExtractedPackageData(
                package=pkg,
                testsheet_data=data,
                ir_photos=(),
                dg_photos=(),
                substation_folder_name=substation_folder_name,
            )

        ir_photos, dg_photos = self._scan_photos(unsorted_dir)
        us_tev_archives = tuple(self._scan_us_tev_archives(unsorted_dir))
        return ExtractedPackageData(
            package=pkg,
            testsheet_data=data,
            ir_photos=ir_photos,
            dg_photos=dg_photos,
            substation_folder_name=substation_folder_name,
            us_tev_archives=us_tev_archives,
        )

    def _load_testsheet_data(
        self, pkg: SubstationTestsheetPackage, warnings: list[str]
    ) -> TestsheetData | None:
        data = pkg.data
        if data is None:
            try:
                data = self.extractor.extract_testsheet_data(pkg.testsheet_path)
            except Exception as ex:
                warnings.append(f"Failed to extract testsheet data from {pkg.testsheet_path}: {ex}")
        return data

    def _scan_photos(self, unsorted_dir: Path) -> tuple[tuple[Path, ...], tuple[Path, ...]]:
        ir_source_dir = unsorted_dir / "IR" if (unsorted_dir / "IR").exists() else unsorted_dir
        dg_source_dir = unsorted_dir / "DG" if (unsorted_dir / "DG").exists() else unsorted_dir

        ir_photos = tuple(self._scan_directory_for_images(ir_source_dir))
        dg_photos = tuple(self._scan_directory_for_images(dg_source_dir))
        return ir_photos, dg_photos

    def _scan_us_tev_archives(self, unsorted_dir: Path) -> list[Path]:
        us_tev_source_dir = unsorted_dir / "US+TEV" if (unsorted_dir / "US+TEV").exists() else unsorted_dir
        if not us_tev_source_dir.exists():
            return []
        return [
            p for p in us_tev_source_dir.iterdir()
            if p.is_file() and p.suffix.lower() == ".zip"
        ]

    def _scan_directory_for_images(self, directory: Path) -> list[Path]:
        return [
            p for p in directory.rglob("*")
            if p.is_file() and p.suffix.lower() in [".jpg", ".jpeg", ".png"]
        ]



class RawMaterialFilter:
    """Pure validation and predicate logic stage for Raw Material workflow."""

    def verify_pe_alignment(
        self, packages: list[SubstationTestsheetPackage], existing_tuples: set[tuple[str, str]]
    ) -> None:
        """Verify all target packages exist in TOTAL PE.xlsx."""
        unmatched_packages = [
            pkg for pkg in packages if not self._is_package_aligned(pkg, existing_tuples)
        ]

        if unmatched_packages:
            unmatched_desc = ", ".join(f"PE {p.substation_number} ({p.date_str})" for p in unmatched_packages)
            raise RuntimeError(
                f"TOTAL PE.xlsx alignment pre-check failed: Target records missing for: {unmatched_desc}. "
                "Please run 'Populate TOTAL PE' workflow first."
            )

    def _is_package_aligned(
        self, pkg: SubstationTestsheetPackage, existing_tuples: set[tuple[str, str]]
    ) -> bool:
        keys = self._get_package_alignment_keys(pkg)
        return any(k in existing_tuples for k in keys)

    def _get_package_alignment_keys(self, pkg: SubstationTestsheetPackage) -> list[tuple[str, str]]:
        dates = [d for d in (pkg.date_str, normalize_date_str(pkg.date_str)) if d]
        sub_str = str(pkg.substation_number)
        sub_fmt = f"{pkg.substation_number:03d}"

        keys: list[tuple[str, str]] = []
        for d in dates:
            keys.append((sub_str, d))
            keys.append((sub_fmt, d))
            if pkg.data and pkg.data.substation_name_erms:
                keys.append((pkg.data.substation_name_erms.upper(), d))
        return keys

    def extract_photo_number(self, filename: str, prefix: str) -> int | None:
        """Extract sequence number directly after technology prefix or trailing sequence."""
        name_upper = filename.upper()
        prefix_upper = prefix.upper()

        if prefix_upper in ("P", "P1000") and re.match(r"^P(?:1000|10000)?\d+", name_upper):
            m = re.match(r"^P(?:1000|10000)?0*(\d+)", name_upper)
            if m:
                return int(m.group(1))

        if not name_upper.startswith(prefix_upper):
            return None

        after_prefix = name_upper[len(prefix_upper):]
        digits = re.findall(r"\d+", after_prefix)
        if not digits:
            return None

        if len(digits) > 1 and len(digits[0]) == 8:
            return int(digits[-1])
        return int(digits[0])

    def filter_ir_photos(
        self,
        photos: Sequence[Path],
        camera_config: CameraConfig | None,
        photo_range: PhotoRange | None,
        substation_folder_name: str,
        warnings: list[str],
    ) -> list[tuple[Path, int]]:
        """Filter IR photos matching CameraConfig (single or dual_pair) and PhotoRange bounds."""
        if camera_config is None:
            camera_config = CameraConfig()

        bounds = self._resolve_photo_range_bounds(photo_range, "IR", substation_folder_name, warnings)
        if bounds is None:
            return []

        min_val, max_val = bounds
        step = 1 if min_val <= max_val else -1
        expected_nums = list(range(min_val, max_val + step, step))

        matched_photos: list[tuple[Path, int]] = []
        found_nums: set[int] = set()

        if camera_config.ir_mode == "dual_pair":
            ir_prefix = camera_config.ir_prefix
            dc_prefix = camera_config.dc_prefix
            dc_offset = camera_config.dc_offset
            missing_warnings: list[str] = []

            for num in expected_nums:
                paired_num = num + dc_offset
                ir_matched: list[Path] = []
                dc_matched: list[Path] = []

                for p in photos:
                    p_ir_num = self.extract_photo_number(p.name, ir_prefix)
                    if p_ir_num == num and p not in ir_matched:
                        ir_matched.append(p)

                    p_dc_num = self.extract_photo_number(p.name, dc_prefix)
                    if p_dc_num == paired_num and p not in dc_matched:
                        dc_matched.append(p)

                num_matched = ir_matched + [p for p in dc_matched if p not in ir_matched]

                if num_matched:
                    found_nums.add(num)
                    for p in num_matched:
                        matched_photos.append((p, num))

                if not ir_matched:
                    missing_warnings.append(
                        f"IR photo {ir_prefix}{num:04d} missing for PE {substation_folder_name}"
                    )
                if not dc_matched:
                    missing_warnings.append(
                        f"Visual DC photo {dc_prefix}{paired_num:04d} missing for PE {substation_folder_name}"
                    )

            if len(matched_photos) == 0:
                warnings.append(
                    f"PE {substation_folder_name}: No IR photos matched range [{min_val}-{max_val}] with prefix '{ir_prefix}'"
                )
            else:
                warnings.extend(missing_warnings)
        else:
            ir_prefix = camera_config.ir_prefix
            missing_warnings = []
            for num in expected_nums:
                num_matched: list[Path] = []
                for p in photos:
                    p_num = self.extract_photo_number(p.name, ir_prefix)
                    if p_num == num and p not in num_matched:
                        num_matched.append(p)

                if num_matched:
                    found_nums.add(num)
                    for p in num_matched:
                        matched_photos.append((p, num))
                else:
                    missing_warnings.append(
                        f"IR photo {ir_prefix}{num:04d} missing for PE {substation_folder_name}"
                    )

            if len(matched_photos) == 0:
                warnings.append(
                    f"PE {substation_folder_name}: No IR photos matched range [{min_val}-{max_val}] with prefix '{ir_prefix}'"
                )
            else:
                warnings.extend(missing_warnings)

        return matched_photos

    def filter_dg_photos(
        self,
        photos: Sequence[Path],
        camera_config: CameraConfig | None,
        photo_range: PhotoRange | None,
        substation_folder_name: str,
        warnings: list[str],
    ) -> list[tuple[Path, int]]:
        """Filter DG photos matching CameraConfig and PhotoRange bounds."""
        if camera_config is None:
            camera_config = CameraConfig()

        bounds = self._resolve_photo_range_bounds(photo_range, "DG", substation_folder_name, warnings)
        if bounds is None:
            return []

        min_val, max_val = bounds
        step = 1 if min_val <= max_val else -1
        expected_nums = list(range(min_val, max_val + step, step))

        matched_photos: list[tuple[Path, int]] = []
        found_nums: set[int] = set()
        dg_prefix = camera_config.dg_prefix
        missing_warnings: list[str] = []

        for num in expected_nums:
            num_matched: list[Path] = []
            for p in photos:
                p_num = self.extract_photo_number(p.name, dg_prefix)
                if p_num == num and p not in num_matched:
                    num_matched.append(p)

            if num_matched:
                found_nums.add(num)
                for p in num_matched:
                    matched_photos.append((p, num))
            else:
                missing_warnings.append(
                    f"DG photo {dg_prefix}{num:04d} missing for PE {substation_folder_name}"
                )

        if len(matched_photos) == 0:
            warnings.append(
                f"PE {substation_folder_name}: No DG photos matched range [{min_val}-{max_val}] with prefix '{dg_prefix}'"
            )
        else:
            warnings.extend(missing_warnings)

        return matched_photos

    def filter_matching_photos(
        self,
        photos: Sequence[Path],
        prefix: str,
        photo_range: PhotoRange | None,
        tech_name: str,
        substation_folder_name: str,
        warnings: list[str],
    ) -> list[tuple[Path, int]]:
        """Filter photos matching prefix and PhotoRange bounds (compatibility helper)."""
        if tech_name.upper() == "IR":
            return self.filter_ir_photos(
                photos=photos,
                camera_config=CameraConfig(ir_mode="single", ir_prefix=prefix),
                photo_range=photo_range,
                substation_folder_name=substation_folder_name,
                warnings=warnings,
            )
        return self.filter_dg_photos(
            photos=photos,
            camera_config=CameraConfig(dg_prefix=prefix),
            photo_range=photo_range,
            substation_folder_name=substation_folder_name,
            warnings=warnings,
        )

    def _resolve_photo_range_bounds(
        self,
        photo_range: PhotoRange | None,
        tech_name: str,
        substation_folder_name: str,
        warnings: list[str],
    ) -> tuple[int, int] | None:
        if not photo_range or not photo_range.is_valid:
            warnings.append(f"PE {substation_folder_name}: {tech_name} photo range not specified or incomplete.")
            return None

        start_num = photo_range.start_num if photo_range.start_num is not None else photo_range.end_num
        end_num = photo_range.end_num if photo_range.end_num is not None else photo_range.start_num

        if start_num is None or end_num is None:
            warnings.append(f"PE {substation_folder_name}: {tech_name} photo range bounds missing.")
            return None

        return min(start_num, end_num), max(start_num, end_num)

    def filter_us_tev_archive(
        self,
        archives: Sequence[Path],
        substation_number: int,
        warnings: list[str],
    ) -> Path | None:
        """Filter and match a single US+TEV archive for a substation PE."""
        substation_folder_name = f"{substation_number:03d}"
        num_patterns = [
            re.compile(rf"(?:^|[\W_]){substation_number:03d}(?:[\W_]|$)", re.IGNORECASE),
            re.compile(rf"(?:^|[\W_]){substation_number}(?:[\W_]|$)", re.IGNORECASE),
        ]

        matched: list[Path] = []
        for archive in archives:
            stem = archive.stem
            if stem.lower().endswith(("_archive", ".archive")):
                continue
            if any(pattern.search(stem) for pattern in num_patterns):
                if archive not in matched:
                    matched.append(archive)

        if len(matched) > 1:
            raise RuntimeError(
                f"Multiple US+TEV archives matched PE {substation_folder_name}: {[m.name for m in matched]}"
            )

        if not matched:
            warnings.append(f"PE {substation_folder_name}: No US+TEV archive found in unsorted raw data.")
            return None

        return matched[0]


class RawMaterialTransformer:
    """Pure destination path resolution and CopyInstruction plan construction stage."""

    def build_plan(
        self,
        environment: ProjectEnvironment,
        package: SubstationTestsheetPackage,
        filtered_ir: Sequence[tuple[Path, int]],
        filtered_dg: Sequence[tuple[Path, int]],
        matched_us_tev: Path | None = None,
    ) -> TransformationPlan:
        """Construct transformation plan containing destination paths, copy instructions, and extract instructions."""
        substation_folder_name = f"{package.substation_number:03d}"
        substation_dest_dir = self._resolve_substation_dest_dir(environment, package, substation_folder_name)

        ir_dest_dir = substation_dest_dir / "IR"
        dg_dest_dir = substation_dest_dir / "DG"
        us_tev_dest_dir = substation_dest_dir / "US+TEV"

        directories_to_create = (ir_dest_dir, dg_dest_dir, us_tev_dest_dir)

        copy_instructions = (
            *self._create_copy_instructions(filtered_ir, ir_dest_dir, "IR", substation_folder_name),
            *self._create_copy_instructions(filtered_dg, dg_dest_dir, "DG", substation_folder_name),
        )

        extract_instructions: list[ExtractInstruction] = []
        us_tev_count = 0
        if matched_us_tev is not None:
            extract_dest_dir = us_tev_dest_dir / matched_us_tev.stem
            extract_instructions.append(
                ExtractInstruction(
                    source_archive=matched_us_tev,
                    dest_dir=extract_dest_dir,
                    substation_folder_name=substation_folder_name,
                )
            )
            us_tev_count = 1

        return TransformationPlan(
            directories_to_create=directories_to_create,
            copy_instructions=tuple(copy_instructions),
            extract_instructions=tuple(extract_instructions),
            substation_folder_name=substation_folder_name,
            ir_count=len(filtered_ir),
            dg_count=len(filtered_dg),
            us_tev_count=us_tev_count,
        )


    def _resolve_substation_dest_dir(
        self,
        environment: ProjectEnvironment,
        package: SubstationTestsheetPackage,
        substation_folder_name: str,
    ) -> Path:
        raw_material_root = environment.storage.get_raw_material_dir()
        month_folder = format_month_folder(package.month) or format_month_folder(package.date_str) or "01. JANUARY"
        return (
            raw_material_root
            / (package.station or "UNASSIGNED")
            / month_folder
            / (package.date_str or "01-01-2026")
            / substation_folder_name
            / "RAW DATA"
        )

    def _create_copy_instructions(
        self,
        filtered_photos: Sequence[tuple[Path, int]],
        dest_dir: Path,
        tech_name: str,
        substation_folder_name: str,
    ) -> list[CopyInstruction]:
        return [
            CopyInstruction(
                source_path=p_file,
                dest_path=dest_dir / p_file.name,
                tech_name=tech_name,
                substation_folder_name=substation_folder_name,
                photo_number=num,
            )
            for p_file, num in filtered_photos
        ]


class RawMaterialLoader:
    """Pure disk provisioning, shutil.copy2 execution, and archive extraction stage."""

    def execute_plan(
        self, environment: ProjectEnvironment, plan: TransformationPlan
    ) -> tuple[int, int, int]:
        """Provision directories on disk, copy photos, and extract US+TEV archives."""
        self._provision_directories(environment, plan.directories_to_create)
        self._copy_files(plan.copy_instructions)
        self._extract_archives(plan.extract_instructions)
        return plan.ir_count, plan.dg_count, plan.us_tev_count

    def _provision_directories(
        self, environment: ProjectEnvironment, directories: Sequence[Path]
    ) -> None:
        for directory in directories:
            environment.storage.ensure_directory(directory)

    def _copy_files(self, instructions: Sequence[CopyInstruction]) -> None:
        for instruction in instructions:
            shutil.copy2(instruction.source_path, instruction.dest_path)

    def _extract_archives(self, instructions: Sequence[ExtractInstruction]) -> None:
        for instruction in instructions:
            if instruction.dest_dir.exists():
                shutil.rmtree(instruction.dest_dir)
            instruction.dest_dir.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(instruction.source_archive, "r") as z:
                z.extractall(instruction.dest_dir)



class RawMaterialAuditor:
    """Verification and History Logging stage."""

    def audit(
        self,
        substations_count: int,
        total_ir_copied: int,
        total_dg_copied: int,
        total_us_tev_extracted: int,
        warnings: list[str],
        errors: list[str],
    ) -> tuple[AutomatedRawMaterialSummary, RawMaterialResult]:
        """Format AutomatedRawMaterialSummary telemetry and return the result."""
        summary = AutomatedRawMaterialSummary(
            substations_count=substations_count,
            ir_copied_count=total_ir_copied,
            dg_copied_count=total_dg_copied,
            us_tev_extracted_count=total_us_tev_extracted,
            warnings=tuple(warnings),
            errors=tuple(errors),
        )

        result = RawMaterialResult(
            substations_count=substations_count,
            ir_copied_count=total_ir_copied,
            dg_copied_count=total_dg_copied,
            us_tev_extracted_count=total_us_tev_extracted,
            summary=summary,
            warnings=tuple(warnings),
            errors=tuple(errors),
        )
        return summary, result


class RawMaterialWorkflow:
    """Orchestrates TOTAL PE validation, folder provisioning, photo sorting, and US+TEV extraction.

    Resilience Policy: best-effort
        Collect warnings and errors per package while processing remaining items.
    """

    def __init__(
        self,
        preflight_guard: RawMaterialPreflightGuard | None = None,
        extractor: RawMaterialExtractor | None = None,
        filter_stage: RawMaterialFilter | None = None,
        transformer: RawMaterialTransformer | None = None,
        loader: RawMaterialLoader | None = None,
        auditor: RawMaterialAuditor | None = None,
    ) -> None:
        self.preflight_guard = preflight_guard or RawMaterialPreflightGuard()
        self.extractor = extractor or RawMaterialExtractor()
        self.filter_stage = filter_stage or RawMaterialFilter()
        self.transformer = transformer or RawMaterialTransformer()
        self.loader = loader or RawMaterialLoader()
        self.auditor = auditor or RawMaterialAuditor()

    def execute(
        self, environment: ProjectEnvironment, request: RawMaterialRequest
    ) -> RawMaterialResult:
        """Execute the Raw Material workflow."""
        self.preflight_guard.validate(environment, request)

        total_pe_path = environment.storage.get_total_pe_path()
        packages = self.extractor.discover_packages(environment, request)

        existing_tuples = self.extractor.load_pe_alignment_data(total_pe_path)
        self.filter_stage.verify_pe_alignment(packages, existing_tuples)

        camera_config = environment.get_camera_config()

        substations_count = 0
        total_ir_copied = 0
        total_dg_copied = 0
        total_us_tev_extracted = 0
        warnings: list[str] = []
        errors: list[str] = []

        for pkg in packages:
            substations_count += 1
            ir_copied, dg_copied, us_tev_extracted = self._process_package(
                environment, request, pkg, warnings, camera_config=camera_config
            )
            total_ir_copied += ir_copied
            total_dg_copied += dg_copied
            total_us_tev_extracted += us_tev_extracted

        _, result = self.auditor.audit(
            substations_count,
            total_ir_copied,
            total_dg_copied,
            total_us_tev_extracted,
            warnings,
            errors,
        )
        return result

    def _process_package(
        self,
        environment: ProjectEnvironment,
        request: RawMaterialRequest,
        pkg: SubstationTestsheetPackage,
        warnings: list[str],
        camera_config: CameraConfig | None = None,
    ) -> tuple[int, int, int]:
        if camera_config is None:
            camera_config = environment.get_camera_config()

        extracted = self.extractor.extract_package_data(pkg, warnings)

        photo_ranges = extracted.testsheet_data.photo_ranges if extracted.testsheet_data else None
        ir_range = photo_ranges.ir if photo_ranges else None
        dg_range = photo_ranges.dg if photo_ranges else None

        filtered_ir = self.filter_stage.filter_ir_photos(
            photos=extracted.ir_photos,
            camera_config=camera_config,
            photo_range=ir_range,
            substation_folder_name=extracted.substation_folder_name,
            warnings=warnings,
        )

        filtered_dg = self.filter_stage.filter_dg_photos(
            photos=extracted.dg_photos,
            camera_config=camera_config,
            photo_range=dg_range,
            substation_folder_name=extracted.substation_folder_name,
            warnings=warnings,
        )

        matched_us_tev = self.filter_stage.filter_us_tev_archive(
            archives=extracted.us_tev_archives,
            substation_number=pkg.substation_number,
            warnings=warnings,
        )

        plan = self.transformer.build_plan(
            environment=environment,
            package=pkg,
            filtered_ir=filtered_ir,
            filtered_dg=filtered_dg,
            matched_us_tev=matched_us_tev,
        )

        ir_copied, dg_copied, us_tev_extracted = self.loader.execute_plan(environment, plan)

        if request.progress_sink:
            request.progress_sink(
                f"Processed PE {extracted.substation_folder_name}: "
                f"{ir_copied} IR photos, {dg_copied} DG photos copied, "
                f"{us_tev_extracted} US+TEV survey extracted."
            )

        return ir_copied, dg_copied, us_tev_extracted

