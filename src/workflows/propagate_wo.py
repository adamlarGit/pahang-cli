"""Propagate Work Orders workflow implementation for Pahang CLI."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence
import openpyxl

from src.core.normalizers import normalize_date_str
from src.master.total_pe import LocalExcelTotalPeRepository, TotalPeRepository
from src.project.environment import ProjectEnvironment
from src.workflows.models import (
    PropagateWoRequest,
    PropagateWoResult,
)


@dataclass(frozen=True)
class PropagateWoPeRow:
    """Representation of a row in TOTAL PE DataCycle1 sheet."""

    row_index: int
    pe_no: Any
    fl_num: str | None
    substation_name: str | None
    date_str: str | None
    type_c: str | None
    wo: str | None


@dataclass(frozen=True)
class PropagateWoRowUpdate:
    """Instruction to update a single row in Column F."""

    row_index: int
    wo: str
    fl: str


@dataclass(frozen=True)
class PropagateWoRawData:
    """Raw extracted data from DATA MSMS.xlsx and TOTAL PE.xlsx."""

    fl_to_wo: dict[str, str]
    pe_rows: tuple[PropagateWoPeRow, ...]
    total_pe_path: Path
    data_msms_path: Path


@dataclass(frozen=True)
class PropagateWoFilteredResults:
    """Intermediate results of filtering and matching."""

    updates: tuple[PropagateWoRowUpdate, ...]
    matched_count: int
    already_populated_count: int
    unmatched_count: int
    unmatched_fls: tuple[str, ...]


@dataclass(frozen=True)
class PropagateWoPlan:
    """Execution plan for propagating work orders into TOTAL PE."""

    total_pe_path: Path
    updates: tuple[PropagateWoRowUpdate, ...]
    matched_count: int
    already_populated_count: int
    unmatched_count: int
    unmatched_fls: tuple[str, ...]
    target_date: str | None = None


class PropagateWoPreflightGuard:
    """Pre-flight resource guard stage for Propagate Work Orders workflow."""

    def validate(
        self, environment: ProjectEnvironment, request: PropagateWoRequest | None = None
    ) -> None:
        """Validate environmental preconditions before reading data."""
        data_msms_path = environment.storage.get_data_msms_path()
        self._validate_data_msms(data_msms_path)

        total_pe_path = environment.storage.get_total_pe_path()
        self._validate_total_pe(total_pe_path)

    def _validate_data_msms(self, data_msms_path: Path) -> None:
        if not data_msms_path.exists():
            raise FileNotFoundError(f"DATA MSMS.xlsx workbook not found: {data_msms_path}")

        try:
            wb = openpyxl.load_workbook(data_msms_path, read_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                raise RuntimeError("DATA MSMS.xlsx contains no active sheet")

            # Check if there are data rows (header + at least 1 record)
            has_records = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    has_records = True
                    break

            wb.close()
            if not has_records:
                raise RuntimeError("DATA MSMS.xlsx contains no records")
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise RuntimeError(f"Failed to load DATA MSMS.xlsx: {e}")

    def _validate_total_pe(self, total_pe_path: Path) -> None:
        if not total_pe_path.exists():
            raise FileNotFoundError(f"TOTAL PE.xlsx workbook not found: {total_pe_path}")

        try:
            wb = openpyxl.load_workbook(total_pe_path, read_only=True)
            if "DataCycle1" not in wb.sheetnames:
                wb.close()
                raise RuntimeError("DataCycle1 sheet not found in TOTAL PE.xlsx")
            wb.close()
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise RuntimeError(f"Failed to load TOTAL PE.xlsx: {e}")


class PropagateWoExtractor:
    """Pure I/O reading stage for Propagate Work Orders workflow."""

    def extract(self, total_pe_path: Path, data_msms_path: Path) -> PropagateWoRawData:
        """Extract fl_to_wo mapping from DATA MSMS and rows from TOTAL PE DataCycle1."""
        fl_to_wo = self._extract_fl_to_wo(data_msms_path)
        pe_rows = self._extract_pe_rows(total_pe_path)

        return PropagateWoRawData(
            fl_to_wo=fl_to_wo,
            pe_rows=tuple(pe_rows),
            total_pe_path=total_pe_path,
            data_msms_path=data_msms_path,
        )

    def _extract_fl_to_wo(self, data_msms_path: Path) -> dict[str, str]:
        wb_msms = openpyxl.load_workbook(data_msms_path, data_only=True)
        ws_msms = wb_msms.active
        if ws_msms is None:
            wb_msms.close()
            return {}

        fl_to_wo: dict[str, str] = {}
        for r_idx in range(2, (ws_msms.max_row or 1) + 1):
            wo_val = ws_msms.cell(r_idx, 1).value  # Col A: Work Order
            fl_val = ws_msms.cell(r_idx, 5).value  # Col E: FL ERMS
            loc_val = ws_msms.cell(r_idx, 2).value  # Col B: Location

            wo_str = str(wo_val).strip() if wo_val is not None else ""
            if not wo_str or wo_str.lower() in ("none", "nan"):
                continue

            for fl_candidate in (fl_val, loc_val):
                if fl_candidate is not None:
                    fl_str = str(fl_candidate).strip().upper()
                    if fl_str and fl_str.lower() not in ("none", "nan"):
                        fl_to_wo[fl_str] = wo_str
                        fl_norm = fl_str.replace("/", "")
                        if fl_norm:
                            fl_to_wo[fl_norm] = wo_str

        wb_msms.close()
        return fl_to_wo

    def _extract_pe_rows(self, total_pe_path: Path) -> list[PropagateWoPeRow]:
        wb_pe = openpyxl.load_workbook(total_pe_path, data_only=False)
        if "DataCycle1" not in wb_pe.sheetnames:
            wb_pe.close()
            raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")

        ws_pe = wb_pe["DataCycle1"]
        pe_rows: list[PropagateWoPeRow] = []

        for r_idx in range(2, (ws_pe.max_row or 1) + 1):
            pe_no = ws_pe.cell(r_idx, 1).value
            fl_val = ws_pe.cell(r_idx, 2).value
            sub_name = ws_pe.cell(r_idx, 3).value
            date_val = ws_pe.cell(r_idx, 4).value
            type_val = ws_pe.cell(r_idx, 5).value
            wo_val = ws_pe.cell(r_idx, 6).value

            fl_num = str(fl_val).strip() if fl_val is not None else None
            substation_name = str(sub_name).strip() if sub_name is not None else None
            date_str = str(date_val).strip() if date_val is not None else None
            type_c = str(type_val).strip() if type_val is not None else None
            wo = str(wo_val).strip() if wo_val is not None else None

            pe_rows.append(
                PropagateWoPeRow(
                    row_index=r_idx,
                    pe_no=pe_no,
                    fl_num=fl_num,
                    substation_name=substation_name,
                    date_str=date_str,
                    type_c=type_c,
                    wo=wo,
                )
            )

        wb_pe.close()
        return pe_rows


class PropagateWoFilter:
    """Pure predicate logic stage for Propagate Work Orders workflow."""

    def filter(
        self,
        raw_data: PropagateWoRawData,
        target_date: str | None = None,
        overwrite: bool = False,
    ) -> PropagateWoFilteredResults:
        """Filter rows by target_date, identify empty WO cells, match FL to WO."""
        matched_count = 0
        already_populated_count = 0
        unmatched_count = 0
        unmatched_fls: list[str] = []
        updates: list[PropagateWoRowUpdate] = []

        norm_target_date = normalize_date_str(target_date) if target_date else None

        for row in raw_data.pe_rows:
            if target_date is not None:
                if row.date_str is None:
                    continue
                row_date_str = row.date_str
                norm_row_date = normalize_date_str(row_date_str)
                if norm_row_date != norm_target_date and row_date_str != target_date:
                    continue

            if not row.fl_num:
                continue

            fl_str = row.fl_num.upper()
            if not fl_str or fl_str.lower() in ("none", "nan"):
                continue

            # Check if already populated
            if row.wo is not None and not overwrite:
                cur_wo = row.wo.strip()
                if cur_wo and cur_wo.lower() not in ("none", "nan"):
                    already_populated_count += 1
                    continue

            # Lookup in mapping
            wo_match = raw_data.fl_to_wo.get(fl_str) or raw_data.fl_to_wo.get(
                fl_str.replace("/", "")
            )
            if wo_match:
                updates.append(
                    PropagateWoRowUpdate(
                        row_index=row.row_index,
                        wo=wo_match,
                        fl=fl_str,
                    )
                )
                matched_count += 1
            else:
                unmatched_count += 1
                unmatched_fls.append(fl_str)

        return PropagateWoFilteredResults(
            updates=tuple(updates),
            matched_count=matched_count,
            already_populated_count=already_populated_count,
            unmatched_count=unmatched_count,
            unmatched_fls=tuple(unmatched_fls),
        )


class PropagateWoTransformer:
    """Pure transformation plan construction stage for Propagate Work Orders workflow."""

    def build_plan(
        self,
        total_pe_path: Path,
        updates: Sequence[PropagateWoRowUpdate],
        matched_count: int,
        already_populated_count: int,
        unmatched_count: int,
        unmatched_fls: Sequence[str],
        target_date: str | None = None,
    ) -> PropagateWoPlan:
        """Construct transformation execution plan."""
        return PropagateWoPlan(
            total_pe_path=total_pe_path,
            updates=tuple(updates),
            matched_count=matched_count,
            already_populated_count=already_populated_count,
            unmatched_count=unmatched_count,
            unmatched_fls=tuple(unmatched_fls),
            target_date=target_date,
        )


class PropagateWoLoader:
    """Pure write I/O stage for Propagate Work Orders workflow."""

    def __init__(self, total_pe_repository: TotalPeRepository | None = None) -> None:
        self.total_pe_repository = total_pe_repository or LocalExcelTotalPeRepository()

    def load(self, plan: PropagateWoPlan) -> int:
        """Write Column F in TOTAL PE DataCycle1 sheet using openpyxl, preserving formulas."""
        if not plan.updates:
            return 0

        wb = openpyxl.load_workbook(plan.total_pe_path, data_only=False)
        if "DataCycle1" not in wb.sheetnames:
            wb.close()
            raise RuntimeError(f"'DataCycle1' sheet missing in {plan.total_pe_path}")

        ws = wb["DataCycle1"]
        for upd in plan.updates:
            ws.cell(row=upd.row_index, column=6, value=upd.wo)

        wb.save(plan.total_pe_path)
        wb.close()
        return len(plan.updates)


class PropagateWoAuditor:
    """Verification & History Audit stage for Propagate Work Orders workflow."""

    def audit(
        self,
        plan: PropagateWoPlan,
        updated_count: int,
    ) -> PropagateWoResult:
        """Verify output integrity and return workflow telemetry result."""
        self._verify_output(plan.total_pe_path)

        return PropagateWoResult(
            matched_count=plan.matched_count,
            already_populated_count=plan.already_populated_count,
            unmatched_count=plan.unmatched_count,
            unmatched_fls=plan.unmatched_fls,
            updated_count=updated_count,
        )

    def _verify_output(self, total_pe_path: Path) -> None:
        if not total_pe_path.exists():
            raise RuntimeError(f"TOTAL PE.xlsx does not exist after load at {total_pe_path}")
        if total_pe_path.stat().st_size == 0:
            raise RuntimeError(f"TOTAL PE.xlsx is empty (0 bytes) after load at {total_pe_path}")


class PropagateWoWorkflow:
    """Orchestrates WO number propagation from DATA MSMS.xlsx to TOTAL PE.xlsx.

    Resilience Policy: best-effort
        Per-row matching and population. Skips rows that are missing or already populated.
    """

    def __init__(
        self,
        preflight_guard: PropagateWoPreflightGuard | None = None,
        extractor: PropagateWoExtractor | None = None,
        filter_stage: PropagateWoFilter | None = None,
        transformer: PropagateWoTransformer | None = None,
        loader: PropagateWoLoader | None = None,
        auditor: PropagateWoAuditor | None = None,
    ) -> None:
        self.preflight_guard = preflight_guard or PropagateWoPreflightGuard()
        self.extractor = extractor or PropagateWoExtractor()
        self.filter_stage = filter_stage or PropagateWoFilter()
        self.transformer = transformer or PropagateWoTransformer()
        self.loader = loader or PropagateWoLoader()
        self.auditor = auditor or PropagateWoAuditor()

    def execute(
        self,
        environment: ProjectEnvironment,
        request: PropagateWoRequest | None = None,
    ) -> PropagateWoResult:
        """Execute Propagate Work Orders workflow."""
        req = request or PropagateWoRequest()

        if req.progress_sink:
            req.progress_sink("Validating preflight conditions for Propagate WO...")

        self.preflight_guard.validate(environment, req)

        total_pe_path = environment.storage.get_total_pe_path()
        data_msms_path = environment.storage.get_data_msms_path()

        if req.progress_sink:
            req.progress_sink("Extracting work orders and TOTAL PE records...")

        raw_data = self.extractor.extract(total_pe_path, data_msms_path)

        if req.progress_sink:
            req.progress_sink("Filtering and matching work orders by functional location...")

        filtered = self.filter_stage.filter(
            raw_data=raw_data,
            target_date=req.target_date,
            overwrite=req.overwrite,
        )

        plan = self.transformer.build_plan(
            total_pe_path=total_pe_path,
            updates=filtered.updates,
            matched_count=filtered.matched_count,
            already_populated_count=filtered.already_populated_count,
            unmatched_count=filtered.unmatched_count,
            unmatched_fls=filtered.unmatched_fls,
            target_date=req.target_date,
        )

        if req.progress_sink:
            req.progress_sink(
                f"Writing {len(plan.updates)} work orders to TOTAL PE Column F..."
            )

        updated_count = self.loader.load(plan)

        if req.progress_sink:
            req.progress_sink("Auditing TOTAL PE workbook integrity...")

        result = self.auditor.audit(plan, updated_count)

        if req.progress_sink:
            req.progress_sink(
                f"Propagate WO completed: {result.matched_count} matched, "
                f"{result.already_populated_count} already populated, "
                f"{result.unmatched_count} unmatched."
            )

        return result
