"""Enrich MSMS Workflow (TOTAL PE.xlsx -> DATA MSMS.xlsx).

Enriches blank metadata cells in DATA MSMS.xlsx (columns D-G) with substation metadata
from TOTAL PE.xlsx (DataCycle1 sheet) for human verification.

Resilience Policy: best-effort per Work Order.
"""

from __future__ import annotations

from dataclasses import dataclass
import logging
from pathlib import Path
from typing import Any, Sequence

import openpyxl

from src.core.normalizers import normalize_fl_erms
from src.project.environment import ProjectEnvironment
from src.workflows.models import (
    EnrichMsmsRequest,
    EnrichMsmsResult,
)

logger = logging.getLogger(__name__)


class EnrichMsmsPreflightGuard:
    """Pre-flight resource guard stage for Enrich MSMS workflow."""

    def validate(self, environment: ProjectEnvironment) -> None:
        """Validate environmental preconditions before reading data."""
        data_msms_path = environment.storage.get_data_msms_path()
        if not data_msms_path.exists():
            raise FileNotFoundError(f"DATA MSMS workbook not found: {data_msms_path}")

        total_pe_path = environment.storage.get_total_pe_path()
        if not total_pe_path.exists():
            raise FileNotFoundError(f"TOTAL PE workbook not found: {total_pe_path}")

        try:
            wb = openpyxl.load_workbook(total_pe_path, read_only=True)
            if "DataCycle1" not in wb.sheetnames:
                wb.close()
                raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")
            wb.close()
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise RuntimeError(f"Failed to load TOTAL PE workbook: {e}")


@dataclass(frozen=True)
class TotalPeLookups:
    """Lookup index maps derived from TOTAL PE DataCycle1 sheet."""

    by_wo: dict[str, dict[str, Any]]
    by_fl: dict[str, dict[str, Any]]


@dataclass(frozen=True)
class EnrichCellUpdate:
    """Instruction for in-place cell updates on a specific DATA MSMS row."""

    row_index: int  # 1-based Excel row number
    substation_name_erms: str | None = None
    fl_erms: str | None = None
    cycle_date: str | None = None
    substation_number: Any = None


@dataclass(frozen=True)
class EnrichMsmsPlan:
    """Transformation execution plan for Enrich MSMS workflow."""

    data_msms_path: Path
    updates: tuple[EnrichCellUpdate, ...]
    matched_count: int
    unmatched_count: int
    unmatched_wos: tuple[str, ...]
    updated_cells_count: int


class EnrichMsmsExtractor:
    """Pure I/O reading stage for Enrich MSMS workflow."""

    def read_data_msms_rows(self, data_msms_path: Path) -> list[dict[str, Any]]:
        """Read existing rows from DATA MSMS.xlsx with 1-based row index."""
        wb = openpyxl.load_workbook(data_msms_path, read_only=True)
        ws = wb.active

        rows: list[dict[str, Any]] = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
            wo_val = row[0] if len(row) > 0 else None
            loc_val = row[1] if len(row) > 1 else None
            desc_val = row[2] if len(row) > 2 else None
            sub_name_erms = row[3] if len(row) > 3 else None
            fl_erms = row[4] if len(row) > 4 else None
            cycle_date = row[5] if len(row) > 5 else None
            sub_num = row[6] if len(row) > 6 else None

            wo_str = str(wo_val).strip() if wo_val is not None else ""
            if not wo_str and not loc_val and not desc_val:
                continue

            rows.append({
                "row_idx": r_idx,
                "wo": wo_str,
                "location": str(loc_val).strip() if loc_val is not None else "",
                "description": str(desc_val).strip() if desc_val is not None else "",
                "substation_name_erms": sub_name_erms,
                "fl_erms": fl_erms,
                "cycle_date": cycle_date,
                "substation_number": sub_num,
            })

        wb.close()
        return rows

    def read_total_pe_lookups(self, total_pe_path: Path) -> TotalPeLookups:
        """Build WO and FL lookup dictionaries from TOTAL PE DataCycle1 sheet."""
        wb = openpyxl.load_workbook(total_pe_path, data_only=True, read_only=True)
        ws = wb["DataCycle1"]

        by_wo: dict[str, dict[str, Any]] = {}
        by_fl: dict[str, dict[str, Any]] = {}

        for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
            pe_no_val = row[0] if len(row) > 0 else None
            fl_val = row[1] if len(row) > 1 else None
            sub_name_val = row[2] if len(row) > 2 else None
            date_val = row[3] if len(row) > 3 else None
            wo_val = row[5] if len(row) > 5 else None

            pe_data = {
                "substation_number": pe_no_val,
                "fl_erms": str(fl_val).strip() if fl_val is not None else None,
                "substation_name_erms": str(sub_name_val).strip() if sub_name_val is not None else None,
                "cycle_date": str(date_val).strip() if date_val is not None else None,
            }

            if wo_val is not None:
                wo_str = str(wo_val).strip()
                if wo_str and wo_str.lower() not in ("none", "nan"):
                    by_wo[wo_str] = pe_data

            if fl_val is not None:
                fl_str = str(fl_val).strip().upper()
                if fl_str and fl_str.lower() not in ("none", "nan"):
                    by_fl[fl_str] = pe_data
                    fl_norm = fl_str.replace("/", "")
                    if fl_norm:
                        by_fl[fl_norm] = pe_data
                    fl_normalized = normalize_fl_erms(fl_str).upper()
                    if fl_normalized:
                        by_fl[fl_normalized] = pe_data
                        fl_norm_noslash = fl_normalized.replace("/", "")
                        if fl_norm_noslash:
                            by_fl[fl_norm_noslash] = pe_data

        wb.close()
        return TotalPeLookups(by_wo=by_wo, by_fl=by_fl)


class EnrichMsmsFilter:
    """Pure predicate logic stage for Enrich MSMS workflow."""

    def filter_rows(
        self,
        msms_rows: Sequence[dict[str, Any]],
        total_pe_by_wo: dict[str, dict[str, Any]],
        total_pe_by_fl: dict[str, dict[str, Any]],
    ) -> tuple[list[EnrichCellUpdate], int, int, tuple[str, ...], int]:
        """Match MSMS rows against TOTAL PE and identify cells to update.

        Returns tuple of (cell_updates, matched_count, unmatched_count, unmatched_wos, updated_cells_count).
        """
        matched_count = 0
        unmatched_count = 0
        unmatched_wos: list[str] = []
        updated_cells_count = 0
        cell_updates: list[EnrichCellUpdate] = []

        def _is_blank(val: Any) -> bool:
            if val is None:
                return True
            s = str(val).strip()
            return s == "" or s.lower() in ("none", "nan")

        for row in msms_rows:
            wo_str = str(row.get("wo", "")).strip()
            fl_val = row.get("fl_erms")
            fl_str = str(fl_val).strip().upper() if fl_val is not None else ""
            loc_val = row.get("location")
            loc_str = str(loc_val).strip().upper() if loc_val is not None else ""

            if not wo_str and not fl_str and not loc_str:
                continue

            pe_data = None
            if wo_str and wo_str.lower() not in ("none", "nan") and wo_str in total_pe_by_wo:
                pe_data = total_pe_by_wo[wo_str]
            else:
                candidates: list[str] = []
                if fl_str and fl_str.lower() not in ("none", "nan"):
                    if fl_str not in candidates:
                        candidates.append(fl_str)
                    fl_noslash = fl_str.replace("/", "")
                    if fl_noslash and fl_noslash not in candidates:
                        candidates.append(fl_noslash)
                if loc_str and loc_str.lower() not in ("none", "nan", "location"):
                    if loc_str not in candidates:
                        candidates.append(loc_str)
                    loc_noslash = loc_str.replace("/", "")
                    if loc_noslash and loc_noslash not in candidates:
                        candidates.append(loc_noslash)
                    norm_loc = normalize_fl_erms(loc_str).upper()
                    if norm_loc and norm_loc.lower() not in ("none", "nan", "location"):
                        if norm_loc not in candidates:
                            candidates.append(norm_loc)
                        norm_loc_noslash = norm_loc.replace("/", "")
                        if norm_loc_noslash and norm_loc_noslash not in candidates:
                            candidates.append(norm_loc_noslash)

                for cand in candidates:
                    if cand in total_pe_by_fl:
                        pe_data = total_pe_by_fl[cand]
                        break


            if pe_data is not None:
                matched_count += 1
                row_idx = row["row_idx"]

                sub_name_update = None
                fl_update = None
                date_update = None
                sub_num_update = None

                if _is_blank(row.get("substation_name_erms")) and pe_data.get("substation_name_erms"):
                    sub_name_update = pe_data["substation_name_erms"]
                    updated_cells_count += 1

                if _is_blank(row.get("fl_erms")) and pe_data.get("fl_erms"):
                    fl_update = pe_data["fl_erms"]
                    updated_cells_count += 1

                if _is_blank(row.get("cycle_date")) and pe_data.get("cycle_date"):
                    date_update = pe_data["cycle_date"]
                    updated_cells_count += 1

                if _is_blank(row.get("substation_number")) and pe_data.get("substation_number") is not None:
                    sub_num_update = pe_data["substation_number"]
                    updated_cells_count += 1

                if any(x is not None for x in (sub_name_update, fl_update, date_update, sub_num_update)):
                    cell_updates.append(
                        EnrichCellUpdate(
                            row_index=row_idx,
                            substation_name_erms=sub_name_update,
                            fl_erms=fl_update,
                            cycle_date=date_update,
                            substation_number=sub_num_update,
                        )
                    )
            else:
                if wo_str and wo_str.lower() not in ("none", "nan"):
                    unmatched_count += 1
                    unmatched_wos.append(wo_str)

        return cell_updates, matched_count, unmatched_count, tuple(unmatched_wos), updated_cells_count


class EnrichMsmsTransformer:
    """Pure transformation plan construction stage for Enrich MSMS workflow."""

    def build_plan(
        self,
        data_msms_path: Path,
        updates: Sequence[EnrichCellUpdate],
        matched_count: int,
        unmatched_count: int,
        unmatched_wos: Sequence[str],
        updated_cells_count: int,
    ) -> EnrichMsmsPlan:
        """Construct immutable EnrichMsmsPlan."""
        return EnrichMsmsPlan(
            data_msms_path=data_msms_path,
            updates=tuple(updates),
            matched_count=matched_count,
            unmatched_count=unmatched_count,
            unmatched_wos=tuple(unmatched_wos),
            updated_cells_count=updated_cells_count,
        )


class EnrichMsmsLoader:
    """Pure write I/O stage for Enrich MSMS workflow."""

    def load(self, plan: EnrichMsmsPlan) -> None:
        """Apply cell updates in-place to DATA MSMS.xlsx using openpyxl."""
        if not plan.updates:
            return

        wb = openpyxl.load_workbook(plan.data_msms_path)
        ws = wb.active

        for update in plan.updates:
            r = update.row_index
            if update.substation_name_erms is not None:
                ws.cell(row=r, column=4, value=update.substation_name_erms)
            if update.fl_erms is not None:
                ws.cell(row=r, column=5, value=update.fl_erms)
            if update.cycle_date is not None:
                ws.cell(row=r, column=6, value=update.cycle_date)
            if update.substation_number is not None:
                ws.cell(row=r, column=7, value=update.substation_number)

        wb.save(plan.data_msms_path)
        wb.close()


class EnrichMsmsAuditor:
    """Verification & Audit stage for Enrich MSMS workflow."""

    def audit(self, plan: EnrichMsmsPlan) -> EnrichMsmsResult:
        """Verify output integrity and construct telemetry result."""
        self._verify_output(plan.data_msms_path)
        return EnrichMsmsResult(
            matched_count=plan.matched_count,
            unmatched_count=plan.unmatched_count,
            unmatched_wos=plan.unmatched_wos,
            updated_cells_count=plan.updated_cells_count,
        )

    def _verify_output(self, data_msms_path: Path) -> None:
        if not data_msms_path.exists():
            raise RuntimeError(f"DATA MSMS workbook does not exist: {data_msms_path}")
        if data_msms_path.stat().st_size == 0:
            raise RuntimeError(f"DATA MSMS workbook is empty (0 bytes): {data_msms_path}")


class EnrichMsmsWorkflow:
    """6-stage ETL workflow for enriching DATA MSMS.xlsx with TOTAL PE.xlsx metadata.

    Resilience Policy: best-effort per Work Order.
    """

    def __init__(
        self,
        preflight_guard: EnrichMsmsPreflightGuard | None = None,
        extractor: EnrichMsmsExtractor | None = None,
        filter_stage: EnrichMsmsFilter | None = None,
        transformer: EnrichMsmsTransformer | None = None,
        loader: EnrichMsmsLoader | None = None,
        auditor: EnrichMsmsAuditor | None = None,
    ) -> None:
        self.preflight_guard = preflight_guard or EnrichMsmsPreflightGuard()
        self.extractor = extractor or EnrichMsmsExtractor()
        self.filter_stage = filter_stage or EnrichMsmsFilter()
        self.transformer = transformer or EnrichMsmsTransformer()
        self.loader = loader or EnrichMsmsLoader()
        self.auditor = auditor or EnrichMsmsAuditor()

    def execute(
        self,
        environment: ProjectEnvironment,
        request: EnrichMsmsRequest | None = None,
    ) -> EnrichMsmsResult:
        """Execute Enrich MSMS workflow."""
        self.preflight_guard.validate(environment)

        req = request or EnrichMsmsRequest()
        if req.progress_sink:
            req.progress_sink("Starting Enrich MSMS workflow...")

        data_msms_path = environment.storage.get_data_msms_path()
        total_pe_path = environment.storage.get_total_pe_path()

        if req.progress_sink:
            req.progress_sink(f"Reading {data_msms_path.name} and {total_pe_path.name}...")

        msms_rows = self.extractor.read_data_msms_rows(data_msms_path)
        pe_lookups = self.extractor.read_total_pe_lookups(total_pe_path)

        updates, matched, unmatched, unmatched_wos, updated_cells = self.filter_stage.filter_rows(
            msms_rows=msms_rows,
            total_pe_by_wo=pe_lookups.by_wo,
            total_pe_by_fl=pe_lookups.by_fl,
        )

        plan = self.transformer.build_plan(
            data_msms_path=data_msms_path,
            updates=updates,
            matched_count=matched,
            unmatched_count=unmatched,
            unmatched_wos=unmatched_wos,
            updated_cells_count=updated_cells,
        )

        if req.progress_sink:
            req.progress_sink(f"Applying {len(plan.updates)} row updates ({plan.updated_cells_count} cells)...")

        self.loader.load(plan)

        result = self.auditor.audit(plan)

        if req.progress_sink:
            req.progress_sink(
                f"Enrich MSMS complete: {result.matched_count} matched, {result.unmatched_count} unmatched, "
                f"{result.updated_cells_count} cells updated."
            )

        return result
