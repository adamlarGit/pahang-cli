"""Populate Data MSMS workflow implementation for Pahang CLI.

Extracts measurements from testsheets via TestsheetReadingMapper and defect
records from QR03 VI, matching by WONUM + METERNAME to populate CSV reading
columns in place inside MSMS/TO BE FILLED/.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, time
import logging
from pathlib import Path
import re
from typing import Any, Sequence
import openpyxl

from src.core.normalizers import (
    _parse_date_object,
    extract_background_temperature,
    normalize_date_str,
    normalize_for_csv,
)
from src.project.environment import ProjectEnvironment
from src.quick_report.defects import MasterQr03DefectRepository, ViDefectRecord
from src.testsheet.feeder_thermal import (
    FEEDER_CHANNEL_COLUMNS,
    extract_board_average_temperature,
    is_active_feeder_cable,
    parse_feeder_meter,
    synthesize_feeder_thermal_readings,
)
from src.testsheet.mapper import TestsheetReadingMapper, parse_equipment_index
from src.testsheet.models import SubstationTestsheetPackage
from src.testsheet.repository import SubstationTestsheetRepository
from src.workflows.models import (
    PopulateDataMsmsRequest,
    PopulateDataMsmsResult,
    PopulateMode,
)

logger = logging.getLogger(__name__)

# Aligned Defect Matching Sets for VI11_* inspection meters

# Group A: Switchgear
SG_LABELLING_DEFECT_AREAS = {
    "NO LINK NO./PANEL NO./FEEDER NAME",
    "WRONG LINK NO./PANEL NO./FEEDER NAME",
    "LINK NO./PANEL NO./FEEDER NAME LABEL IN POOR CONDITION",
}
SG_PRESGAUGE_DEFECT_AREAS = {
    "LOW SF6 GAS",
    "SF6 GAS INDICATOR BROKEN",
}
SG_VDIS_DEFECT_AREAS = {
    "VCB STATUS LAMP INDICATOR NOT OPERATED",
    "BREAKER INDICATOR LAMP NOT OPERATED",
    "RELAY NOT OPERATED",
}
SG_HEATER_DEFECT_AREAS = {
    "HEATER NO SUPPLY",
    "HEATER INDICATOR LAMP NOT OPERATED",
}
SG_EARTHIN_REMARKS = {"SWG", "SWG1", "SWG ROOM", "SWG - NOT CONNECTED"}

# Group B: Feeder Pillar
FP_LINK_FUSE_DEFECT_AREAS = {
    "FP (J) FUSE HOLDER MISSING",
    "FP (J) FUSE HOLDER BROKEN",
    "FP (J) LINK HOLDER MISSING",
    "FP (J) LINK HOLDER BROKEN",
}
FP_PLOCK_FP_DEFECT_AREAS = {"DOOR BROKEN", "FP DOOR BROKEN"}
FP_PLOCK_SUB_DEFECT_AREAS = {"OLD ABLOY PADLOCK", "NO PADLOCK"}

# Group C: Transformer
TX_TXGUARD_DEFECT_AREAS = {
    "NO TX GUARD",
    "CABLE TERMINATION TOUCH TX GUARD",
}
TX_TXBUSH_DEFECT_AREAS = {
    "NO LV INSULATION BOOT COVER",
    "NO HV INSULATION BOOT COVER",
    "CABLE TERMINATION INSULATION BROKEN",
}
TX_OILLEVEL_DEFECT_AREAS = {
    "LOW OIL LEVEL",
    "OIL INDICATOR BROKEN",
}
TX_OILLEAK_DEFECT_AREAS = {
    "OIL LEAKS",
    "OIL LEAK",
}
TX_CBLCLMP_DEFECT_AREAS = {
    "NO CABLE SUPPORT",
    "CABLE SUPPORT IN POOR CONDITION",
}

# Group E: Substation / Civil
SUB_RETROOF_DEFECT_AREAS = {"ROOF BROKEN", "CEILING BROKEN"}
SUB_CLEANLINESS_DEFECT_AREAS = {
    "BUSHES & CREEPERS",
    "UNUSED ITEM LEFTOVER",
    "SARANG BINATANG",
    "CARCASS (BANGKAI BUSUK)",
    "VERMIN",
}
SUB_VANDALISM_DEFECT_AREAS = {
    "FENCE BROKEN",
    "GATE BROKEN",
    "DOOR BROKEN",
    "WALL BROKEN",
    "WINDOW BROKEN",
    "VANDALISME",
    "OLD ABLOY PADLOCK",
    "NO PADLOCK",
}

# Backward compatibility alias for deprecated keyword dictionary
VI_METER_KEYWORDS: dict[str, list[str]] = {}


def parse_time_tuple(time_val: Any) -> tuple[int, int, int] | None:
    """Parse time value into (hour, minute, second) tuple."""
    if time_val is None:
        return None
    if isinstance(time_val, time):
        return time_val.hour, time_val.minute, time_val.second
    if isinstance(time_val, datetime):
        return time_val.hour, time_val.minute, time_val.second
    if isinstance(time_val, (int, float)):
        int_val = int(time_val)
        s = f"{int_val:04d}"
        if len(s) == 4:
            hh, mm = int(s[:2]), int(s[2:])
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return hh, mm, 0
    s_clean = str(time_val).strip()
    if not s_clean or s_clean in ("-", "None", "N/A"):
        return None
    m = re.match(r"^(\d{1,2})[:.](\d{2})(?:[:.](\d{2}))?", s_clean)
    if m:
        hh, mm = int(m.group(1)), int(m.group(2))
        ss = int(m.group(3)) if m.group(3) else 0
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return hh, mm, ss
    if s_clean.isdigit():
        padded = s_clean.zfill(4)
        if len(padded) == 4:
            hh, mm = int(padded[:2]), int(padded[2:])
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return hh, mm, 0
    return None


def parse_testsheet_datetime(
    date_val: Any,
    time_val: Any = None,
    tz_offset: str = "+08:00",
) -> str | None:
    """Parse testsheet date and optional time into ISO 8601 string.

    Returns format 'YYYY-MM-DDTHH:MM:SS+08:00' or None if date unparseable.
    """
    if date_val is None:
        return None

    d = _parse_date_object(date_val)
    if d is None:
        return None

    t = parse_time_tuple(time_val)
    if t is not None:
        hh, mm, ss = t
        return f"{d.year:04d}-{d.month:02d}-{d.day:02d}T{hh:02d}:{mm:02d}:{ss:02d}{tz_offset}"
    return f"{d.year:04d}-{d.month:02d}-{d.day:02d}T00:00:00{tz_offset}"


def match_vi_defect(
    meter_name: str,
    tnb_loc: str | Sequence[ViDefectRecord] | None = "",
    defects: Sequence[ViDefectRecord] | None = None,
) -> ViDefectRecord | None:
    """Match a VI11_* meter name against a list of VI defect records using aligned business logic."""
    if isinstance(tnb_loc, (list, tuple)):
        defects = tnb_loc
        tnb_loc = ""
    elif tnb_loc is None:
        tnb_loc = ""

    if not defects or not meter_name:
        return None

    meter = meter_name.strip().upper()
    loc = tnb_loc.strip().upper()

    if not meter.startswith("VI11_"):
        return None

    for defect in defects:
        eq = defect.equipment.strip().upper()
        area = defect.defect_area.strip().upper()
        rem = defect.additional_remarks.strip().upper()

        # Group A: Switchgear (TNBLOCATION contains /11KV/ or meter starts with VI11_SG_ or VI11_SWG_)
        if meter == "VI11_SG_LABELLING_RMU":
            if eq in ("SWITCHGEAR", "SWG") and area in SG_LABELLING_DEFECT_AREAS:
                return defect

        elif meter == "VI11_SG_PRESGAUGE_RMU":
            if eq in ("SWITCHGEAR", "SWG") and area in SG_PRESGAUGE_DEFECT_AREAS:
                return defect

        elif meter == "VI11_SG_VDIS_RMU":
            if eq in ("SWITCHGEAR", "SWG") and area in SG_VDIS_DEFECT_AREAS:
                return defect

        elif meter in ("VI11_SG_COVERDOOR_RMU", "VI11_SWG_DOOR_VCB"):
            if eq in ("SWITCHGEAR", "SWG") and area == "DOOR BROKEN":
                return defect

        elif meter == "VI11_SG_HEATER_VCB":
            if eq in ("SWITCHGEAR", "SWG") and area in SG_HEATER_DEFECT_AREAS:
                return defect

        elif meter in ("VI11_SG_EARTHIN_RMU", "VI11_SWG_EARTH_VCB"):
            if eq in ("EARTHING", "EARTH") and any(k in rem for k in SG_EARTHIN_REMARKS):
                return defect

        elif meter == "VI11_SG_HANDLE_RMU":
            if eq in ("SWITCHGEAR", "SWG") and "HANDLE" in area:
                return defect

        elif meter == "VI11_SG_OILLEAK_RMU":
            if eq in ("SWITCHGEAR", "SWG") and "OIL LEAK" in area:
                return defect

        # Group B: Feeder Pillar (TNBLOCATION contains /FP/ or meter starts with VI11_FP_)
        elif meter == "VI11_FP_LVDBGUARD_RMU":
            if eq in ("FP/LVDB", "LVDB", "FP") and area == "NO LVDB GUARD":
                return defect

        elif meter == "VI11_FP_LINK/FUSE_RMU":
            if eq in ("FP/LVDB", "LVDB", "FP") and area in FP_LINK_FUSE_DEFECT_AREAS:
                return defect

        elif meter == "VI11_FP_PLOCK_RMU":
            if eq in ("FP/LVDB", "LVDB", "FP") and area in FP_PLOCK_FP_DEFECT_AREAS:
                return defect
            if eq in ("SUBSTATION", "SUB") and area in FP_PLOCK_SUB_DEFECT_AREAS:
                if "FP" in rem or "LVDB" in rem:
                    return defect

        elif meter == "VI11_FP_TDI_RMU":
            if eq in ("FP/LVDB", "LVDB", "FP") and area == "TDI BROKEN":
                return defect

        # Group C: Transformer (TNBLOCATION contains /TX/ or meter starts with VI11_TX_)
        elif meter in (
            "VI11_TX_TXGUARD_RMU",
            "VI11_TX_TXBUSH_RMU",
            "VI11_TX_OILLEVEL_RMU",
            "VI11_TX_OILLEAK_RMU",
            "VI11_TX_CBLCLMP_RMU",
        ):
            # Multi-transformer disambiguation
            has_tx1_rem = "TX1" in rem or "DTX1" in rem
            has_tx2_rem = "TX2" in rem or "DTX2" in rem
            if "DTX1" in loc and has_tx2_rem and not has_tx1_rem:
                continue
            if "DTX2" in loc and has_tx1_rem and not has_tx2_rem:
                continue

            if eq in ("LTX/DTX", "DTX", "LTX", "TX"):
                if meter == "VI11_TX_TXGUARD_RMU" and area in TX_TXGUARD_DEFECT_AREAS:
                    return defect
                if meter == "VI11_TX_TXBUSH_RMU" and area in TX_TXBUSH_DEFECT_AREAS:
                    return defect
                if meter == "VI11_TX_OILLEVEL_RMU" and area in TX_OILLEVEL_DEFECT_AREAS:
                    return defect
                if meter == "VI11_TX_OILLEAK_RMU" and area in TX_OILLEAK_DEFECT_AREAS:
                    return defect
                if meter == "VI11_TX_CBLCLMP_RMU" and area in TX_CBLCLMP_DEFECT_AREAS:
                    return defect

        # Group D: Secondary Equipment (VI11_SEC_)
        elif meter == "VI11_SEC_BATTERY_RMU":
            if eq in ("BATTERY CHARGER", "BATTERY"):
                return defect

        elif meter == "VI11_SEC_BADEFI_RMU":
            if eq == "EFI":
                return defect

        elif meter == "VI11_SEC_BADRTU_RMU":
            # Do NOT match from BATTERY CHARGER. Keep unmapped.
            pass

        elif meter == "VI11_SEC_BADRCB_RMU":
            if eq == "RCB":
                return defect

        elif meter == "VI11_SEC_MCORE_RMU":
            if eq in ("MULTICORE", "MCORE"):
                return defect

        # Group E: Substation / Civil (VI11_SUB_)
        elif meter == "VI11_SUB_SIGNBOARD_RMU":
            if eq in ("SIGNBOARD", "SIGN BOARD"):
                return defect

        elif meter == "VI11_SUB_LIGHT_CSU":
            if eq in ("LIGHTING", "LIGHT"):
                return defect

        elif meter == "VI11_SUB_RETROOF_CSU":
            if eq in ("SUBSTATION", "SUB") and area in SUB_RETROOF_DEFECT_AREAS:
                return defect

        elif meter == "VI11_SUB_CLEANLINESS_RMU":
            if eq in ("SUBSTATION", "SUB") and area in SUB_CLEANLINESS_DEFECT_AREAS:
                return defect

        elif meter == "VI11_SUB_VANDALISM_RMU":
            if eq in ("SUBSTATION", "SUB") and area in SUB_VANDALISM_DEFECT_AREAS:
                if not ("FP" in rem or "LVDB" in rem):
                    return defect

    return None


class PopulateDataMsmsPreflightGuard:
    """Pre-flight resource guard stage for Populate Data MSMS workflow."""

    def validate(self, environment: ProjectEnvironment, request: PopulateDataMsmsRequest) -> None:
        """Validate environmental preconditions before running workflow."""
        total_pe_path = environment.storage.get_total_pe_path()
        if not total_pe_path.exists():
            raise FileNotFoundError(f"TOTAL PE.xlsx not found: {total_pe_path}")

        wb = openpyxl.load_workbook(total_pe_path, read_only=True)
        try:
            if "DataCycle1" not in wb.sheetnames:
                raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")
            ws = wb["DataCycle1"]
            has_wo = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 6:
                    wo_val = str(row[5] or "").strip()
                    if wo_val and wo_val.lower() not in ("none", "nan", "-"):
                        has_wo = True
                        break
            if not has_wo:
                raise RuntimeError(
                    "No populated Work Orders found in TOTAL PE.xlsx (DataCycle1). Please run Propagate WO first."
                )
        finally:
            wb.close()

        csv_files = environment.storage.list_msms_to_be_filled_csv_files()
        to_be_filled_dir = environment.storage.get_msms_to_be_filled_dir()
        if not csv_files:
            if not to_be_filled_dir.exists() or not to_be_filled_dir.is_dir():
                raise FileNotFoundError(f"MSMS TO BE FILLED directory not found: {to_be_filled_dir}")
            raise FileNotFoundError(f"No CSV files found in TO BE FILLED directory: {to_be_filled_dir}")

        testsheet_dir = environment.storage.get_testsheet_dir()
        if not testsheet_dir.exists() or not testsheet_dir.is_dir():
            raise FileNotFoundError(f"TESTSHEET directory not found: {testsheet_dir}")

        if request.mode == PopulateMode.SPECIFIC_FOLDERS and request.target_folder_names:
            all_folders = [p.name for p in testsheet_dir.rglob("*") if p.is_dir()]
            found = any(
                target in f or f == target
                for target in request.target_folder_names
                for f in all_folders
            )
            if not found:
                raise FileNotFoundError(
                    f"Target date folder(s) {request.target_folder_names} not found in TESTSHEET directory: {testsheet_dir}"
                )


@dataclass
class PopulateDataMsmsExtractedData:
    """Extracted data container for Populate Data MSMS workflow."""

    wo_to_package: dict[str, SubstationTestsheetPackage]
    fl_to_package: dict[str, SubstationTestsheetPackage]
    wo_to_fl: dict[str, str]
    csv_files: list[Path]
    packages: list[SubstationTestsheetPackage]


class PopulateDataMsmsExtractor:
    """Pure I/O reading stage for Populate Data MSMS workflow."""

    def __init__(
        self,
        testsheet_repo: SubstationTestsheetRepository | None = None,
        defect_repo: MasterQr03DefectRepository | None = None,
    ) -> None:
        self.testsheet_repo = testsheet_repo or SubstationTestsheetRepository()
        self.defect_repo = defect_repo

    def extract(
        self, environment: ProjectEnvironment, request: PopulateDataMsmsRequest
    ) -> PopulateDataMsmsExtractedData:
        """Extract TOTAL PE lookup, testsheet packages, and CSV paths."""
        total_pe_path = environment.storage.get_total_pe_path()
        testsheet_dir = environment.storage.get_testsheet_dir()

        # 1. Read TOTAL PE (DataCycle1)
        wb_pe = openpyxl.load_workbook(total_pe_path, data_only=True)
        ws_pe = wb_pe["DataCycle1"]

        pe_records: list[dict[str, Any]] = []
        wo_to_fl: dict[str, str] = {}

        for r_idx in range(2, ws_pe.max_row + 1):
            pe_no = ws_pe.cell(r_idx, 1).value
            fl_num = str(ws_pe.cell(r_idx, 2).value or "").strip().upper()
            sub_name = str(ws_pe.cell(r_idx, 3).value or "").strip().upper()
            date_val = str(ws_pe.cell(r_idx, 4).value or "").strip()
            type_c = str(ws_pe.cell(r_idx, 5).value or "").strip().upper()
            wo_val = str(ws_pe.cell(r_idx, 6).value or "").strip()

            if wo_val and wo_val.lower() not in ("none", "nan", "-"):
                wo_to_fl[wo_val] = fl_num
                pe_records.append({
                    "pe_no": pe_no,
                    "fl_erms": fl_num,
                    "sub_name": sub_name,
                    "date_str": date_val,
                    "norm_date": normalize_date_str(date_val),
                    "type": type_c,
                    "wo": wo_val,
                })

        wb_pe.close()

        # 2. Discover testsheet packages
        packages = self.testsheet_repo.discover_packages(testsheet_dir)

        wo_to_package: dict[str, SubstationTestsheetPackage] = {}
        fl_to_package: dict[str, SubstationTestsheetPackage] = {}

        for pkg in packages:
            pkg_data = pkg.data
            pkg_fl = (pkg_data.fl_erms if pkg_data else "").strip().upper()
            pkg_sub = (pkg_data.substation_name_erms if pkg_data else "").strip().upper()
            pkg_date = normalize_date_str(pkg.date_str)
            pkg_pe = pkg.substation_number

            if pkg_fl:
                fl_to_package[pkg_fl] = pkg
                fl_to_package[pkg_fl.replace("/", "")] = pkg

            if pkg_data and pkg_data.wo_number:
                wo_to_package[pkg_data.wo_number] = pkg

            # Match against TOTAL PE records
            for rec in pe_records:
                matched = False
                if pkg_fl and rec["fl_erms"] and (pkg_fl == rec["fl_erms"] or pkg_fl.replace("/", "") == rec["fl_erms"].replace("/", "")):
                    matched = True
                elif pkg_pe and rec["pe_no"] and str(pkg_pe) == str(rec["pe_no"]) and pkg_date == rec["norm_date"]:
                    matched = True
                elif pkg_sub and rec["sub_name"] and pkg_sub == rec["sub_name"] and pkg_date == rec["norm_date"]:
                    matched = True

                if matched:
                    wo_to_package[rec["wo"]] = pkg

        # 3. CSV files in TO BE FILLED
        csv_files = environment.storage.list_msms_to_be_filled_csv_files()

        return PopulateDataMsmsExtractedData(
            wo_to_package=wo_to_package,
            fl_to_package=fl_to_package,
            wo_to_fl=wo_to_fl,
            csv_files=csv_files,
            packages=packages,
        )



class PopulateDataMsmsFilter:
    """Filter stage identifying relevant CSV files and candidate rows."""

    def filter_csv_files(
        self,
        csv_files: Sequence[Path],
        mode: PopulateMode,
        target_folder_names: Sequence[str] = (),
    ) -> list[Path]:
        """Filter target CSV files according to PopulateMode."""
        if mode != PopulateMode.SPECIFIC_FOLDERS or not target_folder_names:
            return list(csv_files)

        filtered: list[Path] = []
        for csv_path in csv_files:
            matched = any(
                target in csv_path.name or target in normalize_date_str(csv_path.name)
                for target in target_folder_names
            )
            if matched:
                filtered.append(csv_path)

        return filtered if filtered else list(csv_files)


@dataclass
class CsvRowEvaluation:
    """Evaluation result for a single CSV row."""

    row_dict: dict[str, str]
    is_populated: bool = False
    is_skipped_already_filled: bool = False
    is_skipped_no_testsheet: bool = False
    is_unmapped_meter: bool = False


@dataclass
class CsvFileTransformation:
    """Transformation plan for a single CSV file."""

    csv_path: Path
    fieldnames: list[str]
    delimiter: str
    lineterminator: str
    rows: list[dict[str, str]]
    evaluations: list[CsvRowEvaluation]


class PopulateDataMsmsTransformer:
    """Transformer stage applying TestsheetReadingMapper and defect rules."""

    def __init__(
        self,
        mapper: TestsheetReadingMapper | None = None,
        defect_repo: MasterQr03DefectRepository | None = None,
    ) -> None:
        self.mapper = mapper or TestsheetReadingMapper()
        self.defect_repo = defect_repo

    def transform(
        self,
        extracted: PopulateDataMsmsExtractedData,
        csv_files: Sequence[Path],
        overwrite: bool = False,
    ) -> list[CsvFileTransformation]:
        """Transform CSV rows in place using extracted testsheets and defects."""
        wb_cache: dict[Path, openpyxl.Workbook] = {}
        vi_defects_cache: dict[str, list[ViDefectRecord]] = {}
        transformations: list[CsvFileTransformation] = []

        try:
            for csv_path in csv_files:
                trans = self._transform_single_csv(
                    csv_path=csv_path,
                    extracted=extracted,
                    wb_cache=wb_cache,
                    vi_defects_cache=vi_defects_cache,
                    overwrite=overwrite,
                )
                transformations.append(trans)
        finally:
            for wb in wb_cache.values():
                wb.close()

        return transformations

    def _transform_single_csv(
        self,
        csv_path: Path,
        extracted: PopulateDataMsmsExtractedData,
        wb_cache: dict[Path, openpyxl.Workbook],
        vi_defects_cache: dict[str, list[ViDefectRecord]],
        overwrite: bool,
    ) -> CsvFileTransformation:
        fieldnames, delimiter, lineterminator, raw_rows = self._read_csv_rows(csv_path)
        evaluations: list[CsvRowEvaluation] = []

        for row in raw_rows:
            eval_res = self._evaluate_row(
                row=row,
                extracted=extracted,
                wb_cache=wb_cache,
                vi_defects_cache=vi_defects_cache,
                overwrite=overwrite,
            )
            evaluations.append(eval_res)

        return CsvFileTransformation(
            csv_path=csv_path,
            fieldnames=fieldnames,
            delimiter=delimiter,
            lineterminator=lineterminator,
            rows=[e.row_dict for e in evaluations],
            evaluations=evaluations,
        )

    def _evaluate_row(
        self,
        row: dict[str, str],
        extracted: PopulateDataMsmsExtractedData,
        wb_cache: dict[Path, openpyxl.Workbook],
        vi_defects_cache: dict[str, list[ViDefectRecord]],
        overwrite: bool,
    ) -> CsvRowEvaluation:
        wo_num = str(row.get("WONUM") or "").strip()
        meter_name = str(row.get("METERNAME") or "").strip()
        tnb_loc = str(row.get("TNBLOCATION") or "").strip()

        # Check existing reading
        current_reading = str(row.get("TNBNEWREADING") or "").strip()
        current_date = str(row.get("TNBNEWREADINGDATE") or "").strip()
        has_existing = bool(current_reading or current_date)

        if has_existing and not overwrite:
            return CsvRowEvaluation(
                row_dict=row,
                is_skipped_already_filled=True,
            )

        # Match package
        pkg = extracted.wo_to_package.get(wo_num)
        if not pkg and tnb_loc:
            # Fallback by location prefix
            fl_prefix = tnb_loc.split("/")[0] if "/" in tnb_loc else tnb_loc
            pkg = extracted.fl_to_package.get(fl_prefix) or extracted.fl_to_package.get(tnb_loc)
        if not pkg and wo_num in extracted.wo_to_fl:
            fl_target = extracted.wo_to_fl[wo_num]
            pkg = extracted.fl_to_package.get(fl_target)

        if not pkg:
            return CsvRowEvaluation(
                row_dict=row,
                is_skipped_no_testsheet=True,
            )

        fl_erms = (pkg.data.fl_erms if pkg.data else "").strip().upper() or extracted.wo_to_fl.get(wo_num, "")

        # Open workbook (cached)
        wb = self._get_cached_workbook(pkg.testsheet_path, wb_cache)

        # Extract timestamps
        act_start, act_finish, reading_date = self._extract_timestamps(wb, pkg)

        # 1. Visual Inspection (VI11_*)
        if meter_name.upper().startswith("VI11_"):
            defects = self._get_vi_defects(fl_erms, vi_defects_cache)
            matched_defect = match_vi_defect(meter_name, tnb_loc, defects)
            if matched_defect is not None:
                row["TNBNEWREADING"] = "YES"
                row["TNBCOMMENTS"] = matched_defect.additional_remarks
                row["TNBNEWREADINGDATE"] = reading_date
                if "ACTSTART" in row and act_start:
                    row["ACTSTART"] = act_start
                if "ACTFINISH" in row and act_finish:
                    row["ACTFINISH"] = act_finish
                return CsvRowEvaluation(row_dict=row, is_populated=True)
            else:
                # No defect found: skip completely (leave untouched)
                return CsvRowEvaluation(row_dict=row, is_populated=False)

        # 2. Hardcoded / Special Meters
        if meter_name.upper() == "US_S11_RMU_PE13R":
            row["TNBNEWREADING"] = "0"
            row["TNBNEWREADINGDATE"] = reading_date
            if "ACTSTART" in row and act_start:
                row["ACTSTART"] = act_start
            if "ACTFINISH" in row and act_finish:
                row["ACTFINISH"] = act_finish
            return CsvRowEvaluation(row_dict=row, is_populated=True)

        # 3. Feeder Pillar (LVDB / FP) Thermal Synthesis
        feeder_parsed = parse_feeder_meter(meter_name)
        if feeder_parsed is not None and "PCE Testsheet" in wb.sheetnames:
            feeder_channel, metric_suffix = feeder_parsed
            ws_ts = wb["PCE Testsheet"]

            # Determine whether equipment is FP1 or FP2
            fp_idx = 1
            if tnb_loc:
                try:
                    cat, idx = parse_equipment_index(tnb_loc)
                    if cat == "FP":
                        fp_idx = idx
                except ValueError:
                    pass

            # Cable type row: 45 for FP1, 47 for FP2
            # Board average cell: R50 for FP1, R54 for FP2
            col_letter = FEEDER_CHANNEL_COLUMNS.get(feeder_channel)
            if col_letter:
                cable_row = 45 if fp_idx == 1 else 47
                cable_val = ws_ts[f"{col_letter}{cable_row}"].value

                if is_active_feeder_cable(cable_val):
                    board_temp_cell = "R50" if fp_idx == 1 else "R54"
                    board_temp_val = ws_ts[board_temp_cell].value
                    board_temp = extract_board_average_temperature(board_temp_val)

                    if board_temp is not None:
                        sub_type = pkg.data.substation_type if (pkg.data and pkg.data.substation_type) else ""
                        seed_key = f"{fl_erms}:{wo_num}:{tnb_loc}"
                        synth_readings = synthesize_feeder_thermal_readings(
                            board_avg_temp=board_temp,
                            feeder_id=feeder_channel,
                            substation_type=sub_type,
                            seed_key=seed_key,
                        )
                        reading_val = synth_readings.get(metric_suffix)
                        if reading_val is not None:
                            row["TNBNEWREADING"] = normalize_for_csv(reading_val)
                            row["TNBNEWREADINGDATE"] = reading_date
                            if "ACTSTART" in row and act_start:
                                row["ACTSTART"] = act_start
                            if "ACTFINISH" in row and act_finish:
                                row["ACTFINISH"] = act_finish
                            return CsvRowEvaluation(row_dict=row, is_populated=True)

            # Inactive, spare, or unconfigured feeder stays blank
            return CsvRowEvaluation(row_dict=row, is_populated=False)

        # 4. Numeric GAUGE meters
        target = self.mapper.get_target(meter_name, tnb_loc)
        if target is None:
            # Stub meter (e.g. Earth meters or LV compartment)
            return CsvRowEvaluation(row_dict=row, is_unmapped_meter=True)

        sheet_name, cell_coord = target
        if sheet_name not in wb.sheetnames:
            return CsvRowEvaluation(row_dict=row, is_unmapped_meter=True)

        cell_val = wb[sheet_name][cell_coord].value
        if meter_name.upper() == "BG_ROOM_TEM":
            extracted_temp = extract_background_temperature(cell_val)
            norm_val = normalize_for_csv(extracted_temp) if extracted_temp is not None else ""
        else:
            norm_val = normalize_for_csv(cell_val)

        if norm_val != "":
            row["TNBNEWREADING"] = norm_val
            row["TNBNEWREADINGDATE"] = reading_date
            if "ACTSTART" in row and act_start:
                row["ACTSTART"] = act_start
            if "ACTFINISH" in row and act_finish:
                row["ACTFINISH"] = act_finish
            return CsvRowEvaluation(row_dict=row, is_populated=True)

        return CsvRowEvaluation(row_dict=row, is_populated=False)

    def _get_cached_workbook(
        self, path: Path, cache: dict[Path, openpyxl.Workbook]
    ) -> openpyxl.Workbook:
        if path not in cache:
            cache[path] = openpyxl.load_workbook(path, data_only=True)
        return cache[path]

    def _get_vi_defects(
        self, fl_erms: str, cache: dict[str, list[ViDefectRecord]]
    ) -> list[ViDefectRecord]:
        if not self.defect_repo or not fl_erms:
            return []
        if fl_erms not in cache:
            try:
                cache[fl_erms] = self.defect_repo.fetch_vi_defects(fl_erms)
            except Exception as e:
                logger.warning(f"Failed to fetch VI defects for {fl_erms}: {e}")
                cache[fl_erms] = []
        return cache[fl_erms]

    def _extract_timestamps(
        self, wb: openpyxl.Workbook, pkg: SubstationTestsheetPackage
    ) -> tuple[str, str, str]:
        p4_date = None
        p5_time = None
        s5_time = None

        if "PCE Testsheet" in wb.sheetnames:
            ws = wb["PCE Testsheet"]
            p4_date = ws["P4"].value
            p5_time = ws["P5"].value
            s5_time = ws["S5"].value

        if not p4_date:
            p4_date = pkg.date_str

        act_start = parse_testsheet_datetime(p4_date, p5_time)
        act_finish = parse_testsheet_datetime(p4_date, s5_time)
        reading_date = act_start if act_start else parse_testsheet_datetime(p4_date, None)

        return act_start, act_finish, reading_date

    def _read_csv_rows(
        self, csv_path: Path
    ) -> tuple[list[str], str, str, list[dict[str, str]]]:
        """Read CSV file detecting fieldnames, delimiter, and line terminator."""
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(csv_path, mode="r", encoding=encoding, newline="") as f:
                    sample = f.read(4096)
                    f.seek(0)
                    delimiter = ","
                    if ";" in sample and sample.count(";") > sample.count(","):
                        delimiter = ";"
                    elif "\t" in sample and sample.count("\t") > sample.count(","):
                        delimiter = "\t"

                    lineterminator = "\r\n" if "\r\n" in sample else "\n"

                    reader = csv.DictReader(f, delimiter=delimiter)
                    fieldnames = list(reader.fieldnames or [])
                    rows: list[dict[str, str]] = []
                    for r in reader:
                        clean_r = {k: (v if v is not None else "") for k, v in r.items()}
                        rows.append(clean_r)

                    return fieldnames, delimiter, lineterminator, rows
            except UnicodeDecodeError:
                continue

        raise ValueError(f"Could not decode CSV file: {csv_path.name}")


class PopulateDataMsmsLoader:
    """Loader stage executing in-place CSV updates."""

    def load(self, transformations: Sequence[CsvFileTransformation]) -> None:
        """Write modified CSV rows back to disk in place."""
        for trans in transformations:
            temp_path = trans.csv_path.with_suffix(".tmp")
            with open(temp_path, mode="w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=trans.fieldnames,
                    delimiter=trans.delimiter,
                    lineterminator=trans.lineterminator,
                )
                writer.writeheader()
                writer.writerows(trans.rows)

            temp_path.replace(trans.csv_path)


class PopulateDataMsmsAuditor:
    """Auditing stage calculating telemetry and verifying output."""

    def audit(
        self, transformations: Sequence[CsvFileTransformation]
    ) -> PopulateDataMsmsResult:
        """Calculate and return PopulateDataMsmsResult telemetry."""
        csv_files_processed = len(transformations)
        total_rows_evaluated = 0
        rows_populated = 0
        rows_skipped_already_filled = 0
        rows_skipped_no_testsheet = 0
        unmapped_meters_count = 0
        warnings: list[str] = []
        errors: list[str] = []

        for trans in transformations:
            for ev in trans.evaluations:
                total_rows_evaluated += 1
                if ev.is_populated:
                    rows_populated += 1
                elif ev.is_skipped_already_filled:
                    rows_skipped_already_filled += 1
                elif ev.is_skipped_no_testsheet:
                    rows_skipped_no_testsheet += 1
                    wo = ev.row_dict.get("WONUM", "")
                    if wo and f"No testsheet for WO: {wo}" not in warnings:
                        warnings.append(f"No testsheet for WO: {wo}")
                elif ev.is_unmapped_meter:
                    unmapped_meters_count += 1

        return PopulateDataMsmsResult(
            csv_files_processed=csv_files_processed,
            total_rows_evaluated=total_rows_evaluated,
            rows_populated=rows_populated,
            rows_skipped_already_filled=rows_skipped_already_filled,
            rows_skipped_no_testsheet=rows_skipped_no_testsheet,
            unmapped_meters_count=unmapped_meters_count,
            warnings=tuple(warnings),
            errors=tuple(errors),
        )


class PopulateDataMsmsWorkflow:
    """6-stage ETL workflow for populating testsheet data into MSMS CSVs."""

    def __init__(
        self,
        preflight_guard: PopulateDataMsmsPreflightGuard | None = None,
        extractor: PopulateDataMsmsExtractor | None = None,
        filter_stage: PopulateDataMsmsFilter | None = None,
        transformer: PopulateDataMsmsTransformer | None = None,
        loader: PopulateDataMsmsLoader | None = None,
        auditor: PopulateDataMsmsAuditor | None = None,
    ) -> None:
        self.preflight_guard = preflight_guard or PopulateDataMsmsPreflightGuard()
        self.extractor = extractor or PopulateDataMsmsExtractor()
        self.filter_stage = filter_stage or PopulateDataMsmsFilter()
        self.transformer = transformer or PopulateDataMsmsTransformer()
        self.loader = loader or PopulateDataMsmsLoader()
        self.auditor = auditor or PopulateDataMsmsAuditor()

    def execute(
        self, environment: ProjectEnvironment, request: PopulateDataMsmsRequest | None = None
    ) -> PopulateDataMsmsResult:
        """Execute the Populate Data MSMS workflow."""
        req = request or PopulateDataMsmsRequest()
        if req.progress_sink:
            req.progress_sink("Validating TOTAL PE, CSVs, and testsheets preconditions...")

        self.preflight_guard.validate(environment, req)

        if req.progress_sink:
            req.progress_sink("Extracting TOTAL PE mappings and testsheet packages...")

        # Wire defect repo from environment if available
        try:
            defect_repo = MasterQr03DefectRepository(environment=environment)
        except Exception:
            defect_repo = None

        self.extractor.defect_repo = defect_repo
        self.transformer.defect_repo = defect_repo

        extracted = self.extractor.extract(environment, req)

        if req.progress_sink:
            req.progress_sink(
                f"Discovered {len(extracted.packages)} testsheet package(s), {len(extracted.csv_files)} CSV file(s)..."
            )

        target_csvs = self.filter_stage.filter_csv_files(
            csv_files=extracted.csv_files,
            mode=req.mode,
            target_folder_names=req.target_folder_names,
        )

        if req.progress_sink:
            req.progress_sink(f"Transforming and matching readings for {len(target_csvs)} CSV file(s)...")

        transformations = self.transformer.transform(
            extracted=extracted,
            csv_files=target_csvs,
            overwrite=req.overwrite,
        )

        if req.progress_sink:
            req.progress_sink("Writing updated CSV files in place...")

        self.loader.load(transformations)

        result = self.auditor.audit(transformations)

        if req.progress_sink:
            req.progress_sink(
                f"Population completed: {result.rows_populated} row(s) populated across {result.csv_files_processed} CSV file(s)."
            )

        return result
