"""Workflow for replacing inspection photos or placeholder text in Excel testsheets.

Core functions (replace_pce_images, _select_signature_path) are reused by postprocessing_pipeline_workflow.py.
"""

from __future__ import annotations

import os
import random
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

from src import cli_selectors


@dataclass(frozen=True)
class ReplaceImagesSummary:
    target: Path
    photo1: Path
    photo2: Path | None
    mode: str
    processed_count: int
    failed_count: int


def replace_pce_images(
    xlsx_path: str | Path,
    photo1_path: str | Path,
    photo2_path: str | Path | None = None,
    output_path: str | Path | None = None,
    mode: str = "placeholder",
) -> str:
    """Replace inspection photos or placeholders on 'PCE Testsheet' and 'PCE VI' inside the given Excel file."""
    path = Path(xlsx_path).expanduser().resolve()
    p1 = Path(photo1_path).expanduser().resolve()
    p2 = Path(photo2_path).expanduser().resolve() if photo2_path else None

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    if not p1.exists():
        raise FileNotFoundError(f"New image 1 not found: {p1}")
    if p2 and not p2.exists():
        raise FileNotFoundError(f"New image 2 not found: {p2}")

    if p2 is None:
        p2 = p1

    def get_img_file(path_arg: Path) -> Path:
        if path_arg.is_dir():
            png_files = [f for f in path_arg.iterdir() if f.is_file() and f.suffix.lower() == ".png"]
            if not png_files:
                raise FileNotFoundError(f"No .png files found inside directory: {path_arg}")
            return random.choice(png_files)
        return path_arg

    def get_format(img_path: Path) -> str:
        ext = img_path.suffix.lower().replace(".", "")
        return "jpeg" if ext in ["jpg", "jpeg"] else "png"

    print(f"Loading Excel file: {path} (Mode: {mode}) ...")
    wb = openpyxl.load_workbook(str(path))

    if mode == "placeholder":
        for ws in wb.worksheets:
            placeholders_replaced = 0
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell) or not cell.value or not isinstance(cell.value, str):
                        continue
                    val_lower = cell.value.lower()
                    has_vendor = "{{signvendor}}" in val_lower
                    has_tnb = "{{signtnb}}" in val_lower

                    if not has_vendor and not has_tnb:
                        continue

                    col_idx = cell.column - 1
                    row_idx = cell.row - 1

                    val_clean = str(cell.value)
                    val_clean = re.sub(r"\{\{\s*signvendor\s*\}\}", "", val_clean, flags=re.IGNORECASE)
                    val_clean = re.sub(r"\{\{\s*signtnb\s*\}\}", "", val_clean, flags=re.IGNORECASE)
                    val_clean = val_clean.strip()
                    cell.value = val_clean if val_clean else None

                    if has_vendor:
                        chosen_p1 = get_img_file(p1)
                        img1 = Image(str(chosen_p1))
                        if get_format(chosen_p1) == "png":
                            img1.format = "png"
                        if ws.title == "PCE Testsheet":
                            if col_idx <= 6 and 60 <= row_idx <= 72:
                                marker_from = AnchorMarker(col=3, colOff=1449204, row=65, rowOff=12650)
                                marker_to = AnchorMarker(col=6, colOff=826923, row=70, rowOff=62917)
                            else:
                                marker_from = AnchorMarker(col=col_idx, colOff=1449204, row=row_idx, rowOff=12650)
                                marker_to = AnchorMarker(col=col_idx + 3, colOff=826923, row=row_idx + 5, rowOff=62917)
                        elif ws.title == "PCE VI":
                            if col_idx <= 6 and 42 <= row_idx <= 55:
                                marker_from = AnchorMarker(col=2, colOff=14608, row=47, rowOff=164752)
                                marker_to = AnchorMarker(col=5, colOff=663339, row=51, rowOff=50601)
                            else:
                                marker_from = AnchorMarker(col=col_idx, colOff=14608, row=row_idx, rowOff=164752)
                                marker_to = AnchorMarker(col=col_idx + 3, colOff=663339, row=row_idx + 4, rowOff=50601)
                        else:
                            marker_from = AnchorMarker(col=col_idx, colOff=1449204, row=row_idx, rowOff=12650)
                            marker_to = AnchorMarker(col=col_idx + 3, colOff=826923, row=row_idx + 5, rowOff=62917)

                        img1.anchor = TwoCellAnchor(_from=marker_from, to=marker_to)
                        ws.add_image(img1)
                        placeholders_replaced += 1

                    if has_tnb:
                        chosen_p2 = get_img_file(p2)
                        img2 = Image(str(chosen_p2))
                        if get_format(chosen_p2) == "png":
                            img2.format = "png"
                        if ws.title == "PCE Testsheet":
                            if col_idx >= 15 and 60 <= row_idx <= 72:
                                marker_from = AnchorMarker(col=19, colOff=502432, row=64, rowOff=24928)
                                marker_to = AnchorMarker(col=24, colOff=23064, row=71, rowOff=12017)
                            else:
                                marker_from = AnchorMarker(col=col_idx, colOff=502432, row=row_idx, rowOff=24928)
                                marker_to = AnchorMarker(col=col_idx + 5, colOff=23064, row=row_idx + 7, rowOff=12017)
                        elif ws.title == "PCE VI":
                            if col_idx >= 7 and 42 <= row_idx <= 55:
                                marker_from = AnchorMarker(col=10, colOff=463652, row=46, rowOff=114300)
                                marker_to = AnchorMarker(col=14, colOff=197220, row=51, rowOff=126503)
                            else:
                                marker_from = AnchorMarker(col=col_idx, colOff=463652, row=row_idx, rowOff=114300)
                                marker_to = AnchorMarker(col=col_idx + 4, colOff=197220, row=row_idx + 5, rowOff=126503)
                        else:
                            marker_from = AnchorMarker(col=col_idx, colOff=502432, row=row_idx, rowOff=24928)
                            marker_to = AnchorMarker(col=col_idx + 5, colOff=23064, row=row_idx + 7, rowOff=12017)

                        img2.anchor = TwoCellAnchor(_from=marker_from, to=marker_to)
                        ws.add_image(img2)
                        placeholders_replaced += 1

            if ws.title in ["PCE Testsheet", "PCE VI"]:
                if placeholders_replaced > 0:
                    print(f" -> Replaced {placeholders_replaced} placeholder(s) on '{ws.title}' successfully.")
                else:
                    print(f" -> Warning: No '{{{{signvendor}}}}' or '{{{{signtnb}}}}' placeholders found on '{ws.title}'.")
            elif placeholders_replaced > 0:
                print(f" -> Replaced {placeholders_replaced} placeholder(s) on '{ws.title}' successfully.")

    else:
        if "PCE Testsheet" in wb.sheetnames:
            ws_test = wb["PCE Testsheet"]
            if len(ws_test._images) >= 2:
                chosen_p2 = get_img_file(p2)
                ws_test._images[0].ref = str(chosen_p2)
                ws_test._images[0].format = get_format(chosen_p2)

                chosen_p1 = get_img_file(p1)
                ws_test._images[1].ref = str(chosen_p1)
                ws_test._images[1].format = get_format(chosen_p1)
                print(" -> Replaced both existing images on 'PCE Testsheet' successfully.")
            else:
                print(f" -> Warning: Expected 2 images on 'PCE Testsheet', found {len(ws_test._images)}")

        if "PCE VI" in wb.sheetnames:
            ws_vi = wb["PCE VI"]
            if len(ws_vi._images) >= 3:
                chosen_p1 = get_img_file(p1)
                ws_vi._images[1].ref = str(chosen_p1)
                ws_vi._images[1].format = get_format(chosen_p1)

                chosen_p2 = get_img_file(p2)
                ws_vi._images[2].ref = str(chosen_p2)
                ws_vi._images[2].format = get_format(chosen_p2)
                print(" -> Replaced both existing inspection photos on 'PCE VI' successfully (Header logo preserved).")
            else:
                print(f" -> Warning: Expected 3 images on 'PCE VI', found {len(ws_vi._images)}")

    save_target = str(output_path) if output_path else str(path)
    wb.save(save_target)
    print(f"Saved updated workbook to: {save_target}\n")
    return save_target


def batch_replace_pce_images(
    folder_path: str | Path,
    photo1_path: str | Path,
    photo2_path: str | Path | None = None,
    output_folder: str | Path | None = None,
    mode: str = "placeholder",
) -> tuple[int, int]:
    """Iterates through all .xlsx files in a folder and replaces their inspection photos or placeholders."""
    fpath = Path(folder_path).expanduser().resolve()
    if not fpath.exists() or not fpath.is_dir():
        raise NotADirectoryError(f"Folder not found or not a directory: {fpath}")

    if output_folder:
        outpath = Path(output_folder).expanduser().resolve()
    elif fpath.name == "processed_testsheet":
        outpath = fpath
    else:
        outpath = fpath / "processed_testsheet"

    outpath.mkdir(parents=True, exist_ok=True)

    processed_count = 0
    failed_count = 0

    files = sorted(os.listdir(fpath))
    print(f"=== Starting Batch Image Replacement ({mode.upper()} MODE) in Folder: {fpath} ===")
    for filename in files:
        if filename.startswith("~$") or not filename.lower().endswith(".xlsx"):
            continue

        full_path = fpath / filename
        out_path = outpath / filename

        print(f"\n[{processed_count + failed_count + 1}] Processing file: {filename}")
        try:
            replace_pce_images(full_path, photo1_path, photo2_path, output_path=out_path, mode=mode)
            processed_count += 1
        except Exception as e:
            print(f" -> ERROR processing {filename}: {str(e)}")
            failed_count += 1

    print("\n=== Batch Processing Completed ===")
    print(f"Successfully processed: {processed_count} files")
    if failed_count > 0:
        print(f"Failed: {failed_count} files")
    return processed_count, failed_count


def _select_signature_path(
    prompt_message: str,
    sign_base_dir: Path,
    default_folder: str | None = None,
) -> tuple[Path | None, str | None]:
    """Scan `sign_base_dir` for subfolders with .png files and prompt user to select one or enter custom path."""
    valid_folders: list[tuple[str, Path, list[Path]]] = []
    if sign_base_dir.exists() and sign_base_dir.is_dir():
        for item in sorted(os.listdir(sign_base_dir)):
            subfolder = sign_base_dir / item
            if subfolder.is_dir():
                png_files = sorted([
                    f for f in subfolder.iterdir()
                    if f.is_file() and f.suffix.lower() == ".png"
                ])
                if png_files:
                    valid_folders.append((item, subfolder, png_files))

    options: list[cli_selectors.SelectOption[str]] = []
    for folder_name, _, png_files in valid_folders:
        options.append(
            cli_selectors.SelectOption(
                f"{folder_name} ({len(png_files)} .png files)",
                folder_name,
            )
        )
    options.append(cli_selectors.SelectOption("Enter custom image path (.png)", "__custom__"))
    options.append(cli_selectors.SelectOption("Cancel", "__cancel__", shortcut_key="c"))

    # Determine default selection
    chosen_default = None
    if default_folder and any(f[0] == default_folder for f in valid_folders):
        chosen_default = default_folder
    elif valid_folders:
        chosen_default = valid_folders[0][0]

    selection = cli_selectors.select_one(prompt_message, options, default_value=chosen_default)
    if selection in ("__cancel__", None):
        return None, None

    if selection == "__custom__":
        while True:
            custom_str = input("Enter path to .png signature image (or 'q' to cancel): ").strip().strip('"')
            if not custom_str or custom_str.lower() == "q":
                return None, None
            custom_path = Path(custom_str).expanduser().resolve()
            if custom_path.is_file() and custom_path.suffix.lower() == ".png":
                return custom_path, "__custom__"
            print("Invalid file path or not a .png file. Please try again.")

    for folder_name, subfolder, png_files in valid_folders:
        if folder_name == selection:
            print(f" -> Selected signature person: {folder_name} ({len(png_files)} signature variations will be dynamically randomized per replacement)")
            return subfolder, folder_name

    return None, None


def run_replace_images() -> ReplaceImagesSummary:
    """Interactive entrypoint for inspection photo / signature stamp replacement."""
    project_root = Path(__file__).resolve().parent.parent
    sign_dir = project_root / "OTHERS" / "SIGN"

    # Robust target path validation loop
    target: Path | None = None
    while True:
        target_str = input("Enter path to Excel file (.xlsx) or folder containing .xlsx files (or 'q' to cancel): ").strip().strip('"')
        if not target_str or target_str.lower() == "q":
            print("Cancelled.")
            return ReplaceImagesSummary(target=Path("."), photo1=Path("."), photo2=None, mode="cancelled", processed_count=0, failed_count=0)
        candidate = Path(target_str).expanduser().resolve()
        if candidate.exists() and (candidate.is_file() or candidate.is_dir()):
            target = candidate
            break
        print(f"[ERROR] Path not found: {candidate}. Please try again.")

    mode_options = [
        cli_selectors.SelectOption("Placeholder Replacement Mode (Replace {{signvendor}} / {{signtnb}} text)", "placeholder"),
        cli_selectors.SelectOption("Image Replacement Mode (Replace existing embedded images)", "image"),
        cli_selectors.SelectOption("Cancel", "__cancel__", shortcut_key="c"),
    ]
    mode = cli_selectors.select_one("Select replacement mode:", mode_options, default_value="placeholder")
    if mode in ("__cancel__", None):
        print("Cancelled.")
        return ReplaceImagesSummary(target=target, photo1=Path("."), photo2=None, mode="cancelled", processed_count=0, failed_count=0)

    # Select signature for img1 (Tested by / {{signvendor}})
    photo1_path, folder1 = _select_signature_path(
        "Select signature person for img1 (Tested by / {{signvendor}}):",
        sign_dir,
    )
    if not photo1_path:
        print("Cancelled signature selection for img1.")
        return ReplaceImagesSummary(target=target, photo1=Path("."), photo2=None, mode=mode, processed_count=0, failed_count=0)

    # Select signature for img2 (TNB Supervisor / {{signtnb}})
    photo2_path, folder2 = _select_signature_path(
        "Select signature person for img2 (TNB Supervisor / {{signtnb}}):",
        sign_dir,
        default_folder=folder1,
    )
    if not photo2_path:
        print("Cancelled signature selection for img2.")
        return ReplaceImagesSummary(target=target, photo1=photo1_path, photo2=None, mode=mode, processed_count=0, failed_count=0)

    print(f"\n=== Running {mode.upper()} replacement ===")
    print(f"Target : {target}")
    print(f"Img1   : {photo1_path}")
    print(f"Img2   : {photo2_path}\n")

    if target.is_dir():
        proc, fail = batch_replace_pce_images(target, photo1_path, photo2_path, mode=mode)
        return ReplaceImagesSummary(target=target, photo1=photo1_path, photo2=photo2_path, mode=mode, processed_count=proc, failed_count=fail)
    elif target.is_file():
        try:
            if target.parent.name == "processed_testsheet":
                out_path = target
            else:
                out_dir = target.parent / "processed_testsheet"
                out_dir.mkdir(parents=True, exist_ok=True)
                out_path = out_dir / target.name
            replace_pce_images(target, photo1_path, photo2_path, output_path=out_path, mode=mode)
            return ReplaceImagesSummary(target=target, photo1=photo1_path, photo2=photo2_path, mode=mode, processed_count=1, failed_count=0)
        except Exception as e:
            print(f"[ERROR] Failed to process file {target}: {e}")
            return ReplaceImagesSummary(target=target, photo1=photo1_path, photo2=photo2_path, mode=mode, processed_count=0, failed_count=1)
    else:
        raise FileNotFoundError(f"Target not found: {target}")

