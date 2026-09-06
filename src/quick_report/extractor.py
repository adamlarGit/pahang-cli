"""Package and defect extraction stage for Quick Report workflow."""

from __future__ import annotations

import dataclasses
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any, TYPE_CHECKING
import warnings

import openpyxl

from src.core.normalizers import (
    FL_PREFIX_TO_STATION,
    format_month_folder,
    resolve_station_from_fl,
)
from src.quick_report.defects import MasterQr03DefectRepository
from src.quick_report.utils import normalize_functional_location_input
from src.testsheet.extractor import to_excel_date
from src.testsheet.models import SubstationTestsheetPackage
from src.testsheet.repository import SubstationTestsheetRepository

if TYPE_CHECKING:
    from src.project.environment import ProjectEnvironment
    from src.quick_report.defects import CbmDefectRecord, ViDefectRecord
    from src.workflows.models import QuickReportRequest


_resolve_station_from_fl = resolve_station_from_fl


class QuickReportExtractor:
    """Pure Read I/O discovery and per-station defect extraction stage for Quick Report."""

    def __init__(self, repository: SubstationTestsheetRepository | None = None) -> None:
        self.repository = repository or SubstationTestsheetRepository()
        self._defect_repo: MasterQr03DefectRepository | None = None

    def extract(
        self, environment: ProjectEnvironment, request: QuickReportRequest
    ) -> list[SubstationTestsheetPackage]:
        """Discover testsheet packages strictly via read I/O (without domain filtering)."""
        mode_val = getattr(request.mode, "value", str(request.mode)).lower()
        if mode_val == "fl":
            return self._extract_fl_mode(environment, request)

        if mode_val == "folder":
            packages: list[SubstationTestsheetPackage] = []
            for folder_str in request.target_folders:
                candidate = Path(folder_str)
                if candidate.exists():
                    folder_path = candidate
                else:
                    folder_path = environment.get_testsheet_dir() / folder_str
                    if not folder_path.exists():
                        raise FileNotFoundError(
                            f"Requested target folder does not exist: '{folder_str}' "
                            f"(checked: {candidate}, {folder_path})"
                        )
                packages.extend(self.repository.discover_packages(folder_path))
            return packages

        return self.repository.discover_packages(environment.get_testsheet_dir())

    def _extract_fl_mode(
        self, environment: ProjectEnvironment, request: QuickReportRequest
    ) -> list[SubstationTestsheetPackage]:
        """Targeted package discovery for FL mode with lazy hydration."""
        target_fls = {
            normalize_functional_location_input(name)
            for name in (request.target_package_names or ())
            if name and str(name).strip()
        }
        if not target_fls:
            return []

        matched_packages: list[SubstationTestsheetPackage] = []
        matched_fls: set[str] = set()

        # Tier 1: Check TOTAL PE.xlsx (DataCycle1 sheet)
        total_pe_path: Path | None = None
        if hasattr(environment, "storage") and hasattr(environment.storage, "get_total_pe_path"):
            try:
                total_pe_path = environment.storage.get_total_pe_path()
            except Exception:
                pass
        elif hasattr(environment, "get_total_pe_path"):
            try:
                total_pe_path = environment.get_total_pe_path()
            except Exception:
                pass

        if total_pe_path is not None and isinstance(total_pe_path, (str, Path)) and Path(total_pe_path).is_file():
            t1_pkgs, t1_fls = self._discover_via_total_pe(
                Path(total_pe_path), environment.get_testsheet_dir(), target_fls
            )
            matched_packages.extend(t1_pkgs)
            matched_fls.update(t1_fls)

        # Tier 2 (Fallback): Route remaining FLs by prefix to station folder
        remaining_fls = target_fls - matched_fls
        if remaining_fls:
            t2_pkgs, t2_fls = self._discover_via_station_prefix(
                environment.get_testsheet_dir(), remaining_fls
            )
            matched_packages.extend(t2_pkgs)
            matched_fls.update(t2_fls)

        # Hydration: Only perform full extract_testsheet_data for matched package(s)
        hydrated_packages: list[SubstationTestsheetPackage] = []
        for pkg in matched_packages:
            if pkg.data is None:
                try:
                    full_data = self.repository.extractor.extract_testsheet_data(
                        pkg.testsheet_path,
                        station_hint=pkg.station,
                        date_hint=pkg.date_str,
                    )
                    pkg = dataclasses.replace(pkg, data=full_data)
                except Exception:
                    pass
            hydrated_packages.append(pkg)

        return hydrated_packages

    def _discover_via_total_pe(
        self,
        total_pe_path: Path,
        testsheet_root: Path,
        target_fls: set[str],
    ) -> tuple[list[SubstationTestsheetPackage], set[str]]:
        """Look up target FLs in TOTAL PE DataCycle1 sheet and locate matching testsheets."""
        matched_packages: list[SubstationTestsheetPackage] = []
        matched_fls: set[str] = set()

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(total_pe_path, data_only=True, read_only=True)
                if "DataCycle1" not in wb.sheetnames:
                    wb.close()
                    return [], set()

                ws = wb["DataCycle1"]
                rows = ws.iter_rows(values_only=True)
                header = next(rows, None)
                if not header:
                    wb.close()
                    return [], set()

                pe_col = 0
                fl_col = 1
                name_col = 2
                date_col = 3
                for idx, cell in enumerate(header):
                    if cell is None:
                        continue
                    c_upper = str(cell).strip().upper()
                    if "PE" in c_upper and "NO" in c_upper:
                        pe_col = idx
                    elif "FL" in c_upper or "FUNCTIONAL" in c_upper:
                        fl_col = idx
                    elif "SUBSTATION" in c_upper or "NAME" in c_upper:
                        name_col = idx
                    elif "DATE" in c_upper:
                        date_col = idx

                fl_matches: dict[str, dict[str, Any]] = {}
                for row in rows:
                    if not row or len(row) <= max(pe_col, fl_col):
                        continue
                    fl_val = row[fl_col]
                    if not fl_val:
                        continue
                    norm_fl = normalize_functional_location_input(fl_val)
                    if norm_fl in target_fls:
                        pe_raw = row[pe_col] if pe_col < len(row) else None
                        if pe_raw is None:
                            continue
                        m = re.match(r"^(\d+)", str(pe_raw).strip())
                        if not m:
                            continue
                        pe_no = int(m.group(1))
                        if pe_no <= 0:
                            continue

                        date_raw = row[date_col] if date_col < len(row) else None
                        date_str = ""
                        if isinstance(date_raw, (datetime, date)):
                            date_str = date_raw.strftime("%d-%m-%Y")
                        elif date_raw:
                            dt_obj = to_excel_date(date_raw)
                            if dt_obj:
                                date_str = dt_obj.strftime("%d-%m-%Y")
                            elif re.match(r"^\d{2}-\d{2}-\d{4}$", str(date_raw).strip()):
                                date_str = str(date_raw).strip()

                        station = resolve_station_from_fl(norm_fl) or ""
                        fl_matches[norm_fl] = {
                            "pe_no": pe_no,
                            "date_str": date_str,
                            "station": station,
                        }
                wb.close()
        except Exception:
            return [], set()

        for norm_fl, info in fl_matches.items():
            pkg = self._find_testsheet_file(
                testsheet_root=testsheet_root,
                pe_no=info["pe_no"],
                date_str=info["date_str"],
                station=info["station"],
            )
            if pkg is not None:
                matched_packages.append(pkg)
                matched_fls.add(norm_fl)

        return matched_packages, matched_fls

    def _find_testsheet_file(
        self,
        testsheet_root: Path,
        pe_no: int,
        date_str: str,
        station: str,
    ) -> SubstationTestsheetPackage | None:
        """Find the specific testsheet file for a given PE number, date folder, and station."""
        if not testsheet_root.exists():
            return None

        candidate_date_dirs: list[Path] = []
        station_dir: Path | None = None

        if station:
            candidate = testsheet_root / station
            if candidate.is_dir():
                station_dir = candidate
            else:
                for child in testsheet_root.iterdir():
                    if child.is_dir() and child.name.upper() == station.upper():
                        station_dir = child
                        break

        if station_dir is not None:
            if date_str:
                direct_date = station_dir / date_str
                if direct_date.is_dir():
                    candidate_date_dirs.append(direct_date)
                for month_folder in sorted(station_dir.iterdir()):
                    if month_folder.is_dir():
                        d_dir = month_folder / date_str
                        if d_dir.is_dir():
                            candidate_date_dirs.append(d_dir)
            if not candidate_date_dirs:
                candidate_date_dirs = self.repository._find_date_folders(station_dir)
        else:
            if date_str:
                for p in testsheet_root.rglob(date_str):
                    if p.is_dir():
                        candidate_date_dirs.append(p)
            if not candidate_date_dirs:
                candidate_date_dirs = self.repository._find_date_folders(testsheet_root)

        for d_dir in candidate_date_dirs:
            xlsx_files = [
                f for f in sorted(d_dir.glob("*.xlsx"))
                if not f.name.startswith("~$") and not f.name.startswith("processed_")
            ]
            for f in xlsx_files:
                match = re.match(r"^(\d+)", f.name)
                if match and int(match.group(1)) == pe_no:
                    resolved_station = station or ""
                    raw_month = ""
                    for idx, part in enumerate(d_dir.parts):
                        if part.upper() in ("TESTSHEET", "RAW MATERIAL") and idx + 2 < len(d_dir.parts):
                            resolved_station = d_dir.parts[idx + 1]
                            raw_month = d_dir.parts[idx + 2]
                            break
                    if not raw_month and d_dir.parent and not d_dir.parent.name.upper().startswith("TESTSHEET"):
                        raw_month = d_dir.parent.name
                    if not resolved_station and d_dir.parent and d_dir.parent.parent and not d_dir.parent.parent.name.upper().startswith("TESTSHEET"):
                        resolved_station = d_dir.parent.parent.name

                    month = format_month_folder(raw_month) or format_month_folder(d_dir.name)
                    unsorted_dir = d_dir / "UNSORTED RAW DATA"
                    return SubstationTestsheetPackage(
                        testsheet_path=f,
                        unsorted_raw_data_dir=unsorted_dir,
                        station=resolved_station,
                        month=month,
                        date_str=d_dir.name,
                        substation_number=pe_no,
                        data=None,
                    )
        return None

    def _discover_via_station_prefix(
        self,
        testsheet_dir: Path,
        remaining_fls: set[str],
    ) -> tuple[list[SubstationTestsheetPackage], set[str]]:
        """Fallback discovery: route by station prefix, scan with eager_extract=False, and match metadata."""
        matched_packages: list[SubstationTestsheetPackage] = []
        matched_fls: set[str] = set()

        by_station: dict[str, list[str]] = {}
        for fl in remaining_fls:
            st = resolve_station_from_fl(fl) or ""
            by_station.setdefault(st, []).append(fl)

        for st, fl_list in by_station.items():
            scan_dir = testsheet_dir
            if st and testsheet_dir.exists():
                cand = testsheet_dir / st
                if cand.is_dir():
                    scan_dir = cand
                else:
                    for child in testsheet_dir.iterdir():
                        if child.is_dir() and child.name.upper() == st.upper():
                            scan_dir = child
                            break

            try:
                candidates = self.repository.discover_packages(scan_dir, eager_extract=False)
            except TypeError:
                candidates = self.repository.discover_packages(scan_dir)

            fl_set = set(fl_list)
            for cand_pkg in candidates:
                cand_fl = ""
                if cand_pkg.data is not None and cand_pkg.data.fl_erms:
                    cand_fl = normalize_functional_location_input(cand_pkg.data.fl_erms)
                else:
                    try:
                        meta = self.repository.extractor.extract_testsheet_metadata(
                            cand_pkg.testsheet_path,
                            station_hint=cand_pkg.station,
                            date_hint=cand_pkg.date_str,
                        )
                        if meta and meta.fl_erms:
                            cand_fl = normalize_functional_location_input(meta.fl_erms)
                    except Exception:
                        pass

                if cand_fl in fl_set:
                    matched_packages.append(cand_pkg)
                    matched_fls.add(cand_fl)
                    fl_set.remove(cand_fl)
                    if not fl_set:
                        break

        return matched_packages, matched_fls

    def _get_defect_repo(
        self, environment: ProjectEnvironment
    ) -> MasterQr03DefectRepository:
        """Get or initialize cached MasterQr03DefectRepository instance."""
        if self._defect_repo is None:
            self._defect_repo = MasterQr03DefectRepository(environment=environment)
        return self._defect_repo

    def extract_defects(
        self, pkg: SubstationTestsheetPackage, environment: ProjectEnvironment
    ) -> tuple[list[CbmDefectRecord], list[ViDefectRecord]]:
        """Fetch CBM and VI defects for a single substation package.

        Returns (cbm_defects, vi_defects).

        Raises FileNotFoundError if the ENGR directory or workbooks are missing.
        Raises RuntimeError if required sheets are missing or workbooks are unreadable.
        Returns empty lists only when the source is valid but no matching defect rows exist.
        """
        if not pkg.data or not pkg.data.fl_erms:
            return [], []

        defect_repo = self._get_defect_repo(environment)
        station_hint = pkg.station if isinstance(pkg.station, str) else None
        cbm_defects = defect_repo.fetch_cbm_defects(pkg.data.fl_erms, station=station_hint)
        vi_defects = defect_repo.fetch_vi_defects(pkg.data.fl_erms, station=station_hint)
        return cbm_defects, vi_defects
