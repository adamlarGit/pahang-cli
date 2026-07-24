"""Populate TOTAL PE workflow implementation for Pahang CLI."""

from __future__ import annotations

from pathlib import Path
import re
import openpyxl

from src.project.environment import ProjectEnvironment
from src.testsheet.repository import SubstationTestsheetRepository
from src.workflows.models import (
    PopulateMode,
    PopulateTotalPeRequest,
    PopulateTotalPeResult,
)


def normalize_date_str(date_input: str) -> str:
    """Normalize date strings (e.g., '01/05/2026' -> '01-05-2026')."""
    if not date_input:
        return ""
    s = str(date_input).strip().replace("/", "-")
    match = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        return f"{day:02d}-{month:02d}-{year:04d}"
    return s


class PopulateTotalPeWorkflow:
    """Scans testsheet packages and upserts PE information into TOTAL PE.xlsx."""

    def __init__(self, repository: SubstationTestsheetRepository | None = None) -> None:
        self.repository = repository or SubstationTestsheetRepository()

    def execute(
        self, environment: ProjectEnvironment, request: PopulateTotalPeRequest
    ) -> PopulateTotalPeResult:
        testsheet_dir = environment.storage.get_testsheet_dir()
        total_pe_path = environment.storage.get_total_pe_path()

        if request.progress_sink:
            request.progress_sink(f"Scanning testsheet packages in {testsheet_dir}...")

        packages = self.repository.discover_packages(testsheet_dir)

        if request.mode == PopulateMode.SPECIFIC_FOLDERS and request.target_folder_names:
            packages = [
                pkg for pkg in packages
                if any(target in pkg.date_str or target in str(pkg.testsheet_path) for target in request.target_folder_names)
            ]
        elif request.mode == PopulateMode.AUTO and total_pe_path.exists():
            wb_check = openpyxl.load_workbook(total_pe_path, data_only=True)
            try:
                ws_check = wb_check["DataCycle1"] if "DataCycle1" in wb_check.sheetnames else wb_check.active
                existing_auto_keys = set()
                for r_idx in range(2, ws_check.max_row + 1):
                    pe_val = str(ws_check.cell(r_idx, 1).value or "").strip()
                    sub_val = str(ws_check.cell(r_idx, 3).value or "").strip().upper()
                    dt_val = str(ws_check.cell(r_idx, 4).value or "").strip()
                    norm_dt = normalize_date_str(dt_val)

                    for d in (dt_val, norm_dt):
                        if not d:
                            continue
                        if pe_val:
                            existing_auto_keys.add((pe_val, d))
                            try:
                                num_int = int(pe_val)
                                existing_auto_keys.add((str(num_int), d))
                                existing_auto_keys.add((f"{num_int:03d}", d))
                            except ValueError:
                                pass
                        if sub_val:
                            existing_auto_keys.add((sub_val, d))

                packages = [
                    pkg for pkg in packages
                    if (str(pkg.pe_num), pkg.date_str) not in existing_auto_keys
                    and (str(pkg.pe_num), normalize_date_str(pkg.date_str)) not in existing_auto_keys
                    and (f"{pkg.pe_num:03d}", pkg.date_str) not in existing_auto_keys
                    and (f"{pkg.pe_num:03d}", normalize_date_str(pkg.date_str)) not in existing_auto_keys
                    and (pkg.data is None or (pkg.data.substation_name.upper(), pkg.date_str) not in existing_auto_keys)
                ]
            finally:
                wb_check.close()

        if not packages:
            if request.progress_sink:
                request.progress_sink("No testsheet packages found to process.")
            return PopulateTotalPeResult(new_rows_added=0)

        if not total_pe_path.exists():
            environment.storage.ensure_directory(total_pe_path.parent)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DataCycle1"
            ws.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
        else:
            wb = openpyxl.load_workbook(total_pe_path)
            if "DataCycle1" in wb.sheetnames:
                ws = wb["DataCycle1"]
            else:
                ws = wb.active

        new_count = 0
        updated_count = 0

        existing_rows: dict[tuple[str, str], int] = {}
        for r_idx in range(2, ws.max_row + 1):
            pe_val_raw = str(ws.cell(r_idx, 1).value or "").strip()
            sub_val = str(ws.cell(r_idx, 3).value or "").strip().upper()
            date_val = str(ws.cell(r_idx, 4).value or "").strip()
            norm_dt = normalize_date_str(date_val)

            for d in (date_val, norm_dt):
                if not d:
                    continue
                if pe_val_raw:
                    existing_rows[(pe_val_raw, d)] = r_idx
                    try:
                        pe_int = int(pe_val_raw)
                        existing_rows[(str(pe_int), d)] = r_idx
                        existing_rows[(f"{pe_int:03d}", d)] = r_idx
                    except ValueError:
                        pass
                if sub_val:
                    existing_rows[(sub_val, d)] = r_idx

        for pkg in packages:
            data = pkg.data
            if data is None:
                continue

            pe_no = data.pe_number
            fl_num = data.fl_number
            sub_name = data.substation_name
            dt_str = data.date_str or pkg.date_str
            norm_pkg_dt = normalize_date_str(dt_str)
            type_c = data.type_code
            wo = data.wo_number

            lookup_key = (str(pe_no), dt_str)
            norm_lookup_key = (str(pe_no), norm_pkg_dt)
            padded_key = (f"{pe_no:03d}", dt_str)
            norm_padded_key = (f"{pe_no:03d}", norm_pkg_dt)
            sub_key = (sub_name.upper(), dt_str)
            norm_sub_key = (sub_name.upper(), norm_pkg_dt)

            target_row = (
                existing_rows.get(lookup_key)
                or existing_rows.get(norm_lookup_key)
                or existing_rows.get(padded_key)
                or existing_rows.get(norm_padded_key)
                or existing_rows.get(sub_key)
                or existing_rows.get(norm_sub_key)
            )

            if target_row is not None:
                ws.cell(target_row, 1, pe_no)
                ws.cell(target_row, 2, fl_num)
                ws.cell(target_row, 3, sub_name)
                ws.cell(target_row, 4, dt_str)
                ws.cell(target_row, 5, type_c)
                ws.cell(target_row, 6, wo)
                updated_count += 1
            else:
                ws.append([pe_no, fl_num, sub_name, dt_str, type_c, wo, ""])
                new_row_idx = ws.max_row
                existing_rows[lookup_key] = new_row_idx
                existing_rows[sub_key] = new_row_idx
                new_count += 1

        wb.save(total_pe_path)
        wb.close()

        if request.progress_sink:
            request.progress_sink(
                f"TOTAL PE populated successfully: {new_count} new rows, {updated_count} updated."
            )

        return PopulateTotalPeResult(new_rows_added=new_count)
