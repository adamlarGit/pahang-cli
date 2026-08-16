"""MSMS Repository."""
from __future__ import annotations

import logging
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from src.core.normalizers import normalize_fl_erms
from src.msms.models import (
    MSMS_COLUMN_MAPPING,
    ConsolidateResult,
    EnrichResult,
)

logger = logging.getLogger(__name__)


def col_to_index(col_letter: str) -> int:
    """Convert Excel column letter to 0-based index for pandas iloc."""
    return column_index_from_string(col_letter) - 1


def read_col(df: pd.DataFrame, col_letter: str):
    """Read a DataFrame column by Excel column letter."""
    return df.iloc[:, col_to_index(col_letter)]


def write_cell(ws, row: int, col_letter: str, value):
    """Write to a worksheet cell by column letter."""
    ws[f"{col_letter}{row}"] = value


class MsmsRepository(ABC):
    """Abstract interface for MSMS repository."""

    @abstractmethod
    def read_data_msms(self, path: Path) -> pd.DataFrame:
        """Read DATA MSMS.xlsx master table into DataFrame."""

    @abstractmethod
    def consolidate_xls_files(
        self, xls_paths: Sequence[Path], target_data_msms: Path
    ) -> ConsolidateResult:
        """Consolidate scattered Maximo .xls files into master DATA MSMS.xlsx."""

    @abstractmethod
    def enrich_from_engr(
        self, data_msms_path: Path, total_pe_path: Path
    ) -> EnrichResult:
        """Enrich blank metadata cells in DATA MSMS.xlsx using TOTAL PE.xlsx."""

    @abstractmethod
    def get_work_order_by_fl(self, data_msms_path: Path, functional_location: str) -> str | None:
        """Lookup Work Order by Functional Location."""


class LocalExcelMsmsRepository(MsmsRepository):

    """Local filesystem implementation of MsmsRepository using openpyxl and pandas."""

    def read_data_msms(self, path: Path) -> pd.DataFrame:
        if not path.exists():
            raise FileNotFoundError(f"DATA MSMS workbook not found at '{path}'")
        return pd.read_excel(path)

    @staticmethod
    def _is_non_empty_row(row_vals: Sequence[Any]) -> bool:
        """Check if any cell in row values contains non-empty, non-sentinel data."""
        for val in row_vals:
            if val is not None:
                val_str = str(val).strip()
                if val_str and val_str.lower() not in ("none", "nan"):
                    return True
        return False

    def consolidate_xls_files(
        self, xls_paths: Sequence[Path], target_data_msms: Path
    ) -> ConsolidateResult:
        if not target_data_msms.exists():
            raise FileNotFoundError(f"Target DATA MSMS workbook not found at '{target_data_msms}'")

        if not xls_paths:
            return ConsolidateResult(files_processed=0, rows_appended=0, duplicates_skipped=0)

        wb = load_workbook(target_data_msms)
        ws = wb.active

        orig_max_row = ws.max_row or 1
        num_cols = max(ws.max_column or 7, 7)

        # Build existing work order index and collect existing non-empty rows
        existing_records: list[list[Any]] = []
        existing_wos: set[str] = set()
        for r_idx in range(2, orig_max_row + 1):
            row_vals = [ws.cell(r_idx, c).value for c in range(1, num_cols + 1)]
            if self._is_non_empty_row(row_vals):
                existing_records.append(row_vals)
                val = row_vals[0]
                if val is not None:
                    val_str = str(val).strip()
                    if val_str and val_str.lower() not in ("none", "nan"):
                        existing_wos.add(val_str)

        files_processed = 0
        rows_appended = 0
        duplicates_skipped = 0
        errors: list[str] = []
        new_records: list[list[Any]] = []

        for xls_path in xls_paths:
            if not xls_path.exists():
                errors.append(f"File not found: '{xls_path}'")
                continue

            try:
                tables = pd.read_html(str(xls_path), flavor="lxml")
            except ValueError as e:
                logger.warning(f"No tables found in '{xls_path.name}': {e}")
                errors.append(f"Failed to read HTML table in '{xls_path.name}': {e}")
                continue
            except Exception as e:
                errors.append(f"Failed to read HTML table in '{xls_path.name}': {e}")
                continue

            if not tables or tables[0].empty:
                logger.warning(f"No table data found in '{xls_path.name}'")
                continue

            df = tables[0]

            # Detect header row vs data row
            # Maximo columns: Work Order (Col 0), Status (Col 1), Location (Col 2), Description (Col 3)
            for _, row in df.iterrows():
                wo_raw = row.iloc[0] if len(row) > 0 else None
                if pd.isna(wo_raw):
                    continue
                wo_str = str(wo_raw).strip()
                if not wo_str or wo_str.upper() in ("WONUM", "WORK ORDER", "WO", "NAN", "NONE"):
                    continue

                if wo_str in existing_wos:
                    duplicates_skipped += 1
                    continue

                loc_raw = row.iloc[2] if len(row) > 2 else ""
                loc_str = str(loc_raw).strip() if not pd.isna(loc_raw) else ""
                if loc_str.upper() in ("LOCATION", "NAN", "NONE"):
                    loc_str = ""

                desc_raw = row.iloc[3] if len(row) > 3 else ""
                desc_str = str(desc_raw).strip() if not pd.isna(desc_raw) else ""
                if desc_str.upper() in ("DESCRIPTION", "NAN", "NONE"):
                    desc_str = ""

                fl_erms = normalize_fl_erms(loc_str)

                # Col A=WO, Col B=Location, Col C=Description, Col D=None, Col E=FL ERMS, Col F=None, Col G=None
                new_records.append([wo_str, loc_str, desc_str, None, fl_erms, None, None])
                existing_wos.add(wo_str)
                rows_appended += 1

            files_processed += 1

        all_records = existing_records + new_records

        # Write contiguously starting from row 2
        for r_idx, r_data in enumerate(all_records, start=2):
            for c_idx in range(1, num_cols + 1):
                val = r_data[c_idx - 1] if c_idx - 1 < len(r_data) else None
                ws.cell(row=r_idx, column=c_idx).value = val

        last_row = 1 + len(all_records)
        if orig_max_row > last_row:
            ws.delete_rows(last_row + 1, orig_max_row - last_row)

        wb.save(target_data_msms)
        wb.close()

        return ConsolidateResult(
            files_processed=files_processed,
            rows_appended=rows_appended,
            duplicates_skipped=duplicates_skipped,
            errors=tuple(errors),
        )

    def enrich_from_engr(
        self, data_msms_path: Path, total_pe_path: Path
    ) -> EnrichResult:
        if not data_msms_path.exists():
            raise FileNotFoundError(f"DATA MSMS workbook not found at '{data_msms_path}'")
        if not total_pe_path.exists():
            raise FileNotFoundError(f"TOTAL PE workbook not found at '{total_pe_path}'")

        wb_pe = load_workbook(total_pe_path, data_only=True)
        if "DataCycle1" not in wb_pe.sheetnames:
            wb_pe.close()
            raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")
        ws_pe = wb_pe["DataCycle1"]

        # Build lookup table from DataCycle1 in TOTAL PE:
        # Col 1 (A): PE NO / Substation Number
        # Col 2 (B): FL NUMBER / FL ERMS
        # Col 3 (C): SUBSTATION NAME / Substation Name ERMS
        # Col 4 (D): DATE / Cycle Date
        # Col 6 (F): WO / Work Order
        total_pe_by_wo: dict[str, dict[str, Any]] = {}
        total_pe_by_fl: dict[str, dict[str, Any]] = {}

        for r_idx in range(2, (ws_pe.max_row or 1) + 1):
            pe_no_val = ws_pe.cell(r_idx, 1).value
            fl_val = ws_pe.cell(r_idx, 2).value
            sub_name_val = ws_pe.cell(r_idx, 3).value
            date_val = ws_pe.cell(r_idx, 4).value
            wo_val = ws_pe.cell(r_idx, 6).value

            pe_data = {
                "substation_number": pe_no_val,
                "fl_erms": str(fl_val).strip() if fl_val is not None else None,
                "substation_name_erms": str(sub_name_val).strip() if sub_name_val is not None else None,
                "cycle_date": str(date_val).strip() if date_val is not None else None,
            }

            if wo_val is not None:
                wo_str = str(wo_val).strip()
                if wo_str and wo_str.lower() not in ("none", "nan"):
                    total_pe_by_wo[wo_str] = pe_data

            if fl_val is not None:
                fl_str = str(fl_val).strip().upper()
                if fl_str and fl_str.lower() not in ("none", "nan"):
                    total_pe_by_fl[fl_str] = pe_data
                    fl_norm = fl_str.replace("/", "")
                    if fl_norm:
                        total_pe_by_fl[fl_norm] = pe_data
                    fl_normalized = normalize_fl_erms(fl_str).upper()
                    if fl_normalized:
                        total_pe_by_fl[fl_normalized] = pe_data
                        fl_norm_noslash = fl_normalized.replace("/", "")
                        if fl_norm_noslash:
                            total_pe_by_fl[fl_norm_noslash] = pe_data

        wb_pe.close()

        # Update DATA MSMS.xlsx
        wb_msms = load_workbook(data_msms_path)
        ws_msms = wb_msms.active

        matched_count = 0
        unmatched_count = 0
        unmatched_wos: list[str] = []
        updated_cells_count = 0

        for r_idx in range(2, (ws_msms.max_row or 1) + 1):
            wo_val = ws_msms.cell(r_idx, 1).value
            wo_str = str(wo_val).strip() if wo_val is not None else ""
            loc_val = ws_msms.cell(r_idx, 2).value
            loc_str = str(loc_val).strip().upper() if loc_val is not None else ""
            fl_val = ws_msms.cell(r_idx, 5).value
            fl_str = str(fl_val).strip().upper() if fl_val is not None else ""

            if not wo_str and not fl_str and not loc_str:
                continue

            pe_data = None
            if wo_str and wo_str.lower() not in ("none", "nan") and wo_str in total_pe_by_wo:
                pe_data = total_pe_by_wo[wo_str]
            else:
                candidates: list[str] = []
                if fl_str and fl_str.lower() not in ("none", "nan"):
                    if fl_str not in candidates:
                        candidates.append(fl_str)
                    fl_noslash = fl_str.replace("/", "")
                    if fl_noslash and fl_noslash not in candidates:
                        candidates.append(fl_noslash)
                if loc_str and loc_str.lower() not in ("none", "nan", "location"):
                    if loc_str not in candidates:
                        candidates.append(loc_str)
                    loc_noslash = loc_str.replace("/", "")
                    if loc_noslash and loc_noslash not in candidates:
                        candidates.append(loc_noslash)
                    norm_loc = normalize_fl_erms(loc_str).upper()
                    if norm_loc and norm_loc.lower() not in ("none", "nan", "location"):
                        if norm_loc not in candidates:
                            candidates.append(norm_loc)
                        norm_loc_noslash = norm_loc.replace("/", "")
                        if norm_loc_noslash and norm_loc_noslash not in candidates:
                            candidates.append(norm_loc_noslash)

                for cand in candidates:
                    if cand in total_pe_by_fl:
                        pe_data = total_pe_by_fl[cand]
                        break

            if pe_data is not None:
                matched_count += 1


                # Col D (4): Substation Name ERMS
                c4 = ws_msms.cell(r_idx, 4)
                if (c4.value is None or str(c4.value).strip() in ("", "None", "nan")) and pe_data.get("substation_name_erms"):
                    c4.value = pe_data["substation_name_erms"]
                    updated_cells_count += 1

                # Col E (5): FL ERMS
                c5 = ws_msms.cell(r_idx, 5)
                if (c5.value is None or str(c5.value).strip() in ("", "None", "nan")) and pe_data.get("fl_erms"):
                    c5.value = pe_data["fl_erms"]
                    updated_cells_count += 1

                # Col F (6): Cycle Date
                c6 = ws_msms.cell(r_idx, 6)
                if (c6.value is None or str(c6.value).strip() in ("", "None", "nan")) and pe_data.get("cycle_date"):
                    c6.value = pe_data["cycle_date"]
                    updated_cells_count += 1

                # Col G (7): Substation Number
                c7 = ws_msms.cell(r_idx, 7)
                if (c7.value is None or str(c7.value).strip() in ("", "None", "nan")) and pe_data.get("substation_number") is not None:
                    c7.value = pe_data["substation_number"]
                    updated_cells_count += 1
            else:
                if wo_str and wo_str.lower() not in ("none", "nan"):
                    unmatched_count += 1
                    unmatched_wos.append(wo_str)

        wb_msms.save(data_msms_path)
        wb_msms.close()

        return EnrichResult(
            matched_count=matched_count,
            unmatched_count=unmatched_count,
            unmatched_wos=tuple(unmatched_wos),
            updated_cells_count=updated_cells_count,
        )

    def get_work_order_by_fl(self, data_msms_path: Path, functional_location: str) -> str | None:
        if not data_msms_path.exists():
            return None
        df = pd.read_excel(data_msms_path)
        fl_idx = col_to_index(MSMS_COLUMN_MAPPING["functional_location"])
        wo_idx = col_to_index(MSMS_COLUMN_MAPPING["wo"])
        
        fl_col = df.iloc[:, fl_idx].astype(str)
        match = df[fl_col == functional_location]
        if not match.empty:
            return str(match.iloc[0, wo_idx])
        return None


MsmsRepo = MsmsRepository

