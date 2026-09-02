"""Extractor for parsing testsheet Excel workbooks in Pahang CLI."""

from __future__ import annotations

from datetime import date, datetime
import re
import warnings
from pathlib import Path
import openpyxl

from src.core.normalizers import format_testsheet_time, format_humidity_str, parse_background_temp
from src.testsheet.models import (
    BatteryBankSpec,
    FireExtinguisherSpec,
    LVDBSpec,
    PhotoRange,
    RawPhotoRanges,
    SubstationEquipmentPackage,
    SwitchgearPanelSpec,
    SwitchgearSpec,
    TestsheetData,
    TransformerSpec,
)


def normalize_fl_erms(val: object) -> str:
    """Strip whitespace, handle .0 float suffix from FL ERMS values."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace("\t", "").strip()


def clean_val(val: object) -> str | None:
    """Strip tabs/spaces, return None if empty/dash/NONE."""
    if val is None:
        return None
    s = str(val).replace("\t", "").strip()
    if not s or s in ("-", "None", "NONE", "N/A", "#REF!", "nan"):
        return None
    return s


def is_marked(val: object) -> bool:
    """Checkbox detection — True if cell has a non-negative marker."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return s not in ("", "NONE", "NO", "N/A", "0", "FALSE", "-", "NAN")


def normalize_building_type(val: object) -> str | None:
    """Normalize building type strings to ATTACH, INDOOR, or OUTDOOR."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if not s or s in ("-", "NONE", "N/A"):
        return None
    if "ATTACH" in s:
        return "ATTACH"
    if "INDOOR" in s or "DALAMAN" in s:
        return "INDOOR"
    if "OUTDOOR" in s or "LUARAN" in s:
        return "OUTDOOR"
    return s


def to_excel_date(val: object) -> datetime | None:
    """Parse date from cell value — handles datetime, date, and string formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    s = str(val).strip()
    if not s or s in ("-", "None", "N/A"):
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%d %b %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


class TestsheetExtractor:
    """Extracts metadata and photo range bounds from Pahang testsheet workbooks."""

    def extract_photo_ranges(self, workbook_path: Path | str) -> RawPhotoRanges:
        """Extract IR and DG photo range bounds from testsheet RAW DATA sheet."""
        data = self.extract_testsheet_data(workbook_path)
        return data.photo_ranges

    def extract_testsheet_metadata(
        self,
        workbook_path: Path | str,
        station_hint: str = "",
        date_hint: str = "",
    ) -> TestsheetData:
        """Fast read-only extraction of header metadata required for TOTAL PE."""
        path = Path(workbook_path)
        if not path.exists():
            raise FileNotFoundError(f"Testsheet workbook not found: {path}")

        substation_number = 1
        num_match = re.match(r"^(\d+)", path.name)
        if num_match:
            substation_number = int(num_match.group(1))

        fl_erms = ""
        substation_name_erms = ""
        cycle_1: datetime | None = None
        date_str = date_hint
        station_name = station_hint
        wo_number = ""
        substation_type = ""

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                sheet_names = wb.sheetnames

                # 1. PCE Testsheet
                if "PCE Testsheet" in sheet_names:
                    ws_pce = wb["PCE Testsheet"]
                    fl_erms = normalize_fl_erms(ws_pce["W5"].value)
                    cleaned_sub_erms = clean_val(ws_pce["C5"].value)
                    substation_name_erms = cleaned_sub_erms if cleaned_sub_erms is not None else ""
                    cycle_1 = to_excel_date(ws_pce["P4"].value)
                    if cycle_1 is not None and not date_str:
                        date_str = cycle_1.strftime("%d-%m-%Y")

                    y1_val = ws_pce["Y1"].value
                    if y1_val is not None and str(y1_val).strip() not in ("", "-", "None", "N/A"):
                        try:
                            substation_number = int(float(str(y1_val).strip()))
                        except (ValueError, TypeError):
                            pass

                # 2. PCE VI (for substation_type)
                pce_vi_name = None
                for sname in sheet_names:
                    s_lower = sname.strip().lower()
                    if (
                        s_lower == "pce vi"
                        or s_lower.startswith("pce vi ")
                        or s_lower.startswith("pce vi(")
                        or s_lower.startswith("pce vi_")
                        or s_lower.startswith("pce vi-")
                    ):
                        pce_vi_name = sname
                        break

                if pce_vi_name is not None:
                    ws_vi = wb[pce_vi_name]
                    cleaned_type = clean_val(ws_vi["N1"].value)
                    substation_type = cleaned_type if cleaned_type is not None else ""
            finally:
                wb.close()

        return TestsheetData(
            substation_number=substation_number,
            station_name=station_name,
            date_str=date_str,
            fl_erms=fl_erms,
            wo_number=wo_number,
            substation_name_erms=substation_name_erms,
            substation_type=substation_type,
            cycle_1=cycle_1,
        )

    def extract_testsheet_data(
        self,
        workbook_path: Path | str,
        station_hint: str = "",
        date_hint: str = "",
    ) -> TestsheetData:
        """Parse testsheet workbook to extract TestsheetData and photo ranges."""
        path = Path(workbook_path)
        if not path.exists():
            raise FileNotFoundError(f"Testsheet workbook not found: {path}")

        substation_number = 1
        num_match = re.match(r"^(\d+)", path.name)
        if num_match:
            substation_number = int(num_match.group(1))

        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            fl_erms = ""
            fl_site = ""
            substation_name_erms = ""
            cycle_1: datetime | None = None

            substation_name_site = ""
            gps_coordinate = ""
            substation_type = ""
            building_type: str | None = None

            date_str = date_hint
            station_name = station_hint
            wo_number = ""

            ir_start: int | None = None
            ir_end: int | None = None
            dg_start: int | None = None
            dg_end: int | None = None
            
            ambient = "-"
            humidity = "-"
            time_str = "-"
            tev_background = "-"

            # Phase 1: PCE Testsheet (fixed cells)
            if "PCE Testsheet" in wb.sheetnames:
                ws_pce = wb["PCE Testsheet"]
                fl_erms = normalize_fl_erms(ws_pce["W5"].value)
                cleaned_sub_erms = clean_val(ws_pce["C5"].value)
                substation_name_erms = cleaned_sub_erms if cleaned_sub_erms is not None else ""
                cycle_1 = to_excel_date(ws_pce["P4"].value)
                
                time_str = format_testsheet_time(ws_pce["P5"].value)
                humidity = format_humidity_str(ws_pce["S6"].value)
                ambient = parse_background_temp(ws_pce["W6"].value)
                tev_bg_val = clean_val(ws_pce["P6"].value)
                if tev_bg_val is not None:
                    tev_background = str(tev_bg_val)

                y1_val = ws_pce["Y1"].value
                if y1_val is not None and str(y1_val).strip() not in ("", "-", "None", "N/A"):
                    try:
                        substation_number = int(float(str(y1_val).strip()))
                    except (ValueError, TypeError):
                        pass

            # Phase 2: PCE VI (fixed cells, optional)
            pce_vi_sheet = None
            for sname in wb.sheetnames:
                s_lower = sname.strip().lower()
                if (
                    s_lower == "pce vi"
                    or s_lower.startswith("pce vi ")
                    or s_lower.startswith("pce vi(")
                    or s_lower.startswith("pce vi_")
                    or s_lower.startswith("pce vi-")
                ):
                    pce_vi_sheet = wb[sname]
                    break

            if pce_vi_sheet is not None:
                ws_vi = pce_vi_sheet
                cleaned_sub_site = clean_val(ws_vi["C7"].value)
                substation_name_site = cleaned_sub_site if cleaned_sub_site is not None else ""
                cleaned_gps = clean_val(ws_vi["C8"].value)
                gps_coordinate = cleaned_gps if cleaned_gps is not None else ""
                cleaned_type = clean_val(ws_vi["N1"].value)
                substation_type = cleaned_type if cleaned_type is not None else ""
                cleaned_fl_site = clean_val(ws_vi["K7"].value)
                fl_site = cleaned_fl_site if cleaned_fl_site is not None else ""

                if is_marked(ws_vi["D9"].value):
                    building_type = normalize_building_type(ws_vi["C9"].value)
                elif is_marked(ws_vi["G9"].value):
                    building_type = normalize_building_type(ws_vi["F9"].value)
                elif is_marked(ws_vi["I9"].value):
                    building_type = normalize_building_type(ws_vi["H9"].value)
                elif is_marked(ws_vi["K9"].value):
                    building_type = normalize_building_type(ws_vi["J9"].value)
                elif is_marked(ws_vi["M9"].value):
                    building_type = normalize_building_type(ws_vi["L9"].value)
                elif is_marked(ws_vi["O9"].value):
                    o_val = ws_vi["P9"].value if ws_vi["P9"].value is not None else ws_vi["N9"].value
                    building_type = normalize_building_type(o_val)

            # Phase 3: RAW DATA sheet (photo range extraction)
            if "RAW DATA" in wb.sheetnames:
                ws_raw = wb["RAW DATA"]
                for row in ws_raw.iter_rows(values_only=True):
                    row_cells = [str(c).strip() if c is not None else "" for c in row]
                    if not row_cells:
                        continue

                    # 1. Grid table schema (e.g. Row 1: [None, 'START', 'END'], Row 2: ['IR', 49, 66], Row 3: ['DG', 1715, 1739])
                    first_cell = row_cells[0].upper()
                    if first_cell in ("IR", "FLIR") and len(row_cells) >= 3:
                        if ir_start is None and row_cells[1]:
                            ir_start = self._parse_int_safe(row_cells[1])
                        if ir_end is None and row_cells[2]:
                            ir_end = self._parse_int_safe(row_cells[2])
                    elif first_cell in ("DG", "IMG") and len(row_cells) >= 3:
                        if dg_start is None and row_cells[1]:
                            dg_start = self._parse_int_safe(row_cells[1])
                        if dg_end is None and row_cells[2]:
                            dg_end = self._parse_int_safe(row_cells[2])

                    # 2. Label-based schema (e.g. ['IR START', 10, 'IR END', 12])
                    for idx, text in enumerate(row_cells):
                        text_upper = text.upper()

                        if "IR START" in text_upper or "FLIR START" in text_upper:
                            val = self._find_next_val(row_cells, idx)
                            parsed = self._parse_int_safe(val)
                            if parsed is not None:
                                ir_start = parsed
                        elif "IR END" in text_upper or "FLIR END" in text_upper:
                            val = self._find_next_val(row_cells, idx)
                            parsed = self._parse_int_safe(val)
                            if parsed is not None:
                                ir_end = parsed

                        elif "DG START" in text_upper or "IMG START" in text_upper:
                            val = self._find_next_val(row_cells, idx)
                            parsed = self._parse_int_safe(val)
                            if parsed is not None:
                                dg_start = parsed
                        elif "DG END" in text_upper or "IMG END" in text_upper:
                            val = self._find_next_val(row_cells, idx)
                            parsed = self._parse_int_safe(val)
                            if parsed is not None:
                                dg_end = parsed

                        elif "IR RANGE" in text_upper or "FLIR RANGE" in text_upper or "IR PHOTO" in text_upper:
                            val = self._find_next_val(row_cells, idx)
                            parsed_start, parsed_end = self._parse_range_val(val)
                            if ir_start is None:
                                ir_start = parsed_start
                            if ir_end is None:
                                ir_end = parsed_end
                        elif "DG RANGE" in text_upper or "IMG RANGE" in text_upper or "DG PHOTO" in text_upper:
                            val = self._find_next_val(row_cells, idx)
                            parsed_start, parsed_end = self._parse_range_val(val)
                            if dg_start is None:
                                dg_start = parsed_start
                            if dg_end is None:
                                dg_end = parsed_end

            # Phase 4: Substation Equipment Package Extraction
            pce_sheets: list[openpyxl.worksheet.worksheet.Worksheet] = []
            for sname in wb.sheetnames:
                s_lower = sname.strip().lower()
                if s_lower == "pce testsheet" or (s_lower.startswith("pce testsheet") and "(" in s_lower):
                    pce_sheets.append(wb[sname])
            if not pce_sheets and "PCE Testsheet" in wb.sheetnames:
                pce_sheets.append(wb["PCE Testsheet"])

            pce_sheets = self._sort_pce_sheets(pce_sheets)
            first_pce = pce_sheets[0] if pce_sheets else (wb["PCE Testsheet"] if "PCE Testsheet" in wb.sheetnames else None)

            switchgears = self._extract_switchgear_specs(wb, ws_vi=pce_vi_sheet, pce_sheets=pce_sheets)
            transformers = self._extract_transformer_specs(ws_vi=pce_vi_sheet, ws_pce=first_pce)
            lvdb_specs = self._extract_lvdb_specs(ws_pce=first_pce)
            battery_banks = self._extract_battery_banks(ws_pce=first_pce)
            fire_extinguisher = self._extract_fire_extinguisher_spec(ws_vi=pce_vi_sheet, building_type=building_type)
            has_battery_charger, has_rtu, has_sf6, has_efi = self._extract_auxiliary_flags(
                ws_vi=pce_vi_sheet,
                ws_pce=first_pce,
                switchgears=switchgears,
                battery_banks=battery_banks,
                substation_type=substation_type,
                building_type=building_type,
            )

            equipment = SubstationEquipmentPackage(
                switchgears=switchgears,
                transformers=transformers,
                lvdb_specs=lvdb_specs,
                battery_banks=battery_banks,
                fire_extinguisher=fire_extinguisher,
                has_battery_charger=has_battery_charger,
                has_rtu=has_rtu,
                has_sf6=has_sf6,
                has_efi=has_efi,
            )

        finally:
            wb.close()

        photo_ranges = RawPhotoRanges(
            ir=PhotoRange(start_num=ir_start, end_num=ir_end),
            dg=PhotoRange(start_num=dg_start, end_num=dg_end),
        )

        return TestsheetData(
            substation_number=substation_number,
            station_name=station_name,
            date_str=date_str,
            fl_erms=fl_erms,
            fl_site=fl_site,
            wo_number=wo_number,
            photo_ranges=photo_ranges,
            substation_name_erms=substation_name_erms,
            substation_name_site=substation_name_site,
            gps_coordinate=gps_coordinate,
            substation_type=substation_type,
            building_type=building_type,
            ambient=ambient,
            humidity=humidity,
            time=time_str,
            tev_background=tev_background,
            cycle_1=cycle_1,
            equipment=equipment,
        )

    def _sort_pce_sheets(
        self, sheets: list[openpyxl.worksheet.worksheet.Worksheet]
    ) -> list[openpyxl.worksheet.worksheet.Worksheet]:
        """Sort PCE Testsheet worksheets in natural sequence order."""
        def sheet_key(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
            m = re.search(r"\((\d+)\)", ws.title)
            if m:
                return int(m.group(1))
            return 1

        return sorted(sheets, key=sheet_key)

    def _extract_switchgear_type(
        self, ws_vi: openpyxl.worksheet.worksheet.Worksheet, row: int
    ) -> str:
        """Extract switchgear type from checkboxes on PCE VI for the specified row (11 or 14)."""
        check_symbols = {"/", "X", "V", "YES", "TRUE", "1", "OK", "CHECK", "TICK"}

        # RMU OIL: label C (3), check D (4) / E (5)
        c_val = str(ws_vi.cell(row, 3).value or "").strip().upper()
        d_val = ws_vi.cell(row, 4).value
        e_val = ws_vi.cell(row, 5).value
        if is_marked(d_val) or is_marked(e_val) or c_val in check_symbols:
            return "RMU OIL"

        # RMU SF6: label F (6), check G (7)
        f_val = str(ws_vi.cell(row, 6).value or "").strip().upper()
        g_val = ws_vi.cell(row, 7).value
        if is_marked(g_val) or f_val in check_symbols:
            return "RMU SF6"

        # MRMU: label H (8), check I (9)
        h_val = str(ws_vi.cell(row, 8).value or "").strip().upper()
        i_val = ws_vi.cell(row, 9).value
        if is_marked(i_val) or h_val in check_symbols:
            return "MRMU"

        # VCB: label J (10), check K (11)
        j_val = str(ws_vi.cell(row, 10).value or "").strip().upper()
        k_val = ws_vi.cell(row, 11).value
        if is_marked(k_val) or j_val in check_symbols:
            return "VCB"

        # OCB: label L (12), check M (13)
        l_val = str(ws_vi.cell(row, 12).value or "").strip().upper()
        m_val = ws_vi.cell(row, 13).value
        if is_marked(m_val) or l_val in check_symbols:
            return "OCB"

        # OTHER: label N (14), check O (15) / P (16)
        n_val = str(ws_vi.cell(row, 14).value or "").strip().upper()
        o_val = ws_vi.cell(row, 15).value
        p_val = ws_vi.cell(row, 16).value
        if is_marked(o_val) or is_marked(p_val) or n_val in check_symbols:
            custom_p = clean_val(p_val)
            custom_o = clean_val(o_val)
            if custom_p and custom_p.upper() not in check_symbols and custom_p.upper() != "OTHER":
                return custom_p
            if custom_o and custom_o.upper() not in check_symbols and custom_o.upper() != "OTHER":
                return custom_o
            return "OTHER"

        return ""

    def _extract_panels(
        self, pce_sheets: list[openpyxl.worksheet.worksheet.Worksheet]
    ) -> tuple[SwitchgearPanelSpec, ...]:
        """Extract switchgear panels across all PCE Testsheet worksheets."""
        panels: list[SwitchgearPanelSpec] = []
        panel_idx = 1

        for ws in pce_sheets:
            for r in (10, 14, 18, 22):
                feeder_no = clean_val(ws[f"B{r}"].value) or ""
                name = clean_val(ws[f"C{r}"].value) or ""
                serial_no = clean_val(ws[f"I{r}"].value) or ""

                # Exclude slots where name, panel_feeder_no, and serial_no are all blank
                if not (feeder_no or name or serial_no):
                    continue

                status = clean_val(ws[f"E{r}"].value) or clean_val(ws[f"D{r}"].value) or ""
                load_amp = clean_val(ws[f"F{r}"].value) or ""
                cable_type = clean_val(ws[f"G{r}"].value) or ""
                heater_amp = clean_val(ws[f"H{r}"].value) or ""
                panel_type = clean_val(ws[f"J{r}"].value) or ""

                # Ultrasound (Col Q / Col S) and TEV (Col T / Col U / Col V)
                us_reading = clean_val(ws[f"Q{r}"].value) or ""
                us_char = clean_val(ws[f"S{r}"].value) or ""
                tev_reading = clean_val(ws[f"T{r}"].value) or ""
                tev_ppc = clean_val(ws[f"U{r}"].value) or ""
                tev_char = clean_val(ws[f"V{r}"].value) or ""

                # Fallback to sub-rows (e.g. Breaker, Top Panel, PT) if top row empty
                if not us_reading:
                    for sub_r in range(r, r + 4):
                        if sub_r <= ws.max_row:
                            val = clean_val(ws[f"Q{sub_r}"].value)
                            if val:
                                us_reading = val
                                break
                if not us_char:
                    for sub_r in range(r, r + 4):
                        if sub_r <= ws.max_row:
                            val = clean_val(ws[f"S{sub_r}"].value)
                            if val:
                                us_char = val
                                break
                if not tev_reading:
                    for sub_r in range(r, r + 4):
                        if sub_r <= ws.max_row:
                            val = clean_val(ws[f"T{sub_r}"].value)
                            if val:
                                tev_reading = val
                                break
                if not tev_ppc:
                    for sub_r in range(r, r + 4):
                        if sub_r <= ws.max_row:
                            val = clean_val(ws[f"U{sub_r}"].value)
                            if val:
                                tev_ppc = val
                                break
                if not tev_char:
                    for sub_r in range(r, r + 4):
                        if sub_r <= ws.max_row:
                            val = clean_val(ws[f"V{sub_r}"].value)
                            if val:
                                tev_char = val
                                break

                panels.append(
                    SwitchgearPanelSpec(
                        panel_no=panel_idx,
                        panel_feeder_no=feeder_no,
                        name=name,
                        panel_type=panel_type,
                        serial_no=serial_no,
                        status=status,
                        load_amp=load_amp,
                        cable_type=cable_type,
                        heater_amp=heater_amp,
                        us_reading=us_reading,
                        us_char=us_char,
                        tev_reading=tev_reading,
                        tev_ppc=tev_ppc,
                        tev_char=tev_char,
                    )
                )
                panel_idx += 1

        return tuple(panels)

    def _extract_switchgear_specs(
        self,
        wb: openpyxl.Workbook,
        ws_vi: openpyxl.worksheet.worksheet.Worksheet | None,
        pce_sheets: list[openpyxl.worksheet.worksheet.Worksheet],
    ) -> tuple[SwitchgearSpec, ...]:
        """Extract switchgear specifications and attached panels."""
        panels = self._extract_panels(pce_sheets)

        sg1_type = ""
        sg1_mfg = ""
        sg1_model = ""
        sg1_year = ""
        sg1_rating = ""
        sg1_serial = ""

        sg2_type = ""
        sg2_mfg = ""
        sg2_model = ""
        sg2_year = ""
        sg2_rating = ""
        sg2_serial = ""

        if ws_vi is not None:
            # Switchgear 1 (Rows 11-13)
            sg1_type = self._extract_switchgear_type(ws_vi, 11)
            sg1_mfg = clean_val(ws_vi["C12"].value) or ""
            sg1_model = clean_val(ws_vi["G12"].value) or clean_val(ws_vi["F12"].value) or ""
            if sg1_model in ("Manufacturer/Model:", "Manufacturer/Modal:"):
                sg1_model = ""
            sg1_year = clean_val(ws_vi["C13"].value) or ""
            sg1_rating = clean_val(ws_vi["J13"].value) or clean_val(ws_vi["I13"].value) or ""
            if not sg1_rating and clean_val(ws_vi["H13"].value) not in (None, "Switchgear Rating :"):
                sg1_rating = clean_val(ws_vi["H13"].value) or ""
            sg1_serial = clean_val(ws_vi["O13"].value) or clean_val(ws_vi["N13"].value) or ""
            if not sg1_serial and clean_val(ws_vi["M13"].value) not in (None, "RMU SF6 / OIL S/N :"):
                sg1_serial = clean_val(ws_vi["M13"].value) or ""

            # Switchgear 2 (Rows 14-16)
            sg2_type = self._extract_switchgear_type(ws_vi, 14)
            sg2_mfg = clean_val(ws_vi["C15"].value) or ""
            sg2_model = clean_val(ws_vi["G15"].value) or clean_val(ws_vi["F15"].value) or ""
            if sg2_model in ("Manufacturer/Model:", "Manufacturer/Modal:"):
                sg2_model = ""
            sg2_year = clean_val(ws_vi["C16"].value) or ""
            sg2_rating = clean_val(ws_vi["J16"].value) or clean_val(ws_vi["I16"].value) or ""
            if not sg2_rating and clean_val(ws_vi["H16"].value) not in (None, "Switchgear Rating :"):
                sg2_rating = clean_val(ws_vi["H16"].value) or ""
            sg2_serial = clean_val(ws_vi["O16"].value) or clean_val(ws_vi["N16"].value) or ""
            if not sg2_serial and clean_val(ws_vi["M16"].value) not in (None, "RMU SF6 / OIL S/N :"):
                sg2_serial = clean_val(ws_vi["M16"].value) or ""

        sg1_active = bool(sg1_type or sg1_mfg or sg1_model or sg1_year or sg1_rating or sg1_serial or panels)
        sg2_active = bool(sg2_type or sg2_mfg or sg2_model or sg2_year or sg2_rating or sg2_serial)

        result: list[SwitchgearSpec] = []
        if sg1_active:
            result.append(
                SwitchgearSpec(
                    switchgear_type=sg1_type,
                    manufacturer=sg1_mfg,
                    model=sg1_model,
                    manufactured_year=sg1_year,
                    rating=sg1_rating,
                    serial_no=sg1_serial,
                    panels=panels,
                )
            )
        if sg2_active:
            result.append(
                SwitchgearSpec(
                    switchgear_type=sg2_type,
                    manufacturer=sg2_mfg,
                    model=sg2_model,
                    manufactured_year=sg2_year,
                    rating=sg2_rating,
                    serial_no=sg2_serial,
                    panels=(),
                )
            )

        return tuple(result)

    def _extract_transformer_specs(
        self,
        ws_vi: openpyxl.worksheet.worksheet.Worksheet | None,
        ws_pce: openpyxl.worksheet.worksheet.Worksheet | None = None,
    ) -> tuple[TransformerSpec, ...]:
        """Extract transformer specifications from PCE VI rows 17-21 and PCE Testsheet."""
        if ws_vi is None:
            return ()

        c17_val = ws_vi["C17"].value
        if c17_val is None:
            return ()

        c17_str = str(c17_val).strip().upper()
        if not c17_str or c17_str in ("0", "0.0", "NONE", "NO", "N/A", "-", "NOT ACCESSIBLE", "NAN"):
            return ()

        match = re.search(r"(\d+)", c17_str)
        if not match:
            return ()

        n_tx = int(match.group(1))
        if n_tx <= 0:
            return ()
        n_tx = min(n_tx, 4)

        transformers: list[TransformerSpec] = []
        for i in range(1, n_tx + 1):
            r = 17 + i  # Rows 18, 19, 20, 21
            tx_type = clean_val(ws_vi[f"D{r}"].value) or clean_val(ws_vi[f"C{r}"].value) or ""
            rating_kva = clean_val(ws_vi[f"F{r}"].value) or clean_val(ws_vi[f"E{r}"].value) or ""
            const_year = clean_val(ws_vi[f"I{r}"].value) or clean_val(ws_vi[f"H{r}"].value) or clean_val(ws_vi[f"G{r}"].value) or ""
            mfg = clean_val(ws_vi[f"L{r}"].value) or clean_val(ws_vi[f"K{r}"].value) or clean_val(ws_vi[f"J{r}"].value) or ""
            serial_no = clean_val(ws_vi[f"O{r}"].value) or clean_val(ws_vi[f"N{r}"].value) or ""

            # Filter out static header labels if accidentally captured
            if tx_type.upper() == "TYPE":
                tx_type = ""
            if rating_kva.upper() == "KVA":
                rating_kva = ""
            if const_year.upper() in ("CONSTRUCTION YR:", "CONSTRUCTION YR", "CONSTRUCTION YR.", "CONSTRUCTION YEAR"):
                const_year = ""
            if mfg.upper() in ("MANUFACTURER:", "MANUFACTURER"):
                mfg = ""
            if serial_no.upper() in ("S/N :", "S/N:", "S/N"):
                serial_no = ""

            # Exclusions: NOT ACCESSIBLE anywhere or completely empty row
            row_strs = [str(ws_vi.cell(r, c).value or "").strip().upper() for c in range(1, 17)]
            if any("NOT ACCESSIBLE" in s for s in row_strs):
                continue
            if not (tx_type or rating_kva or const_year or mfg or serial_no):
                continue

            # Extract ultrasound measurements from PCE Testsheet (TX1/TX2: Col K/L; TX3/TX4: Col V/X)
            tx_us_reading = ""
            tx_us_char = ""
            if ws_pce is not None:
                if i in (1, 2):
                    start_r = 33 if i == 1 else 38
                    col_db = 11  # K
                    col_char = 12  # L
                else:
                    start_r = 33 if i == 3 else 38
                    col_db = 22  # V
                    col_char = 24  # X

                for row_idx in range(start_r, start_r + 5):
                    if row_idx <= ws_pce.max_row:
                        db_val = clean_val(ws_pce.cell(row_idx, col_db).value)
                        if db_val and not tx_us_reading:
                            tx_us_reading = db_val
                        char_val = clean_val(ws_pce.cell(row_idx, col_char).value)
                        if char_val and not tx_us_char:
                            tx_us_char = char_val

            transformers.append(
                TransformerSpec(
                    tx_id=f"Tx {i}",
                    rating_kva=rating_kva,
                    construction_year=const_year,
                    manufacturer=mfg,
                    serial_no=serial_no,
                    type=tx_type,
                    us_reading=tx_us_reading,
                    us_char=tx_us_char,
                )
            )

        return tuple(transformers)

    def _extract_lvdb_specs(
        self, ws_pce: openpyxl.worksheet.worksheet.Worksheet | None
    ) -> tuple[LVDBSpec, ...]:
        """Extract LVDB and Feeder Pillar specifications from PCE Testsheet."""
        if ws_pce is None:
            return ()

        lvdb_specs: list[LVDBSpec] = []

        # Slot 1: Rows 48-51
        label1_raw = clean_val(ws_pce["R48"].value) or ""
        source1_raw = clean_val(ws_pce["T48"].value) or ""
        photo1 = ws_pce["S49"].value
        mfg1 = clean_val(ws_pce["V49"].value) or clean_val(ws_pce["U49"].value) or ""
        if not mfg1 and clean_val(ws_pce["T49"].value) not in (None, "Manufacturer :", "Manufacturer:"):
            mfg1 = clean_val(ws_pce["T49"].value) or ""
        sn1 = clean_val(ws_pce["V50"].value) or clean_val(ws_pce["U50"].value) or ""
        if not sn1 and clean_val(ws_pce["T50"].value) not in (None, "Serial No. :", "Serial No.:", "Serial No"):
            sn1 = clean_val(ws_pce["T50"].value) or ""
        rating1 = clean_val(ws_pce["V51"].value) or clean_val(ws_pce["U51"].value) or ""
        if not rating1 and clean_val(ws_pce["R51"].value) not in (None, "Rating :", "Rating:"):
            rating1 = clean_val(ws_pce["R51"].value) or ""

        photo1_active = photo1 is not None and str(photo1).strip() not in ("", "-", "None", "nan", "N/A")
        slot1_active = bool(photo1_active or mfg1 or sn1 or rating1)

        if slot1_active:
            label1 = "FP" if "FP" in label1_raw.upper() else "LVDB"
            source1 = source1_raw if source1_raw else "TX1"
            name1 = f"{label1} 1"
            lvdb_specs.append(
                LVDBSpec(
                    name=name1,
                    label=label1,
                    source=source1,
                    manufacturer=mfg1,
                    serial_no=sn1,
                    rating=rating1,
                )
            )

        # Slot 2: Rows 52-55
        label2_raw = clean_val(ws_pce["R52"].value) or ""
        source2_raw = clean_val(ws_pce["T52"].value) or ""
        photo2 = ws_pce["S53"].value
        mfg2 = clean_val(ws_pce["V53"].value) or clean_val(ws_pce["U53"].value) or ""
        if not mfg2 and clean_val(ws_pce["T53"].value) not in (None, "Manufacturer :", "Manufacturer:"):
            mfg2 = clean_val(ws_pce["T53"].value) or ""
        sn2 = clean_val(ws_pce["V54"].value) or clean_val(ws_pce["U54"].value) or ""
        if not sn2 and clean_val(ws_pce["R54"].value) not in (None, "Serial No. :", "Serial No.:", "Serial No") and clean_val(ws_pce["T54"].value) not in (None, "Serial No. :"):
            sn2 = clean_val(ws_pce["R54"].value) or clean_val(ws_pce["T54"].value) or ""
        rating2 = clean_val(ws_pce["V55"].value) or clean_val(ws_pce["U55"].value) or ""
        if not rating2 and clean_val(ws_pce["R55"].value) not in (None, "Rating :", "Rating:"):
            rating2 = clean_val(ws_pce["R55"].value) or ""

        photo2_active = photo2 is not None and str(photo2).strip() not in ("", "-", "None", "nan", "N/A")
        slot2_active = bool(photo2_active or mfg2 or sn2 or rating2)

        if slot2_active:
            label2 = "FP" if "FP" in label2_raw.upper() else "LVDB"
            source2 = source2_raw if source2_raw else "TX2"
            name2 = f"{label2} 2"
            lvdb_specs.append(
                LVDBSpec(
                    name=name2,
                    label=label2,
                    source=source2,
                    manufacturer=mfg2,
                    serial_no=sn2,
                    rating=rating2,
                )
            )

        return tuple(lvdb_specs)

    def _extract_battery_banks(
        self, ws_pce: openpyxl.worksheet.worksheet.Worksheet | None
    ) -> tuple[BatteryBankSpec, ...]:
        """Extract battery bank and charger specifications from PCE Testsheet rows 59-65."""
        if ws_pce is None:
            return ()

        battery_banks: list[BatteryBankSpec] = []
        for r in range(59, 66):
            col_b = clean_val(ws_pce[f"B{r}"].value)
            if not col_b:
                continue
            col_b_upper = col_b.upper()
            if any(k in col_b_upper for k in ("BATTERY", "BATERI", "CHARGER")):
                mfg = clean_val(ws_pce[f"J{r}"].value) or ""
                model = clean_val(ws_pce[f"K{r}"].value) or ""
                sn = clean_val(ws_pce[f"L{r}"].value) or ""

                if mfg or model or sn or clean_val(ws_pce[f"C{r}"].value) or clean_val(ws_pce[f"E{r}"].value) or clean_val(ws_pce[f"F{r}"].value):
                    battery_banks.append(
                        BatteryBankSpec(
                            name=col_b.upper(),
                            manufacturer=mfg,
                            model=model,
                            serial_no=sn,
                        )
                    )

        return tuple(battery_banks)

    def _extract_fire_extinguisher_spec(
        self,
        ws_vi: openpyxl.worksheet.worksheet.Worksheet | None,
        building_type: str | None,
    ) -> FireExtinguisherSpec:
        """Extract fire extinguisher condition and expiry date."""
        b_type_norm = normalize_building_type(building_type)
        if b_type_norm == "OUTDOOR":
            return FireExtinguisherSpec(has_fire_extinguisher=False, expiry_date="", status="")

        status = ""
        expiry_date = ""
        is_valid = False
        is_expired = False

        if ws_vi is not None:
            is_valid = is_marked(ws_vi["D42"].value) or is_marked(ws_vi["E42"].value)
            is_expired = is_marked(ws_vi["F42"].value) or is_marked(ws_vi["G42"].value)
            if is_valid:
                status = "VALID"
            elif is_expired:
                status = "EXPIRED"

            expiry_raw = ws_vi["J42"].value if ws_vi["J42"].value is not None else ws_vi["I42"].value
            if isinstance(expiry_raw, (datetime, date)):
                expiry_date = expiry_raw.strftime("%d/%m/%Y")
            else:
                cleaned_expiry = clean_val(expiry_raw)
                if cleaned_expiry and cleaned_expiry != "Expiry Date:":
                    expiry_date = cleaned_expiry

        has_fe = bool(
            b_type_norm in ("INDOOR", "ATTACH", "COMPACT")
            or is_valid
            or is_expired
            or expiry_date
        )

        if not has_fe:
            return FireExtinguisherSpec(has_fire_extinguisher=False, expiry_date="", status="")

        return FireExtinguisherSpec(
            has_fire_extinguisher=True,
            expiry_date=expiry_date,
            status=status,
        )

    def _extract_auxiliary_flags(
        self,
        ws_vi: openpyxl.worksheet.worksheet.Worksheet | None,
        ws_pce: openpyxl.worksheet.worksheet.Worksheet | None,
        switchgears: tuple[SwitchgearSpec, ...] = (),
        battery_banks: tuple[BatteryBankSpec, ...] = (),
        substation_type: str = "",
        building_type: str | None = None,
    ) -> tuple[bool, bool, bool, bool]:
        """Extract auxiliary flags (has_battery_charger, has_rtu, has_sf6, has_efi)."""
        has_efi = False
        has_sf6 = False
        has_rtu = False
        has_battery_charger = False

        # 1. EFI: Row 25 on PCE VI
        # Marked Good (Col E) or Not Good (Col G) indicates EFI exists on site.
        # Unavailable only if unchecked or remarks (Col I) contains "N/A" or "MISSING".
        if ws_vi is not None:
            e25 = ws_vi["E25"].value
            g25 = ws_vi["G25"].value
            i25 = str(ws_vi["I25"].value or "").strip().upper()
            d25 = ws_vi["D25"].value
            f25 = ws_vi["F25"].value

            d25_marked = d25 is not None and str(d25).strip().upper() in ("/", "X", "V", "✅", "TRUE", "1", "OK", "CHECK", "TICK")
            f25_marked = f25 is not None and str(f25).strip().upper() in ("/", "X", "V", "✅", "TRUE", "1", "OK", "CHECK", "TICK")

            if (is_marked(e25) or is_marked(g25) or d25_marked or f25_marked) and not any(k in i25 for k in ("N/A", "MISSING")):
                has_efi = True

        # 2. SF6: Row 29 on PCE VI or Switchgear type is SF6/MRMU
        for sg in switchgears:
            sg_upper = sg.switchgear_type.upper()
            if "SF6" in sg_upper or "MRMU" in sg_upper:
                has_sf6 = True
                break
        if not has_sf6 and ws_vi is not None:
            e29 = ws_vi["E29"].value
            g29 = ws_vi["G29"].value
            i29 = str(ws_vi["I29"].value or "").strip().upper()
            d29 = ws_vi["D29"].value
            f29 = ws_vi["F29"].value

            d29_marked = d29 is not None and str(d29).strip().upper() in ("/", "X", "V", "✅", "TRUE", "1", "OK", "CHECK", "TICK")
            f29_marked = f29 is not None and str(f29).strip().upper() in ("/", "X", "V", "✅", "TRUE", "1", "OK", "CHECK", "TICK")

            if (is_marked(e29) or is_marked(g29) or d29_marked or f29_marked) and not any(k in i29 for k in ("N/A", "MISSING")):
                has_sf6 = True

        # 3. Battery Charger: battery_banks exist or PCE VI Row 39 (Good/Not Good checkmark, not N/A)
        if len(battery_banks) > 0:
            has_battery_charger = True
        elif ws_vi is not None:
            e39 = ws_vi["E39"].value
            g39 = ws_vi["G39"].value
            i39 = str(ws_vi["I39"].value or "").strip().upper()
            d39 = ws_vi["D39"].value
            f39 = ws_vi["F39"].value

            d39_marked = d39 is not None and str(d39).strip().upper() in ("/", "X", "V", "✅", "TRUE", "1", "OK", "CHECK", "TICK")
            f39_marked = f39 is not None and str(f39).strip().upper() in ("/", "X", "V", "✅", "TRUE", "1", "OK", "CHECK", "TICK")

            if (is_marked(e39) or is_marked(g39) or d39_marked or f39_marked) and not any(k in i39 for k in ("N/A", "MISSING")):
                has_battery_charger = True

        # 4. RTU: If battery bank / charger is present, assume RTU is present,
        # unless substation is COMPACT / CS.
        b_type_norm = normalize_building_type(building_type)
        sub_type_upper = (substation_type or "").strip().upper()
        is_compact_or_cs = (
            b_type_norm == "COMPACT"
            or sub_type_upper == "CS"
            or "COMPACT" in sub_type_upper
            or "CS " in sub_type_upper
        )
        if has_battery_charger and not is_compact_or_cs:
            has_rtu = True

        return has_battery_charger, has_rtu, has_sf6, has_efi

    def _find_next_val(self, cells: list[str], current_idx: int) -> str:
        """Find the next non-empty string cell value in row after current_idx."""
        for i in range(current_idx + 1, len(cells)):
            val = cells[i].strip()
            if val:
                return val
        return ""

    def _parse_int_safe(self, val: object) -> int | None:
        """Safely convert a cell value or string into an integer."""
        if val is None or val == "":
            return None
        match = re.search(r"(\d+)", str(val))
        if match:
            return int(match.group(1))
        return None

    def _parse_range_val(self, val: str) -> tuple[int | None, int | None]:
        """Parse start and end photo integers from range string."""
        if not val:
            return None, None
        nums = [int(n) for n in re.findall(r"\d+", val)]
        if len(nums) >= 2:
            return nums[0], nums[1]
        if len(nums) == 1:
            return nums[0], nums[0]
        return None, None

