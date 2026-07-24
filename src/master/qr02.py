"""QR02 CBA repository and unit of work implementations for Pahang CLI."""

from __future__ import annotations

from abc import ABC, abstractmethod
from collections.abc import Callable, Sequence
from datetime import datetime
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.project.storage import WorkspaceStorage
from src.testsheet.extractor import normalize_building_type, normalize_fl_erms, to_excel_date


def _get_field(rec: Any, *names: str, default: Any = None) -> Any:
    """Helper to extract value from object attribute or dictionary key."""
    for name in names:
        if isinstance(rec, dict):
            if name in rec and rec[name] is not None and rec[name] != "":
                return rec[name]
        else:
            val = getattr(rec, name, None)
            if val is not None and val != "":
                return val
    return default


def _fuzzy_normalize_name(name: str | None) -> str:
    """Fuzzy normalize substation name by removing prefixes, NO., and non-alphanumeric chars."""
    if not name:
        return ""
    s = str(name).strip().upper()
    raw_alphanum = re.sub(r"[^A-Z0-9]", "", s)

    # Remove numerical prefixes like "001." or "1-"
    s = re.sub(r"^\d+[\.\s_-]*", "", s)

    # Strip prefixes PE/PDT/P-E/SSU
    prev = None
    while s != prev:
        prev = s
        s = re.sub(r"^(?:P-E|PE|PDT|SSU)\b\s*", "", s)

    # Strip NO. or NO
    s = re.sub(r"\bNO\.?\b", "", s)

    # Remove non-alphanumeric characters
    cleaned = re.sub(r"[^A-Z0-9]", "", s)
    return cleaned if cleaned else raw_alphanum


class Qr02Transaction(ABC):
    """Abstract unit of work for QR02 CBA operations."""

    @abstractmethod
    def __enter__(self) -> Qr02Transaction:
        ...

    @abstractmethod
    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: Any | None,
    ) -> bool | None:
        ...

    @abstractmethod
    def upsert_qr02_cba_records(self, records: Sequence[Any]) -> int:
        ...


class Qr02Repository(ABC):
    """Abstract repository for QR02 CBA workbook management."""

    @abstractmethod
    def transaction(self) -> Qr02Transaction:
        ...


class LocalExcelQr02Transaction(Qr02Transaction):
    """Excel-backed transaction for QR02 CBA workbook updates."""

    def __init__(
        self,
        cba_path: Path | str,
        on_commit: Callable[[], None] | None = None,
    ) -> None:
        self.cba_path = Path(cba_path)
        self.on_commit = on_commit
        self.wb: openpyxl.Workbook | None = None
        self.ws: Worksheet | None = None
        self.fl_to_row: dict[str, int] = {}
        self._max_row = 0

    def __enter__(self) -> LocalExcelQr02Transaction:
        if not self.cba_path.exists():
            raise FileNotFoundError(f"ENGR CBA workbook file not found at '{self.cba_path}'")
        self.wb = openpyxl.load_workbook(self.cba_path)

        if "QR02 CBA" in self.wb.sheetnames:
            self.ws = self.wb["QR02 CBA"]
        else:
            if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) == 1:
                self.ws = self.wb["Sheet"]
                self.ws.title = "QR02 CBA"
            else:
                self.ws = self.wb.create_sheet("QR02 CBA")

        self._sanitize_ghost_formatting(self.ws)
        self._build_index()
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: Any | None,
    ) -> bool | None:
        try:
            if exc_type is None and self.ws is not None:
                self._sanitize_ghost_formatting(self.ws)
                self.atomic_save()
                if self.on_commit is not None:
                    self.on_commit()
        finally:
            if self.wb is not None:
                self.wb.close()
                self.wb = None
                self.ws = None
        return None

    def _get_real_dimensions(self, ws: Worksheet) -> tuple[int, int]:
        """Scan ws._cells for true maximum row and column with actual values."""
        max_r = 0
        max_c = 0
        for (r, c), cell in ws._cells.items():
            if cell.value is not None and str(cell.value).strip() != "":
                if r > max_r:
                    max_r = r
                if c > max_c:
                    max_c = c
        return max_r, max_c

    def _sanitize_ghost_formatting(self, ws: Worksheet) -> None:
        """Purge ghost cells/rows/columns beyond real data bounds."""
        max_r, max_c = self._get_real_dimensions(ws)
        if max_r == 0 and max_c == 0:
            ws._cells.clear()
            return
        ghost_keys = [
            (r, c)
            for (r, c), cell in ws._cells.items()
            if (r > max_r or c > max_c)
            and (cell.value is None or str(cell.value).strip() == "")
        ]
        for key in ghost_keys:
            del ws._cells[key]

    def _is_header_row(self, val_i: Any, val_j: Any) -> bool:
        headers = {
            "FL",
            "FL ERMS",
            "FUNCTION LOCATION",
            "NAME",
            "SUBSTATION NAME",
            "ERMS NAME",
            "STATION NAME",
            "NO.",
            "PE NAME",
        }
        s_i = str(val_i).strip().upper() if val_i is not None else ""
        s_j = str(val_j).strip().upper() if val_j is not None else ""
        return s_i in headers or s_j in headers

    def _build_index(self) -> None:
        self.fl_to_row.clear()

        if self.ws is None:
            return

        max_r, _ = self._get_real_dimensions(self.ws)
        self._max_row = max_r

        for r in range(1, max_r + 1):
            val_i = self.ws.cell(row=r, column=9).value
            val_j = self.ws.cell(row=r, column=10).value

            if r <= 5 and self._is_header_row(val_i, val_j):
                continue

            fl_norm = normalize_fl_erms(val_i)
            if fl_norm and fl_norm not in self.fl_to_row:
                self.fl_to_row[fl_norm] = r

    def upsert_qr02_cba_records(self, records: Sequence[Any]) -> int:
        if self.ws is None:
            raise RuntimeError("Transaction is not active. Must be used as context manager.")

        updated_count = 0
        for rec in records:
            rec_fl = normalize_fl_erms(_get_field(rec, "fl_erms", "fl_number", "fl", default=""))
            rec_name = str(
                _get_field(
                    rec,
                    "substation_name_erms",
                    "substation_name_site",
                    "substation_name",
                    "name",
                    default="",
                )
            ).strip()

            target_row: int | None = None
            if rec_fl and rec_fl in self.fl_to_row:
                target_row = self.fl_to_row[rec_fl]
            else:
                if self._max_row == 0:
                    # Write header row at row 1 if sheet is empty
                    self.ws.cell(row=1, column=9, value="FL ERMS")
                    self.ws.cell(row=1, column=10, value="SUBSTATION NAME")
                    self.ws.cell(row=1, column=12, value="GPS")
                    self.ws.cell(row=1, column=13, value="TYPE")
                    self.ws.cell(row=1, column=14, value="BUILDING TYPE")
                    self.ws.cell(row=1, column=15, value="CYCLE 1")
                    self.ws.cell(row=1, column=16, value="VENDOR")
                    self._max_row = 1

                target_row = self._max_row + 1
                self._max_row = target_row

            # Fill Col I (FL) and Col J (Name) if missing or new row
            cell_i = self.ws.cell(row=target_row, column=9)
            if rec_fl and (cell_i.value is None or str(cell_i.value).strip() == ""):
                cell_i.value = rec_fl

            cell_j = self.ws.cell(row=target_row, column=10)
            if rec_name and (cell_j.value is None or str(cell_j.value).strip() == ""):
                cell_j.value = rec_name

            # Col L (12): GPS Coordinate
            rec_gps = _get_field(rec, "gps_coordinate", "gps", default=None)
            if rec_gps is not None and str(rec_gps).strip():
                self.ws.cell(row=target_row, column=12, value=str(rec_gps).strip())

            # Col M (13): Type
            rec_type = _get_field(rec, "substation_type", "type_code", "type", default=None)
            if rec_type is not None and str(rec_type).strip():
                self.ws.cell(row=target_row, column=13, value=str(rec_type).strip())

            # Col N (14): Building Type
            rec_bldg = _get_field(rec, "building_type", "bldg_type", default=None)
            if rec_bldg is not None:
                norm_bldg = normalize_building_type(rec_bldg)
                val_bldg = norm_bldg if norm_bldg is not None else str(rec_bldg).strip()
                if val_bldg:
                    self.ws.cell(row=target_row, column=14, value=val_bldg)

            # Col O (15): Cycle 1 date
            rec_cycle = _get_field(rec, "cycle_1", "cycle1", "date", default=None)
            if rec_cycle is not None:
                dt_cycle = to_excel_date(rec_cycle)
                if dt_cycle is not None:
                    cell_o = self.ws.cell(
                        row=target_row,
                        column=15,
                        value=dt_cycle.date() if isinstance(dt_cycle, datetime) else dt_cycle,
                    )
                    cell_o.number_format = "DD-MMM-YYYY"

            # Col P (16): Vendor = "EET"
            self.ws.cell(row=target_row, column=16, value="EET")

            # Update index for subsequent records in the transaction
            if rec_fl:
                self.fl_to_row[rec_fl] = target_row

            updated_count += 1

        return updated_count

    def atomic_save(self) -> None:
        """Write workbook to tempfile, shutil.copy2 to destination path."""
        if self.wb is None:
            return
        cba_path = self.cba_path
        cba_path.parent.mkdir(parents=True, exist_ok=True)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=cba_path.parent) as tmp:
            tmp_path = Path(tmp.name)
        try:
            self.wb.save(tmp_path)
            shutil.copy2(tmp_path, cba_path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink()


class LocalExcelQr02Repository(Qr02Repository):
    """Excel-backed repository resolving per-station ENGR CBA workbooks."""

    def __init__(self, storage: WorkspaceStorage, station: str, year: str) -> None:
        self.storage = storage
        self.station = station
        self.year = year

    def _get_cba_path(self) -> Path:
        return self.storage.get_engr_cba_path(self.station, self.year)

    def invalidate_caches(self) -> None:
        """Callback for post-transaction cleanup."""
        pass

    def transaction(self) -> Qr02Transaction:
        return LocalExcelQr02Transaction(
            cba_path=self._get_cba_path(),
            on_commit=self.invalidate_caches,
        )


class FakeQr02Transaction(Qr02Transaction):
    """Test double for Qr02Transaction storing records in-memory."""

    def __init__(self, repo: FakeQr02Repository) -> None:
        self.repo = repo
        self.is_active = False

    def __enter__(self) -> FakeQr02Transaction:
        self.is_active = True
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> bool | None:
        self.is_active = False
        return None

    def upsert_qr02_cba_records(self, records: Sequence[Any]) -> int:
        if not self.is_active:
            raise RuntimeError("Transaction is not active")
        count = 0
        for rec in records:
            self.repo.records.append(rec)
            count += 1
        return count


class FakeQr02Repository(Qr02Repository):
    """Test double for Qr02Repository."""

    def __init__(self) -> None:
        self.records: list[Any] = []

    def transaction(self) -> FakeQr02Transaction:
        return FakeQr02Transaction(self)
