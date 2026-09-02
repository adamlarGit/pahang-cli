"""Total PE repository and domain logic for Pahang CLI."""

from __future__ import annotations

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Sequence
import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string

from src.core.normalizers import normalize_date_str, normalize_for_excel
from src.testsheet.extractor import to_excel_date
from src.testsheet.models import SubstationTestsheetPackage
from src.msms.models import PropagateResult

def col_to_index(col_letter: str) -> int:
    return column_index_from_string(col_letter) - 1

def read_col(df: pd.DataFrame, col_letter: str) -> pd.Series:
    return df.iloc[:, col_to_index(col_letter)]

def write_cell(ws: Any, row: int, col_letter: str, value: Any) -> None:
    ws[f"{col_letter}{row}"] = value


class TotalPeRepository(ABC):
    """Abstract repository for TOTAL PE workbook management."""

    @abstractmethod
    def get_existing_auto_keys(self, total_pe_path: Path) -> set[tuple[str, str]]:
        """Return set of existing (PE/Station, Date) keys in TOTAL PE.xlsx."""
        ...

    @abstractmethod
    def upsert_packages(
        self, total_pe_path: Path, packages: Sequence[SubstationTestsheetPackage]
    ) -> tuple[int, int]:
        """Upsert testsheet packages into TOTAL PE.xlsx.

        Returns (new_count, updated_count).
        """
        ...

    @abstractmethod
    def propagate_work_orders(
        self, total_pe_path: Path, data_msms_path: Path, target_date: str | None = None
    ) -> PropagateResult:
        """Propagate WONUM from DATA MSMS.xlsx to Column F in TOTAL PE.xlsx."""
        ...


class LocalExcelTotalPeRepository(TotalPeRepository):
    """Excel-backed repository for TOTAL PE workbook management."""

    @staticmethod
    def _get_real_dimensions(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int]:
        max_r = 1
        max_c = 1
        if hasattr(ws, "_cells"):
            for (r, c), cell in ws._cells.items():
                if cell.value is not None:
                    if isinstance(cell.value, str) and str(cell.value).strip() == "":
                        continue
                    if r > max_r:
                        max_r = r
                    if c > max_c:
                        max_c = c
        else:
            max_r = ws.max_row
            max_c = ws.max_column
        return max_r, max_c

    @classmethod
    def _sanitize_ghost_formatting(cls, wb: openpyxl.Workbook) -> None:
        for ws in wb.worksheets:
            real_max_r, real_max_c = cls._get_real_dimensions(ws)
            reported_max_r = ws.max_row or 1
            reported_max_c = ws.max_column or 1

            if reported_max_r > real_max_r + 1:
                ghost_rows = reported_max_r - real_max_r
                ws.delete_rows(real_max_r + 1, ghost_rows)

            if reported_max_c > real_max_c + 1:
                ghost_cols = reported_max_c - real_max_c
                ws.delete_cols(real_max_c + 1, ghost_cols)

    @classmethod
    def _sort_datacycle_sheet(cls, ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        max_r, max_c = cls._get_real_dimensions(ws)
        if max_r < 2:
            return False

        real_max_col = max(max_c, 7)

        last_row = 1
        for row in range(2, max_r + 1):
            pe_val = ws.cell(row=row, column=1).value
            fl_val = ws.cell(row=row, column=2).value
            sub_val = ws.cell(row=row, column=3).value
            if pe_val is not None or fl_val is not None or sub_val is not None:
                last_row = row

        if last_row < 2:
            return False

        rows_data = []
        for r in range(2, last_row + 1):
            vals = []
            for c in range(1, real_max_col + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and val.startswith("="):
                    vals.append(None)
                else:
                    vals.append(val)
            substation_val = vals[0]
            try:
                substation_key = int(float(str(substation_val).strip())) if substation_val is not None and str(substation_val).strip() != "" else 999999
            except (ValueError, TypeError):
                substation_key = 999999
            rows_data.append((substation_key, r, vals))

        sorted_rows = sorted(rows_data, key=lambda x: (x[0], x[1]))

        for idx, (_, _, vals) in enumerate(sorted_rows, start=2):
            for c, val in enumerate(vals, start=1):
                existing_val = ws.cell(row=idx, column=c).value
                if isinstance(existing_val, str) and existing_val.startswith("="):
                    continue
                cell = ws.cell(row=idx, column=c, value=val)
                if c == 1 and val is not None:
                    try:
                        cell.value = int(float(str(val).strip()))
                    except (ValueError, TypeError):
                        pass

        return True

    def get_existing_auto_keys(self, total_pe_path: Path) -> set[tuple[str, str]]:
        if not total_pe_path.exists():
            return set()

        wb_check = openpyxl.load_workbook(total_pe_path, data_only=True, read_only=True)
        try:
            if "DataCycle1" not in wb_check.sheetnames:
                raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")
            ws_check = wb_check["DataCycle1"]
            existing_keys: set[tuple[str, str]] = set()
            for row in ws_check.iter_rows(min_row=2, values_only=True):
                pe_val = str(row[0] or "").strip() if len(row) > 0 and row[0] is not None else ""
                sub_val = str(row[2] or "").strip().upper() if len(row) > 2 and row[2] is not None else ""
                dt_val = str(row[3] or "").strip() if len(row) > 3 and row[3] is not None else ""
                norm_dt = normalize_date_str(dt_val)

                for d in (dt_val, norm_dt):
                    if not d:
                        continue
                    if pe_val:
                        existing_keys.add((pe_val, d))
                        try:
                            num_int = int(pe_val)
                            existing_keys.add((str(num_int), d))
                            existing_keys.add((f"{num_int:03d}", d))
                        except ValueError:
                            pass
                    if sub_val:
                        existing_keys.add((sub_val, d))
            return existing_keys
        finally:
            wb_check.close()

    def upsert_packages(
        self, total_pe_path: Path, packages: Sequence[SubstationTestsheetPackage]
    ) -> tuple[int, int]:
        total_pe_path.parent.mkdir(parents=True, exist_ok=True)

        if not total_pe_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DataCycle1"
            ws.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
        else:
            wb = openpyxl.load_workbook(total_pe_path)
            if "DataCycle1" not in wb.sheetnames:
                raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")
            ws = wb["DataCycle1"]

        new_count = 0
        updated_count = 0

        existing_rows: dict[tuple[str, str], int] = {}
        for r_idx in range(2, ws.max_row + 1):
            pe_val_raw = str(ws.cell(r_idx, 1).value or "").strip()
            sub_val = str(ws.cell(r_idx, 3).value or "").strip().upper()
            date_val = str(ws.cell(r_idx, 4).value or "").strip()
            norm_dt = normalize_date_str(date_val)

            for d in (date_val, norm_dt):
                if not d:
                    continue
                if pe_val_raw:
                    existing_rows[(pe_val_raw, d)] = r_idx
                    try:
                        pe_int = int(pe_val_raw)
                        existing_rows[(str(pe_int), d)] = r_idx
                        existing_rows[(f"{pe_int:03d}", d)] = r_idx
                    except ValueError:
                        pass
                if sub_val:
                    existing_rows[(sub_val, d)] = r_idx

        for pkg in packages:
            data = pkg.data
            if data is None:
                continue

            pe_no = normalize_for_excel(data.substation_number)
            fl_num = normalize_for_excel(data.fl_erms)
            sub_name = normalize_for_excel(data.substation_name_erms)
            dt_str = data.date_str or pkg.date_str
            norm_pkg_dt = normalize_date_str(dt_str)
            type_c = normalize_for_excel(data.substation_type)
            wo = normalize_for_excel(data.wo_number)

            dt_obj = to_excel_date(dt_str)
            date_cell_val = dt_obj.date() if dt_obj is not None else norm_pkg_dt

            lookup_key = (str(pe_no), dt_str)
            norm_lookup_key = (str(pe_no), norm_pkg_dt)
            padded_key = (f"{pe_no:03d}" if isinstance(pe_no, int) else str(pe_no), dt_str)
            norm_padded_key = (f"{pe_no:03d}" if isinstance(pe_no, int) else str(pe_no), norm_pkg_dt)
            sub_key = (str(sub_name).upper() if sub_name else "", dt_str)
            norm_sub_key = (str(sub_name).upper() if sub_name else "", norm_pkg_dt)

            target_row = (
                existing_rows.get(lookup_key)
                or existing_rows.get(norm_lookup_key)
                or existing_rows.get(padded_key)
                or existing_rows.get(norm_padded_key)
                or existing_rows.get(sub_key)
                or existing_rows.get(norm_sub_key)
            )

            if target_row is not None:
                ws.cell(target_row, 1, pe_no)
                ws.cell(target_row, 2, fl_num)
                ws.cell(target_row, 3, sub_name)
                c_dt = ws.cell(target_row, 4, date_cell_val)
                c_dt.number_format = "DD-MMM-YYYY"
                ws.cell(target_row, 5, type_c)
                ws.cell(target_row, 6, wo)
                updated_count += 1
            else:
                next_row = 2
                while ws.cell(next_row, 1).value is not None and str(ws.cell(next_row, 1).value).strip() != "":
                    next_row += 1

                ws.cell(next_row, 1, pe_no)
                ws.cell(next_row, 2, fl_num)
                ws.cell(next_row, 3, sub_name)
                c_dt = ws.cell(next_row, 4, date_cell_val)
                c_dt.number_format = "DD-MMM-YYYY"
                ws.cell(next_row, 5, type_c)
                ws.cell(next_row, 6, wo)
                existing_rows[lookup_key] = next_row
                existing_rows[sub_key] = next_row
                new_count += 1

        self._sort_datacycle_sheet(ws)
        self._sanitize_ghost_formatting(wb)
        wb.save(total_pe_path)
        wb.close()

        return new_count, updated_count

    def propagate_work_orders(
        self, total_pe_path: Path, data_msms_path: Path, target_date: str | None = None
    ) -> PropagateResult:
        """Propagate WO numbers from DATA MSMS.xlsx to Column F in TOTAL PE.xlsx."""
        if not total_pe_path.exists():
            raise FileNotFoundError(f"TOTAL PE workbook not found at '{total_pe_path}'")
        if not data_msms_path.exists():
            raise FileNotFoundError(f"DATA MSMS workbook not found at '{data_msms_path}'")

        # Read DATA MSMS.xlsx mapping fl_erms -> WO
        wb_msms = openpyxl.load_workbook(data_msms_path, data_only=True)
        ws_msms = wb_msms.active

        fl_to_wo: dict[str, str] = {}
        for r_idx in range(2, (ws_msms.max_row or 1) + 1):
            wo_val = ws_msms.cell(r_idx, 1).value  # Col A: Work Order
            fl_val = ws_msms.cell(r_idx, 5).value  # Col E: FL ERMS
            loc_val = ws_msms.cell(r_idx, 2).value  # Col B: Location

            wo_str = str(wo_val).strip() if wo_val is not None else ""
            if not wo_str or wo_str.lower() in ("none", "nan"):
                continue

            for fl_candidate in (fl_val, loc_val):
                if fl_candidate is not None:
                    fl_str = str(fl_candidate).strip().upper()
                    if fl_str and fl_str.lower() not in ("none", "nan"):
                        fl_to_wo[fl_str] = wo_str
                        fl_norm = fl_str.replace("/", "")
                        if fl_norm:
                            fl_to_wo[fl_norm] = wo_str
        wb_msms.close()

        # Open TOTAL PE with data_only=False to preserve all existing formulas and non-WO columns!
        wb_pe = openpyxl.load_workbook(total_pe_path, data_only=False)
        if "DataCycle1" not in wb_pe.sheetnames:
            wb_pe.close()
            raise RuntimeError(f"'DataCycle1' sheet missing in {total_pe_path}")

        ws_pe = wb_pe["DataCycle1"]

        matched_count = 0
        already_populated_count = 0
        unmatched_count = 0
        unmatched_fls: list[str] = []
        updated_count = 0

        norm_target_date = normalize_date_str(target_date) if target_date else None

        for r_idx in range(2, (ws_pe.max_row or 1) + 1):
            # If target_date is given, check row date in Col D (4)
            if target_date is not None:
                row_date_val = ws_pe.cell(r_idx, 4).value
                if row_date_val is None:
                    continue
                row_date_str = str(row_date_val).strip()
                norm_row_date = normalize_date_str(row_date_str)
                if norm_row_date != norm_target_date and row_date_str != target_date:
                    continue

            fl_val = ws_pe.cell(r_idx, 2).value  # Col B: FL NUMBER
            if fl_val is None:
                continue
            fl_str = str(fl_val).strip().upper()
            if not fl_str or fl_str.lower() in ("none", "nan"):
                continue

            # Check if Column F (6: WO) is already populated
            current_wo_val = ws_pe.cell(r_idx, 6).value
            if current_wo_val is not None:
                current_wo_str = str(current_wo_val).strip()
                if current_wo_str and current_wo_str.lower() not in ("none", "nan"):
                    already_populated_count += 1
                    continue

            # Look up WO in fl_to_wo
            wo_match = fl_to_wo.get(fl_str) or fl_to_wo.get(fl_str.replace("/", ""))
            if wo_match:
                ws_pe.cell(r_idx, 6, wo_match)  # Update ONLY Column F!
                matched_count += 1
                updated_count += 1
            else:
                unmatched_count += 1
                unmatched_fls.append(fl_str)

        wb_pe.save(total_pe_path)
        wb_pe.close()

        return PropagateResult(
            matched_count=matched_count,
            already_populated_count=already_populated_count,
            unmatched_count=unmatched_count,
            unmatched_fls=tuple(unmatched_fls),
            updated_count=updated_count,
        )


TotalPeRepo = TotalPeRepository

