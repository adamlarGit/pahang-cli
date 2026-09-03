"""Unit tests for Testsheet OpenPyXL Equipment Extractor (Ticket 083 / 085)."""

from __future__ import annotations

from pathlib import Path
import openpyxl
import pytest

from src.testsheet.extractor import TestsheetExtractor
from src.testsheet.models import (
    BatteryBankSpec,
    SubstationEquipmentPackage,
    SwitchgearSpec,
)


@pytest.fixture
def variant_a_workbook(tmp_path: Path) -> Path:
    """Variant A: 1 TX Indoor PCE Substation."""
    file_path = tmp_path / "001. PE_VARIANT_A.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: PCE Testsheet
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["C5"] = "PE INDOOR 1TX"
    ws_pce["W5"] = "FL-VAR-A"
    ws_pce["P4"] = "2026-05-10"

    # Panels on PCE Testsheet (Rows 10, 14, 18, 22)
    # Panel 1
    ws_pce["B10"] = "F01"
    ws_pce["C10"] = "INCOMING 1"
    ws_pce["D10"] = "CLOSE"
    ws_pce["F10"] = "150A"
    ws_pce["G10"] = "XLPE 3C 240mm2"
    ws_pce["H10"] = "0.5A"
    ws_pce["I10"] = "SN-P01"
    ws_pce["J10"] = "VCB"

    # Panel 2
    ws_pce["B14"] = "F02"
    ws_pce["C14"] = "OUTGOING 1"
    ws_pce["D14"] = "CLOSE"
    ws_pce["F14"] = "80A"
    ws_pce["G14"] = "XLPE 3C 240mm2"
    ws_pce["H14"] = "0.5A"
    ws_pce["I14"] = "SN-P02"
    ws_pce["J14"] = "VCB"

    # Panel 3
    ws_pce["B18"] = "F03"
    ws_pce["C18"] = "TX 1 FEEDER"
    ws_pce["D18"] = "CLOSE"
    ws_pce["F18"] = "50A"
    ws_pce["G18"] = "XLPE 3C 240mm2"
    ws_pce["H18"] = "0.5A"
    ws_pce["I18"] = "SN-P03"
    ws_pce["J18"] = "VCB"

    # Panel 4 - Empty slot (should be excluded)
    ws_pce["B22"] = None
    ws_pce["C22"] = ""
    ws_pce["I22"] = None

    # LVDB Slot 1 (rows 48-51)
    ws_pce["R48"] = "LVDB"
    ws_pce["T48"] = "TX1"
    ws_pce["S49"] = 101
    ws_pce["U49"] = "TAMCO"
    ws_pce["U50"] = "LV-SN-001"
    ws_pce["U51"] = "1600A"

    # Sheet 2: PCE VI
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "PE INDOOR 1TX SITE"
    ws_vi["N1"] = "PE"
    ws_vi["C9"] = "INDOOR"
    ws_vi["D9"] = "/"

    # Switchgear 1 (Rows 11-13)
    ws_vi["J11"] = "/"  # VCB marked
    ws_vi["C12"] = "TAMCO"
    ws_vi["G12"] = "GV3"
    ws_vi["C13"] = "2020"
    ws_vi["J13"] = "12kV 630A"
    ws_vi["O13"] = "SG-2020-001"

    # Transformer Specs (Row 17 count = 1, Row 18 Tx 1)
    ws_vi["C17"] = 1
    ws_vi["D18"] = "HERMETICALLY SEALED"
    ws_vi["F18"] = "1000kVA"
    ws_vi["I18"] = "2020"
    ws_vi["L18"] = "TAMCO"
    ws_vi["O18"] = "TX-2020-001"

    # Auxiliaries & Fire Extinguisher
    ws_vi["E25"] = "/"  # EFI Good
    ws_vi["E29"] = "/"  # SF6 Good
    ws_vi["D42"] = "/"  # Fire Extinguisher Valid
    ws_vi["J42"] = "31/12/2026"

    # RAW DATA
    ws_raw = wb.create_sheet(title="RAW DATA")
    ws_raw.cell(2, 1, "IR")
    ws_raw.cell(2, 2, 100)
    ws_raw.cell(2, 3, 110)
    ws_raw.cell(3, 1, "DG")
    ws_raw.cell(3, 2, 200)
    ws_raw.cell(3, 3, 210)

    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def variant_b_workbook(tmp_path: Path) -> Path:
    """Variant B: 2 TX Attach PCE with multi-sheet panel rollovers."""
    file_path = tmp_path / "002. PE_VARIANT_B.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: PCE Testsheet (Panels 1-4)
    ws_pce1 = wb.active
    ws_pce1.title = "PCE Testsheet"
    ws_pce1["C5"] = "PE ATTACH 2TX"
    ws_pce1["W5"] = "FL-VAR-B"

    for idx, r in enumerate([10, 14, 18, 22], start=1):
        ws_pce1[f"B{r}"] = f"F0{idx}"
        ws_pce1[f"C{r}"] = f"PANEL {idx}"
        ws_pce1[f"D{r}"] = "CLOSE"
        ws_pce1[f"F{r}"] = "100A"
        ws_pce1[f"G{r}"] = "XLPE"
        ws_pce1[f"H{r}"] = "0.5A"
        ws_pce1[f"I{r}"] = f"SN-P0{idx}"
        ws_pce1[f"J{r}"] = "RMU SF6"

    # LVDB Slot 1 & Slot 2
    ws_pce1["R48"] = "LVDB"
    ws_pce1["T48"] = "TX1"
    ws_pce1["S49"] = 201
    ws_pce1["U49"] = "ABB"
    ws_pce1["U50"] = "LV-ABB-01"
    ws_pce1["U51"] = "1600A"

    ws_pce1["R52"] = "LVDB"
    ws_pce1["T52"] = "TX2"
    ws_pce1["S53"] = 202
    ws_pce1["U53"] = "ABB"
    ws_pce1["U54"] = "LV-ABB-02"
    ws_pce1["U55"] = "1600A"

    # Sheet 2: PCE Testsheet (2) (Panels 5-6)
    ws_pce2 = wb.create_sheet(title="PCE Testsheet (2)")
    ws_pce2["B10"] = "F05"
    ws_pce2["C10"] = "PANEL 5"
    ws_pce2["D10"] = "CLOSE"
    ws_pce2["F10"] = "120A"
    ws_pce2["G10"] = "XLPE"
    ws_pce2["H10"] = "0.5A"
    ws_pce2["I10"] = "SN-P05"
    ws_pce2["J10"] = "RMU SF6"

    ws_pce2["B14"] = "F06"
    ws_pce2["C14"] = "PANEL 6"
    ws_pce2["D14"] = "CLOSE"
    ws_pce2["F14"] = "120A"
    ws_pce2["G14"] = "XLPE"
    ws_pce2["H14"] = "0.5A"
    ws_pce2["I14"] = "SN-P06"
    ws_pce2["J14"] = "RMU SF6"

    # Sheet 3: PCE VI
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "PE ATTACH 2TX SITE"
    ws_vi["N1"] = "PE"
    ws_vi["C9"] = "ATTACHED"
    ws_vi["D9"] = "/"

    # Switchgear 1 (RMU SF6)
    ws_vi["G11"] = "/"  # RMU SF6
    ws_vi["C12"] = "ABB"
    ws_vi["G12"] = "SAFEPLUS"
    ws_vi["C13"] = "2021"
    ws_vi["J13"] = "12kV 630A"
    ws_vi["O13"] = "ABB-SG-001"

    # 2 Transformers (Row 17 count = 2)
    ws_vi["C17"] = 2
    # Tx 1 (Row 18)
    ws_vi["D18"] = "CONSERVATOR"
    ws_vi["F18"] = "1000kVA"
    ws_vi["I18"] = "2019"
    ws_vi["L18"] = "MALONEY"
    ws_vi["O18"] = "TX-MAL-001"
    # Tx 2 (Row 19)
    ws_vi["D19"] = "HERMETICALLY SEALED"
    ws_vi["F19"] = "1000kVA"
    ws_vi["I19"] = "2021"
    ws_vi["L19"] = "TAMCO"
    ws_vi["O19"] = "TX-TAM-002"

    # Fire Extinguisher (Expired)
    ws_vi["F42"] = "/"  # Expired
    ws_vi["J42"] = "01/01/2024"

    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def variant_c_workbook(tmp_path: Path) -> Path:
    """Variant C: SSU Switching Station (0 TX, 2 Battery Chargers, RTU)."""
    file_path = tmp_path / "003. SSU_VARIANT_C.xlsx"
    wb = openpyxl.Workbook()

    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["C5"] = "SSU TELUK CEMPEDAK"
    ws_pce["W5"] = "FL-VAR-C"

    # Panels
    ws_pce["B10"] = "F01"
    ws_pce["C10"] = "INCOMING"
    ws_pce["I10"] = "SN-SSU-01"

    # Battery Banks on PCE Testsheet (Rows 59, 60)
    ws_pce["B59"] = "Battery Bank 1"
    ws_pce["J59"] = "SAFT"
    ws_pce["K59"] = "NIFE-110"
    ws_pce["L59"] = "BB-SN-001"

    ws_pce["B60"] = "Battery Charger 2"
    ws_pce["J60"] = "CHLORIDE"
    ws_pce["K60"] = "ED-110"
    ws_pce["L60"] = "BC-SN-002"

    # PCE VI
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "SSU TELUK CEMPEDAK SITE"
    ws_vi["N1"] = "SSU RTU"
    ws_vi["C9"] = "INDOOR"
    ws_vi["D9"] = "/"

    # Switchgear 1 (MRMU)
    ws_vi["I11"] = "/"  # MRMU
    ws_vi["C12"] = "SCHNEIDER"
    ws_vi["G12"] = "RM6"
    ws_vi["C13"] = "2022"
    ws_vi["J13"] = "12kV 630A"
    ws_vi["O13"] = "SCH-001"

    # Authoritative 0 Tx
    ws_vi["C17"] = 0

    # Auxiliaries: Battery Voltage Good (Row 39)
    ws_vi["E39"] = "/"

    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def variant_d_workbook(tmp_path: Path) -> Path:
    """Variant D: CS Compact Substation (Outdoor / Feeder Pillar)."""
    file_path = tmp_path / "004. CS_VARIANT_D.xlsx"
    wb = openpyxl.Workbook()

    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["C5"] = "CS TAMAN MEWAH"
    ws_pce["W5"] = "FL-VAR-D"

    # 1 Panel
    ws_pce["B10"] = "F01"
    ws_pce["C10"] = "TX FEEDER"
    ws_pce["I10"] = "CS-P01"

    # Feeder Pillar Slot 1
    ws_pce["R48"] = "FP"
    ws_pce["T48"] = "TX1"
    ws_pce["S49"] = 401
    ws_pce["U49"] = "TAMCO"
    ws_pce["U50"] = "FP-SN-100"
    ws_pce["U51"] = "800A"

    # PCE VI
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "CS TAMAN MEWAH SITE"
    ws_vi["N1"] = "CS"
    ws_vi["C9"] = "OUTDOOR"
    ws_vi["D9"] = "/"

    # Switchgear 1 (RMU OIL)
    ws_vi["D11"] = "/"  # RMU OIL
    ws_vi["C12"] = "LUCY"
    ws_vi["G12"] = "VRN"
    ws_vi["C13"] = "2015"
    ws_vi["J13"] = "11kV 630A"
    ws_vi["O13"] = "LUCY-OIL-01"

    # 1 Transformer
    ws_vi["C17"] = "1"
    ws_vi["D18"] = "HERMETICALLY"
    ws_vi["F18"] = "500kVA"
    ws_vi["I18"] = "2015"
    ws_vi["L18"] = "TAMCO"
    ws_vi["O18"] = "TX-500-01"

    # EFI Good (Row 25)
    ws_vi["E25"] = "/"

    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def variant_e_workbook(tmp_path: Path) -> Path:
    """Variant E: Dual Switchgear Substation (Switchgear 1 & Switchgear 2)."""
    file_path = tmp_path / "005. DUAL_SG_VARIANT_E.xlsx"
    wb = openpyxl.Workbook()

    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["C5"] = "PE DUAL SWITCHGEAR"
    ws_pce["W5"] = "FL-VAR-E"

    # Panels
    ws_pce["B10"] = "F01"
    ws_pce["C10"] = "INCOMING"
    ws_pce["I10"] = "SN-SG1-01"

    # PCE VI
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "PE DUAL SWITCHGEAR SITE"
    ws_vi["N1"] = "PE"
    ws_vi["C9"] = "INDOOR"
    ws_vi["D9"] = "/"

    # Switchgear 1 (RMU SF6)
    ws_vi["G11"] = "/"  # RMU SF6
    ws_vi["C12"] = "ABB"
    ws_vi["G12"] = "SAFEPLUS"
    ws_vi["C13"] = "2019"
    ws_vi["J13"] = "12kV 630A"
    ws_vi["O13"] = "ABB-SG1-001"

    # Switchgear 2 (VCB)
    ws_vi["K14"] = "/"  # VCB
    ws_vi["C15"] = "TAMCO"
    ws_vi["G15"] = "GV3"
    ws_vi["C16"] = "2023"
    ws_vi["J16"] = "12kV 1250A"
    ws_vi["O16"] = "TAM-SG2-002"

    wb.save(file_path)
    wb.close()
    return file_path


def test_extract_variant_a_1tx_indoor_pce(variant_a_workbook: Path) -> None:
    """Verify Variant A: 1 TX Indoor PCE equipment extraction."""
    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(variant_a_workbook)

    assert isinstance(data.equipment, SubstationEquipmentPackage)
    eq = data.equipment

    # Switchgear
    assert eq.has_switchgear is True
    assert len(eq.switchgears) == 1
    sg = eq.switchgear
    assert sg.switchgear_type == "VCB"
    assert sg.manufacturer == "TAMCO"
    assert sg.model == "GV3"
    assert sg.manufactured_year == "2020"
    assert sg.rating == "12kV 630A"
    assert sg.serial_no == "SG-2020-001"

    # Panels (3 extracted, 1 empty excluded)
    assert len(sg.panels) == 3
    assert sg.panels[0].panel_no == 1
    assert sg.panels[0].panel_feeder_no == "F01"
    assert sg.panels[0].name == "INCOMING 1"
    assert sg.panels[0].status == "CLOSE"
    assert sg.panels[0].load_amp == "150A"
    assert sg.panels[0].cable_type == "XLPE 3C 240mm2"
    assert sg.panels[0].heater_amp == "0.5A"
    assert sg.panels[0].serial_no == "SN-P01"
    assert sg.panels[0].panel_type == "VCB"

    assert sg.panels[1].panel_no == 2
    assert sg.panels[1].name == "OUTGOING 1"

    assert sg.panels[2].panel_no == 3
    assert sg.panels[2].name == "TX 1 FEEDER"

    # Transformers
    assert eq.transformer_count == 1
    tx = eq.transformers[0]
    assert tx.tx_id == "Tx 1"
    assert tx.type == "HERMETICALLY SEALED"
    assert tx.rating_kva == "1000kVA"
    assert tx.construction_year == "2020"
    assert tx.manufacturer == "TAMCO"
    assert tx.serial_no == "TX-2020-001"

    # LVDB
    assert eq.lvdb_count == 1
    lvdb = eq.lvdb_specs[0]
    assert lvdb.name == "LVDB 1"
    assert lvdb.label == "LVDB"
    assert lvdb.source == "TX1"
    assert lvdb.manufacturer == "TAMCO"
    assert lvdb.serial_no == "LV-SN-001"
    assert lvdb.rating == "1600A"

    # Fire Extinguisher
    fe = eq.fire_extinguisher
    assert fe.has_fire_extinguisher is True
    assert fe.status == "VALID"
    assert fe.expiry_date == "31/12/2026"

    # Auxiliary flags
    assert eq.has_efi is True
    assert eq.has_sf6 is True
    assert eq.has_rtu is False
    assert eq.has_battery_charger is False


def test_extract_variant_b_2tx_attach_pce_multisheet(variant_b_workbook: Path) -> None:
    """Verify Variant B: 2 TX Attach PCE with multi-sheet panel rollovers."""
    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(variant_b_workbook)
    eq = data.equipment

    # Switchgear & Panels (4 from Sheet 1, 2 from Sheet 2 = 6 total)
    assert len(eq.switchgears) == 1
    sg = eq.switchgear
    assert sg.switchgear_type == "RMU SF6"
    assert sg.manufacturer == "ABB"
    assert sg.model == "SAFEPLUS"
    assert len(sg.panels) == 6
    for i, p in enumerate(sg.panels, start=1):
        assert p.panel_no == i
        assert p.panel_feeder_no == f"F0{i}"
        assert p.name == f"PANEL {i}"

    # 2 Transformers
    assert eq.transformer_count == 2
    assert eq.transformers[0].tx_id == "Tx 1"
    assert eq.transformers[0].manufacturer == "MALONEY"
    assert eq.transformers[0].rating_kva == "1000kVA"
    assert eq.transformers[1].tx_id == "Tx 2"
    assert eq.transformers[1].manufacturer == "TAMCO"
    assert eq.transformers[1].rating_kva == "1000kVA"

    # 2 LVDBs
    assert eq.lvdb_count == 2
    assert eq.lvdb_specs[0].name == "LVDB 1"
    assert eq.lvdb_specs[0].source == "TX1"
    assert eq.lvdb_specs[1].name == "LVDB 2"
    assert eq.lvdb_specs[1].source == "TX2"

    # Fire Extinguisher (Expired)
    assert eq.fire_extinguisher.has_fire_extinguisher is True
    assert eq.fire_extinguisher.status == "EXPIRED"
    assert eq.fire_extinguisher.expiry_date == "01/01/2024"


def test_extract_variant_c_ssu_switching_station(variant_c_workbook: Path) -> None:
    """Verify Variant C: SSU Switching Station (0 TX, 2 Battery Chargers, RTU)."""
    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(variant_c_workbook)
    eq = data.equipment

    # 0 Transformers
    assert eq.transformer_count == 0
    assert eq.transformers == ()

    # Battery Banks
    assert len(eq.battery_banks) == 2
    assert eq.battery_banks[0].name == "BATTERY BANK 1"
    assert eq.battery_banks[0].manufacturer == "SAFT"
    assert eq.battery_banks[0].model == "NIFE-110"
    assert eq.battery_banks[0].serial_no == "BB-SN-001"

    assert eq.battery_banks[1].name == "BATTERY CHARGER 2"
    assert eq.battery_banks[1].manufacturer == "CHLORIDE"
    assert eq.battery_banks[1].model == "ED-110"
    assert eq.battery_banks[1].serial_no == "BC-SN-002"

    # Auxiliary flags
    assert eq.has_rtu is True
    assert eq.has_battery_charger is True
    assert eq.has_sf6 is True  # MRMU implies SF6


def test_extract_variant_d_cs_compact_substation(variant_d_workbook: Path) -> None:
    """Verify Variant D: CS Compact Substation (Outdoor, Feeder Pillar, No Fire Extinguisher)."""
    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(variant_d_workbook)
    eq = data.equipment

    # Outdoor -> Fire Extinguisher must be False
    assert eq.fire_extinguisher.has_fire_extinguisher is False
    assert eq.fire_extinguisher.status == ""
    assert eq.fire_extinguisher.expiry_date == ""

    # Switchgear RMU OIL
    assert eq.switchgear.switchgear_type == "RMU OIL"
    assert eq.switchgear.manufacturer == "LUCY"
    assert eq.has_sf6 is False  # OIL is not SF6

    # Feeder Pillar
    assert eq.lvdb_count == 1
    assert eq.lvdb_specs[0].label == "FP"
    assert eq.lvdb_specs[0].name == "FP 1"
    assert eq.lvdb_specs[0].rating == "800A"


def test_extract_variant_e_dual_switchgear(variant_e_workbook: Path) -> None:
    """Verify Variant E: Dual Switchgear Substation (Switchgear 1 and Switchgear 2)."""
    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(variant_e_workbook)
    eq = data.equipment

    assert len(eq.switchgears) == 2
    sg1, sg2 = eq.switchgears

    assert sg1.switchgear_type == "RMU SF6"
    assert sg1.manufacturer == "ABB"
    assert sg1.model == "SAFEPLUS"
    assert sg1.manufactured_year == "2019"
    assert sg1.serial_no == "ABB-SG1-001"
    assert len(sg1.panels) == 1

    assert sg2.switchgear_type == "VCB"
    assert sg2.manufacturer == "TAMCO"
    assert sg2.model == "GV3"
    assert sg2.manufactured_year == "2023"
    assert sg2.rating == "12kV 1250A"
    assert sg2.serial_no == "TAM-SG2-002"
    assert len(sg2.panels) == 0


def test_transformer_specs_not_accessible_and_empty_exclusions(tmp_path: Path) -> None:
    """Verify Transformer extraction handles N/A, NOT ACCESSIBLE, and missing C17 values."""
    file_path = tmp_path / "006. TX_NA.xlsx"
    wb = openpyxl.Workbook()
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_vi = wb.create_sheet(title="PCE VI")

    # Quantity C17 is "N/A"
    ws_vi["C17"] = "N/A"
    ws_vi["D18"] = "OIL"
    ws_vi["F18"] = "1000kVA"

    wb.save(file_path)
    wb.close()

    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(file_path)
    assert data.equipment.transformers == ()
    assert data.equipment.transformer_count == 0


def test_transformer_specs_individual_row_not_accessible(tmp_path: Path) -> None:
    """Verify individual transformer row marked NOT ACCESSIBLE is skipped."""
    file_path = tmp_path / "007. TX_PARTIAL.xlsx"
    wb = openpyxl.Workbook()
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_vi = wb.create_sheet(title="PCE VI")

    ws_vi["C17"] = 2
    # Tx 1 valid
    ws_vi["D18"] = "HERMETICALLY SEALED"
    ws_vi["F18"] = "1000kVA"
    ws_vi["L18"] = "TAMCO"

    # Tx 2 NOT ACCESSIBLE
    ws_vi["D19"] = "NOT ACCESSIBLE"
    ws_vi["F19"] = "-"

    wb.save(file_path)
    wb.close()

    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(file_path)
    assert data.equipment.transformer_count == 1
    assert data.equipment.transformers[0].tx_id == "Tx 1"


def test_missing_and_unparseable_dataclass_fields_policy(tmp_path: Path) -> None:
    """Verify missing or unparseable cells normalize to empty strings in dataclasses."""
    file_path = tmp_path / "008. BLANK_SPECS.xlsx"
    wb = openpyxl.Workbook()
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"

    # Panel with only name populated
    ws_pce["C10"] = "FEEDER 1"

    wb.save(file_path)
    wb.close()

    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(file_path)
    assert data.equipment.has_switchgear is True
    sg = data.equipment.switchgear
    assert sg.switchgear_type == ""
    assert sg.manufacturer == ""
    assert sg.model == ""
    assert sg.manufactured_year == ""
    assert sg.rating == ""
    assert sg.serial_no == ""
    assert len(sg.panels) == 1
    p = sg.panels[0]
    assert p.panel_no == 1
    assert p.name == "FEEDER 1"
    assert p.panel_feeder_no == ""
    assert p.panel_type == ""
    assert p.serial_no == ""
    assert p.status == ""
    assert p.load_amp == ""
    assert p.cable_type == ""
    assert p.heater_amp == ""


def test_switchgear_types_checkbox_detection() -> None:
    """Verify _extract_switchgear_type correctly identifies each switchgear type."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws = wb.active

    # RMU OIL via D11
    ws["D11"] = "/"
    assert extractor._extract_switchgear_type(ws, 11) == "RMU OIL"

    # RMU SF6 via G11
    ws["D11"] = None
    ws["G11"] = "X"
    assert extractor._extract_switchgear_type(ws, 11) == "RMU SF6"

    # MRMU via I11
    ws["G11"] = None
    ws["I11"] = "1"
    assert extractor._extract_switchgear_type(ws, 11) == "MRMU"

    # VCB via K11
    ws["I11"] = None
    ws["K11"] = "YES"
    assert extractor._extract_switchgear_type(ws, 11) == "VCB"

    # OCB via M11
    ws["K11"] = None
    ws["M11"] = "TRUE"
    assert extractor._extract_switchgear_type(ws, 11) == "OCB"

    # OTHER via O11 with default "OTHER"
    ws["M11"] = None
    ws["O11"] = "/"
    assert extractor._extract_switchgear_type(ws, 11) == "OTHER"

    # OTHER with custom name in P11
    ws["P11"] = "GIS 22kV"
    assert extractor._extract_switchgear_type(ws, 11) == "GIS 22kV"


def test_transformer_specs_up_to_four_units() -> None:
    """Verify _extract_transformer_specs parses up to 4 transformer units."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["C17"] = 4
    for i in range(1, 5):
        r = 17 + i
        ws[f"D{r}"] = f"TYPE {i}"
        ws[f"F{r}"] = f"{i * 500}kVA"
        ws[f"I{r}"] = f"201{i}"
        ws[f"L{r}"] = f"MFG {i}"
        ws[f"O{r}"] = f"SN-TX-{i}"

    specs = extractor._extract_transformer_specs(ws)
    assert len(specs) == 4
    for i, tx in enumerate(specs, start=1):
        assert tx.tx_id == f"Tx {i}"
        assert tx.type == f"TYPE {i}"
        assert tx.rating_kva == f"{i * 500}kVA"
        assert tx.construction_year == f"201{i}"
        assert tx.manufacturer == f"MFG {i}"
        assert tx.serial_no == f"SN-TX-{i}"


def test_extract_transformer_specs_with_cables_and_thermal() -> None:
    """Verify _extract_transformer_specs parses cable types and thermal readings from PCE Testsheet."""
    extractor = TestsheetExtractor()
    wb_vi = openpyxl.Workbook()
    ws_vi = wb_vi.active
    ws_vi["C17"] = 2
    ws_vi["D18"] = "HERMETIC"
    ws_vi["F18"] = "1000kVA"
    ws_vi["D19"] = "HERMETIC"
    ws_vi["F19"] = "500kVA"

    wb_pce = openpyxl.Workbook()
    ws_pce = wb_pce.active

    # TX1: row 33 to 37
    ws_pce["C33"] = "XLPE"
    ws_pce["C35"] = "XLPE"
    ws_pce["F33"] = 40.1  # HT CABLE Tmin
    ws_pce["G33"] = 40.6  # Tmax
    ws_pce["H33"] = 0.5   # Delta T
    ws_pce["I33"] = 40.35 # Avg
    ws_pce["F34"] = 42.7  # HT BUSHING
    ws_pce["G34"] = 43.1
    ws_pce["H34"] = 0.4
    ws_pce["I34"] = 42.9
    ws_pce["F35"] = 38.7  # LV CABLE
    ws_pce["G35"] = 39.2
    ws_pce["H35"] = 0.5
    ws_pce["I35"] = 38.95
    ws_pce["F36"] = 42.2  # LV BUSHING
    ws_pce["G36"] = 42.8
    ws_pce["H36"] = 0.6
    ws_pce["I36"] = 42.5
    ws_pce["F37"] = 41.6  # BODY
    ws_pce["G37"] = 45.1
    ws_pce["H37"] = 3.5
    ws_pce["I37"] = 43.35

    # TX2: row 38 to 42
    ws_pce["C38"] = "PILC"
    ws_pce["C40"] = "XLPE"
    ws_pce["F38"] = 35.0
    ws_pce["G38"] = 36.0
    ws_pce["H38"] = 1.0
    ws_pce["I38"] = 35.5

    specs = extractor._extract_transformer_specs(ws_vi, ws_pce=ws_pce)
    assert len(specs) == 2

    tx1 = specs[0]
    assert tx1.tx_id == "Tx 1"
    assert tx1.hv_cable_type == "XLPE"
    assert tx1.lv_cable_type == "XLPE"
    assert tx1.hv_cable_thermal.tmin == "40.1"
    assert tx1.hv_cable_thermal.tmax == "40.6"
    assert tx1.hv_cable_thermal.delta_t == "0.5"
    assert tx1.hv_cable_thermal.avg == "40.35"
    assert tx1.hv_bushing_thermal.tmin == "42.7"
    assert tx1.lv_cable_thermal.tmin == "38.7"
    assert tx1.lv_bushing_thermal.tmin == "42.2"
    assert tx1.body_thermal.tmin == "41.6"
    assert tx1.body_thermal.delta_t == "3.5"

    tx2 = specs[1]
    assert tx2.tx_id == "Tx 2"
    assert tx2.hv_cable_type == "PILC"
    assert tx2.lv_cable_type == "XLPE"
    assert tx2.hv_cable_thermal.tmin == "35.0"
    assert tx2.hv_cable_thermal.delta_t == "1.0"


def test_lvdb_specs_empty_template_ignored() -> None:
    """Verify _extract_lvdb_specs returns empty tuple when no photo or fields are filled."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws = wb.active

    # Standard empty template layout
    ws["R48"] = "LVDB"
    ws["T48"] = "TX1"
    ws["T49"] = "Manufacturer :"
    ws["T50"] = "Serial No. :"
    ws["R51"] = "Rating :"

    specs = extractor._extract_lvdb_specs(ws)
    assert specs == ()


def test_extract_lvdb_specs_with_feeders_and_cable_type() -> None:
    """Verify _extract_lvdb_specs extracts feeder ways and calculates board cable type."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws = wb.active

    # Slot 1: row 45 feeders
    ws["D45"] = "XLPE"
    ws["E45"] = "XLPE"
    ws["G45"] = "-"       # Inactive sentinel
    ws["I45"] = "XLPE"
    ws["J45"] = "PILC"
    ws["K45"] = "XLPE"
    ws["L45"] = "SPARE"   # Inactive sentinel

    ws["R48"] = "FP"
    ws["T48"] = "TX1"
    ws["V49"] = "SSE"
    ws["V50"] = "FPPO-1628"
    ws["V51"] = "1600A"

    # Slot 2: row 47 feeders
    ws["D47"] = "PILC"
    ws["I47"] = "PILC"
    ws["R52"] = "LVDB"
    ws["T52"] = "TX2"
    ws["V53"] = "TAMCO"
    ws["V54"] = "SN-22"
    ws["V55"] = "800A"

    specs = extractor._extract_lvdb_specs(ws)
    assert len(specs) == 2

    fp1 = specs[0]
    assert fp1.name == "FP 1"
    assert fp1.label == "FP"
    assert fp1.cable_type == "XLPE"
    assert len(fp1.feeders) == 5
    # Channel checks
    channels = [f.channel for f in fp1.feeders]
    assert channels == ["IN1", "IN2", "OT1", "OT2", "OT3"]
    assert fp1.get_feeder_cable("OT1") == "XLPE"
    assert fp1.get_feeder_cable("OT2") == "PILC"
    assert fp1.get_feeder_cable("OT3") == "XLPE"
    assert fp1.get_feeder_cable("OT4") == "XLPE"  # Fallback to board cable_type

    fp2 = specs[1]
    assert fp2.name == "LVDB 2"
    assert fp2.cable_type == "PILC"
    assert len(fp2.feeders) == 2
    assert fp2.get_feeder_cable("IN1") == "PILC"


def test_battery_banks_empty_template_ignored() -> None:
    """Verify _extract_battery_banks returns empty tuple when rows only contain template labels."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 59 has template text but no actual equipment data
    ws["B59"] = "Battery Bank"

    specs = extractor._extract_battery_banks(ws)
    assert specs == ()


def test_fire_extinguisher_spec_outdoor_always_false() -> None:
    """Verify _extract_fire_extinguisher_spec returns False for OUTDOOR even if checkboxes marked."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["D42"] = "/"
    ws["J42"] = "31/12/2026"

    spec = extractor._extract_fire_extinguisher_spec(ws, building_type="OUTDOOR")
    assert spec.has_fire_extinguisher is False
    assert spec.status == ""
    assert spec.expiry_date == ""


def test_auxiliary_flags_combinations() -> None:
    """Verify _extract_auxiliary_flags evaluates EFI, SF6, RTU, and Battery Charger properly."""
    extractor = TestsheetExtractor()
    wb = openpyxl.Workbook()
    ws_vi = wb.active

    # 1. When nothing marked
    flags = extractor._extract_auxiliary_flags(ws_vi, None, switchgears=(), battery_banks=(), substation_type="PE", building_type="INDOOR")
    assert flags == (False, False, False, False)

    # 2. EFI marked on E25
    ws_vi["E25"] = "/"
    flags = extractor._extract_auxiliary_flags(ws_vi, None, switchgears=(), battery_banks=(), substation_type="PE", building_type="INDOOR")
    assert flags == (False, False, False, True)

    # 3. SF6 from switchgear type
    sg_sf6 = SwitchgearSpec(switchgear_type="RMU SF6")
    flags = extractor._extract_auxiliary_flags(ws_vi, None, switchgears=(sg_sf6,), battery_banks=(), substation_type="PE", building_type="INDOOR")
    assert flags == (False, False, True, True)

    # 4. Battery charger from battery banks -> assumes RTU for standard substation
    bb = BatteryBankSpec(name="BATTERY BANK 1")
    flags = extractor._extract_auxiliary_flags(ws_vi, None, switchgears=(sg_sf6,), battery_banks=(bb,), substation_type="PE", building_type="INDOOR")
    assert flags == (True, True, True, True)

    # 5. CS / Compact exception -> has battery charger but NO RTU
    flags_cs = extractor._extract_auxiliary_flags(ws_vi, None, switchgears=(sg_sf6,), battery_banks=(bb,), substation_type="CS", building_type="COMPACT")
    assert flags_cs == (True, False, True, True)

    # 6. EFI Not Good on G25 (defect/anomaly, but EFI is present)
    ws_vi2 = wb.create_sheet(title="PCE VI 2")
    ws_vi2["G25"] = "✅"
    flags_ng = extractor._extract_auxiliary_flags(ws_vi2, None, switchgears=(), battery_banks=(), substation_type="PE", building_type="INDOOR")
    assert flags_ng == (False, False, False, True)

    # 7. EFI N/A or MISSING on I25
    ws_vi2["G25"] = None
    ws_vi2["I25"] = "N/A"
    flags_na = extractor._extract_auxiliary_flags(ws_vi2, None, switchgears=(), battery_banks=(), substation_type="PE", building_type="INDOOR")
    assert flags_na == (False, False, False, False)

    ws_vi2["I25"] = "MISSING"
    flags_missing = extractor._extract_auxiliary_flags(ws_vi2, None, switchgears=(), battery_banks=(), substation_type="PE", building_type="INDOOR")
    assert flags_missing == (False, False, False, False)



