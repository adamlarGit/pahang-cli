"""Unit tests for MsmsRepo (read_data_msms, consolidate_xls_files, enrich_from_engr)."""
from pathlib import Path
import openpyxl
import pandas as pd
import pytest

from src.repositories.msms import (
    MsmsRepo,
    MsmsRepository,
    LocalExcelMsmsRepository,
    ConsolidateResult,
    EnrichResult,
)


def _create_sample_data_msms(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append([
        "Work Order",
        "Location",
        "Description",
        "Substation Name ERMS",
        "FL ERMS",
        "Cycle Date",
        "Substation Number",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _create_sample_total_pe(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataCycle1"
    ws.append([
        "PE NO",
        "FL NUMBER",
        "SUBSTATION NAME",
        "DATE",
        "TYPE",
        "WO",
        "SCOPE",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _create_sample_maximo_html_xls(path: Path, rows: list[tuple[str, str, str, str]]) -> None:
    html_content = ["<html><body><table border='1'>"]
    html_content.append("<tr><th>Wonum</th><th>Status</th><th>Location</th><th>Description</th></tr>")
    for wonum, status, location, desc in rows:
        html_content.append(f"<tr><td>{wonum}</td><td>{status}</td><td>{location}</td><td>{desc}</td></tr>")
    html_content.append("</table></body></html>")
    path.write_text("\n".join(html_content), encoding="utf-8")


def test_read_data_msms_missing_file_raises_error(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    with pytest.raises(FileNotFoundError):
        repo.read_data_msms(tmp_path / "NON_EXISTENT.xlsx")


def test_read_data_msms_success(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [
        ["45001111", "CKTN0001XXXX", "PE TELUK SISIK", "PE TELUK SISIK", "CKTN0001/XXXX", "2026-01-10", 1],
    ])
    
    df = repo.read_data_msms(msms_file)
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 1
    assert str(df.iloc[0, 0]) == "45001111"


def test_consolidate_xls_files_missing_target_raises(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    xls_file = tmp_path / "sample.xls"
    xls_file.touch()
    with pytest.raises(FileNotFoundError):
        repo.consolidate_xls_files([xls_file], tmp_path / "MISSING_DATA_MSMS.xlsx")


def test_consolidate_xls_files_success(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    # Pre-existing entry
    _create_sample_data_msms(msms_file, [
        ["45001000", "CKTN0001AAAA", "PE EXISTING", None, None, None, None],
    ])

    xls1 = tmp_path / "45501001.xls"
    _create_sample_maximo_html_xls(xls1, [
        ("45001000", "APPR", "CKTN0001AAAA", "PE EXISTING"),  # Duplicate, should skip
        ("45001001", "APPR", "CKTN0002BBBB", "PE NEW ONE"),   # New, position 8 slash needed
    ])

    xls2 = tmp_path / "45501002.xls"
    _create_sample_maximo_html_xls(xls2, [
        ("45001001", "APPR", "CKTN0002BBBB", "PE NEW ONE"),   # Duplicate across files
        ("45001002", "APPR", "CKTN0003/CCCC", "PE NEW TWO"),  # New, already has slash
    ])

    res = repo.consolidate_xls_files([xls1, xls2], msms_file)
    assert isinstance(res, ConsolidateResult)
    assert res.files_processed == 2
    assert res.rows_appended == 2
    assert res.duplicates_skipped == 2

    # Verify content in DATA MSMS.xlsx
    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 4  # header + 1 existing + 2 new

    # Row 3: 45001001 with CKTN0002/BBBB in Col E
    assert str(rows[2][0]) == "45001001"
    assert rows[2][1] == "CKTN0002BBBB"
    assert rows[2][2] == "PE NEW ONE"
    assert rows[2][4] == "CKTN0002/BBBB"

    # Row 4: 45001002 with CKTN0003/CCCC in Col E
    assert str(rows[3][0]) == "45001002"
    assert rows[3][1] == "CKTN0003/CCCC"
    assert rows[3][2] == "PE NEW TWO"
    assert rows[3][4] == "CKTN0003/CCCC"


def test_enrich_from_engr_missing_files_raises(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    msms_file.touch()
    with pytest.raises(FileNotFoundError):
        repo.enrich_from_engr(msms_file, tmp_path / "MISSING_TOTAL_PE.xlsx")


def test_enrich_from_engr_success(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001BBBB", "RAW DESC 1", None, None, None, None],
        ["45001002", "CKTN0002CCCC", "RAW DESC 2", "MANUAL SUB", "MANUAL/FL", "2026-01-05", 99],
        ["45009999", "CKTN0003DDDD", "UNMATCHED DESC", None, None, None, None],
    ])

    pe_file = tmp_path / "TOTAL PE.xlsx"
    _create_sample_total_pe(pe_file, [
        [12, "CKTN0001/BBBB", "PE BANDAR RAYA", "2026-01-15", "ATTACHED", "45001001", "FULL"],
        [15, "CKTN0002/CCCC", "PE BUKIT MEWAH", "2026-01-16", "COMPACT", "45001002", "FULL"],
    ])

    res = repo.enrich_from_engr(msms_file, pe_file)
    assert isinstance(res, EnrichResult)
    assert res.matched_count == 2
    assert res.unmatched_count == 1
    assert "45009999" in res.unmatched_wos

    # Verify openpyxl update in DATA MSMS.xlsx
    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Row 2 (45001001): was all None in cols D-G, should be enriched from Total PE
    assert rows[1][0] == "45001001"
    assert rows[1][3] == "PE BANDAR RAYA"
    assert rows[1][4] == "CKTN0001/BBBB"
    assert str(rows[1][5]) == "2026-01-15"
    assert rows[1][6] == 12

    # Row 3 (45001002): already had manual entries in D-G, should not be overwritten
    assert rows[2][3] == "MANUAL SUB"
    assert rows[2][4] == "MANUAL/FL"
    assert str(rows[2][5]) == "2026-01-05"
    assert rows[2][6] == 99


def test_msms_repo_aliases() -> None:
    from src.repositories.msms import MsmsRepo as AliasedRepo
    assert AliasedRepo is MsmsRepository or issubclass(AliasedRepo, MsmsRepository)


def test_consolidate_xls_files_contiguous_append_with_blank_formatted_rows(tmp_path: Path) -> None:
    """Ensure repository consolidate_xls_files overwrites/clears 100 blank formatted rows contiguously."""
    from openpyxl.styles import PatternFill

    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append(["Work Order", "Location", "Description", "Substation Name ERMS", "FL ERMS", "Cycle Date", "Substation Number"])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in range(2, 102):
        ws.cell(row=r, column=1).fill = yellow_fill
    wb.save(msms_file)
    wb.close()

    xls1 = tmp_path / "45501001.xls"
    _create_sample_maximo_html_xls(xls1, [
        ("45001001", "APPR", "CKTN0001AAAA", "PE ONE"),
        ("45001002", "APPR", "CKTN0002BBBB", "PE TWO"),
    ])

    res = repo.consolidate_xls_files([xls1], msms_file)
    assert res.files_processed == 1
    assert res.rows_appended == 2

    wb_res = openpyxl.load_workbook(msms_file)
    ws_res = wb_res.active
    assert ws_res.max_row == 3  # Header + 2 rows

    rows = list(ws_res.iter_rows(values_only=True))
    assert len(rows) == 3
    assert rows[1][0] == "45001001"
    assert rows[1][1] == "CKTN0001AAAA"
    assert rows[1][4] == "CKTN0001/AAAA"
    assert rows[2][0] == "45001002"
    assert rows[2][1] == "CKTN0002BBBB"
    assert rows[2][4] == "CKTN0002/BBBB"
    wb_res.close()


def test_consolidate_xls_files_compacts_disconnected_rows(tmp_path: Path) -> None:
    """Ensure repository consolidate_xls_files compacts disconnected rows separated by blank blocks."""
    from openpyxl.styles import PatternFill

    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append(["Work Order", "Location", "Description", "Substation Name ERMS", "FL ERMS", "Cycle Date", "Substation Number"])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in range(2, 102):
        ws.cell(row=r, column=1).fill = yellow_fill

    # Existing row at row 25
    ws.cell(row=25, column=1).value = "45001000"
    ws.cell(row=25, column=2).value = "CKTN0000AAAA"
    ws.cell(row=25, column=3).value = "PE EXISTING"
    ws.cell(row=25, column=5).value = "CKTN0000/AAAA"
    wb.save(msms_file)
    wb.close()

    xls1 = tmp_path / "45501001.xls"
    _create_sample_maximo_html_xls(xls1, [
        ("45001000", "APPR", "CKTN0000AAAA", "PE EXISTING"),  # Duplicate
        ("45001001", "APPR", "CKTN0001AAAA", "PE ONE"),       # New
    ])

    res = repo.consolidate_xls_files([xls1], msms_file)
    assert res.files_processed == 1
    assert res.rows_appended == 1
    assert res.duplicates_skipped == 1

    wb_res = openpyxl.load_workbook(msms_file)
    ws_res = wb_res.active
    assert ws_res.max_row == 3  # Header + 1 existing + 1 new

    rows = list(ws_res.iter_rows(values_only=True))
    assert len(rows) == 3
    assert rows[1][0] == "45001000"
    assert rows[1][1] == "CKTN0000AAAA"
    assert rows[2][0] == "45001001"
    assert rows[2][1] == "CKTN0001AAAA"
    wb_res.close()


def test_enrich_from_engr_matches_via_location_when_fl_erms_and_wo_blank(tmp_path: Path) -> None:
    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [
        ["45001001", "CRAU/PCEJ00232", "RAW DESC 1", None, None, None, None],
    ])

    pe_file = tmp_path / "TOTAL PE.xlsx"
    _create_sample_total_pe(pe_file, [
        [35, "CRAU/PCE/J00232", "PE CRAU JAYA", "2026-03-01", "ATTACHED", None, "FULL"],
    ])

    res = repo.enrich_from_engr(msms_file, pe_file)
    assert isinstance(res, EnrichResult)
    assert res.matched_count == 1
    assert res.unmatched_count == 0
    assert res.updated_cells_count == 4

    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    assert rows[1][0] == "45001001"
    assert rows[1][1] == "CRAU/PCEJ00232"
    assert rows[1][3] == "PE CRAU JAYA"
    assert rows[1][4] == "CRAU/PCE/J00232"
    assert str(rows[1][5]) == "2026-03-01"
    assert rows[1][6] == 35


