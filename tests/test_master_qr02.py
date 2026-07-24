"""Unit tests for Master QR02 Repository and Transactions in Pahang CLI."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import openpyxl
import pytest

from src.master.qr02 import (
    FakeQr02Repository,
    LocalExcelQr02Repository,
    LocalExcelQr02Transaction,
    _fuzzy_normalize_name,
)
from src.project.storage import LocalWorkspaceStorage
from src.testsheet.models import TestsheetData


@pytest.fixture
def tmp_workspace(tmp_path: Path) -> LocalWorkspaceStorage:
    root = tmp_path / "workspace"
    python_dir = root / "PYTHON"
    engr_dir = python_dir / "ENGR FROM DRIVE"
    engr_dir.mkdir(parents=True)
    return LocalWorkspaceStorage(root)


def test_fuzzy_normalize_name() -> None:
    assert _fuzzy_normalize_name("001. PE SSU CHEROH") == "CHEROH"
    assert _fuzzy_normalize_name("P-E PDT CHEROH NO. 2") == "CHEROH2"
    assert _fuzzy_normalize_name("SSU CHEROH 2") == "CHEROH2"
    assert _fuzzy_normalize_name("") == ""


def test_local_excel_qr02_repository_cba_path(tmp_workspace: LocalWorkspaceStorage) -> None:
    repo = LocalExcelQr02Repository(tmp_workspace, station="RAUB", year="2026")
    expected_path = tmp_workspace.get_engr_folder() / "ENGR-750-36-CBA-RAU-2026.xlsx"
    assert repo._get_cba_path() == expected_path


def test_exact_fl_matching(tmp_path: Path) -> None:
    cba_path = tmp_path / "test_cba.xlsx"

    # Setup workbook with existing row
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QR02 CBA"
    ws.cell(row=1, column=9, value="FL ERMS")
    ws.cell(row=1, column=10, value="SUBSTATION NAME")

    ws.cell(row=2, column=9, value="75001234.0")
    ws.cell(row=2, column=10, value="SSU CHEROH")
    wb.save(cba_path)
    wb.close()

    rec = TestsheetData(
        pe_number=1,
        substation_name="SSU CHEROH",
        fl_erms="75001234",
        gps_coordinate="3.81, 101.80",
        substation_type="SSU",
        building_type="INDOOR",
        cycle_1=datetime(2026, 5, 1, 0, 0, 0),
    )

    tx = LocalExcelQr02Transaction(cba_path)
    with tx:
        count = tx.upsert_qr02_cba_records([rec])
        assert count == 1

    # Reload and verify
    wb_read = openpyxl.load_workbook(cba_path)
    ws_read = wb_read["QR02 CBA"]
    assert ws_read.cell(row=2, column=12).value == "3.81, 101.80"
    assert ws_read.cell(row=2, column=13).value == "SSU"
    assert ws_read.cell(row=2, column=14).value == "INDOOR"
    assert ws_read.cell(row=2, column=15).value in (datetime(2026, 5, 1, 0, 0, 0), datetime(2026, 5, 1, 0, 0, 0).date())
    assert ws_read.cell(row=2, column=15).number_format == "DD-MMM-YYYY"
    assert ws_read.cell(row=2, column=16).value == "EET"
    wb_read.close()


def test_unmatched_fl_appends_new_row(tmp_path: Path) -> None:
    cba_path = tmp_path / "test_cba.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QR02 CBA"
    ws.cell(row=1, column=9, value="FL")
    ws.cell(row=1, column=10, value="NAME")
    ws.cell(row=2, column=9, value="75001234")
    ws.cell(row=2, column=10, value="PE SSU CHEROH")
    wb.save(cba_path)
    wb.close()

    rec = TestsheetData(
        pe_number=1,
        substation_name="PE UNMATCHED",
        fl_erms="75009999",
        gps_coordinate="3.82, 101.81",
        substation_type="PE",
    )

    with LocalExcelQr02Transaction(cba_path) as tx:
        tx.upsert_qr02_cba_records([rec])

    wb_read = openpyxl.load_workbook(cba_path)
    ws_read = wb_read["QR02 CBA"]
    # Row 2 remains intact
    assert ws_read.cell(row=2, column=9).value == "75001234"
    # Row 3 is appended for unmatched FL
    assert ws_read.cell(row=3, column=9).value == "75009999"
    assert ws_read.cell(row=3, column=10).value == "PE UNMATCHED"
    assert ws_read.cell(row=3, column=12).value == "3.82, 101.81"
    wb_read.close()


def test_fallback_append_new_row(tmp_path: Path) -> None:
    cba_path = tmp_path / "test_cba.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QR02 CBA"
    ws.cell(row=1, column=9, value="FL")
    ws.cell(row=1, column=10, value="NAME")
    ws.cell(row=2, column=9, value="75001111")
    ws.cell(row=2, column=10, value="EXISTING PE")
    wb.save(cba_path)
    wb.close()

    rec = TestsheetData(
        pe_number=3,
        substation_name="NEW PE STATION",
        fl_erms="75008888",
        gps_coordinate="3.90, 101.90",
        building_type="ATTACHED",
    )

    with LocalExcelQr02Transaction(cba_path) as tx:
        tx.upsert_qr02_cba_records([rec])

    wb_read = openpyxl.load_workbook(cba_path)
    ws_read = wb_read["QR02 CBA"]
    # Row 3 should be appended
    assert ws_read.cell(row=3, column=9).value == "75008888"
    assert ws_read.cell(row=3, column=10).value == "NEW PE STATION"
    assert ws_read.cell(row=3, column=12).value == "3.90, 101.90"
    assert ws_read.cell(row=3, column=14).value == "ATTACH"
    assert ws_read.cell(row=3, column=16).value == "EET"
    wb_read.close()


def test_ghost_cell_cleanup(tmp_path: Path) -> None:
    cba_path = tmp_path / "test_cba.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QR02 CBA"
    ws.cell(row=1, column=9, value="FL")
    ws.cell(row=1, column=10, value="NAME")
    ws.cell(row=2, column=9, value="75001234")
    ws.cell(row=2, column=10, value="SSU CHEROH")
    # Add ghost empty cell at row 50, col 50
    ghost_cell = ws.cell(row=50, column=50)
    ghost_cell.number_format = "0.00"
    wb.save(cba_path)
    wb.close()

    with LocalExcelQr02Transaction(cba_path) as tx:
        assert (50, 50) not in tx.ws._cells

    wb_read = openpyxl.load_workbook(cba_path)
    ws_read = wb_read["QR02 CBA"]
    assert (50, 50) not in ws_read._cells
    wb_read.close()


def test_fake_qr02_repository() -> None:
    repo = FakeQr02Repository()

    rec1 = TestsheetData(pe_number=1, substation_name="PE 1")
    rec2 = TestsheetData(pe_number=2, substation_name="PE 2")

    with repo.transaction() as tx:
        tx.upsert_qr02_cba_records([rec1, rec2])

    assert len(repo.records) == 2
    assert repo.records[0].substation_name == "PE 1"
    assert repo.records[1].substation_name == "PE 2"
