"""Integration and unit tests for Populate Data MSMS Workflow (src/workflows/populate_data_msms.py)."""

from __future__ import annotations

from datetime import date, time
from pathlib import Path
import openpyxl
import pandas as pd
import pytest

from src.project.environment import ProjectEnvironment
from src.project.models import ProjectMetadata
from src.project.storage import LocalWorkspaceStorage
from src.quick_report.defects import MasterQr03DefectRepository, ViDefectRecord

pytest.importorskip("src.workflows.populate_data_msms")

from src.workflows.models import (
    PopulateDataMsmsRequest,
    PopulateMode,
)

from src.workflows.populate_data_msms import (
    PopulateDataMsmsPreflightGuard,
    PopulateDataMsmsWorkflow,
    match_vi_defect,
    parse_testsheet_datetime,
)


@pytest.fixture
def mock_env(tmp_path: Path) -> ProjectEnvironment:
    meta = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    storage = LocalWorkspaceStorage(tmp_path)
    return ProjectEnvironment(metadata=meta, storage=storage)


def test_parse_testsheet_datetime() -> None:
    # Date string + int time 1033 -> 10:33
    res1 = parse_testsheet_datetime("09-06-2026", 1033)
    assert res1 == "2026-06-09T10:33:00+08:00"

    # Date object + time object
    res2 = parse_testsheet_datetime(date(2026, 6, 9), time(14, 17, 6))
    assert res2 == "2026-06-09T14:17:06+08:00"

    # Date string + None time -> 00:00:00
    res3 = parse_testsheet_datetime("09-06-2026", None)
    assert res3 == "2026-06-09T00:00:00+08:00"

    # 3-digit int time 945 -> 09:45
    res4 = parse_testsheet_datetime("09-06-2026", 945)
    assert res4 == "2026-06-09T09:45:00+08:00"


def test_match_vi_defect() -> None:
    defects = [
        ViDefectRecord(
            equipment="SWITCHGEAR",
            defect_area="LOW SF6 GAS",
            additional_remarks="Pressure drop below green mark",
        ),
        ViDefectRecord(
            equipment="SUBSTATION",
            defect_area="BUSHES & CREEPERS",
            additional_remarks="High grass around compound",
        ),
    ]

    # Matched meter
    matched = match_vi_defect("VI11_SG_PRESGAUGE_RMU", "CCHL/PCEJ00024/11KV/1", defects)
    assert matched is not None
    assert matched.additional_remarks == "Pressure drop below green mark"

    # Matched cleanliness
    matched_clean = match_vi_defect("VI11_SUB_CLEANLINESS_RMU", "CCHL/PCEJ00024", defects)
    assert matched_clean is not None
    assert matched_clean.defect_area == "BUSHES & CREEPERS"

    # Backward compatibility: 2-arg call (meter_name, defects)
    matched_2arg = match_vi_defect("VI11_SUB_CLEANLINESS_RMU", defects)
    assert matched_2arg is not None
    assert matched_2arg.defect_area == "BUSHES & CREEPERS"

    # Defensive handling: None tnb_loc
    matched_none_loc = match_vi_defect("VI11_SUB_CLEANLINESS_RMU", None, defects)
    assert matched_none_loc is not None

    # Backward compatibility: VI_METER_KEYWORDS alias
    from src.workflows.populate_data_msms import VI_METER_KEYWORDS
    assert isinstance(VI_METER_KEYWORDS, dict)

    # Unmatched meter
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "CCHL/PCEJ00024/TX/DTX1", defects) is None


def test_row_845_signboard_defect_does_not_match_sg_labelling_rmu() -> None:
    """Regression test for row 845: SIGNBOARD defect must NOT match VI11_SG_LABELLING_RMU."""
    signboard_defect = ViDefectRecord(
        equipment="SIGNBOARD",
        defect_area="NO FUNCTIONAL LOCATION",
        additional_remarks="SUBSTATION SIGNBOARD NO FL NUMBER",
    )

    # Must NOT match Switchgear Labelling meter
    matched_sg = match_vi_defect(
        "VI11_SG_LABELLING_RMU", "CCHL/PCEJ00024/11KV/1", [signboard_defect]
    )
    assert matched_sg is None

    # Must match Substation Signboard meter
    matched_sub = match_vi_defect(
        "VI11_SUB_SIGNBOARD_RMU", "CCHL/PCEJ00024", [signboard_defect]
    )
    assert matched_sub is not None
    assert matched_sub.equipment == "SIGNBOARD"


def test_match_vi_defect_all_groups_and_rules() -> None:
    """Comprehensive test for aligned defect matching rules across Groups A-E."""
    # Group A: Switchgear
    sg_label1 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="NO LINK NO./PANEL NO./FEEDER NAME", additional_remarks="Panel 1")
    sg_label2 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="WRONG LINK NO./PANEL NO./FEEDER NAME", additional_remarks="Panel 2")
    sg_label3 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="LINK NO./PANEL NO./FEEDER NAME LABEL IN POOR CONDITION", additional_remarks="Panel 3")
    assert match_vi_defect("VI11_SG_LABELLING_RMU", "LOC/11KV/1", [sg_label1]) == sg_label1
    assert match_vi_defect("VI11_SG_LABELLING_RMU", "LOC/11KV/1", [sg_label2]) == sg_label2
    assert match_vi_defect("VI11_SG_LABELLING_RMU", "LOC/11KV/1", [sg_label3]) == sg_label3

    sg_gas1 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="LOW SF6 GAS", additional_remarks="Gas low")
    sg_gas2 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="SF6 GAS INDICATOR BROKEN", additional_remarks="Gauge cracked")
    assert match_vi_defect("VI11_SG_PRESGAUGE_RMU", "LOC/11KV/1", [sg_gas1]) == sg_gas1
    assert match_vi_defect("VI11_SG_PRESGAUGE_RMU", "LOC/11KV/1", [sg_gas2]) == sg_gas2

    sg_vdis1 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="VCB STATUS LAMP INDICATOR NOT OPERATED")
    sg_vdis2 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="BREAKER INDICATOR LAMP NOT OPERATED")
    sg_vdis3 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="RELAY NOT OPERATED")
    assert match_vi_defect("VI11_SG_VDIS_RMU", "LOC/11KV/1", [sg_vdis1]) == sg_vdis1
    assert match_vi_defect("VI11_SG_VDIS_RMU", "LOC/11KV/1", [sg_vdis2]) == sg_vdis2
    assert match_vi_defect("VI11_SG_VDIS_RMU", "LOC/11KV/1", [sg_vdis3]) == sg_vdis3

    sg_door = ViDefectRecord(equipment="SWITCHGEAR", defect_area="DOOR BROKEN")
    sg_rust = ViDefectRecord(equipment="SWITCHGEAR", defect_area="BODY RUST")
    assert match_vi_defect("VI11_SG_COVERDOOR_RMU", "LOC/11KV/1", [sg_door]) == sg_door
    assert match_vi_defect("VI11_SWG_DOOR_VCB", "LOC/11KV/1", [sg_door]) == sg_door
    assert match_vi_defect("VI11_SG_COVERDOOR_RMU", "LOC/11KV/1", [sg_rust]) is None

    sg_htr1 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="HEATER NO SUPPLY")
    sg_htr2 = ViDefectRecord(equipment="SWITCHGEAR", defect_area="HEATER INDICATOR LAMP NOT OPERATED")
    assert match_vi_defect("VI11_SG_HEATER_VCB", "LOC/11KV/1", [sg_htr1]) == sg_htr1
    assert match_vi_defect("VI11_SG_HEATER_VCB", "LOC/11KV/1", [sg_htr2]) == sg_htr2

    sg_earth = ViDefectRecord(equipment="EARTHING", additional_remarks="SWG ROOM EARTH DISCONNECTED")
    tx_earth = ViDefectRecord(equipment="EARTHING", additional_remarks="TX1 BODY EARTH")
    assert match_vi_defect("VI11_SG_EARTHIN_RMU", "LOC/11KV/1", [sg_earth]) == sg_earth
    assert match_vi_defect("VI11_SWG_EARTH_VCB", "LOC/11KV/1", [sg_earth]) == sg_earth
    assert match_vi_defect("VI11_SG_EARTHIN_RMU", "LOC/11KV/1", [tx_earth]) is None

    sg_handle = ViDefectRecord(equipment="SWITCHGEAR", defect_area="OPERATING HANDLE BROKEN")
    assert match_vi_defect("VI11_SG_HANDLE_RMU", "LOC/11KV/1", [sg_handle]) == sg_handle

    sg_oilleak = ViDefectRecord(equipment="SWITCHGEAR", defect_area="OIL LEAK AT TANK")
    assert match_vi_defect("VI11_SG_OILLEAK_RMU", "LOC/11KV/1", [sg_oilleak]) == sg_oilleak

    # Group B: Feeder Pillar
    fp_guard = ViDefectRecord(equipment="FP/LVDB", defect_area="NO LVDB GUARD")
    assert match_vi_defect("VI11_FP_LVDBGUARD_RMU", "LOC/FP/FP1", [fp_guard]) == fp_guard

    fp_fuse1 = ViDefectRecord(equipment="FP/LVDB", defect_area="FP (J) FUSE HOLDER MISSING")
    fp_fuse2 = ViDefectRecord(equipment="FP/LVDB", defect_area="FP (J) LINK HOLDER BROKEN")
    fp_casing = ViDefectRecord(equipment="FP/LVDB", defect_area="FP (D) CASING MISSING / BROKEN")
    assert match_vi_defect("VI11_FP_LINK/FUSE_RMU", "LOC/FP/FP1", [fp_fuse1]) == fp_fuse1
    assert match_vi_defect("VI11_FP_LINK/FUSE_RMU", "LOC/FP/FP1", [fp_fuse2]) == fp_fuse2
    assert match_vi_defect("VI11_FP_LINK/FUSE_RMU", "LOC/FP/FP1", [fp_casing]) is None

    fp_door = ViDefectRecord(equipment="FP/LVDB", defect_area="FP DOOR BROKEN")
    sub_plock_fp = ViDefectRecord(equipment="SUBSTATION", defect_area="OLD ABLOY PADLOCK", additional_remarks="FP DOOR PADLOCK")
    assert match_vi_defect("VI11_FP_PLOCK_RMU", "LOC/FP/FP1", [fp_door]) == fp_door
    assert match_vi_defect("VI11_FP_PLOCK_RMU", "LOC/FP/FP1", [sub_plock_fp]) == sub_plock_fp

    fp_tdi = ViDefectRecord(equipment="FP/LVDB", defect_area="TDI BROKEN")
    assert match_vi_defect("VI11_FP_TDI_RMU", "LOC/FP/FP1", [fp_tdi]) == fp_tdi

    # Group C: Transformer & Disambiguation
    tx_guard = ViDefectRecord(equipment="LTX/DTX", defect_area="NO TX GUARD")
    tx_bush = ViDefectRecord(equipment="LTX/DTX", defect_area="NO LV INSULATION BOOT COVER")
    tx_level = ViDefectRecord(equipment="LTX/DTX", defect_area="LOW OIL LEVEL")
    tx_leak = ViDefectRecord(equipment="LTX/DTX", defect_area="OIL LEAKS")
    tx_clamp = ViDefectRecord(equipment="LTX/DTX", defect_area="NO CABLE SUPPORT")
    assert match_vi_defect("VI11_TX_TXGUARD_RMU", "LOC/TX/DTX1", [tx_guard]) == tx_guard
    assert match_vi_defect("VI11_TX_TXBUSH_RMU", "LOC/TX/DTX1", [tx_bush]) == tx_bush
    assert match_vi_defect("VI11_TX_OILLEVEL_RMU", "LOC/TX/DTX1", [tx_level]) == tx_level
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX1", [tx_leak]) == tx_leak
    assert match_vi_defect("VI11_TX_CBLCLMP_RMU", "LOC/TX/DTX1", [tx_clamp]) == tx_clamp

    # Transformer disambiguation
    tx2_only_defect = ViDefectRecord(equipment="LTX/DTX", defect_area="OIL LEAKS", additional_remarks="DTX2 LOW OIL")
    tx1_only_defect = ViDefectRecord(equipment="LTX/DTX", defect_area="OIL LEAKS", additional_remarks="TX1 OIL LEAK")
    both_tx_defect = ViDefectRecord(equipment="LTX/DTX", defect_area="OIL LEAKS", additional_remarks="TX1 & TX2 OIL LEAK")

    # DTX1 location
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX1", [tx2_only_defect]) is None
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX1", [tx1_only_defect]) == tx1_only_defect
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX1", [both_tx_defect]) == both_tx_defect

    # DTX2 location
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX2", [tx1_only_defect]) is None
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX2", [tx2_only_defect]) == tx2_only_defect
    assert match_vi_defect("VI11_TX_OILLEAK_RMU", "LOC/TX/DTX2", [both_tx_defect]) == both_tx_defect

    # Group D: Secondary Equipment
    sec_bat = ViDefectRecord(equipment="BATTERY CHARGER", defect_area="FAULT", additional_remarks="Charger bad")
    assert match_vi_defect("VI11_SEC_BATTERY_RMU", "LOC/SEC", [sec_bat]) == sec_bat
    assert match_vi_defect("VI11_SEC_BADRTU_RMU", "LOC/SEC", [sec_bat]) is None

    sec_efi = ViDefectRecord(equipment="EFI", defect_area="FAULT")
    sec_rcb = ViDefectRecord(equipment="RCB", defect_area="FAULT")
    sec_mc = ViDefectRecord(equipment="MULTICORE", defect_area="FAULT")
    assert match_vi_defect("VI11_SEC_BADEFI_RMU", "LOC/SEC", [sec_efi]) == sec_efi
    assert match_vi_defect("VI11_SEC_BADRCB_RMU", "LOC/SEC", [sec_rcb]) == sec_rcb
    assert match_vi_defect("VI11_SEC_MCORE_RMU", "LOC/SEC", [sec_mc]) == sec_mc

    # Group E: Substation / Civil
    sub_sign = ViDefectRecord(equipment="SIGNBOARD", defect_area="DAMAGED")
    sub_light = ViDefectRecord(equipment="LIGHTING", defect_area="FAULT")
    sub_roof = ViDefectRecord(equipment="SUBSTATION", defect_area="ROOF BROKEN")
    sub_clean = ViDefectRecord(equipment="SUBSTATION", defect_area="SARANG BINATANG")
    sub_vandal = ViDefectRecord(equipment="SUBSTATION", defect_area="FENCE BROKEN")
    sub_padlock_gate = ViDefectRecord(equipment="SUBSTATION", defect_area="OLD ABLOY PADLOCK", additional_remarks="MAIN GATE")
    sub_padlock_fp = ViDefectRecord(equipment="SUBSTATION", defect_area="OLD ABLOY PADLOCK", additional_remarks="FP 1")

    assert match_vi_defect("VI11_SUB_SIGNBOARD_RMU", "LOC", [sub_sign]) == sub_sign
    assert match_vi_defect("VI11_SUB_LIGHT_CSU", "LOC", [sub_light]) == sub_light
    assert match_vi_defect("VI11_SUB_RETROOF_CSU", "LOC", [sub_roof]) == sub_roof
    assert match_vi_defect("VI11_SUB_CLEANLINESS_RMU", "LOC", [sub_clean]) == sub_clean
    assert match_vi_defect("VI11_SUB_VANDALISM_RMU", "LOC", [sub_vandal]) == sub_vandal
    assert match_vi_defect("VI11_SUB_VANDALISM_RMU", "LOC", [sub_padlock_gate]) == sub_padlock_gate
    # Substation padlock with FP remarks excluded from VANDALISM
    assert match_vi_defect("VI11_SUB_VANDALISM_RMU", "LOC", [sub_padlock_fp]) is None


def test_master_qr03_defect_repository_report_by_eet_filter(tmp_path: Path) -> None:
    """Verify MasterQr03DefectRepository filters VI defects to only REPORT BY == 'EET'."""
    engr_wb = openpyxl.Workbook()
    ws_cba = engr_wb.active
    ws_cba.title = "QR03 CBA"
    ws_cba.append(["FUNCTIONAL LOCATION (ERMS)", "EQUIPMENT"])

    ws_vi = engr_wb.create_sheet("QR03 VI")
    ws_vi.append(["FUNCTIONAL LOCATION (ERMS)", "EQUIPMENT", "DEFECT AREA", "ADDITIONAL REMARKS", "REPORT BY"])
    ws_vi.append(["FL-001", "SWITCHGEAR", "LOW SF6 GAS", "EET row 1", "EET"])
    ws_vi.append(["FL-001", "SWITCHGEAR", "LOW SF6 GAS", "Vendor row", "THIRD_PARTY_VENDOR"])
    ws_vi.append(["FL-001", "SWITCHGEAR", "LOW SF6 GAS", "EET row 2", " eet "])
    engr_wb.save(tmp_path / "ENGR-TEST.xlsx")
    engr_wb.close()

    repo = MasterQr03DefectRepository(tmp_path)
    vi_defects = repo.fetch_vi_defects("FL-001")
    assert len(vi_defects) == 2
    assert vi_defects[0].additional_remarks == "EET row 1"
    assert vi_defects[1].additional_remarks == "EET row 2"



def test_preflight_guard_validations(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    guard = PopulateDataMsmsPreflightGuard()
    request = PopulateDataMsmsRequest(mode=PopulateMode.AUTO)

    # 1. Missing TOTAL PE.xlsx
    with pytest.raises(FileNotFoundError, match="TOTAL PE.xlsx not found"):
        guard.validate(mock_env, request)

    # Create TOTAL PE without DataCycle1
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(total_pe_path)
    wb.close()

    # 2. Missing DataCycle1 sheet
    with pytest.raises(RuntimeError, match="'DataCycle1' sheet missing"):
        guard.validate(mock_env, request)

    # Add DataCycle1 with no WOs
    wb = openpyxl.load_workbook(total_pe_path)
    ws = wb.create_sheet("DataCycle1")
    ws.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws.append([1, "FL01", "SUB01", "01-08-2026", "RMU", None, "SCOPE"])
    wb.save(total_pe_path)
    wb.close()

    # 3. No populated Work Orders
    with pytest.raises(RuntimeError, match=r"No populated Work Orders found in TOTAL PE\.xlsx"):
        guard.validate(mock_env, request)

    # Populate a WO
    wb = openpyxl.load_workbook(total_pe_path)
    ws = wb["DataCycle1"]
    ws.cell(2, 6, "200000000001")
    wb.save(total_pe_path)
    wb.close()

    # 4a. Missing MSMS TO BE FILLED directory
    assert mock_env.storage.get_msms_dir() == mock_env.storage.root_path / "PYTHON" / "MSMS"
    assert mock_env.storage.get_msms_raw_data_dir() == mock_env.storage.root_path / "PYTHON" / "MSMS" / "RAW DATA"
    assert mock_env.storage.get_msms_to_be_filled_dir() == mock_env.storage.root_path / "PYTHON" / "MSMS" / "TO BE FILLED"
    assert mock_env.storage.get_msms_completed_dir() == mock_env.storage.root_path / "PYTHON" / "MSMS" / "COMPLETED"

    with pytest.raises(FileNotFoundError, match=r"MSMS TO BE FILLED directory not found"):
        guard.validate(mock_env, request)

    # 4b. Empty MSMS TO BE FILLED directory
    filled_dir = mock_env.storage.get_msms_to_be_filled_dir()
    mock_env.storage.ensure_directory(filled_dir)
    with pytest.raises(FileNotFoundError, match=r"No CSV files found in TO BE FILLED"):
        guard.validate(mock_env, request)

    # Put a CSV in TO BE FILLED
    (filled_dir / "01-08-2026_001.csv").write_text("WONUM,TNBLOCATION,METERNAME\n", encoding="utf-8")

    # 5. Missing TESTSHEET directory
    with pytest.raises(FileNotFoundError, match=r"TESTSHEET directory not found"):
        guard.validate(mock_env, request)

    # Create TESTSHEET directory
    testsheet_dir = mock_env.storage.get_testsheet_dir()
    mock_env.storage.ensure_directory(testsheet_dir)

    # Should pass now!
    guard.validate(mock_env, request)



def test_filter_and_transformer_overwrite_logic(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # 1. Setup minimal testsheet
    date_dir = tmp_path / "TESTSHEET" / "TEMERLOH" / "06. JUNE" / "09-06-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    ts_wb = openpyxl.Workbook()
    ws_ts = ts_wb.active
    ws_ts.title = "PCE Testsheet"
    ws_ts["W5"] = "CCHL/PCEJ00024"
    ws_ts["C5"] = "CENTERPOINT"
    ws_ts["P4"] = "09-06-2026"
    ws_ts["P5"] = 1000
    ws_ts["S5"] = 1030
    ws_ts["N10"] = 42.0  # Cable Avg
    ts_path = date_dir / "066. CENTERPOINT.xlsx"
    ts_wb.save(ts_path)
    ts_wb.close()

    # 2. Setup TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws_pe.append([66, "CCHL/PCEJ00024", "CENTERPOINT", "09-06-2026", "RMU", "200000000001", "CBM"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # 3. Setup CSV with an already filled row and an unmapped WO row
    filled_dir = mock_env.storage.get_msms_to_be_filled_dir()
    mock_env.storage.ensure_directory(filled_dir)
    csv_path = filled_dir / "09-06-2026_001.csv"

    csv_rows = [
        "WONUM,TNBLOCATION,METERNAME,METER.DESCRIPTION,TNBNEWREADING,TNBNEWREADINGDATE,ACTSTART,ACTFINISH,TNBCOMMENTS\n",
        # Row 1: already filled with 25.0
        "200000000001,CCHL/PCEJ00024/11KV/1,TH_S11_RMUCBL1_AVG_PE13R,PCE: TH RMU: Cable Comp 1 Avg Temp,25.0,2026-06-09T08:00:00+08:00,2026-06-09T08:00:00+08:00,2026-06-09T08:30:00+08:00,\n",
        # Row 2: unmapped WO
        "999999999999,CCHL/UNKNOWN/11KV/1,TH_S11_RMUCBL1_AVG_PE13R,PCE: TH RMU: Cable Comp 1 Avg Temp,,,,,\n",
    ]
    csv_path.write_text("".join(csv_rows), encoding="utf-8")

    workflow = PopulateDataMsmsWorkflow()

    # Run with overwrite=False -> row 1 should be skipped as already filled, row 2 skipped as no testsheet
    req_no_overwrite = PopulateDataMsmsRequest(mode=PopulateMode.AUTO, overwrite=False)
    res1 = workflow.execute(mock_env, req_no_overwrite)

    assert res1.rows_skipped_already_filled == 1
    assert res1.rows_skipped_no_testsheet == 1
    assert res1.rows_populated == 0

    df1 = pd.read_csv(csv_path, dtype=str).fillna("")
    assert df1.iloc[0]["TNBNEWREADING"] == "25.0"

    # Run with overwrite=True -> row 1 should now be overwritten to 42.0
    req_overwrite = PopulateDataMsmsRequest(mode=PopulateMode.AUTO, overwrite=True)
    res2 = workflow.execute(mock_env, req_overwrite)

    assert res2.rows_skipped_already_filled == 0
    assert res2.rows_skipped_no_testsheet == 1
    assert res2.rows_populated == 1

    df2 = pd.read_csv(csv_path, dtype=str).fillna("")
    assert df2.iloc[0]["TNBNEWREADING"] in ("42", "42.0")



def test_rmu_centerpoint_population_end_to_end(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # 1. Setup TESTSHEET for RMU (CENTERPOINT)
    date_dir = tmp_path / "TESTSHEET" / "TEMERLOH" / "06. JUNE" / "09-06-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    ts_wb = openpyxl.Workbook()
    ws_ts = ts_wb.active
    ws_ts.title = "PCE Testsheet"
    ws_ts["W5"] = "CCHL/PCEJ00024"
    ws_ts["C5"] = "CENTERPOINT"
    ws_ts["P4"] = "09-06-2026"
    ws_ts["P5"] = 1033
    ws_ts["S5"] = 1048
    ws_ts["P6"] = 12.0  # Background TEV
    ws_ts["S6"] = 65.0  # Humidity
    ws_ts["W6"] = "BACKGROUND TEMP : 32.5 °C"

    # RMU Cable 1 (Slot 1, row 10)
    ws_ts["N10"] = 28.5  # Avg
    ws_ts["L10"] = 29.0  # Max
    ws_ts["K10"] = 28.0  # Ref
    ws_ts["M10"] = 1.0   # Dif
    ws_ts["T10"] = 15.0  # TEV dB
    ws_ts["U10"] = 0.0   # TEV Pulse
    ws_ts["Q10"] = 0.0   # US dB

    # RMU Body (Overview row 26)
    ws_ts["N26"] = 31.0  # Body Avg

    # TX1 (DTX1)
    ws_ts["I33"] = 45.0  # HT Cable Avg
    ws_ts["G33"] = 46.0  # HT Cable Max
    ws_ts["F33"] = 44.0  # HT Cable Ref
    ws_ts["H33"] = 2.0   # HT Cable Dif
    ws_ts["K33"] = 0.0   # HT Cable US
    ws_ts["I37"] = 38.0  # TX Body Avg

    ts_path = date_dir / "066. CENTERPOINT.xlsx"
    ts_wb.save(ts_path)
    ts_wb.close()

    # 2. Setup TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws_pe.append([66, "CCHL/PCEJ00024", "CENTERPOINT", "09-06-2026", "RMU", "200000000001", "CBM"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # 3. Setup ENGR master file with QR03 VI defect
    engr_dir = mock_env.storage.get_engr_folder()
    mock_env.storage.ensure_directory(engr_dir)
    engr_wb = openpyxl.Workbook()
    # Sheet 1: QR03 CBA
    ws_cba = engr_wb.active
    ws_cba.title = "QR03 CBA"
    ws_cba.append(["FUNCTIONAL LOCATION (ERMS)", "DEFECT AREA", "READING", "REMARKS"])
    # Sheet 2: QR03 VI
    ws_vi = engr_wb.create_sheet("QR03 VI")
    ws_vi.append(["FUNCTIONAL LOCATION (ERMS)", "EQUIPMENT", "DEFECT AREA", "ADDITIONAL REMARKS", "REPORT BY"])
    ws_vi.append(["CCHL/PCEJ00024", "SWITCHGEAR", "LOW SF6 GAS", "Pressure drop below green mark", "EET"])
    engr_path = engr_dir / "ENGR-750-36-CBA-TEM-2026.xlsx"
    engr_wb.save(engr_path)
    engr_wb.close()

    # 4. Setup MSMS CSV in TO BE FILLED/
    filled_dir = mock_env.storage.get_msms_to_be_filled_dir()
    mock_env.storage.ensure_directory(filled_dir)
    csv_path = filled_dir / "09-06-2026_001.csv"

    csv_rows = [
        "WONUM,TNBLOCATION,METERNAME,METER.DESCRIPTION,TNBNEWREADING,TNBNEWREADINGDATE,ACTSTART,ACTFINISH,TNBCOMMENTS\n",
        "200000000001,CCHL/PCEJ00024/TX/DTX1,BG_ROOM_TEM,PCE: PE Room Background Temperature,,,,,\n",
        "200000000001,CCHL/PCEJ00024/TX/DTX1,BG_ROOM_HUM,PCE: PE Room Background Humidity,,,,,\n",
        "200000000001,CCHL/PCEJ00024/TX/DTX1,BG_ROOM_TV,PCE: PE Room Surrounding TEV,,,,,\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,TH_S11_RMUCBL1_AVG_PE13R,PCE: TH RMU: Cable Comp 1 Avg Temp,,,,,\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,TH_S11_RMUCBL1_MAX_PE13R,PCE: TH RMU: Cable Comp 1 Max Temp,,,,,\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,TV_S11_CBL_PE13R,PCE: TEV RMU: Cable Comp 1 Reading,,,,,\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,US_S11_RMU_PE13R,PCE: US RMU: Body,,,,,\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,TH_S11_RMU_AVG_PE13R,PCE: TH RMU Body: Avg Temp,,,,,\n",
        "200000000001,CCHL/PCEJ00024/TX/DTX1,TH_DTX_HV_AVG_PE13R,PCE: Thermo TX HV : Avg Temp,,,,,\n",
        "200000000001,CCHL/PCEJ00024/TX/DTX1,TH_TX_RMU_AVG_PE13R,PCE: Thermo TX Body : Avg Temp,,,,,\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,VI11_SG_PRESGAUGE_RMU,SF6 Pressure Gauge,,,,,\n",
        "200000000001,CCHL/PCEJ00024/TX/DTX1,VI11_TX_OILLEAK_RMU,Oil Leak,,,,,\n",
        "200000000001,CCHL/PCEJ00024/FP/FP1,TH_FPIN1_AVG_PE13R,PCE: TH FP/LVDB In 1: Avg Temp,,,,,\n",
    ]
    csv_path.write_text("".join(csv_rows), encoding="utf-8")

    # 5. Execute PopulateDataMsmsWorkflow
    workflow = PopulateDataMsmsWorkflow()
    request = PopulateDataMsmsRequest(mode=PopulateMode.AUTO)
    result = workflow.execute(mock_env, request)

    assert result.csv_files_processed == 1
    assert result.rows_populated > 0

    # 6. Verify populated CSV content
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    
    # BG_ROOM_TEM
    row_bg_tem = df[df["METERNAME"] == "BG_ROOM_TEM"].iloc[0]
    assert row_bg_tem["TNBNEWREADING"] == "32.5"
    assert row_bg_tem["TNBNEWREADINGDATE"] == "2026-06-09T10:33:00+08:00"
    assert row_bg_tem["ACTSTART"] == "2026-06-09T10:33:00+08:00"
    assert row_bg_tem["ACTFINISH"] == "2026-06-09T10:48:00+08:00"

    # BG_ROOM_HUM
    row_bg_hum = df[df["METERNAME"] == "BG_ROOM_HUM"].iloc[0]
    assert row_bg_hum["TNBNEWREADING"] in ("65", "65.0")

    # BG_ROOM_TV
    row_bg_tv = df[df["METERNAME"] == "BG_ROOM_TV"].iloc[0]
    assert row_bg_tv["TNBNEWREADING"] in ("12", "12.0")

    # Cable 1 Avg
    row_cbl = df[df["METERNAME"] == "TH_S11_RMUCBL1_AVG_PE13R"].iloc[0]
    assert row_cbl["TNBNEWREADING"] in ("28.5", "28.50")

    # RMU Body Avg
    row_body = df[df["METERNAME"] == "TH_S11_RMU_AVG_PE13R"].iloc[0]
    assert row_body["TNBNEWREADING"] in ("31", "31.0")

    # RMU Body US (Hardcoded 0)
    row_us = df[df["METERNAME"] == "US_S11_RMU_PE13R"].iloc[0]
    assert row_us["TNBNEWREADING"] in ("0", "0.0")

    # TX HV Avg
    row_tx_hv = df[df["METERNAME"] == "TH_DTX_HV_AVG_PE13R"].iloc[0]
    assert row_tx_hv["TNBNEWREADING"] in ("45", "45.0")

    # VI Defect (Pressure Gauge)
    row_vi_gauge = df[df["METERNAME"] == "VI11_SG_PRESGAUGE_RMU"].iloc[0]
    assert row_vi_gauge["TNBNEWREADING"] == "YES"
    assert row_vi_gauge["TNBCOMMENTS"] == "Pressure drop below green mark"
    assert row_vi_gauge["TNBNEWREADINGDATE"] == "2026-06-09T10:33:00+08:00"

    # VI No Defect (Oil Leak) -> Row untouched / blank
    row_vi_oil = df[df["METERNAME"] == "VI11_TX_OILLEAK_RMU"].iloc[0]
    assert row_vi_oil["TNBNEWREADING"] == ""
    assert row_vi_oil["TNBCOMMENTS"] == ""

    # Feeder Pillar Stub -> TNBNEWREADING blank
    row_fp = df[df["METERNAME"] == "TH_FPIN1_AVG_PE13R"].iloc[0]
    assert row_fp["TNBNEWREADING"] == ""


def test_vcb_bukit_rangin_population_end_to_end(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # 1. Setup TESTSHEET for VCB (SSU BUKIT RANGIN) across 2 sheets
    date_dir = tmp_path / "TESTSHEET" / "KUANTAN" / "07. JULY" / "12-07-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    ts_wb = openpyxl.Workbook()
    # Sheet 1: Panels 1-4
    ws_ts1 = ts_wb.active
    ws_ts1.title = "PCE Testsheet"
    ws_ts1["W5"] = "CRAU/PCEJ00199"
    ws_ts1["C5"] = "SSU BUKIT RANGIN"
    ws_ts1["P4"] = "12-07-2026"
    ws_ts1["P5"] = 915
    ws_ts1["S5"] = 945

    # Panel 1 (Slot 1, rows 10-13)
    ws_ts1["N10"] = 30.0  # CBL Avg
    ws_ts1["N11"] = 32.0  # BR Avg
    ws_ts1["N12"] = 31.5  # BB Avg
    ws_ts1["N13"] = 29.5  # PT Avg

    # Sheet 2: Panels 5-8 (Rollover)
    ws_ts2 = ts_wb.create_sheet("PCE Testsheet (2)")
    # Panel 7 (Slot 3, rows 18-21 on Sheet 2)
    ws_ts2["N18"] = 33.0  # CBL Avg
    ws_ts2["N19"] = 34.0  # BR Avg
    ws_ts2["N20"] = 33.5  # BB Avg
    ws_ts2["N21"] = 31.0  # PT Avg
    ws_ts2["L21"] = 32.0  # PT Max
    ws_ts2["K21"] = 30.0  # PT Ref
    ws_ts2["M21"] = 2.0   # PT Dif

    ts_path = date_dir / "064. SSU BUKIT RANGIN.xlsx"
    ts_wb.save(ts_path)
    ts_wb.close()

    # 2. Setup TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws_pe.append([64, "CRAU/PCEJ00199", "SSU BUKIT RANGIN", "12-07-2026", "VCB", "200000000002", "CBM"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # 3. Setup MSMS CSV in TO BE FILLED/
    filled_dir = mock_env.storage.get_msms_to_be_filled_dir()
    mock_env.storage.ensure_directory(filled_dir)
    csv_path = filled_dir / "12-07-2026_001.csv"

    csv_rows = [
        "WONUM,TNBLOCATION,METERNAME,METER.DESCRIPTION,TNBNEWREADING,TNBNEWREADINGDATE,ACTSTART,ACTFINISH,TNBCOMMENTS\n",
        # Panel 1
        "200000000002,CRAU/PCEJ00199/11KV/1,TH_S11_CBL_AVG_PE13V,PCE: TH VCB Cable Comp: Avg Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/1,TH_S11_BR_AVG_PE13V,PCE: TH VCB Breaker Comp: Avg Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/1,TH_S11_BB_AVG_PE13V,PCE: TH VCB Busbar Comp: Avg Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/1,TH_S11_PT_AVG_PE13V2,PCE: TH VCB PT Comp: Avg Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/1,TH_S11_LV_AVG_PE13V,PCE: TH VCB LV Comp: Avg Temp,,,,,\n",
        # Panel 7
        "200000000002,CRAU/PCEJ00199/11KV/7,TH_S11_CBL_AVG_PE13V,PCE: TH VCB Cable Comp: Avg Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/7,TH_S11_PT_AVG_PE13V2,PCE: TH VCB PT Comp: Avg Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/7,TH_S11_PT_MAX_PE13V2,PCE: TH VCB PT Comp: Max Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/7,TH_S11_PT_REF_PE13V2,PCE: TH VCB PT Comp: Ref Temp,,,,,\n",
        "200000000002,CRAU/PCEJ00199/11KV/7,TH_S11_PT_DIF_PE13V2,PCE: TH VCB PT Comp: Temp Diff,,,,,\n",
    ]
    csv_path.write_text("".join(csv_rows), encoding="utf-8")

    # 4. Execute PopulateDataMsmsWorkflow
    workflow = PopulateDataMsmsWorkflow()
    request = PopulateDataMsmsRequest(mode=PopulateMode.AUTO)
    result = workflow.execute(mock_env, request)

    assert result.csv_files_processed == 1

    # 5. Verify populated CSV content
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    # Panel 1 readings
    p1_cbl = df[(df["TNBLOCATION"].str.endswith("/11KV/1")) & (df["METERNAME"] == "TH_S11_CBL_AVG_PE13V")].iloc[0]
    assert p1_cbl["TNBNEWREADING"] in ("30", "30.0")
    assert p1_cbl["ACTSTART"] == "2026-07-12T09:15:00+08:00"

    p1_pt = df[(df["TNBLOCATION"].str.endswith("/11KV/1")) & (df["METERNAME"] == "TH_S11_PT_AVG_PE13V2")].iloc[0]
    assert p1_pt["TNBNEWREADING"] in ("29.5", "29.50")

    # LV stub is blank
    p1_lv = df[(df["TNBLOCATION"].str.endswith("/11KV/1")) & (df["METERNAME"] == "TH_S11_LV_AVG_PE13V")].iloc[0]
    assert p1_lv["TNBNEWREADING"] == ""

    # Panel 7 readings (Rollover to Sheet 2)
    p7_cbl = df[(df["TNBLOCATION"].str.endswith("/11KV/7")) & (df["METERNAME"] == "TH_S11_CBL_AVG_PE13V")].iloc[0]
    assert p7_cbl["TNBNEWREADING"] in ("33", "33.0")

    p7_pt_max = df[(df["TNBLOCATION"].str.endswith("/11KV/7")) & (df["METERNAME"] == "TH_S11_PT_MAX_PE13V2")].iloc[0]
    assert p7_pt_max["TNBNEWREADING"] in ("32", "32.0")

    p7_pt_dif = df[(df["TNBLOCATION"].str.endswith("/11KV/7")) & (df["METERNAME"] == "TH_S11_PT_DIF_PE13V2")].iloc[0]
    assert p7_pt_dif["TNBNEWREADING"] in ("2", "2.0")


def test_populate_data_msms_from_python_to_be_filled(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # 1. Setup minimal testsheet
    date_dir = tmp_path / "TESTSHEET" / "TEMERLOH" / "06. JUNE" / "09-06-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    ts_wb = openpyxl.Workbook()
    ws_ts = ts_wb.active
    ws_ts.title = "PCE Testsheet"
    ws_ts["W5"] = "CCHL/PCEJ00024"
    ws_ts["C5"] = "CENTERPOINT"
    ws_ts["P4"] = "09-06-2026"
    ws_ts["P5"] = 1000
    ws_ts["S5"] = 1030
    ws_ts["N10"] = 42.0
    ts_path = date_dir / "066. CENTERPOINT.xlsx"
    ts_wb.save(ts_path)
    ts_wb.close()

    # 2. Setup TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws_pe.append([66, "CCHL/PCEJ00024", "CENTERPOINT", "09-06-2026", "RMU", "200000000001", "CBM"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # 3. Setup CSV ONLY in PYTHON/MSMS/TO BE FILLED
    python_filled_dir = tmp_path / "PYTHON" / "MSMS" / "TO BE FILLED"
    python_filled_dir.mkdir(parents=True)
    csv_path = python_filled_dir / "09-06-2026_001.csv"

    csv_rows = [
        "WONUM,TNBLOCATION,METERNAME,METER.DESCRIPTION,TNBNEWREADING,TNBNEWREADINGDATE,ACTSTART,ACTFINISH,TNBCOMMENTS\n",
        "200000000001,CCHL/PCEJ00024/11KV/1,TH_S11_RMUCBL1_AVG_PE13R,PCE: TH RMU: Cable Comp 1 Avg Temp,,,,,\n",
    ]
    csv_path.write_text("".join(csv_rows), encoding="utf-8")

    # 4. Execute workflow
    workflow = PopulateDataMsmsWorkflow()
    req = PopulateDataMsmsRequest(mode=PopulateMode.AUTO)
    result = workflow.execute(mock_env, req)

    assert result.csv_files_processed == 1
    assert result.rows_populated == 1

    df = pd.read_csv(csv_path, dtype=str).fillna("")
    assert df.iloc[0]["TNBNEWREADING"] in ("42", "42.0")
    assert df.iloc[0]["ACTSTART"] == "2026-06-09T10:00:00+08:00"


def test_populate_data_msms_feeder_pillar_synthesis(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # 1. Setup testsheet with LVDB 1 and LVDB 2 feeder details
    date_dir = tmp_path / "TESTSHEET" / "TEMERLOH" / "06. JUNE" / "09-06-2026"
    date_dir.mkdir(parents=True, exist_ok=True)
    (date_dir / "UNSORTED RAW DATA").mkdir(exist_ok=True)

    ts_wb = openpyxl.Workbook()
    ws_ts = ts_wb.active
    ws_ts.title = "PCE Testsheet"
    ws_ts["W5"] = "CCHL/PCEJ00099"
    ws_ts["C5"] = "TEST SUBSTATION"
    ws_ts["P4"] = "09-06-2026"
    ws_ts["P5"] = 1000
    ws_ts["S5"] = 1030

    # LVDB 1 (Row 44 config, Row 45 cable type, R50 avg temp)
    ws_ts["D44"] = "TX1"
    ws_ts["D45"] = "XLPE"           # IN1 active
    ws_ts["I44"] = "FEEDER 1"
    ws_ts["I45"] = "PILC"           # OT1 active
    ws_ts["J44"] = "SPARE WAY"
    ws_ts["J45"] = "SPARE"          # OT2 spare/inactive
    ws_ts["K44"] = "-"
    ws_ts["K45"] = "-"              # OT3 inactive
    ws_ts["R50"] = "AVG 28.0"       # LVDB 1 base temp (string formatted as in real testsheets)

    # LVDB 2 (Row 46 config, Row 47 cable type, R54 avg temp)
    ws_ts["D46"] = "TX2"
    ws_ts["D47"] = "BUSBAR"         # IN1 active
    ws_ts["I46"] = "FEEDER 2"
    ws_ts["I47"] = "ABC"            # OT1 active
    ws_ts["R54"] = "AVG 29.5"       # LVDB 2 base temp (string formatted as in real testsheets)

    ts_path = date_dir / "099. TEST SUBSTATION.xlsx"
    ts_wb.save(ts_path)
    ts_wb.close()

    # 2. Setup TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws_pe.append([99, "CCHL/PCEJ00099", "TEST SUBSTATION", "09-06-2026", "PE", "200000000099", "CBM"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # 3. Setup CSV in TO BE FILLED
    to_be_filled_dir = mock_env.storage.get_msms_to_be_filled_dir()
    mock_env.storage.ensure_directory(to_be_filled_dir)
    csv_path = to_be_filled_dir / "09-06-2026_001.csv"

    csv_rows = [
        "WONUM,TNBLOCATION,METERNAME,METER.DESCRIPTION,TNBNEWREADING,TNBNEWREADINGDATE,ACTSTART,ACTFINISH,TNBCOMMENTS\n",
        # FP1 Incomer 1 (Active -> Should be populated)
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPIN1_AVG_PE13R,PCE: TH FP IN1: Avg Temp,,,,,\n",
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPIN1_MAX_PE13R,PCE: TH FP IN1: Max Temp,,,,,\n",
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPIN1_REF_PE13R,PCE: TH FP IN1: Ref Temp,,,,,\n",
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPIN1_DEL_PE13R,PCE: TH FP IN1: Temp Diff,,,,,\n",
        # FP1 Outgoing 1 (Active -> Should be populated)
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPOT1_AVG_PE13R,PCE: TH FP OT1: Avg Temp,,,,,\n",
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPOT1_DEL_PE13R,PCE: TH FP OT1: Temp Diff,,,,,\n",
        # FP1 Outgoing 2 (SPARE -> Must remain blank)
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPOT2_AVG_PE13R,PCE: TH FP OT2: Avg Temp,,,,,\n",
        # FP1 Outgoing 3 (Dash '-' -> Must remain blank)
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_FPOT3_AVG_PE13R,PCE: TH FP OT3: Avg Temp,,,,,\n",
        # FP1 Earth (Must remain blank)
        "200000000099,CCHL/PCEJ00099/FP/FP1,TH_EARTH_AVG_PE13R,PCE: TH FP Earth: Avg Temp,,,,,\n",
        # FP2 Incomer 1 (Active -> Should be populated)
        "200000000099,CCHL/PCEJ00099/FP/FP2,TH_FPIN1_AVG_PE13R,PCE: TH FP2 IN1: Avg Temp,,,,,\n",
    ]
    csv_path.write_text("".join(csv_rows), encoding="utf-8")

    # 4. Execute workflow
    workflow = PopulateDataMsmsWorkflow()
    req = PopulateDataMsmsRequest(mode=PopulateMode.AUTO)
    result = workflow.execute(mock_env, req)
    assert result.rows_populated > 0
    assert not result.errors

    # 5. Verify results
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    # FP1 IN1
    fp1_in1_avg = df[df["METERNAME"] == "TH_FPIN1_AVG_PE13R"].iloc[0]
    assert fp1_in1_avg["TNBNEWREADING"] != ""
    assert 27.5 <= float(fp1_in1_avg["TNBNEWREADING"]) <= 28.5
    assert fp1_in1_avg["ACTSTART"] == "2026-06-09T10:00:00+08:00"

    fp1_in1_max = float(df[df["METERNAME"] == "TH_FPIN1_MAX_PE13R"].iloc[0]["TNBNEWREADING"])
    fp1_in1_ref = float(df[df["METERNAME"] == "TH_FPIN1_REF_PE13R"].iloc[0]["TNBNEWREADING"])
    fp1_in1_del = float(df[df["METERNAME"] == "TH_FPIN1_DEL_PE13R"].iloc[0]["TNBNEWREADING"])
    assert round(fp1_in1_max - fp1_in1_ref, 1) == fp1_in1_del
    assert fp1_in1_del < 1.0

    # FP1 OT1
    fp1_ot1_avg = df[df["METERNAME"] == "TH_FPOT1_AVG_PE13R"].iloc[0]
    assert fp1_ot1_avg["TNBNEWREADING"] != ""

    # Inactive feeders must be blank
    fp1_ot2_avg = df[df["METERNAME"] == "TH_FPOT2_AVG_PE13R"].iloc[0]
    assert fp1_ot2_avg["TNBNEWREADING"] == ""

    fp1_ot3_avg = df[df["METERNAME"] == "TH_FPOT3_AVG_PE13R"].iloc[0]
    assert fp1_ot3_avg["TNBNEWREADING"] == ""

    fp1_earth = df[df["METERNAME"] == "TH_EARTH_AVG_PE13R"].iloc[0]
    assert fp1_earth["TNBNEWREADING"] == ""

    # FP2 IN1 (board avg was 29.5)
    fp2_in1_avg = df[(df["TNBLOCATION"].str.endswith("/FP/FP2")) & (df["METERNAME"] == "TH_FPIN1_AVG_PE13R")].iloc[0]
    assert fp2_in1_avg["TNBNEWREADING"] != ""
    assert 29.0 <= float(fp2_in1_avg["TNBNEWREADING"]) <= 30.0


