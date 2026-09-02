"""Unit tests for Populate TOTAL PE pipeline stage components."""

from __future__ import annotations

from pathlib import Path
import pytest

from src.testsheet.models import SubstationTestsheetPackage, TestsheetData
from src.workflows.models import PopulateMode
from src.workflows.populate_total_pe import (
    PopulateTotalPeExtractor,
    PopulateTotalPeFilter,
    PopulateTotalPeLoader,
    PopulateTotalPePlan,
    PopulateTotalPeTransformer,
)


@pytest.fixture
def sample_packages() -> list[SubstationTestsheetPackage]:
    pkg1 = SubstationTestsheetPackage(
        testsheet_path=Path("/testsheets/RAUB/01. MAY/01-05-2026/001. SSU CHEROH.xlsx"),
        unsorted_raw_data_dir=Path("/testsheets/RAUB/01. MAY/01-05-2026/UNSORTED RAW DATA"),
        station="RAUB",
        month="01. MAY",
        date_str="01-05-2026",
        substation_number=1,
        data=TestsheetData(substation_number=1, substation_name_erms="SSU CHEROH"),
    )
    pkg2 = SubstationTestsheetPackage(
        testsheet_path=Path("/testsheets/RAUB/01. MAY/02-05-2026/002. PPU BENTA.xlsx"),
        unsorted_raw_data_dir=Path("/testsheets/RAUB/01. MAY/02-05-2026/UNSORTED RAW DATA"),
        station="RAUB",
        month="01. MAY",
        date_str="02-05-2026",
        substation_number=2,
        data=TestsheetData(substation_number=2, substation_name_erms="PPU BENTA"),
    )
    pkg3 = SubstationTestsheetPackage(
        testsheet_path=Path("/testsheets/LIPIS/02. JUNE/10-06-2026/015. PE TRAS.xlsx"),
        unsorted_raw_data_dir=Path("/testsheets/LIPIS/02. JUNE/10-06-2026/UNSORTED RAW DATA"),
        station="LIPIS",
        month="02. JUNE",
        date_str="10-06-2026",
        substation_number=15,
        data=None,
    )
    return [pkg1, pkg2, pkg3]


def test_filter_packages_all_mode(sample_packages: list[SubstationTestsheetPackage]) -> None:
    filter_stage = PopulateTotalPeFilter()
    res = filter_stage.filter_packages(sample_packages, mode=PopulateMode.ALL)
    assert len(res) == 3
    assert res == sample_packages


def test_filter_packages_specific_folders_matching(
    sample_packages: list[SubstationTestsheetPackage],
) -> None:
    filter_stage = PopulateTotalPeFilter()
    # Match by date string
    res_date = filter_stage.filter_packages(
        sample_packages,
        mode=PopulateMode.SPECIFIC_FOLDERS,
        target_folder_names=["01-05-2026"],
    )
    assert len(res_date) == 1
    assert res_date[0].substation_number == 1

    # Match by path substring
    res_path = filter_stage.filter_packages(
        sample_packages,
        mode=PopulateMode.SPECIFIC_FOLDERS,
        target_folder_names=["LIPIS"],
    )
    assert len(res_path) == 1
    assert res_path[0].substation_number == 15


def test_filter_packages_specific_folders_no_match(
    sample_packages: list[SubstationTestsheetPackage],
) -> None:
    filter_stage = PopulateTotalPeFilter()
    res = filter_stage.filter_packages(
        sample_packages,
        mode=PopulateMode.SPECIFIC_FOLDERS,
        target_folder_names=["NONEXISTENT_FOLDER"],
    )
    assert len(res) == 0


def test_filter_packages_auto_mode_filtering(
    sample_packages: list[SubstationTestsheetPackage],
) -> None:
    filter_stage = PopulateTotalPeFilter()
    # Mock existing keys in TOTAL PE containing PE 1 ("1", "01-05-2026")
    existing_keys = {("1", "01-05-2026")}

    res = filter_stage.filter_packages(
        sample_packages,
        mode=PopulateMode.AUTO,
        existing_auto_keys=existing_keys,
    )
    # Package 1 filtered out, Packages 2 & 3 kept
    assert len(res) == 2
    assert [p.substation_number for p in res] == [2, 15]


def test_filter_packages_auto_mode_substation_key(
    sample_packages: list[SubstationTestsheetPackage],
) -> None:
    filter_stage = PopulateTotalPeFilter()
    # Mock existing keys containing substation name "PPU BENTA" (uppercase)
    existing_keys = {("PPU BENTA", "02-05-2026")}

    res = filter_stage.filter_packages(
        sample_packages,
        mode=PopulateMode.AUTO,
        existing_auto_keys=existing_keys,
    )
    # Package 2 filtered out, Packages 1 & 3 kept
    assert len(res) == 2
    assert [p.substation_number for p in res] == [1, 15]


def test_transformer_build_plan(sample_packages: list[SubstationTestsheetPackage]) -> None:
    transformer = PopulateTotalPeTransformer()
    total_pe_path = Path("/path/to/TOTAL PE.xlsx")

    plan = transformer.build_plan(total_pe_path, sample_packages)

    assert isinstance(plan, PopulateTotalPePlan)
    assert plan.total_pe_path == total_pe_path
    assert plan.packages == tuple(sample_packages)


def test_total_pe_repository_raises_runtime_error_when_datacycle1_missing(
    tmp_path: Path, sample_packages: list[SubstationTestsheetPackage]
) -> None:
    import openpyxl
    from src.master.total_pe import LocalExcelTotalPeRepository

    excel_path = tmp_path / "TOTAL PE.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WrongSheetName"
    wb.save(excel_path)
    wb.close()

    repo = LocalExcelTotalPeRepository()

    with pytest.raises(RuntimeError, match="'DataCycle1' sheet missing"):
        repo.get_existing_auto_keys(excel_path)

    with pytest.raises(RuntimeError, match="'DataCycle1' sheet missing"):
        repo.upsert_packages(excel_path, sample_packages)


def test_extractor_delegates_to_repositories(tmp_path: Path, sample_packages: list[SubstationTestsheetPackage]) -> None:
    class DummyTestsheetRepo:
        def discover_packages(self, path: Path) -> list[SubstationTestsheetPackage]:
            return sample_packages

    class DummyTotalPeRepo:
        def get_existing_auto_keys(self, path: Path) -> set[tuple[str, str]]:
            return {("1", "01-05-2026")}
        def upsert_packages(self, path: Path, packages: Sequence[SubstationTestsheetPackage]) -> tuple[int, int]:
            return 1, 0

    extractor = PopulateTotalPeExtractor(
        repository=DummyTestsheetRepo(),  # type: ignore[arg-type]
        total_pe_repository=DummyTotalPeRepo(),  # type: ignore[arg-type]
    )

    discovered = extractor.discover_packages(tmp_path)
    assert len(discovered) == 3

    keys = extractor.get_existing_auto_keys(tmp_path / "TOTAL PE.xlsx")
    assert ("1", "01-05-2026") in keys


def test_loader_delegates_to_repository(tmp_path: Path, sample_packages: list[SubstationTestsheetPackage]) -> None:
    class DummyTotalPeRepo:
        def get_existing_auto_keys(self, path: Path) -> set[tuple[str, str]]:
            return set()
        def upsert_packages(self, path: Path, packages: Sequence[SubstationTestsheetPackage]) -> tuple[int, int]:
            return len(packages), 2


    loader = PopulateTotalPeLoader(total_pe_repository=DummyTotalPeRepo())  # type: ignore[arg-type]
    plan = PopulateTotalPePlan(total_pe_path=tmp_path / "TOTAL PE.xlsx", packages=tuple(sample_packages))
    new_cnt, upd_cnt = loader.upsert_packages(plan)

    assert new_cnt == 3
    assert upd_cnt == 2


def test_preflight_guard_raises_file_not_found_on_missing_dir(tmp_path: Path) -> None:
    from src.project.environment import ProjectEnvironment
    from src.project.models import ProjectMetadata
    from src.project.storage import LocalWorkspaceStorage
    from src.workflows.populate_total_pe import PopulateTotalPePreflightGuard

    meta = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    env = ProjectEnvironment(metadata=meta, storage=LocalWorkspaceStorage(tmp_path))
    guard = PopulateTotalPePreflightGuard()

    with pytest.raises(FileNotFoundError, match="TESTSHEET directory not found"):
        guard.validate(env)


def test_auditor_raises_runtime_error_on_missing_or_empty_output(
    tmp_path: Path, sample_packages: list[SubstationTestsheetPackage]
) -> None:
    from src.project.environment import ProjectEnvironment
    from src.project.models import ProjectMetadata
    from src.project.storage import LocalWorkspaceStorage
    from src.workflows.populate_total_pe import PopulateTotalPeAuditor

    meta = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    env = ProjectEnvironment(metadata=meta, storage=LocalWorkspaceStorage(tmp_path))
    auditor = PopulateTotalPeAuditor()

    missing_plan = PopulateTotalPePlan(
        total_pe_path=tmp_path / "NON_EXISTENT.xlsx", packages=tuple(sample_packages)
    )
    with pytest.raises(RuntimeError, match="does not exist after load"):
        auditor.audit(env, missing_plan, (1, 0))

    empty_file = tmp_path / "EMPTY.xlsx"
    empty_file.touch()
    empty_plan = PopulateTotalPePlan(total_pe_path=empty_file, packages=tuple(sample_packages))
    with pytest.raises(RuntimeError, match="empty \\(0 bytes\\) after load"):
        auditor.audit(env, empty_plan, (1, 0))


def test_filter_helper_is_package_in_keys(
    sample_packages: list[SubstationTestsheetPackage],
) -> None:
    filter_stage = PopulateTotalPeFilter()
    pkg = sample_packages[0]

    # Matching normalized date or padded sub num or ERMS name
    assert filter_stage._is_package_in_keys(pkg, {("1", "01/05/2026")}) is True
    assert filter_stage._is_package_in_keys(pkg, {("001", "01-05-2026")}) is True
    assert filter_stage._is_package_in_keys(pkg, {("SSU CHEROH", "01-05-2026")}) is True
    assert filter_stage._is_package_in_keys(pkg, {("999", "01-05-2026")}) is False


def test_resolve_target_folders(tmp_path: Path) -> None:
    from src.workflows.populate_total_pe import resolve_target_folders

    date_dir1 = tmp_path / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir1.mkdir(parents=True)
    date_dir2 = tmp_path / "RAUB" / "01. MAY" / "02-05-2026"
    date_dir2.mkdir(parents=True)

    # 1. Resolve by date string name
    res1 = resolve_target_folders(tmp_path, ["01-05-2026"])
    assert len(res1) == 1
    assert res1[0] == date_dir1.resolve()

    # 2. Resolve by absolute path and folder name deduplication
    res2 = resolve_target_folders(tmp_path, ["01-05-2026", str(date_dir1)])
    assert len(res2) == 1
    assert res2[0] == date_dir1.resolve()

    # 3. Non-existent folder returns empty
    res3 = resolve_target_folders(tmp_path, ["NON_EXISTENT"])
    assert len(res3) == 0


def test_extractor_scoped_discovery_specific_folders(tmp_path: Path) -> None:
    from src.testsheet.repository import SubstationTestsheetRepository
    import openpyxl

    date_dir1 = tmp_path / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir1.mkdir(parents=True)
    (date_dir1 / "UNSORTED RAW DATA").mkdir()
    wb1 = openpyxl.Workbook()
    wb1.save(date_dir1 / "001. SSU CHEROH.xlsx")
    wb1.close()

    date_dir2 = tmp_path / "RAUB" / "01. MAY" / "02-05-2026"
    date_dir2.mkdir(parents=True)
    (date_dir2 / "UNSORTED RAW DATA").mkdir()
    wb2 = openpyxl.Workbook()
    wb2.save(date_dir2 / "002. PPU BENTA.xlsx")
    wb2.close()

    extractor = PopulateTotalPeExtractor(repository=SubstationTestsheetRepository())

    # In SPECIFIC_FOLDERS mode, scoped to 01-05-2026
    pkgs = extractor.discover_packages(
        testsheet_dir=tmp_path,
        target_folder_names=["01-05-2026"],
        mode=PopulateMode.SPECIFIC_FOLDERS,
        eager_extract=False,
    )
    assert len(pkgs) == 1
    assert pkgs[0].substation_number == 1
    assert pkgs[0].date_str == "01-05-2026"
    assert pkgs[0].data is None


def test_extractor_hydrate_packages_metadata(tmp_path: Path) -> None:
    import openpyxl
    from src.testsheet.repository import SubstationTestsheetRepository

    date_dir = tmp_path / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    wb = openpyxl.Workbook()
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["W5"] = "CRAU-S001"
    ws_pce["C5"] = "SSU CHEROH"
    ws_pce["P4"] = "01-05-2026"
    ws_pce["Y1"] = "1"
    wb.save(date_dir / "001. SSU CHEROH.xlsx")
    wb.close()

    extractor = PopulateTotalPeExtractor(repository=SubstationTestsheetRepository())
    pkgs = extractor.discover_packages(testsheet_dir=tmp_path, eager_extract=False)
    assert len(pkgs) == 1
    assert pkgs[0].data is None

    hydrated = extractor.hydrate_packages_metadata(pkgs)
    assert len(hydrated) == 1
    assert hydrated[0].data is not None
    assert hydrated[0].data.substation_number == 1
    assert hydrated[0].data.substation_name_erms == "SSU CHEROH"
    assert hydrated[0].data.fl_erms == "CRAU-S001"


def test_filter_auto_mode_pre_filters_without_data() -> None:
    filter_stage = PopulateTotalPeFilter()
    pkg = SubstationTestsheetPackage(
        testsheet_path=Path("/testsheets/01-05-2026/001. SSU CHEROH.xlsx"),
        unsorted_raw_data_dir=Path("/testsheets/01-05-2026/UNSORTED RAW DATA"),
        station="RAUB",
        month="01. MAY",
        date_str="01-05-2026",
        substation_number=1,
        data=None,
    )

    existing_keys = {("1", "01-05-2026")}
    res = filter_stage.filter_packages([pkg], mode=PopulateMode.AUTO, existing_auto_keys=existing_keys)
    assert len(res) == 0
