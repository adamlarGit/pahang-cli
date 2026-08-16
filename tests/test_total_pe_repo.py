"""Unit tests for TotalPeRepo (propagate_work_orders, formula preservation)."""
from pathlib import Path
import openpyxl
import pytest

from src.repositories.total_pe import (
    TotalPeRepo,
    TotalPeRepository,
    LocalExcelTotalPeRepository,
    PropagateResult,
)


def _create_sample_data_msms(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append([
        "Work Order",
        "Location",
        "Description",
        "Substation Name ERMS",
        "FL ERMS",
        "Cycle Date",
        "Substation Number",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _create_sample_total_pe(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataCycle1"
    ws.append([
        "PE NO",
        "FL NUMBER",
        "SUBSTATION NAME",
        "DATE",
        "TYPE",
        "WO",
        "SCOPE",
        "CUSTOM FORMULA",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def test_propagate_work_orders_missing_files_raises(tmp_path: Path) -> None:
    repo = LocalExcelTotalPeRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    pe_file = tmp_path / "TOTAL PE.xlsx"
    
    with pytest.raises(FileNotFoundError):
        repo.propagate_work_orders(pe_file, msms_file)
        
    pe_file.touch()
    with pytest.raises(FileNotFoundError):
        repo.propagate_work_orders(pe_file, msms_file)


def test_propagate_work_orders_success_and_formula_preservation(tmp_path: Path) -> None:
    repo = LocalExcelTotalPeRepository()
    pe_file = tmp_path / "TOTAL PE.xlsx"
    msms_file = tmp_path / "DATA MSMS.xlsx"

    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE TELUK SISIK", "PE TELUK SISIK", "CKTN0001/AAAA", "2026-01-10", 1],
        ["45001002", "CKTN0002BBBB", "PE BUKIT UBI", "PE BUKIT UBI", "CKTN0002/BBBB", "2026-01-11", 2],
        ["45001003", "CKTN0003CCCC", "PE ALOR AKAR", "PE ALOR AKAR", "CKTN0003/CCCC", "2026-01-12", 3],
    ])

    # Row 2: FL matches, WO is empty -> should be filled with 45001001
    # Row 3: FL matches, WO is already 45009999 -> should be preserved (already populated)
    # Row 4: FL has no match in DATA MSMS -> unmatched
    # Col H (8): contains formula "=IF(A2>0, 1, 0)" -> must NOT be destroyed!
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE TELUK SISIK", "2026-01-10", "ATTACHED", None, "FULL", "=IF(A2>0, 1, 0)"],
        [2, "CKTN0002/BBBB", "PE BUKIT UBI", "2026-01-11", "ATTACHED", "45009999", "FULL", "=IF(A3>0, 1, 0)"],
        [3, "CKTN0009/ZZZZ", "PE UNKNOWN", "2026-01-12", "COMPACT", None, "FULL", "=IF(A4>0, 1, 0)"],
    ])

    res = repo.propagate_work_orders(pe_file, msms_file)
    assert isinstance(res, PropagateResult)
    assert res.matched_count == 1
    assert res.already_populated_count == 1
    assert res.unmatched_count == 1
    assert "CKTN0009/ZZZZ" in res.unmatched_fls
    assert res.updated_count == 1

    # Verify openpyxl output
    wb = openpyxl.load_workbook(pe_file)
    ws = wb["DataCycle1"]

    # Row 2
    assert ws.cell(2, 6).value == "45001001"
    assert ws.cell(2, 8).value == "=IF(A2>0, 1, 0)"
    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 2).value == "CKTN0001/AAAA"
    assert ws.cell(2, 3).value == "PE TELUK SISIK"

    # Row 3
    assert ws.cell(3, 6).value == "45009999"  # Untouched
    assert ws.cell(3, 8).value == "=IF(A3>0, 1, 0)"

    # Row 4
    assert ws.cell(4, 6).value is None  # Unmatched remained None
    assert ws.cell(4, 8).value == "=IF(A4>0, 1, 0)"


def test_propagate_work_orders_target_date_filter(tmp_path: Path) -> None:
    repo = LocalExcelTotalPeRepository()
    pe_file = tmp_path / "TOTAL PE.xlsx"
    msms_file = tmp_path / "DATA MSMS.xlsx"

    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE A", "PE A", "CKTN0001/AAAA", "2026-01-10", 1],
        ["45001002", "CKTN0002BBBB", "PE B", "PE B", "CKTN0002/BBBB", "2026-01-11", 2],
    ])

    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE A", "2026-01-10", "ATTACHED", None, "FULL", None],
        [2, "CKTN0002/BBBB", "PE B", "2026-01-11", "ATTACHED", None, "FULL", None],
    ])

    # Filter only 2026-01-10
    res = repo.propagate_work_orders(pe_file, msms_file, target_date="2026-01-10")
    assert res.matched_count == 1
    assert res.updated_count == 1

    wb = openpyxl.load_workbook(pe_file)
    ws = wb["DataCycle1"]
    assert ws.cell(2, 6).value == "45001001"
    assert ws.cell(3, 6).value is None  # Row on 2026-01-11 skipped


def test_total_pe_repo_aliases() -> None:
    from src.repositories.total_pe import TotalPeRepo as AliasedRepo
    assert AliasedRepo is TotalPeRepository or issubclass(AliasedRepo, TotalPeRepository)
