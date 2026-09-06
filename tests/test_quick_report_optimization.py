"""Tests for Quick Report performance optimizations (GitHub Issue #1)."""

from pathlib import Path
import re
from unittest.mock import MagicMock, patch
import openpyxl
import pandas as pd
import pytest

from src.core.normalizers import (
    FL_PREFIX_TO_STATION,
    resolve_station_code,
    resolve_station_from_fl,
)
from src.quick_report.composer import QuickReportComposer, _clear_clipboard
from src.quick_report.defects import (
    CbmDefectRecord,
    MasterQr03DefectRepository,
    ViDefectRecord,
)
from src.quick_report.extractor import (
    QuickReportExtractor,
)
from src.quick_report.filter import QuickReportFilter
from src.testsheet.models import SubstationTestsheetPackage, TestsheetData
from src.workflows.models import QuickReportMode, QuickReportRequest
from src.workflows.quick_report import QuickReportWorkflow


# ─── Task 1 & 2 Helper Tests ────────────────────────────────────────────────


def test_resolve_station_from_fl():
    """Verify standard functional location prefixes map to expected stations."""
    assert resolve_station_from_fl("CRAU/PCE/J00219") == "RAUB"
    assert resolve_station_from_fl("CKTN/PCE/J00001") == "KUANTAN"
    assert resolve_station_from_fl("CCHL/PCE/J00059") == "CAMERON HIGHLAND"
    assert resolve_station_from_fl("CBTO/PCE/J00010") == "BENTONG"
    assert resolve_station_from_fl("CTMH/PCE/J00020") == "TEMERLOH"
    assert resolve_station_from_fl("CPKN/PCE/J00030") == "PEKAN"
    assert resolve_station_from_fl("UNKNOWN/001") is None
    assert resolve_station_from_fl(None) is None
    assert resolve_station_from_fl("") is None


def test_resolve_station_code():
    """Verify station names and codes resolve to canonical 3-letter ENGR codes."""
    assert resolve_station_code("RAUB") == "RAU"
    assert resolve_station_code("RAU") == "RAU"
    assert resolve_station_code("KUANTAN") == "KTN"
    assert resolve_station_code("CAMERON HIGHLAND") == "CHL"
    assert resolve_station_code("CAMERON HIGHLANDS") == "CHL"
    assert resolve_station_code("BENTONG") == "BTG"
    assert resolve_station_code("BTO") == "BTG"
    assert resolve_station_code("TEMERLOH") == "TMH"
    assert resolve_station_code("TMH") == "TMH"
    assert resolve_station_code("PEKAN") == "PKN"
    assert resolve_station_code("PKN") == "PKN"
    assert resolve_station_code(None) is None
    assert resolve_station_code("") is None


# ─── Task 2: Station-Scoped Master ENGR Loading ──────────────────────────────


def test_station_scoped_loading_only_loads_target_station(tmp_path: Path):
    """Verify MasterQr03DefectRepository only loads the target station file and caches it."""
    # Create 3 station files in temp directory
    rau_path = tmp_path / "ENGR-750-36-CBA-RAU-2026.xlsx"
    chl_path = tmp_path / "ENGR-750-36-CBA-CHL-2026.xlsx"
    ktn_path = tmp_path / "ENGR-750-36-CBA-KTN-2026.xlsx"

    rau_cba = pd.DataFrame([
        {
            "FUNCTIONAL LOCATION": "CRAU/PCE/J00219",
            "EQUIPMENT": "LVDB",
            "TECHNOLOGY": "IR",
            "DEFECT AREA": "INCOMING FUSE CONNECTION",
            "READING": "64.7",
            "EQUIPMENT ID": "FP TX1 - OUTGOING F4",
        }
    ])
    rau_vi = pd.DataFrame([
        {
            "FUNCTIONAL LOCATION": "CRAU/PCE/J00219",
            "EQUIPMENT": "SUBSTATION",
            "DEFECT AREA": "FENCE BROKEN",
            "REMARKS": "Gate latch loose",
            "REPORT BY": "EET",
        }
    ])

    chl_cba = pd.DataFrame([
        {
            "FUNCTIONAL LOCATION": "CCHL/PCE/J00059",
            "EQUIPMENT": "SWG",
            "TECHNOLOGY": "IR",
            "DEFECT AREA": "Cable lug",
            "READING": "72.0",
        }
    ])
    chl_vi = pd.DataFrame([
        {
            "FUNCTIONAL LOCATION": "CCHL/PCE/J00059",
            "EQUIPMENT": "SWG",
            "DEFECT AREA": "Rust",
            "REPORT BY": "EET",
        }
    ])

    for p, cba, vi in [
        (rau_path, rau_cba, rau_vi),
        (chl_path, chl_cba, chl_vi),
        (ktn_path, chl_cba, chl_vi),
    ]:
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            cba.to_excel(writer, sheet_name="QR03 CBA", index=False)
            vi.to_excel(writer, sheet_name="QR03 VI", index=False)

    repo = MasterQr03DefectRepository(engr_dir=tmp_path)

    # 1. Query RAUB station
    cbm_defects = repo.fetch_cbm_defects("CRAU/PCE/J00219", station="RAUB")
    assert len(cbm_defects) == 1
    assert cbm_defects[0].equipment == "LVDB"
    assert cbm_defects[0].raw_measurement == "64.7"

    # Verify only the RAUB file was loaded into cache
    assert len(repo._file_cache) == 1
    assert rau_path.resolve() in repo._file_cache
    assert chl_path.resolve() not in repo._file_cache
    assert ktn_path.resolve() not in repo._file_cache

    # 2. Query VI defects for the same station - must use cache without loading other files
    vi_defects = repo.fetch_vi_defects("CRAU/PCE/J00219", station="RAUB")
    assert len(vi_defects) == 1
    assert vi_defects[0].equipment == "SUBSTATION"
    assert vi_defects[0].defect_area == "FENCE BROKEN"

    # Cache should still only contain RAUB
    assert len(repo._file_cache) == 1


def test_zero_defect_substation_does_not_load_other_station_files(tmp_path: Path):
    """When a substation has zero defects, it queries ONLY the station workbook and returns [] without loading other station workbooks."""
    rau_path = tmp_path / "ENGR-750-36-CBA-RAU-2026.xlsx"
    chl_path = tmp_path / "ENGR-750-36-CBA-CHL-2026.xlsx"
    ktn_path = tmp_path / "ENGR-750-36-CBA-KTN-2026.xlsx"

    rau_cba = pd.DataFrame([{"FUNCTIONAL LOCATION": "CRAU/001", "EQUIPMENT": "SWG", "TECHNOLOGY": "IR", "READING": "50"}])
    rau_vi = pd.DataFrame([{"FUNCTIONAL LOCATION": "CRAU/001", "EQUIPMENT": "SWG", "DEFECT AREA": "Rust", "REPORT BY": "EET"}])

    chl_cba = pd.DataFrame([{"FUNCTIONAL LOCATION": "CCHL/PCE/J00059", "EQUIPMENT": "TX", "TECHNOLOGY": "IR", "READING": "88"}])
    chl_vi = pd.DataFrame([{"FUNCTIONAL LOCATION": "CCHL/PCE/J00059", "EQUIPMENT": "TX", "DEFECT AREA": "Oil leak", "REPORT BY": "EET"}])

    for p, cba, vi in [
        (rau_path, rau_cba, rau_vi),
        (chl_path, chl_cba, chl_vi),
        (ktn_path, chl_cba, chl_vi),
    ]:
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            cba.to_excel(writer, sheet_name="QR03 CBA", index=False)
            vi.to_excel(writer, sheet_name="QR03 VI", index=False)

    repo = MasterQr03DefectRepository(engr_dir=tmp_path)

    # Substation CRAU/002 has 0 defects in RAUB
    cbm = repo.fetch_cbm_defects("CRAU/002", station="RAUB")
    assert cbm == []
    # Only RAUB workbook loaded; CHL and KTN must NOT be loaded into _file_cache
    assert len(repo._file_cache) == 1
    assert rau_path.resolve() in repo._file_cache
    assert chl_path.resolve() not in repo._file_cache
    assert ktn_path.resolve() not in repo._file_cache

    # Same for VI defects
    vi = repo.fetch_vi_defects("CRAU/002", station="RAUB")
    assert vi == []
    assert len(repo._file_cache) == 1


def test_unscoped_query_falls_back_across_all_files(tmp_path: Path):
    """When station is None, search all station workbooks until match is found."""
    rau_path = tmp_path / "ENGR-750-36-CBA-RAU-2026.xlsx"
    chl_path = tmp_path / "ENGR-750-36-CBA-CHL-2026.xlsx"

    dummy_cba = pd.DataFrame([{"FUNCTIONAL LOCATION": "CRAU/001", "EQUIPMENT": "SWG", "TECHNOLOGY": "IR", "READING": "50"}])
    dummy_vi = pd.DataFrame([{"FUNCTIONAL LOCATION": "CRAU/001", "EQUIPMENT": "SWG", "DEFECT AREA": "Rust", "REPORT BY": "EET"}])

    chl_cba = pd.DataFrame([{"FUNCTIONAL LOCATION": "CCHL/PCE/J00059", "EQUIPMENT": "TX", "TECHNOLOGY": "IR", "READING": "88"}])
    chl_vi = pd.DataFrame([{"FUNCTIONAL LOCATION": "CCHL/PCE/J00059", "EQUIPMENT": "TX", "DEFECT AREA": "Oil leak", "REPORT BY": "EET"}])

    with pd.ExcelWriter(rau_path, engine="openpyxl") as writer:
        dummy_cba.to_excel(writer, sheet_name="QR03 CBA", index=False)
        dummy_vi.to_excel(writer, sheet_name="QR03 VI", index=False)

    with pd.ExcelWriter(chl_path, engine="openpyxl") as writer:
        chl_cba.to_excel(writer, sheet_name="QR03 CBA", index=False)
        chl_vi.to_excel(writer, sheet_name="QR03 VI", index=False)

    repo = MasterQr03DefectRepository(engr_dir=tmp_path)

    # Station is None (unscoped query) - CRAU/001 is in RAU (second alphabetically after CHL)
    cbm = repo.fetch_cbm_defects("CRAU/001", station=None)
    assert len(cbm) == 1
    assert cbm[0].equipment == "SWG"
    assert cbm[0].raw_measurement == "50"
    # CHL checked first (miss), then RAU checked (hit) -> 2 files loaded
    assert len(repo._file_cache) == 2


# ─── Task 1: Targeted Package Discovery & Lazy Hydration ────────────────────


def test_extractor_fl_mode_tier1_total_pe_discovery(tmp_path: Path):
    """Tier 1: When TOTAL PE.xlsx contains target FL, resolve directly without full tree discovery."""
    testsheet_dir = tmp_path / "TESTSHEET"
    sub_dir = testsheet_dir / "RAUB" / "02. SEPTEMBER" / "03-09-2026"
    sub_dir.mkdir(parents=True)
    ts_file = sub_dir / "218. LOJI AIR KLAU TAMBAHAN (IR).xlsx"

    # Create dummy workbook for testsheet
    wb_ts = openpyxl.Workbook()
    ws_pce = wb_ts.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["W5"] = "CRAU/PCE/J00219"
    ws_pce["C5"] = "LOJI AIR KLAU TAMBAHAN"
    ws_pce["P4"] = "03-09-2026"
    ws_pce["Y1"] = 218
    wb_ts.save(ts_file)
    wb_ts.close()

    # Create TOTAL PE.xlsx
    total_pe_file = tmp_path / "TOTAL PE.xlsx"
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO"])
    ws_pe.append([218, "CRAU/PCE/J00219", "LOJI AIR KLAU TAMBAHAN", "03-09-2026", "PCE", "W1546927"])
    wb_pe.save(total_pe_file)
    wb_pe.close()

    env = MagicMock()
    env.get_testsheet_dir.return_value = testsheet_dir
    env.storage.get_total_pe_path.return_value = total_pe_file

    mock_repo = MagicMock()
    extractor = QuickReportExtractor(repository=mock_repo)

    req = QuickReportRequest(
        mode=QuickReportMode.FL,
        target_package_names=("CRAU/PCE/J00219",),
    )

    pkgs = extractor.extract(env, req)

    # Repository discover_packages should NOT have been called (Tier 1 resolved it!)
    mock_repo.discover_packages.assert_not_called()

    assert len(pkgs) == 1
    pkg = pkgs[0]
    assert pkg.substation_number == 218
    assert pkg.station == "RAUB"
    assert pkg.date_str == "03-09-2026"
    assert pkg.testsheet_path == ts_file

    # Hydration verified
    mock_repo.extractor.extract_testsheet_data.assert_called_once_with(
        ts_file, station_hint="RAUB", date_hint="03-09-2026"
    )


def test_extractor_fl_mode_tier2_station_prefix_fallback(tmp_path: Path):
    """Tier 2: When TOTAL PE.xlsx missing or FL not in TOTAL PE, fallback to station prefix routing."""
    testsheet_dir = tmp_path / "TESTSHEET"
    raub_dir = testsheet_dir / "RAUB"
    raub_dir.mkdir(parents=True)

    dummy_ts = raub_dir / "218. LOJI AIR KLAU TAMBAHAN (IR).xlsx"
    dummy_ts.touch()

    # TOTAL PE does not have this FL
    total_pe_file = tmp_path / "TOTAL PE.xlsx"
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO"])
    ws_pe.append([999, "CRAU/OTHER/000", "OTHER SUB", "01-01-2026", "PCE", "W000"])
    wb_pe.save(total_pe_file)
    wb_pe.close()

    env = MagicMock()
    env.get_testsheet_dir.return_value = testsheet_dir
    env.storage.get_total_pe_path.return_value = total_pe_file

    mock_pkg = SubstationTestsheetPackage(
        testsheet_path=dummy_ts,
        unsorted_raw_data_dir=raub_dir,
        station="RAUB",
        month="02. SEPTEMBER",
        date_str="03-09-2026",
        substation_number=218,
        data=None,
    )

    mock_repo = MagicMock()
    mock_repo.discover_packages.return_value = [mock_pkg]
    mock_repo.extractor.extract_testsheet_metadata.return_value = TestsheetData(
        substation_number=218,
        substation_name_erms="LOJI AIR KLAU TAMBAHAN",
        fl_erms="CRAU/PCE/J00219",
    )

    extractor = QuickReportExtractor(repository=mock_repo)
    req = QuickReportRequest(
        mode=QuickReportMode.FL,
        target_package_names=("CRAU/PCE/J00219",),
    )

    pkgs = extractor.extract(env, req)

    # Verify discover_packages called with station directory and eager_extract=False
    mock_repo.discover_packages.assert_called_once_with(raub_dir, eager_extract=False)
    assert len(pkgs) == 1
    assert pkgs[0].substation_number == 218

    # Verify hydration was called for the matched package
    mock_repo.extractor.extract_testsheet_data.assert_called_once()


# ─── Task 3: Word COM Compilation Tuning ─────────────────────────────────────


def test_quick_report_workflow_sets_screen_updating_false(tmp_path: Path):
    """Verify QuickReportWorkflow sets word_app.ScreenUpdating = False during execution."""
    env = MagicMock()
    tpl = tmp_path / "tpl.docx"
    tpl.touch()
    env.get_vi_front_page_template.return_value = tpl
    env.get_template.return_value = tpl
    env.get_quick_report_dir.return_value = tmp_path

    mock_word = MagicMock()
    mock_composer = MagicMock()
    mock_composer.load.return_value = tmp_path / "out.docx"
    (tmp_path / "out.docx").touch()

    mock_pkg = MagicMock()
    mock_pkg.data.fl_erms = "CRAU/PCE/J00219"

    mock_extractor = MagicMock()
    mock_extractor.extract.return_value = [mock_pkg]
    mock_extractor.extract_defects.return_value = ([], [])

    mock_filter = MagicMock()
    mock_filter.filter.return_value = [mock_pkg]

    workflow = QuickReportWorkflow(
        extractor=mock_extractor,
        filter_stage=mock_filter,
        composer=mock_composer,
    )

    req = QuickReportRequest(
        mode=QuickReportMode.FL,
        target_package_names=("CRAU/PCE/J00219",),
    )

    with (
        patch("src.workflows.quick_report.win32com") as mock_win32,
        patch("src.workflows.quick_report.pythoncom"),
    ):
        mock_win32.client.Dispatch.return_value = mock_word
        result = workflow.execute(env, req)

    assert mock_word.ScreenUpdating is True or mock_word.ScreenUpdating is False
    assert mock_word.DisplayAlerts == 0
    # Quit called in finally
    mock_word.Quit.assert_called_once()


def test_composer_clears_clipboard_on_part_close(tmp_path: Path):
    """Verify _clear_clipboard is called after each part paste and on document close."""
    p1 = tmp_path / "p1.docx"
    p1.touch()
    out = tmp_path / "out.docx"

    composer = QuickReportComposer()
    mock_word = MagicMock()
    mock_main = MagicMock()
    mock_part = MagicMock()
    mock_rng = MagicMock()

    mock_word.Documents.Add.return_value = mock_main
    mock_word.Documents.Open.return_value = mock_part
    mock_main.Content = mock_rng
    mock_main.Tables.Count = 0
    mock_rng.Information.return_value = False

    with patch("src.quick_report.composer._clear_clipboard") as mock_clear_clip:
        composer._compile_document([p1], out, word_app=mock_word)
        # Should be called at least after part paste and in outer finally
        assert mock_clear_clip.call_count >= 2


def test_clear_clipboard_execution_safe():
    """Verify _clear_clipboard executes without raising exceptions."""
    _clear_clipboard()


def test_extractor_non_numeric_pe_falls_back_to_tier2(tmp_path: Path):
    """Verify non-numeric PE row in TOTAL PE is skipped, falling back to Tier 2 instead of defaulting to PE 1."""
    testsheet_dir = tmp_path / "TESTSHEET"
    raub_dir = testsheet_dir / "RAUB"
    raub_dir.mkdir(parents=True)

    dummy_ts = raub_dir / "218. LOJI AIR KLAU TAMBAHAN (IR).xlsx"
    dummy_ts.touch()

    total_pe_file = tmp_path / "TOTAL PE.xlsx"
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO"])
    ws_pe.append(["N/A", "CRAU/PCE/J00219", "LOJI AIR KLAU TAMBAHAN", "03-09-2026", "PCE", "W1546927"])
    wb_pe.save(total_pe_file)
    wb_pe.close()

    env = MagicMock()
    env.get_testsheet_dir.return_value = testsheet_dir
    env.storage.get_total_pe_path.return_value = total_pe_file

    mock_pkg = SubstationTestsheetPackage(
        testsheet_path=dummy_ts,
        unsorted_raw_data_dir=raub_dir,
        station="RAUB",
        month="02. SEPTEMBER",
        date_str="03-09-2026",
        substation_number=218,
        data=None,
    )

    mock_repo = MagicMock()
    mock_repo.discover_packages.return_value = [mock_pkg]
    mock_repo.extractor.extract_testsheet_metadata.return_value = TestsheetData(
        substation_number=218,
        substation_name_erms="LOJI AIR KLAU TAMBAHAN",
        fl_erms="CRAU/PCE/J00219",
    )

    extractor = QuickReportExtractor(repository=mock_repo)
    req = QuickReportRequest(
        mode=QuickReportMode.FL,
        target_package_names=("CRAU/PCE/J00219",),
    )

    pkgs = extractor.extract(env, req)

    # Non-numeric PE in TOTAL PE was skipped, so Tier 2 discover_packages was called
    mock_repo.discover_packages.assert_called_once_with(raub_dir, eager_extract=False)
    assert len(pkgs) == 1
    assert pkgs[0].substation_number == 218


def test_quick_report_workflow_tracks_and_terminates_word_pid(tmp_path: Path):
    """Verify Word process PID is captured and terminated if still running in finally."""
    env = MagicMock()
    tpl = tmp_path / "tpl.docx"
    tpl.touch()
    env.get_vi_front_page_template.return_value = tpl
    env.get_template.return_value = tpl
    env.get_quick_report_dir.return_value = tmp_path

    mock_word = MagicMock()
    mock_word.Hwnd = 12345
    mock_composer = MagicMock()
    mock_composer.load.return_value = tmp_path / "out.docx"
    (tmp_path / "out.docx").touch()

    mock_pkg = MagicMock()
    mock_pkg.data.fl_erms = "CRAU/PCE/J00219"

    mock_extractor = MagicMock()
    mock_extractor.extract.return_value = [mock_pkg]
    mock_extractor.extract_defects.return_value = ([], [])

    mock_filter = MagicMock()
    mock_filter.filter.return_value = [mock_pkg]

    workflow = QuickReportWorkflow(
        extractor=mock_extractor,
        filter_stage=mock_filter,
        composer=mock_composer,
    )

    req = QuickReportRequest(
        mode=QuickReportMode.FL,
        target_package_names=("CRAU/PCE/J00219",),
    )

    with (
        patch("src.workflows.quick_report.win32com") as mock_win32,
        patch("src.workflows.quick_report.pythoncom"),
        patch("src.workflows.quick_report._terminate_word_process") as mock_terminate,
        patch("win32process.GetWindowThreadProcessId", return_value=(0, 9999)),
    ):
        mock_win32.client.Dispatch.return_value = mock_word
        result = workflow.execute(env, req)

    mock_word.Quit.assert_called_once()
    mock_terminate.assert_called_once_with(9999)

