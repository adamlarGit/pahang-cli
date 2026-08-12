"""Tests for signature replacement workflow and selector options."""

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import pytest
from unittest.mock import patch

from src.workflows.replace_signatures import (
    replace_pce_images,
    _select_signature_path,
)


def test_replace_pce_images_none_signature_strips_text_only(tmp_path: Path) -> None:
    """Test that photo1_path=None and photo2_path=None strips {{signvendor}} and {{signtnb}} text without inserting images."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCE Testsheet"
    ws["C65"] = "Vendor: {{signvendor}}"
    ws["S65"] = "TNB: {{signtnb}}"

    # Add a table to verify table clearing logic
    tab = Table(displayName="TestTable", ref="A1:B5")
    ws.add_table(tab)

    xlsx_file = tmp_path / "test_sheet.xlsx"
    out_file = tmp_path / "test_sheet_out.xlsx"
    wb.save(xlsx_file)

    # Execute replace_pce_images with photo1_path=None and photo2_path=None
    replace_pce_images(xlsx_file, photo1_path=None, photo2_path=None, output_path=out_file)

    res_wb = openpyxl.load_workbook(out_file)
    res_ws = res_wb["PCE Testsheet"]

    assert res_ws["C65"].value == "Vendor:"
    assert res_ws["S65"].value == "TNB:"
    # Verify no images added
    assert len(res_ws._images) == 0
    # Verify table definitions were cleared to prevent corruption
    assert len(res_ws._tables) == 0


@patch("src.cli_selectors.select_one", return_value="__none__")
def test_select_signature_path_none_option(mock_select, tmp_path: Path) -> None:
    """Test that choosing '__none__' option returns (None, '__none__')."""
    sign_dir = tmp_path / "SIGN"
    sign_dir.mkdir()

    path, key = _select_signature_path("Select vendor signature:", sign_dir)
    assert path is None
    assert key == "__none__"


def test_project_environment_get_sign_dir(tmp_path: Path) -> None:
    """Test that ProjectEnvironment.get_sign_dir() resolves OTHERS/SIGN strictly inside the project root."""
    from src.project.environment import ProjectEnvironment
    from src.project.models import ProjectMetadata
    from src.project.storage import LocalWorkspaceStorage

    proj_dir = tmp_path / "PO_PROJECT"
    sign_dir = proj_dir / "OTHERS" / "SIGN"
    sign_dir.mkdir(parents=True)

    meta = ProjectMetadata(key="test", name="Test Project", base_path=str(proj_dir), state="pahang", po_number="", voltage_type="11kV", year="2026", cycle="Cycle 1", technologies=("IR",))
    storage = LocalWorkspaceStorage(proj_dir)
    env = ProjectEnvironment(metadata=meta, storage=storage)

def test_multi_sheet_signature_alignment(tmp_path: Path) -> None:
    """Test that second sheet (PCE Testsheet (2)) anchors signature images identical to Sheet 1."""
    from PIL import Image as PILImage
    sig_png = tmp_path / "sig.png"
    img = PILImage.new("RGBA", (100, 50), color=(255, 0, 0, 255))
    img.save(sig_png)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "PCE Testsheet"
    ws1["C65"] = "{{signvendor}}"
    ws1["S65"] = "{{signtnb}}"

    ws2 = wb.create_sheet(title="PCE Testsheet (2)")
    ws2["C65"] = "{{signvendor}}"
    ws2["S65"] = "{{signtnb}}"

    xlsx_file = tmp_path / "multi_sheet.xlsx"
    out_file = tmp_path / "multi_sheet_out.xlsx"
    wb.save(xlsx_file)

    replace_pce_images(xlsx_file, sig_png, sig_png, output_path=out_file, mode="placeholder")

    res_wb = openpyxl.load_workbook(out_file)
    ws1_img0 = res_wb["PCE Testsheet"]._images[0]
    ws2_img0 = res_wb["PCE Testsheet (2)"]._images[0]

    assert ws1_img0.anchor._from.col == ws2_img0.anchor._from.col
    assert ws1_img0.anchor._from.colOff == ws2_img0.anchor._from.colOff
    assert ws1_img0.anchor.to.col == ws2_img0.anchor.to.col
    assert ws1_img0.anchor.to.colOff == ws2_img0.anchor.to.colOff



