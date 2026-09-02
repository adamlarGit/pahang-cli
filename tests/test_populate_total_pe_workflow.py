"""Integration tests for Populate TOTAL PE workflow in Pahang CLI."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import openpyxl
import pytest

from src.workflows.populate_total_pe import PopulateTotalPeWorkflow
from src.project.environment import ProjectEnvironment
from src.project.models import ProjectMetadata
from src.project.storage import LocalWorkspaceStorage
from src.workflows.models import PopulateMode, PopulateTotalPeRequest


@pytest.fixture
def mock_env(tmp_path: Path) -> ProjectEnvironment:
    meta = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    storage = LocalWorkspaceStorage(tmp_path)
    return ProjectEnvironment(metadata=meta, storage=storage)


def test_populate_total_pe_workflow_creates_and_updates_rows(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # Setup TESTSHEET structure
    date_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    wb = openpyxl.Workbook()
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["W5"] = "CRAU-S001"
    ws_pce["C5"] = "SSU CHEROH"
    ws_pce["P4"] = "01-05-2026"
    wb.save(date_dir / "001. SSU CHEROH.xlsx")
    wb.close()

    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    workflow = PopulateTotalPeWorkflow()
    request = PopulateTotalPeRequest(mode=PopulateMode.AUTO)
    result = workflow.execute(mock_env, request)

    assert result.new_rows_added == 1

    # Verify TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    assert total_pe_path.exists()

    wb_res = openpyxl.load_workbook(total_pe_path)
    ws_res = wb_res["DataCycle1"]
    assert ws_res.cell(2, 1).value == 1
    assert ws_res.cell(2, 2).value == "CRAU-S001"
    assert ws_res.cell(2, 3).value == "SSU CHEROH"
    assert ws_res.cell(2, 4).value in (datetime(2026, 5, 1, 0, 0), datetime(2026, 5, 1, 0, 0).date(), "01-05-2026", "2026-05-01")
    wb_res.close()

    # Re-run AUTO mode to verify it skips already populated (1, "01-05-2026")
    second_res = workflow.execute(mock_env, request)
    assert second_res.new_rows_added == 0


def test_populate_total_pe_targets_column_a_empty_cells(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # Pre-create TOTAL PE.xlsx with empty Column A at row 2 and a note in Column I at row 3
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataCycle1"
    ws.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    ws.append([None, None, None, None, None, None])
    ws.cell(row=3, column=9, value="PASTE HERE AS TEXT")
    wb.save(total_pe_path)
    wb.close()

    # Setup TESTSHEET structure
    date_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "17-06-2026"
    date_dir.mkdir(parents=True)
    (date_dir / "UNSORTED RAW DATA").mkdir()

    wb_ts = openpyxl.Workbook()
    ws_pce = wb_ts.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["W5"] = "CCHL/PCE/J00293"
    ws_pce["C5"] = "CSU KEA FARM"
    ws_pce["P4"] = "17-06-2026"
    wb_ts.save(date_dir / "302. CSU KEA FARM.xlsx")
    wb_ts.close()

    workflow = PopulateTotalPeWorkflow()
    request = PopulateTotalPeRequest(mode=PopulateMode.AUTO)
    result = workflow.execute(mock_env, request)

    assert result.new_rows_added == 1

    wb_res = openpyxl.load_workbook(total_pe_path)
    ws_res = wb_res["DataCycle1"]
    # Record should be placed at Row 2 (first empty cell in Column A), preserving Column I note at Row 3
    assert ws_res.cell(2, 1).value == 302
    assert ws_res.cell(2, 2).value == "CCHL/PCE/J00293"
    assert ws_res.cell(2, 3).value == "CSU KEA FARM"
    assert ws_res.cell(3, 9).value == "PASTE HERE AS TEXT"
    wb_res.close()


def test_populate_total_pe_workflow_specific_folders_mode(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # Setup two folders
    date_dir1 = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir1.mkdir(parents=True)
    (date_dir1 / "UNSORTED RAW DATA").mkdir()

    wb1 = openpyxl.Workbook()
    ws_pce1 = wb1.active
    ws_pce1.title = "PCE Testsheet"
    ws_pce1["W5"] = "CRAU-S001"
    ws_pce1["C5"] = "SSU CHEROH"
    ws_pce1["P4"] = "01-05-2026"
    wb1.save(date_dir1 / "001. SSU CHEROH.xlsx")
    wb1.close()

    date_dir2 = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "02-05-2026"
    date_dir2.mkdir(parents=True)
    (date_dir2 / "UNSORTED RAW DATA").mkdir()

    wb2 = openpyxl.Workbook()
    ws_pce2 = wb2.active
    ws_pce2.title = "PCE Testsheet"
    ws_pce2["W5"] = "CRAU-S002"
    ws_pce2["C5"] = "PPU BENTA"
    ws_pce2["P4"] = "02-05-2026"
    wb2.save(date_dir2 / "002. PPU BENTA.xlsx")
    wb2.close()

    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    workflow = PopulateTotalPeWorkflow()
    # Run in SPECIFIC_FOLDERS mode targeting ONLY date_dir2
    request = PopulateTotalPeRequest(
        mode=PopulateMode.SPECIFIC_FOLDERS,
        target_folder_names=("02-05-2026", str(date_dir2)),
    )
    result = workflow.execute(mock_env, request)

    assert result.new_rows_added == 1

    wb_res = openpyxl.load_workbook(total_pe_path)
    ws_res = wb_res["DataCycle1"]
    assert ws_res.max_row == 2
    assert ws_res.cell(2, 1).value == 2
    assert ws_res.cell(2, 2).value == "CRAU-S002"
    assert ws_res.cell(2, 3).value == "PPU BENTA"
    wb_res.close()
