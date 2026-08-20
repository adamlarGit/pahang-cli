"""Integration tests for Raw Material Creation & Sorting workflow in Pahang CLI."""

from __future__ import annotations

from pathlib import Path
import openpyxl
import pytest

from src.workflows.populate_total_pe import PopulateTotalPeWorkflow
from src.project.environment import ProjectEnvironment
from src.project.models import CameraConfig, ProjectMetadata
from src.project.storage import LocalWorkspaceStorage
from src.workflows.raw_material import RawMaterialWorkflow
from src.workflows.models import PopulateMode, PopulateTotalPeRequest, RawMaterialRequest


@pytest.fixture
def mock_env(tmp_path: Path) -> ProjectEnvironment:
    meta = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    storage = LocalWorkspaceStorage(tmp_path)
    storage.ensure_directory(storage.get_raw_material_dir())
    storage.ensure_directory(storage.get_python_dir())
    return ProjectEnvironment(metadata=meta, storage=storage)


def test_raw_material_precheck_fails_if_total_pe_missing(mock_env: ProjectEnvironment) -> None:
    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=mock_env.base_path)

    with pytest.raises(RuntimeError, match="TOTAL PE.xlsx pre-check failed"):
        workflow.execute(mock_env, request)


def test_raw_material_precheck_fails_if_total_pe_empty(mock_env: ProjectEnvironment) -> None:
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataCycle1"
    ws.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb.save(total_pe_path)
    wb.close()

    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=mock_env.base_path)

    with pytest.raises(RuntimeError, match="TOTAL PE.xlsx pre-check failed"):
        workflow.execute(mock_env, request)


def test_raw_material_precheck_fails_if_unsorted_raw_data_missing(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    date_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir.mkdir(parents=True)
    # Do not create UNSORTED RAW DATA folder

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAW DATA"
    ws.cell(1, 1, "PE NO")
    ws.cell(1, 2, 1)
    wb.save(date_dir / "001. SSU CHEROH.xlsx")
    wb.close()

    # Create a dummy TOTAL PE.xlsx for PopulateTotalPeWorkflow
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # Populate TOTAL PE
    PopulateTotalPeWorkflow().execute(mock_env, PopulateTotalPeRequest(mode=PopulateMode.AUTO))

    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=date_dir)

    with pytest.raises(RuntimeError, match="UNSORTED RAW DATA' directory missing"):
        workflow.execute(mock_env, request)


def test_raw_material_workflow_success_and_photo_sorting(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    date_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir.mkdir(parents=True)
    unsorted_dir = date_dir / "UNSORTED RAW DATA"
    unsorted_dir.mkdir()

    # Create synthetic testsheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAW DATA"
    ws.cell(1, 1, "PE NO")
    ws.cell(1, 2, 1)
    ws.cell(2, 1, "SUBSTATION NAME")
    ws.cell(2, 2, "SSU CHEROH")
    ws.cell(3, 1, "FL NUMBER")
    ws.cell(3, 2, "CRAU-S001")
    ws.cell(4, 1, "DATE")
    ws.cell(4, 2, "01-05-2026")
    ws.cell(8, 1, "IR START")
    ws.cell(8, 2, 10)
    ws.cell(8, 3, "IR END")
    ws.cell(8, 4, 12)
    ws.cell(9, 1, "DG START")
    ws.cell(9, 2, 100)
    ws.cell(9, 3, "DG END")
    ws.cell(9, 4, 102)
    wb.save(date_dir / "001. SSU CHEROH.xlsx")
    wb.close()

    # Create a dummy TOTAL PE.xlsx for PopulateTotalPeWorkflow
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    # Populate TOTAL PE first
    PopulateTotalPeWorkflow().execute(mock_env, PopulateTotalPeRequest(mode=PopulateMode.AUTO))

    # Add raw photos to UNSORTED RAW DATA
    (unsorted_dir / "FLIR0010.jpg").touch()
    (unsorted_dir / "FLIR0011.jpg").touch()
    (unsorted_dir / "FLIR0050.jpg").touch()
    (unsorted_dir / "IMG_0100.jpg").touch()
    (unsorted_dir / "IMG_0101.jpg").touch()

    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=date_dir)
    result = workflow.execute(mock_env, request)

    assert result.substations_count == 1
    assert result.summary.ir_copied_count == 2
    assert result.summary.dg_copied_count == 2

    # Verify directory structure under RAW MATERIAL/RAUB/01. MAY/01-05-2026/001/RAW DATA/
    raw_mat_pe_dir = mock_env.storage.get_raw_material_dir() / "RAUB" / "01. MAY" / "01-05-2026" / "001" / "RAW DATA"
    assert (raw_mat_pe_dir / "IR" / "FLIR0010.jpg").exists()
    assert (raw_mat_pe_dir / "IR" / "FLIR0011.jpg").exists()
    assert not (raw_mat_pe_dir / "IR" / "FLIR0050.jpg").exists()

    assert (raw_mat_pe_dir / "DG" / "IMG_0100.jpg").exists()
    assert (raw_mat_pe_dir / "DG" / "IMG_0101.jpg").exists()

    assert (raw_mat_pe_dir / "US+TEV").exists()


def test_extract_photo_number_trailing_sequence() -> None:
    filter_stage = RawMaterialWorkflow().filter_stage
    assert filter_stage.extract_photo_number("IMG_20260724_0042.jpg", "IMG") == 42
    assert filter_stage.extract_photo_number("FLIR0123.jpg", "FLIR") == 123


def test_raw_material_workflow_dual_pair_camera_config(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    date_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir.mkdir(parents=True)
    unsorted_dir = date_dir / "UNSORTED RAW DATA"
    unsorted_dir.mkdir()

    # Create synthetic testsheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAW DATA"
    ws.cell(1, 1, "PE NO")
    ws.cell(1, 2, 1)
    ws.cell(2, 1, "SUBSTATION NAME")
    ws.cell(2, 2, "SSU CHEROH")
    ws.cell(3, 1, "FL NUMBER")
    ws.cell(3, 2, "CRAU-S001")
    ws.cell(4, 1, "DATE")
    ws.cell(4, 2, "01-05-2026")
    ws.cell(8, 1, "IR START")
    ws.cell(8, 2, 1)
    ws.cell(8, 3, "IR END")
    ws.cell(8, 4, 2)
    ws.cell(9, 1, "DG START")
    ws.cell(9, 2, 10)
    ws.cell(9, 3, "DG END")
    ws.cell(9, 4, 11)
    wb.save(date_dir / "001. SSU CHEROH.xlsx")
    wb.close()

    # Create TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    PopulateTotalPeWorkflow().execute(mock_env, PopulateTotalPeRequest(mode=PopulateMode.AUTO))

    # Add dual pair IR/DC photos and P-series DG photos to UNSORTED RAW DATA
    (unsorted_dir / "IR_0001.jpg").touch()
    (unsorted_dir / "DC_0002.jpg").touch()
    (unsorted_dir / "IR_0002.jpg").touch()
    (unsorted_dir / "DC_0003.jpg").touch()
    (unsorted_dir / "P1000010.JPG").touch()
    (unsorted_dir / "P1000011.JPG").touch()

    import zipfile
    (unsorted_dir / "US+TEV").mkdir()
    with zipfile.ZipFile(unsorted_dir / "US+TEV" / "20260501T090000_001-SSU-CHEROH.zip", "w") as z:
        z.writestr("index.html", "<html>Test</html>")

    # Configure CameraConfig in mock_env repository or project_config.json
    cam_cfg = CameraConfig(
        ir_mode="dual_pair",
        ir_prefix="IR_",
        dc_prefix="DC_",
        dc_offset=1,
        dg_prefix="P",
    )
    mock_env.save_camera_config(cam_cfg)

    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=date_dir)
    result = workflow.execute(mock_env, request)

    assert result.substations_count == 1
    assert result.summary.ir_copied_count == 4  # 2 IR + 2 DC
    assert result.summary.dg_copied_count == 2
    assert result.summary.us_tev_extracted_count == 1
    assert len(result.warnings) == 0

    raw_mat_pe_dir = mock_env.storage.get_raw_material_dir() / "RAUB" / "01. MAY" / "01-05-2026" / "001" / "RAW DATA"
    assert (raw_mat_pe_dir / "IR" / "IR_0001.jpg").exists()
    assert (raw_mat_pe_dir / "IR" / "DC_0002.jpg").exists()
    assert (raw_mat_pe_dir / "IR" / "IR_0002.jpg").exists()
    assert (raw_mat_pe_dir / "IR" / "DC_0003.jpg").exists()
    assert (raw_mat_pe_dir / "DG" / "P1000010.JPG").exists()
    assert (raw_mat_pe_dir / "DG" / "P1000011.JPG").exists()
    assert (raw_mat_pe_dir / "US+TEV" / "20260501T090000_001-SSU-CHEROH" / "index.html").exists()



def test_raw_material_workflow_with_us_tev_sorting(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    import zipfile

    date_dir = tmp_path / "TESTSHEET" / "KUANTAN" / "01. AUGUST" / "15-08-2026"
    date_dir.mkdir(parents=True)
    unsorted_dir = date_dir / "UNSORTED RAW DATA"
    unsorted_dir.mkdir()
    unsorted_ustev_dir = unsorted_dir / "US+TEV"
    unsorted_ustev_dir.mkdir()

    # Create synthetic testsheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAW DATA"
    ws.cell(1, 1, "PE NO")
    ws.cell(1, 2, 83)
    ws.cell(2, 1, "SUBSTATION NAME")
    ws.cell(2, 2, "MEDAN WARISAN 2")
    ws.cell(3, 1, "FL NUMBER")
    ws.cell(3, 2, "CKTN-S083")
    ws.cell(4, 1, "DATE")
    ws.cell(4, 2, "15-08-2026")
    ws.cell(8, 1, "IR START")
    ws.cell(8, 2, 10)
    ws.cell(8, 3, "IR END")
    ws.cell(8, 4, 11)
    ws.cell(9, 1, "DG START")
    ws.cell(9, 2, 100)
    ws.cell(9, 3, "DG END")
    ws.cell(9, 4, 101)
    wb.save(date_dir / "083. MEDAN WARISAN 2.xlsx")
    wb.close()

    # Create TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    PopulateTotalPeWorkflow().execute(mock_env, PopulateTotalPeRequest(mode=PopulateMode.AUTO))

    # Add raw photos and US+TEV zip
    (unsorted_dir / "FLIR0010.jpg").touch()
    (unsorted_dir / "FLIR0011.jpg").touch()
    (unsorted_dir / "IMG_0100.jpg").touch()
    (unsorted_dir / "IMG_0101.jpg").touch()

    zip_name = "20260815T094744_083-MEDAN-WARISAN-2.zip"
    with zipfile.ZipFile(unsorted_ustev_dir / zip_name, "w") as z:
        z.writestr("index.html", "<html>Test</html>")
        z.writestr("survey_metadata.js", "var survey_metadata = {};")
        z.writestr("SWG/feeder.html", "<html>Feeder</html>")

    progress_messages: list[str] = []
    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=date_dir, progress_sink=progress_messages.append)
    result = workflow.execute(mock_env, request)

    assert result.substations_count == 1
    assert result.summary.ir_copied_count == 2
    assert result.summary.dg_copied_count == 2
    assert result.summary.us_tev_extracted_count == 1
    assert result.us_tev_extracted_count == 1
    assert len(result.warnings) == 0

    # Verify extracted directory under RAW MATERIAL/KUANTAN/01. AUGUST/15-08-2026/083/RAW DATA/US+TEV/20260815T094744_083-MEDAN-WARISAN-2
    extracted_pe_dir = (
        mock_env.storage.get_raw_material_dir()
        / "KUANTAN"
        / "01. AUGUST"
        / "15-08-2026"
        / "083"
        / "RAW DATA"
        / "US+TEV"
        / "20260815T094744_083-MEDAN-WARISAN-2"
    )
    assert (extracted_pe_dir / "index.html").exists()
    assert (extracted_pe_dir / "survey_metadata.js").exists()
    assert (extracted_pe_dir / "SWG" / "feeder.html").exists()

    # Verify progress message format
    assert any("1 US+TEV survey extracted" in msg for msg in progress_messages)


def test_raw_material_workflow_resilience_missing_us_tev(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    date_dir = tmp_path / "TESTSHEET" / "KUANTAN" / "01. AUGUST" / "15-08-2026"
    date_dir.mkdir(parents=True)
    unsorted_dir = date_dir / "UNSORTED RAW DATA"
    unsorted_dir.mkdir()

    # Create synthetic testsheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAW DATA"
    ws.cell(1, 1, "PE NO")
    ws.cell(1, 2, 88)
    ws.cell(2, 1, "SUBSTATION NAME")
    ws.cell(2, 2, "HOSPITAL OUTDOOR")
    ws.cell(3, 1, "FL NUMBER")
    ws.cell(3, 2, "CKTN-S088")
    ws.cell(4, 1, "DATE")
    ws.cell(4, 2, "15-08-2026")
    ws.cell(8, 1, "IR START")
    ws.cell(8, 2, 10)
    ws.cell(8, 3, "IR END")
    ws.cell(8, 4, 11)
    ws.cell(9, 1, "DG START")
    ws.cell(9, 2, 100)
    ws.cell(9, 3, "DG END")
    ws.cell(9, 4, 101)
    wb.save(date_dir / "088. HOSPITAL(OUTDOOR) (IR).xlsx")
    wb.close()

    # Create TOTAL PE.xlsx
    total_pe_path = mock_env.storage.get_total_pe_path()
    mock_env.storage.ensure_directory(total_pe_path.parent)
    wb_pe = openpyxl.Workbook()
    ws_pe = wb_pe.active
    ws_pe.title = "DataCycle1"
    ws_pe.append(["PE NO", "FL NUMBER", "SUBSTATION NAME", "DATE", "TYPE", "WO", "SCOPE"])
    wb_pe.save(total_pe_path)
    wb_pe.close()

    PopulateTotalPeWorkflow().execute(mock_env, PopulateTotalPeRequest(mode=PopulateMode.AUTO))

    # Add raw photos but NO US+TEV zip
    (unsorted_dir / "FLIR0010.jpg").touch()
    (unsorted_dir / "FLIR0011.jpg").touch()
    (unsorted_dir / "IMG_0100.jpg").touch()
    (unsorted_dir / "IMG_0101.jpg").touch()

    workflow = RawMaterialWorkflow()
    request = RawMaterialRequest(output_path=date_dir)
    result = workflow.execute(mock_env, request)

    assert result.substations_count == 1
    assert result.summary.ir_copied_count == 2
    assert result.summary.dg_copied_count == 2
    assert result.summary.us_tev_extracted_count == 0
    assert result.us_tev_extracted_count == 0

    # Non-blocking warning logged
    assert len(result.warnings) == 1
    assert "No US+TEV archive found" in result.warnings[0]

    # Empty US+TEV directory still provisioned
    raw_mat_pe_dir = mock_env.storage.get_raw_material_dir() / "KUANTAN" / "01. AUGUST" / "15-08-2026" / "088" / "RAW DATA"
    assert (raw_mat_pe_dir / "US+TEV").exists()



