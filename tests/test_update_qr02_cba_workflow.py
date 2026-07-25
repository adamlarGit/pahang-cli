"""Unit tests for Update QR02 CBA Workflow Orchestrator in Pahang CLI."""

from __future__ import annotations

import json
from pathlib import Path
import openpyxl
import pytest

from src.master.qr02 import FakeQr02Repository, Qr02Repository
from src.project.environment import ProjectEnvironment
from src.project.models import ProjectMetadata
from src.project.storage import LocalWorkspaceStorage, WorkspaceStorage
from src.workflows.models import PopulateMode, UpdateQr02CbaRequest, UpdateQr02CbaResult
from src.workflows.update_qr02_cba import UpdateQr02CbaWorkflow, run_update_qr02_cba


@pytest.fixture
def mock_env(tmp_path: Path) -> ProjectEnvironment:
    meta = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    storage = LocalWorkspaceStorage(tmp_path)
    return ProjectEnvironment(metadata=meta, storage=storage)


def _create_testsheet_file(
    folder_path: Path, filename: str, fl: str, sub_name: str, date_str: str
) -> Path:
    folder_path.mkdir(parents=True, exist_ok=True)
    (folder_path / "UNSORTED RAW DATA").mkdir(exist_ok=True)
    file_path = folder_path / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCE Testsheet"
    ws["W5"] = fl
    ws["C5"] = sub_name
    ws["P4"] = date_str
    wb.save(file_path)
    wb.close()
    return file_path


def test_discover_packages_and_station_grouping(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # Setup packages for two stations
    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    _create_testsheet_file(raub_dir, "001. SSU CHEROH.xlsx", "CRAU-S001", "SSU CHEROH", "01-05-2026")

    ktn_dir = tmp_path / "TESTSHEET" / "KUANTAN" / "01. MAY" / "02-05-2026"
    _create_testsheet_file(ktn_dir, "001. PE KUANTAN.xlsx", "CKTN-S001", "PE KUANTAN", "02-05-2026")

    repos: dict[str, FakeQr02Repository] = {}

    def fake_repo_factory(storage: WorkspaceStorage, station: str, year: str) -> Qr02Repository:
        if station not in repos:
            repos[station] = FakeQr02Repository()
        return repos[station]

    workflow = UpdateQr02CbaWorkflow()
    request = UpdateQr02CbaRequest(mode=PopulateMode.ALL)
    result = workflow.execute(mock_env, request, repository_factory=fake_repo_factory)

    assert result.records_updated == 2
    assert "RAUB" in repos
    assert "KUANTAN" in repos
    assert len(repos["RAUB"].records) == 1
    assert len(repos["KUANTAN"].records) == 1
    assert repos["RAUB"].records[0].fl_number == "CRAU-S001"
    assert repos["KUANTAN"].records[0].fl_number == "CKTN-S001"


def test_auto_mode_filtering_with_history(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # Setup history marking RAUB folder as processed
    python_dir = mock_env.storage.get_python_dir()
    python_dir.mkdir(parents=True, exist_ok=True)
    history_file = python_dir / "qr02_processed_folders.json"

    history_data = {
        "RAUB/01. MAY/01-05-2026": {
            "last_processed": "2026-05-01T10:00:00",
            "files_scanned": 1,
        }
    }
    with history_file.open("w", encoding="utf-8") as f:
        json.dump(history_data, f)

    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    _create_testsheet_file(raub_dir, "001. SSU CHEROH.xlsx", "CRAU-S001", "SSU CHEROH", "01-05-2026")

    ktn_dir = tmp_path / "TESTSHEET" / "KUANTAN" / "01. MAY" / "02-05-2026"
    _create_testsheet_file(ktn_dir, "001. PE KUANTAN.xlsx", "CKTN-S001", "PE KUANTAN", "02-05-2026")

    repos: dict[str, FakeQr02Repository] = {}

    def fake_repo_factory(storage: WorkspaceStorage, station: str, year: str) -> Qr02Repository:
        if station not in repos:
            repos[station] = FakeQr02Repository()
        return repos[station]

    request = UpdateQr02CbaRequest(mode=PopulateMode.AUTO)
    result = run_update_qr02_cba(mock_env, request, repository_factory=fake_repo_factory)

    assert result.records_updated == 1
    assert "RAUB" not in repos
    assert "KUANTAN" in repos


def test_all_mode_reprocesses_history(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    # History exists for RAUB
    python_dir = mock_env.storage.get_python_dir()
    python_dir.mkdir(parents=True, exist_ok=True)
    history_file = python_dir / "qr02_processed_folders.json"
    with history_file.open("w", encoding="utf-8") as f:
        json.dump({"RAUB/01. MAY/01-05-2026": {"last_processed": "old"}}, f)

    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    _create_testsheet_file(raub_dir, "001. SSU CHEROH.xlsx", "CRAU-S001", "SSU CHEROH", "01-05-2026")

    repos: dict[str, FakeQr02Repository] = {}

    def fake_repo_factory(storage: WorkspaceStorage, station: str, year: str) -> Qr02Repository:
        if station not in repos:
            repos[station] = FakeQr02Repository()
        return repos[station]

    request = UpdateQr02CbaRequest(mode=PopulateMode.ALL)
    result = run_update_qr02_cba(mock_env, request, repository_factory=fake_repo_factory)

    assert result.records_updated == 1
    assert "RAUB" in repos


def test_specific_mode_filtering(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    _create_testsheet_file(raub_dir, "001. SSU CHEROH.xlsx", "CRAU-S001", "SSU CHEROH", "01-05-2026")

    ktn_dir = tmp_path / "TESTSHEET" / "KUANTAN" / "01. MAY" / "02-05-2026"
    _create_testsheet_file(ktn_dir, "001. PE KUANTAN.xlsx", "CKTN-S001", "PE KUANTAN", "02-05-2026")

    repos: dict[str, FakeQr02Repository] = {}

    def fake_repo_factory(storage: WorkspaceStorage, station: str, year: str) -> Qr02Repository:
        if station not in repos:
            repos[station] = FakeQr02Repository()
        return repos[station]

    request = UpdateQr02CbaRequest(mode=PopulateMode.SPECIFIC_FOLDERS, target_package_names=("01-05-2026",))
    result = run_update_qr02_cba(mock_env, request, repository_factory=fake_repo_factory)

    assert result.records_updated == 1
    assert "RAUB" in repos
    assert "KUANTAN" not in repos


def test_history_persistence(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    _create_testsheet_file(raub_dir, "001. SSU CHEROH.xlsx", "CRAU-S001", "SSU CHEROH", "01-05-2026")

    repos: dict[str, FakeQr02Repository] = {}

    def fake_repo_factory(storage: WorkspaceStorage, station: str, year: str) -> Qr02Repository:
        if station not in repos:
            repos[station] = FakeQr02Repository()
        return repos[station]

    request = UpdateQr02CbaRequest(mode=PopulateMode.ALL)
    result = run_update_qr02_cba(mock_env, request, repository_factory=fake_repo_factory)

    assert result.records_updated == 1
    assert "RAUB/01. MAY/01-05-2026" in result.processed_folders

    history_file = mock_env.storage.get_python_dir() / "qr02_processed_folders.json"
    assert history_file.exists()

    with history_file.open("r", encoding="utf-8") as f:
        data = json.load(f)

    assert "RAUB/01. MAY/01-05-2026" in data
    assert data["RAUB/01. MAY/01-05-2026"]["files_scanned"] == 1
    assert "last_processed" in data["RAUB/01. MAY/01-05-2026"]


def test_extraction_failure_logs_warning(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    # Create invalid empty file that fails openpyxl parsing
    raub_dir.mkdir(parents=True, exist_ok=True)
    (raub_dir / "UNSORTED RAW DATA").mkdir(exist_ok=True)
    bad_file = raub_dir / "001. BAD.xlsx"
    bad_file.write_bytes(b"invalid data")

    repos: dict[str, FakeQr02Repository] = {}

    def fake_repo_factory(storage: WorkspaceStorage, station: str, year: str) -> Qr02Repository:
        if station not in repos:
            repos[station] = FakeQr02Repository()
        return repos[station]

    request = UpdateQr02CbaRequest(mode=PopulateMode.ALL)
    result = run_update_qr02_cba(mock_env, request, repository_factory=fake_repo_factory)

    assert result.records_updated == 0
    assert len(result.warnings) == 1
    assert "Failed to extract testsheet data" in result.warnings[0]


def test_missing_engr_file_raises_error(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    raub_dir = tmp_path / "TESTSHEET" / "RAUB" / "01. MAY" / "01-05-2026"
    _create_testsheet_file(raub_dir, "001. SSU CHEROH.xlsx", "CRAU-S001", "SSU CHEROH", "01-05-2026")

    # Run without providing fake_repo_factory (so LocalExcelQr02Repository runs)
    request = UpdateQr02CbaRequest(mode=PopulateMode.ALL)
    with pytest.raises(FileNotFoundError):
        run_update_qr02_cba(mock_env, request)
