"""Unit tests for TestsheetExtractor module in Pahang CLI."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import openpyxl
import pytest

from src.core.normalizers import format_month_folder
from src.testsheet.extractor import (
    TestsheetExtractor,
    clean_val,
    is_marked,
    normalize_building_type,
    normalize_fl_erms,
    to_excel_date,
)
from src.testsheet.models import PhotoRange, RawPhotoRanges, TestsheetData


def test_format_month_folder() -> None:
    """Verify format_month_folder converts month inputs to XX. MONTH format."""
    assert format_month_folder("01. JAN") == "01. JANUARY"
    assert format_month_folder("01. JANUARY") == "01. JANUARY"
    assert format_month_folder("2026-01 (Jan)") == "01. JANUARY"
    assert format_month_folder("01-01-2026") == "01. JANUARY"
    assert format_month_folder("JANUARY") == "01. JANUARY"
    assert format_month_folder("january") == "01. JANUARY"
    assert format_month_folder("02. FEB") == "02. FEBRUARY"
    assert format_month_folder("03. MARCH") == "03. MARCH"
    assert format_month_folder("04. APRIL") == "04. APRIL"
    assert format_month_folder("05. MAY") == "05. MAY"
    assert format_month_folder("06. JUNE") == "06. JUNE"
    assert format_month_folder("07. JULY") == "07. JULY"
    assert format_month_folder("08. AUGUST") == "08. AUGUST"
    assert format_month_folder("09. SEPTEMBER") == "09. SEPTEMBER"
    assert format_month_folder("10. OCTOBER") == "10. OCTOBER"
    assert format_month_folder("11. NOVEMBER") == "11. NOVEMBER"
    assert format_month_folder("12. DECEMBER") == "12. DECEMBER"
    assert format_month_folder(date(2026, 3, 15)) == "03. MARCH"
    assert format_month_folder(datetime(2026, 12, 1, 10, 0)) == "12. DECEMBER"



def test_normalize_fl_erms() -> None:
    """Test normalize_fl_erms strips whitespace, handles .0 suffix, and handles None."""
    assert normalize_fl_erms("  CRAU-S001.0  ") == "CRAU-S001"
    assert normalize_fl_erms("CRAU-S001.0\t") == "CRAU-S001"
    assert normalize_fl_erms(None) == ""
    assert normalize_fl_erms("  TEST-FL  ") == "TEST-FL"
    assert normalize_fl_erms(12345) == "12345"


def test_clean_val() -> None:
    """Test clean_val returns None for empty/dash/NONE and strips whitespace."""
    assert clean_val(None) is None
    assert clean_val("") is None
    assert clean_val("   \t ") is None
    assert clean_val("-") is None
    assert clean_val("NONE") is None
    assert clean_val("None") is None
    assert clean_val("N/A") is None
    assert clean_val("#REF!") is None
    assert clean_val("nan") is None
    assert clean_val("  RM CHEROH \t ") == "RM CHEROH"


def test_is_marked() -> None:
    """Test is_marked returns True for checkmarks and False for empty/NO/N-A/0."""
    assert is_marked("/") is True
    assert is_marked("X") is True
    assert is_marked("YES") is True
    assert is_marked("TRUE") is True
    assert is_marked("v") is True

    assert is_marked(None) is False
    assert is_marked("") is False
    assert is_marked("  ") is False
    assert is_marked("NO") is False
    assert is_marked("N/A") is False
    assert is_marked("0") is False
    assert is_marked("-") is False
    assert is_marked("NONE") is False
    assert is_marked("FALSE") is False


def test_normalize_building_type() -> None:
    """Test normalize_building_type maps to ATTACH/INDOOR/OUTDOOR and None for empty/dash."""
    assert normalize_building_type("ATTACHED PE") == "ATTACH"
    assert normalize_building_type("INDOOR PE") == "INDOOR"
    assert normalize_building_type("BANGUNAN DALAMAN") == "INDOOR"
    assert normalize_building_type("OUTDOOR SUBSTATION") == "OUTDOOR"
    assert normalize_building_type("STESEN LUARAN") == "OUTDOOR"

    assert normalize_building_type(None) is None
    assert normalize_building_type("") is None
    assert normalize_building_type("-") is None
    assert normalize_building_type("NONE") is None
    assert normalize_building_type("N/A") is None


def test_to_excel_date() -> None:
    """Test to_excel_date handles datetime, date, string formats, and None."""
    dt = datetime(2026, 5, 1, 14, 30)
    assert to_excel_date(dt) == dt

    d = date(2026, 5, 1)
    assert to_excel_date(d) == datetime(2026, 5, 1, 0, 0)

    assert to_excel_date("01/05/2026") == datetime(2026, 5, 1)
    assert to_excel_date("01-05-2026") == datetime(2026, 5, 1)
    assert to_excel_date("2026-05-01") == datetime(2026, 5, 1)
    assert to_excel_date("01.05.2026") == datetime(2026, 5, 1)
    assert to_excel_date("01 May 2026") == datetime(2026, 5, 1)
    assert to_excel_date("01-May-2026") == datetime(2026, 5, 1)

    assert to_excel_date(None) is None
    assert to_excel_date("") is None
    assert to_excel_date("-") is None
    assert to_excel_date("invalid-date") is None


@pytest.fixture
def sample_testsheet_file(tmp_path: Path) -> Path:
    file_path = tmp_path / "001. RM CHEROH.xlsx"
    wb = openpyxl.Workbook()

    # PCE Testsheet sheet with fixed cells
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["W5"] = "CRAU-S001"
    ws_pce["C5"] = "RM CHEROH"
    ws_pce["P4"] = "01-05-2026"
    ws_pce["P5"] = "1430"
    ws_pce["S6"] = "65.0"
    ws_pce["W6"] = "BACKGROUND TEMP : 23.2 °C"

    # PCE VI sheet with fixed cells
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "RM CHEROH SITE"
    ws_vi["C8"] = "3.8, 102.1"
    ws_vi["N1"] = "RM"
    ws_vi["C9"] = "OUTDOOR"
    ws_vi["D9"] = "/"

    # RAW DATA sheet with photo ranges only (schema: Row 1=[None, START, END], Row 2=[IR, start, end], Row 3=[DG, start, end])
    ws = wb.create_sheet(title="RAW DATA")
    ws.cell(1, 2, "START")
    ws.cell(1, 3, "END")
    ws.cell(2, 1, "IR")
    ws.cell(2, 2, 100)
    ws.cell(2, 3, 105)
    ws.cell(3, 1, "DG")
    ws.cell(3, 2, 500)
    ws.cell(3, 3, 510)

    wb.save(file_path)
    wb.close()
    return file_path


def test_extract_testsheet_data(sample_testsheet_file: Path) -> None:
    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(sample_testsheet_file)

    assert isinstance(data, TestsheetData)
    assert data.substation_number == 1
    assert data.substation_name_erms == "RM CHEROH"
    assert data.fl_erms == "CRAU-S001"
    assert data.date_str == ""
    assert data.cycle_1 == datetime(2026, 5, 1)
    assert data.substation_type == "RM"
    assert data.substation_name_site == "RM CHEROH SITE"
    assert data.gps_coordinate == "3.8, 102.1"
    assert data.building_type == "OUTDOOR"
    assert data.time == "02:30 PM"
    assert data.humidity == "65%"
    assert data.ambient == "23.2 °C"

    assert data.photo_ranges.ir == PhotoRange(start_num=100, end_num=105)
    assert data.photo_ranges.dg == PhotoRange(start_num=500, end_num=510)
    assert data.photo_ranges.ir.contains(102) is True
    assert data.photo_ranges.ir.contains(200) is False


def test_fixed_cell_extraction(tmp_path: Path) -> None:
    """Test fixed cell extraction from PCE Testsheet, PCE VI, and RAW DATA sheets."""
    file_path = tmp_path / "002. PPU BENTA.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: PCE Testsheet
    ws_pce = wb.active
    ws_pce.title = "PCE Testsheet"
    ws_pce["W5"] = "  CRAU-S002.0  "
    ws_pce["C5"] = "PPU BENTA ERMS"
    ws_pce["P4"] = "15-06-2026"

    # Sheet 2: PCE VI
    ws_vi = wb.create_sheet(title="PCE VI")
    ws_vi["C7"] = "PPU BENTA SITE"
    ws_vi["C8"] = "3.8123, 102.1234"
    ws_vi["N1"] = "PPU"
    ws_vi["C9"] = "ATTACHED"
    ws_vi["D9"] = "/"

    # Sheet 3: RAW DATA
    ws_raw = wb.create_sheet(title="RAW DATA")
    ws_raw.cell(1, 2, "START")
    ws_raw.cell(1, 3, "END")
    ws_raw.cell(2, 1, "IR")
    ws_raw.cell(2, 2, 200)
    ws_raw.cell(2, 3, 210)
    ws_raw.cell(3, 1, "DG")
    ws_raw.cell(3, 2, 600)
    ws_raw.cell(3, 3, 615)

    wb.save(file_path)
    wb.close()

    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(file_path)

    assert data.substation_number == 2
    assert data.fl_erms == "CRAU-S002"
    assert data.substation_name_erms == "PPU BENTA ERMS"
    assert data.substation_name_erms == "PPU BENTA ERMS"
    assert data.cycle_1 == datetime(2026, 6, 15)

    assert data.substation_name_site == "PPU BENTA SITE"
    assert data.gps_coordinate == "3.8123, 102.1234"
    assert data.substation_type == "PPU"
    assert data.substation_type == "PPU"
    assert data.building_type == "ATTACH"

    assert data.fl_erms == "CRAU-S002"
    assert data.date_str == ""

    assert data.photo_ranges.ir == PhotoRange(start_num=200, end_num=210)
    assert data.photo_ranges.dg == PhotoRange(start_num=600, end_num=615)


def test_extract_photo_ranges(sample_testsheet_file: Path) -> None:
    extractor = TestsheetExtractor()
    ranges = extractor.extract_photo_ranges(sample_testsheet_file)

    assert isinstance(ranges, RawPhotoRanges)
    assert ranges.ir.start_num == 100
    assert ranges.ir.end_num == 105
    assert ranges.dg.start_num == 500
    assert ranges.dg.end_num == 510


def test_single_photo_range() -> None:
    single_start = PhotoRange(start_num=42, end_num=None)
    assert single_start.is_valid is True
    assert single_start.contains(42) is True
    assert single_start.contains(43) is False

    single_end = PhotoRange(start_num=None, end_num=99)
    assert single_end.is_valid is True
    assert single_end.contains(99) is True
    assert single_end.contains(98) is False


def test_grid_table_photo_range_extraction(tmp_path: Path) -> None:
    """Test extracting photo ranges from grid table format (Row 1: [None, 'START', 'END'], Row 2: ['IR', 49, 66])."""
    file_path = tmp_path / "001. GRID_TABLE.xlsx"
    wb = openpyxl.Workbook()

    ws_test = wb.active
    ws_test.title = "PCE Testsheet"
    ws_test.cell(1, 1, "PE NO")
    ws_test.cell(1, 2, 289)

    ws_raw = wb.create_sheet(title="RAW DATA")
    ws_raw.cell(1, 2, "START")
    ws_raw.cell(1, 3, "END")
    ws_raw.cell(2, 1, "IR")
    ws_raw.cell(2, 2, 49)
    ws_raw.cell(2, 3, 66)
    ws_raw.cell(3, 1, "DG")
    ws_raw.cell(3, 2, 1715)
    ws_raw.cell(3, 3, 1739)

    wb.save(file_path)
    wb.close()

    extractor = TestsheetExtractor()
    data = extractor.extract_testsheet_data(file_path)

    assert data.photo_ranges.ir == PhotoRange(start_num=49, end_num=66)
    assert data.photo_ranges.dg == PhotoRange(start_num=1715, end_num=1739)

