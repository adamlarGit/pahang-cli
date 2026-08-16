"""Unit and integration tests for Propagate Work Orders Workflow."""

from __future__ import annotations

from pathlib import Path
from typing import Any
import openpyxl
import pytest

from src.project.environment import ProjectEnvironment
from src.project.storage import LocalWorkspaceStorage
from src.workflows.models import (
    PropagateWoRequest,
    PropagateWoResult,
)
from src.workflows.propagate_wo import (
    PropagateWoAuditor,
    PropagateWoExtractor,
    PropagateWoFilter,
    PropagateWoLoader,
    PropagateWoPeRow,
    PropagateWoPlan,
    PropagateWoPreflightGuard,
    PropagateWoRawData,
    PropagateWoRowUpdate,
    PropagateWoTransformer,
    PropagateWoWorkflow,
)


def _create_sample_data_msms(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append([
        "Work Order",
        "Location",
        "Description",
        "Substation Name ERMS",
        "FL ERMS",
        "Cycle Date",
        "Substation Number",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _create_sample_total_pe(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataCycle1"
    ws.append([
        "PE NO",
        "FL NUMBER",
        "SUBSTATION NAME",
        "DATE",
        "TYPE",
        "WO",
        "SCOPE",
        "CUSTOM FORMULA",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _setup_test_env(tmp_path: Path) -> tuple[ProjectEnvironment, Path, Path]:
    from src.project.models import ProjectMetadata
    meta = ProjectMetadata(
        key="test_proj",
        name="Test Project",
        po_number="PO12345",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="1",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    storage = LocalWorkspaceStorage(tmp_path)
    env = ProjectEnvironment(
        metadata=meta,
        storage=storage,
    )
    python_dir = storage.get_python_dir()
    testsheet_dir = storage.get_testsheet_dir()
    python_dir.mkdir(parents=True, exist_ok=True)
    testsheet_dir.mkdir(parents=True, exist_ok=True)

    msms_file = storage.get_data_msms_path()
    pe_file = storage.get_total_pe_path()
    return env, msms_file, pe_file


# ---------------------------------------------------------------------------
# Stage 1: PreflightGuard Tests
# ---------------------------------------------------------------------------


def test_preflight_guard_missing_data_msms_raises(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_total_pe(pe_file, [[1, "FL1", "PE 1", "2026-01-10", "ATT", None, "FULL", None]])
    
    guard = PropagateWoPreflightGuard()
    with pytest.raises(FileNotFoundError, match=r"DATA MSMS\.xlsx"):
        guard.validate(env)


def test_preflight_guard_missing_total_pe_raises(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [["45001", "CKTN01", "PE 1", "PE 1", "CKTN01", "2026-01-10", 1]])
    
    guard = PropagateWoPreflightGuard()
    with pytest.raises(FileNotFoundError, match=r"TOTAL PE\.xlsx"):
        guard.validate(env)


def test_preflight_guard_missing_datacycle1_sheet_raises(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [["45001", "CKTN01", "PE 1", "PE 1", "CKTN01", "2026-01-10", 1]])
    
    # Create TOTAL PE without DataCycle1
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WrongSheet"
    ws.append(["Col1"])
    wb.save(pe_file)
    wb.close()

    guard = PropagateWoPreflightGuard()
    with pytest.raises(RuntimeError, match=r"DataCycle1"):
        guard.validate(env)


def test_preflight_guard_empty_data_msms_raises(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [])  # No data rows
    _create_sample_total_pe(pe_file, [[1, "FL1", "PE 1", "2026-01-10", "ATT", None, "FULL", None]])

    guard = PropagateWoPreflightGuard()
    with pytest.raises(RuntimeError, match=r"DATA MSMS.*records"):
        guard.validate(env)


def test_preflight_guard_success(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [["45001", "CKTN01", "PE 1", "PE 1", "CKTN01", "2026-01-10", 1]])
    _create_sample_total_pe(pe_file, [[1, "FL1", "PE 1", "2026-01-10", "ATT", None, "FULL", None]])

    guard = PropagateWoPreflightGuard()
    guard.validate(env)  # Should not raise


# ---------------------------------------------------------------------------
# Stage 2: Extractor Tests
# ---------------------------------------------------------------------------


def test_extractor_reads_fl_to_wo_and_pe_rows(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE TELUK SISIK", "PE TELUK SISIK", "CKTN0001/AAAA", "2026-01-10", 1],
        ["45001002", "CKTN0002BBBB", "PE BUKIT UBI", "PE BUKIT UBI", "CKTN0002/BBBB", "2026-01-11", 2],
    ])
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE TELUK SISIK", "2026-01-10", "ATTACHED", None, "FULL", "=IF(A2>0, 1, 0)"],
        [2, "CKTN0002/BBBB", "PE BUKIT UBI", "2026-01-11", "ATTACHED", "45009999", "FULL", "=IF(A3>0, 1, 0)"],
    ])

    extractor = PropagateWoExtractor()
    raw_data = extractor.extract(pe_file, msms_file)

    assert isinstance(raw_data, PropagateWoRawData)
    assert raw_data.fl_to_wo.get("CKTN0001/AAAA") == "45001001"
    assert raw_data.fl_to_wo.get("CKTN0001AAAA") == "45001001"
    assert raw_data.fl_to_wo.get("CKTN0002/BBBB") == "45001002"
    assert len(raw_data.pe_rows) == 2
    assert raw_data.pe_rows[0].row_index == 2
    assert raw_data.pe_rows[0].fl_num == "CKTN0001/AAAA"
    assert raw_data.pe_rows[0].wo is None
    assert raw_data.pe_rows[1].wo == "45009999"


# ---------------------------------------------------------------------------
# Stage 3: Filter Tests
# ---------------------------------------------------------------------------


def test_filter_matches_fl_and_populates_empty_wo() -> None:
    raw_data = PropagateWoRawData(
        fl_to_wo={"CKTN0001/AAAA": "45001001", "CKTN0001AAAA": "45001001"},
        pe_rows=(
            PropagateWoPeRow(
                row_index=2,
                pe_no=1,
                fl_num="CKTN0001/AAAA",
                substation_name="PE TELUK SISIK",
                date_str="2026-01-10",
                type_c="ATTACHED",
                wo=None,
            ),
        ),
        total_pe_path=Path("TOTAL PE.xlsx"),
        data_msms_path=Path("DATA MSMS.xlsx"),
    )

    flt = PropagateWoFilter()
    filtered = flt.filter(raw_data)

    assert filtered.matched_count == 1
    assert filtered.already_populated_count == 0
    assert filtered.unmatched_count == 0
    assert len(filtered.updates) == 1
    assert filtered.updates[0] == PropagateWoRowUpdate(row_index=2, wo="45001001", fl="CKTN0001/AAAA")


def test_filter_skips_already_populated_wo() -> None:
    raw_data = PropagateWoRawData(
        fl_to_wo={"CKTN0001/AAAA": "45001001", "CKTN0001AAAA": "45001001"},
        pe_rows=(
            PropagateWoPeRow(
                row_index=2,
                pe_no=1,
                fl_num="CKTN0001/AAAA",
                substation_name="PE TELUK SISIK",
                date_str="2026-01-10",
                type_c="ATTACHED",
                wo="45009999",
            ),
        ),
        total_pe_path=Path("TOTAL PE.xlsx"),
        data_msms_path=Path("DATA MSMS.xlsx"),
    )

    flt = PropagateWoFilter()
    filtered = flt.filter(raw_data)

    assert filtered.matched_count == 0
    assert filtered.already_populated_count == 1
    assert filtered.unmatched_count == 0
    assert len(filtered.updates) == 0


def test_filter_flags_unmatched_fl() -> None:
    raw_data = PropagateWoRawData(
        fl_to_wo={"CKTN0001/AAAA": "45001001"},
        pe_rows=(
            PropagateWoPeRow(
                row_index=2,
                pe_no=2,
                fl_num="CKTN9999/ZZZZ",
                substation_name="PE UNKNOWN",
                date_str="2026-01-10",
                type_c="ATTACHED",
                wo=None,
            ),
        ),
        total_pe_path=Path("TOTAL PE.xlsx"),
        data_msms_path=Path("DATA MSMS.xlsx"),
    )

    flt = PropagateWoFilter()
    filtered = flt.filter(raw_data)

    assert filtered.matched_count == 0
    assert filtered.already_populated_count == 0
    assert filtered.unmatched_count == 1
    assert "CKTN9999/ZZZZ" in filtered.unmatched_fls
    assert len(filtered.updates) == 0


def test_filter_with_target_date() -> None:
    raw_data = PropagateWoRawData(
        fl_to_wo={
            "CKTN0001/AAAA": "45001001",
            "CKTN0002/BBBB": "45001002",
        },
        pe_rows=(
            PropagateWoPeRow(
                row_index=2,
                pe_no=1,
                fl_num="CKTN0001/AAAA",
                substation_name="PE 1",
                date_str="2026-01-10",
                type_c="ATTACHED",
                wo=None,
            ),
            PropagateWoPeRow(
                row_index=3,
                pe_no=2,
                fl_num="CKTN0002/BBBB",
                substation_name="PE 2",
                date_str="2026-01-11",
                type_c="ATTACHED",
                wo=None,
            ),
        ),
        total_pe_path=Path("TOTAL PE.xlsx"),
        data_msms_path=Path("DATA MSMS.xlsx"),
    )

    flt = PropagateWoFilter()
    filtered = flt.filter(raw_data, target_date="2026-01-10")

    assert filtered.matched_count == 1
    assert len(filtered.updates) == 1
    assert filtered.updates[0].row_index == 2


# ---------------------------------------------------------------------------
# Stage 4: Transformer Tests
# ---------------------------------------------------------------------------


def test_transformer_builds_immutable_plan() -> None:
    transformer = PropagateWoTransformer()
    total_pe_path = Path("TOTAL PE.xlsx")
    updates = (PropagateWoRowUpdate(row_index=2, wo="45001001", fl="CKTN0001/AAAA"),)

    plan = transformer.build_plan(
        total_pe_path=total_pe_path,
        updates=updates,
        matched_count=1,
        already_populated_count=2,
        unmatched_count=3,
        unmatched_fls=("CKTN9999/ZZZZ",),
        target_date="2026-01-10",
    )

    assert isinstance(plan, PropagateWoPlan)
    assert plan.total_pe_path == total_pe_path
    assert len(plan.updates) == 1
    assert plan.matched_count == 1
    assert plan.already_populated_count == 2
    assert plan.unmatched_count == 3
    assert plan.unmatched_fls == ("CKTN9999/ZZZZ",)
    assert plan.target_date == "2026-01-10"


# ---------------------------------------------------------------------------
# Stage 5: Loader Tests
# ---------------------------------------------------------------------------


def test_loader_preserves_formulas_and_updates_col_f(tmp_path: Path) -> None:
    pe_file = tmp_path / "TOTAL PE.xlsx"
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE TELUK SISIK", "2026-01-10", "ATTACHED", None, "FULL", "=IF(A2>0, 1, 0)"],
        [2, "CKTN0002/BBBB", "PE BUKIT UBI", "2026-01-11", "ATTACHED", "45009999", "FULL", "=IF(A3>0, 1, 0)"],
    ])

    loader = PropagateWoLoader()
    plan = PropagateWoPlan(
        total_pe_path=pe_file,
        updates=(PropagateWoRowUpdate(row_index=2, wo="45001001", fl="CKTN0001/AAAA"),),
        matched_count=1,
        already_populated_count=1,
        unmatched_count=0,
        unmatched_fls=(),
    )

    updated_count = loader.load(plan)
    assert updated_count == 1

    wb = openpyxl.load_workbook(pe_file)
    ws = wb["DataCycle1"]
    assert ws.cell(2, 6).value == "45001001"
    assert ws.cell(2, 8).value == "=IF(A2>0, 1, 0)"
    assert ws.cell(2, 1).value == 1
    assert ws.cell(3, 6).value == "45009999"
    assert ws.cell(3, 8).value == "=IF(A3>0, 1, 0)"
    wb.close()


# ---------------------------------------------------------------------------
# Stage 6: Auditor Tests
# ---------------------------------------------------------------------------


def test_auditor_success(tmp_path: Path) -> None:
    pe_file = tmp_path / "TOTAL PE.xlsx"
    pe_file.write_text("dummy")

    auditor = PropagateWoAuditor()
    plan = PropagateWoPlan(
        total_pe_path=pe_file,
        updates=(PropagateWoRowUpdate(row_index=2, wo="45001001", fl="CKTN0001/AAAA"),),
        matched_count=1,
        already_populated_count=2,
        unmatched_count=1,
        unmatched_fls=("FL_UNMATCHED",),
    )

    result = auditor.audit(plan, updated_count=1)
    assert isinstance(result, PropagateWoResult)
    assert result.matched_count == 1
    assert result.already_populated_count == 2
    assert result.unmatched_count == 1
    assert result.unmatched_fls == ("FL_UNMATCHED",)
    assert result.updated_count == 1


def test_auditor_missing_file_raises(tmp_path: Path) -> None:
    pe_file = tmp_path / "NON_EXISTENT.xlsx"
    auditor = PropagateWoAuditor()
    plan = PropagateWoPlan(
        total_pe_path=pe_file,
        updates=(),
        matched_count=0,
        already_populated_count=0,
        unmatched_count=0,
        unmatched_fls=(),
    )
    with pytest.raises(RuntimeError, match=r"does not exist"):
        auditor.audit(plan, updated_count=0)


def test_auditor_empty_file_raises(tmp_path: Path) -> None:
    pe_file = tmp_path / "EMPTY.xlsx"
    pe_file.touch()
    auditor = PropagateWoAuditor()
    plan = PropagateWoPlan(
        total_pe_path=pe_file,
        updates=(),
        matched_count=0,
        already_populated_count=0,
        unmatched_count=0,
        unmatched_fls=(),
    )
    with pytest.raises(RuntimeError, match=r"empty"):
        auditor.audit(plan, updated_count=0)


# ---------------------------------------------------------------------------
# End-to-End Workflow Integration Tests
# ---------------------------------------------------------------------------


def test_propagate_wo_workflow_e2e_all(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE TELUK SISIK", "PE TELUK SISIK", "CKTN0001/AAAA", "2026-01-10", 1],
        ["45001002", "CKTN0002BBBB", "PE BUKIT UBI", "PE BUKIT UBI", "CKTN0002/BBBB", "2026-01-11", 2],
        ["45001003", "CKTN0003CCCC", "PE ALOR AKAR", "PE ALOR AKAR", "CKTN0003/CCCC", "2026-01-12", 3],
    ])
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE TELUK SISIK", "2026-01-10", "ATTACHED", None, "FULL", "=IF(A2>0, 1, 0)"],
        [2, "CKTN0002/BBBB", "PE BUKIT UBI", "2026-01-11", "ATTACHED", "45009999", "FULL", "=IF(A3>0, 1, 0)"],
        [3, "CKTN0009/ZZZZ", "PE UNKNOWN", "2026-01-12", "COMPACT", None, "FULL", "=IF(A4>0, 1, 0)"],
    ])

    progress_messages: list[str] = []
    request = PropagateWoRequest(progress_sink=progress_messages.append)
    workflow = PropagateWoWorkflow()

    result = workflow.execute(env, request)

    assert isinstance(result, PropagateWoResult)
    assert result.matched_count == 1
    assert result.already_populated_count == 1
    assert result.unmatched_count == 1
    assert "CKTN0009/ZZZZ" in result.unmatched_fls
    assert result.updated_count == 1
    assert len(progress_messages) > 0

    # Verify openpyxl output
    wb = openpyxl.load_workbook(pe_file)
    ws = wb["DataCycle1"]
    assert ws.cell(2, 6).value == "45001001"
    assert ws.cell(2, 8).value == "=IF(A2>0, 1, 0)"
    assert ws.cell(3, 6).value == "45009999"
    assert ws.cell(4, 6).value is None
    wb.close()


def test_propagate_wo_workflow_e2e_target_date(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE A", "PE A", "CKTN0001/AAAA", "2026-01-10", 1],
        ["45001002", "CKTN0002BBBB", "PE B", "PE B", "CKTN0002/BBBB", "2026-01-11", 2],
    ])
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE A", "2026-01-10", "ATTACHED", None, "FULL", None],
        [2, "CKTN0002/BBBB", "PE B", "2026-01-11", "ATTACHED", None, "FULL", None],
    ])

    request = PropagateWoRequest(target_date="2026-01-10")
    workflow = PropagateWoWorkflow()
    result = workflow.execute(env, request)

    assert result.matched_count == 1
    assert result.updated_count == 1

    wb = openpyxl.load_workbook(pe_file)
    ws = wb["DataCycle1"]
    assert ws.cell(2, 6).value == "45001001"
    assert ws.cell(3, 6).value is None  # Skipped
    wb.close()


def test_propagate_wo_workflow_idempotency(tmp_path: Path) -> None:
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE A", "PE A", "CKTN0001/AAAA", "2026-01-10", 1],
    ])
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE A", "2026-01-10", "ATTACHED", None, "FULL", None],
    ])

    workflow = PropagateWoWorkflow()
    res1 = workflow.execute(env, PropagateWoRequest())
    assert res1.matched_count == 1
    assert res1.already_populated_count == 0
    assert res1.updated_count == 1

    # Second run: should be skipped as already populated
    res2 = workflow.execute(env, PropagateWoRequest())
    assert res2.matched_count == 0
    assert res2.already_populated_count == 1
    assert res2.updated_count == 0

    # Third run with overwrite=True: should overwrite
    res3 = workflow.execute(env, PropagateWoRequest(overwrite=True))
    assert res3.matched_count == 1
    assert res3.already_populated_count == 0
    assert res3.updated_count == 1


def test_workflow_service_run_propagate_wo(tmp_path: Path) -> None:
    from src.workflows.service import WorkflowService
    env, msms_file, pe_file = _setup_test_env(tmp_path)
    _create_sample_data_msms(msms_file, [
        ["45001001", "CKTN0001AAAA", "PE A", "PE A", "CKTN0001/AAAA", "2026-01-10", 1],
    ])
    _create_sample_total_pe(pe_file, [
        [1, "CKTN0001/AAAA", "PE A", "2026-01-10", "ATTACHED", None, "FULL", None],
    ])

    service = WorkflowService()
    logs: list[str] = []
    res = service.run_propagate_wo(env, PropagateWoRequest(progress_sink=logs.append))
    assert res.matched_count == 1
    assert res.updated_count == 1
    assert any("Propagate" in msg for msg in logs)

