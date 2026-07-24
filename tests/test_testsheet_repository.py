"""Unit tests for SubstationTestsheetRepository in Pahang CLI."""

from __future__ import annotations

from pathlib import Path
import openpyxl
import pytest

from src.testsheet.repository import SubstationTestsheetRepository


@pytest.fixture
def pahang_testsheet_tree(tmp_path: Path) -> Path:
    testsheet_root = tmp_path / "TESTSHEET"
    date_dir = testsheet_root / "RAUB" / "01. MAY" / "01-05-2026"
    date_dir.mkdir(parents=True)

    unsorted_dir = date_dir / "UNSORTED RAW DATA"
    unsorted_dir.mkdir()

    # Create 001.xlsx and 002.xlsx
    for pe_num, name in [(1, "001. SSU CHEROH.xlsx"), (2, "002. RM CHEROH.xlsx")]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RAW DATA"
        ws.cell(1, 1, "PE NO")
        ws.cell(1, 2, pe_num)
        ws.cell(2, 1, "SUBSTATION NAME")
        ws.cell(2, 2, f"PE_{pe_num}")
        wb.save(date_dir / name)
        wb.close()

    return testsheet_root


def test_discover_packages(pahang_testsheet_tree: Path) -> None:
    repo = SubstationTestsheetRepository()
    packages = repo.discover_packages(pahang_testsheet_tree)

    assert len(packages) == 2

    pkg1 = packages[0]
    assert pkg1.station == "RAUB"
    assert pkg1.month == "01. MAY"
    assert pkg1.date_str == "01-05-2026"
    assert pkg1.pe_num in (1, 2)
    assert pkg1.unsorted_raw_data_dir.name == "UNSORTED RAW DATA"
    assert pkg1.testsheet_path.exists()


def test_discover_packages_direct_date_folder(pahang_testsheet_tree: Path) -> None:
    date_dir = pahang_testsheet_tree / "RAUB" / "01. MAY" / "01-05-2026"
    repo = SubstationTestsheetRepository()
    packages = repo.discover_packages(date_dir)

    assert len(packages) == 2
    pkg = packages[0]
    assert pkg.station == "RAUB"
    assert pkg.month == "01. MAY"
    assert pkg.date_str == "01-05-2026"
