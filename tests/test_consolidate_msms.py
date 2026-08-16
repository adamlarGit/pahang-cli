"""Unit and integration tests for ConsolidateMsmsWorkflow."""
from pathlib import Path
import openpyxl
import pytest

from src.project.environment import ProjectEnvironment
from src.project.models import ProjectMetadata
from src.project.storage import LocalWorkspaceStorage
from src.workflows.models import ConsolidateMsmsRequest, ConsolidateMsmsResult
from src.workflows.consolidate_msms import (
    ConsolidateMsmsPreflightGuard,
    ConsolidateMsmsExtractor,
    ConsolidateMsmsFilter,
    ConsolidateMsmsTransformer,
    ConsolidateMsmsLoader,
    ConsolidateMsmsAuditor,
    ConsolidateMsmsWorkflow,
    ConsolidateMsmsRow,
    ConsolidateMsmsPlan,
)


def _create_sample_data_msms(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append([
        "Work Order",
        "Location",
        "Description",
        "Substation Name ERMS",
        "FL ERMS",
        "Cycle Date",
        "Substation Number",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _create_sample_maximo_html_xls(path: Path, rows: list[tuple[str, str, str, str]]) -> None:
    html_content = ["<html><body><table border='1'>"]
    html_content.append("<tr><th>Wonum</th><th>Status</th><th>Location</th><th>Description</th></tr>")
    for wonum, status, location, desc in rows:
        html_content.append(f"<tr><td>{wonum}</td><td>{status}</td><td>{location}</td><td>{desc}</td></tr>")
    html_content.append("</table></body></html>")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(html_content), encoding="utf-8")


def _setup_project_environment(tmp_path: Path) -> ProjectEnvironment:
    storage = LocalWorkspaceStorage(tmp_path)
    storage.get_python_dir().mkdir(parents=True, exist_ok=True)
    storage.get_python_msms_dir().mkdir(parents=True, exist_ok=True)
    metadata = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    return ProjectEnvironment(metadata=metadata, storage=storage)


# --- PreflightGuard Tests ---

def test_consolidate_msms_preflight_missing_data_msms(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    xls_file = env.storage.get_python_msms_dir() / "sample.xls"
    xls_file.touch()

    guard = ConsolidateMsmsPreflightGuard()
    with pytest.raises(FileNotFoundError, match="DATA MSMS workbook not found"):
        guard.validate(env)


def test_consolidate_msms_preflight_missing_msms_dir(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    _create_sample_data_msms(msms_file, [])
    # Remove msms dir
    msms_dir = env.storage.get_python_msms_dir()
    import shutil
    shutil.rmtree(msms_dir)

    guard = ConsolidateMsmsPreflightGuard()
    with pytest.raises(FileNotFoundError, match="PYTHON/MSMS directory not found"):
        guard.validate(env)


def test_consolidate_msms_preflight_no_xls_files(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    _create_sample_data_msms(msms_file, [])

    guard = ConsolidateMsmsPreflightGuard()
    with pytest.raises(FileNotFoundError, match="No .xls files found"):
        guard.validate(env)


def test_consolidate_msms_preflight_success(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    _create_sample_data_msms(msms_file, [])
    (env.storage.get_python_msms_dir() / "45501001.xls").touch()

    guard = ConsolidateMsmsPreflightGuard()
    guard.validate(env)  # Should not raise


# --- Extractor Tests ---

def test_consolidate_msms_extractor(tmp_path: Path):
    extractor = ConsolidateMsmsExtractor()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [
        ["45001000", "LOC0", "DESC0", None, None, None, None],
    ])

    existing_wos = extractor.get_existing_wos(msms_file)
    assert existing_wos == {"45001000"}

    xls_path = tmp_path / "45501001.xls"
    _create_sample_maximo_html_xls(xls_path, [
        ("45001001", "APPR", "CKTN/PCEJ01565", "PE TELUK SISIK"),
        ("45001002", "APPR", "CKTN0002AAAA", "PE TANJUNG LUMPUR"),
    ])

    records, error = extractor.read_xls_file(xls_path)
    assert error is None
    assert len(records) == 2
    assert records[0]["wo"] == "45001001"
    assert records[0]["location"] == "CKTN/PCEJ01565"
    assert records[0]["description"] == "PE TELUK SISIK"


# --- Filter Tests ---

def test_consolidate_msms_filter_dedup_and_skip():
    filter_stage = ConsolidateMsmsFilter()
    existing_wos = {"45001000"}

    file1 = Path("file1.xls")
    rows1 = [
        {"wo": "45001000", "location": "LOC0", "description": "DESC0"},  # already in existing_wos
        {"wo": "45001001", "location": "LOC1", "description": "DESC1"},  # new
        {"wo": "WONUM", "location": "LOCATION", "description": "DESCRIPTION"},  # header row
    ]

    file2 = Path("file2.xls")
    rows2 = [
        {"wo": "45001001", "location": "LOC1", "description": "DESC1"},  # duplicate of file1
        {"wo": "45001002", "location": "LOC2", "description": "DESC2"},  # new
    ]

    extracted_files = [(file1, rows1), (file2, rows2)]
    filtered, duplicates_skipped = filter_stage.filter_rows(extracted_files, existing_wos)

    assert duplicates_skipped == 2
    # file1 has 1 valid row (45001001)
    # file2 has 1 valid row (45001002)
    assert len(filtered) == 2
    assert len(filtered[0][1]) == 1
    assert filtered[0][1][0]["wo"] == "45001001"
    assert len(filtered[1][1]) == 1
    assert filtered[1][1][0]["wo"] == "45001002"


# --- Transformer Tests ---

def test_consolidate_msms_transformer_plan():
    transformer = ConsolidateMsmsTransformer()
    target_data_msms = Path("target.xlsx")
    completed_dir = Path("completed")

    file1 = Path("file1.xls")
    rows1 = [
        {"wo": "45001001", "location": "CKTN/PCEJ01565", "description": "PE TELUK SISIK"},
        {"wo": "45001002", "location": "CKTN0002AAAA", "description": "PE TANJUNG"},
    ]

    plan = transformer.build_plan(
        target_data_msms=target_data_msms,
        completed_dir=completed_dir,
        filtered_files=[(file1, rows1)],
        duplicates_skipped=1,
        files_processed=1,
        errors=(),
    )

    assert isinstance(plan, ConsolidateMsmsPlan)
    assert len(plan.rows_to_append) == 2
    assert plan.rows_to_append[0].wo == "45001001"
    assert plan.rows_to_append[0].location == "CKTN/PCEJ01565"
    assert plan.rows_to_append[0].fl_erms == "CKTN/PCE/J01565"  # Position 8 slash inserted
    assert plan.rows_to_append[1].fl_erms == "CKTN0002/AAAA"
    assert plan.files_to_move == (file1,)


# --- Loader Tests ---

def test_consolidate_msms_loader(tmp_path: Path):
    loader = ConsolidateMsmsLoader()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [])

    completed_dir = tmp_path / "COMPLETED"
    xls_file = tmp_path / "45501001.xls"
    xls_file.write_text("dummy", encoding="utf-8")

    plan = ConsolidateMsmsPlan(
        target_data_msms=msms_file,
        completed_dir=completed_dir,
        rows_to_append=(
            ConsolidateMsmsRow(
                wo="45001001",
                location="CKTN/PCEJ01565",
                description="PE TELUK SISIK",
                fl_erms="CKTN/PCE/J01565",
            ),
        ),
        files_to_move=(xls_file,),
        duplicates_skipped=0,
        files_processed=1,
        errors=(),
    )

    loader.load(plan)

    # Verify rows appended to Excel
    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # header + 1 row
    assert rows[1][0] == "45001001"
    assert rows[1][1] == "CKTN/PCEJ01565"
    assert rows[1][2] == "PE TELUK SISIK"
    assert rows[1][4] == "CKTN/PCE/J01565"

    # Verify file moved to COMPLETED
    assert not xls_file.exists()
    assert (completed_dir / "45501001.xls").exists()


# --- End-to-End Integration Workflow Tests ---

def test_consolidate_msms_workflow_e2e(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    _create_sample_data_msms(msms_file, [
        ["45001000", "CKTN0001AAAA", "PE EXISTING", None, "CKTN0001/AAAA", None, None],
    ])

    msms_dir = env.storage.get_python_msms_dir()
    xls1 = msms_dir / "45501001.xls"
    _create_sample_maximo_html_xls(xls1, [
        ("45001000", "APPR", "CKTN0001AAAA", "PE EXISTING"),  # duplicate
        ("45001001", "APPR", "CKTN/PCEJ01565", "PE TELUK SISIK"),  # new
    ])

    xls2 = msms_dir / "45501002.xls"
    _create_sample_maximo_html_xls(xls2, [
        ("45001001", "APPR", "CKTN/PCEJ01565", "PE TELUK SISIK"),  # duplicate across files
        ("45001002", "APPR", "CKTN0003/CCCC", "PE BALOK"),  # new, already has slash
    ])

    workflow = ConsolidateMsmsWorkflow()
    result = workflow.execute(env, ConsolidateMsmsRequest())

    assert isinstance(result, ConsolidateMsmsResult)
    assert result.files_processed == 2
    assert result.rows_appended == 2
    assert result.duplicates_skipped == 2
    assert len(result.errors) == 0

    # Verify DATA MSMS.xlsx
    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 4  # header + 1 existing + 2 new

    assert rows[2][0] == "45001001"
    assert rows[2][4] == "CKTN/PCE/J01565"

    assert rows[3][0] == "45001002"
    assert rows[3][4] == "CKTN0003/CCCC"

    # Verify files moved to COMPLETED
    completed_dir = env.storage.get_python_msms_completed_dir()
    assert not xls1.exists()
    assert not xls2.exists()
    assert (completed_dir / "45501001.xls").exists()
    assert (completed_dir / "45501002.xls").exists()


def test_consolidate_msms_resilience_handles_corrupt_file(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    _create_sample_data_msms(msms_file, [])

    msms_dir = env.storage.get_python_msms_dir()
    xls_corrupt = msms_dir / "corrupt.xls"
    xls_corrupt.write_text("NOT VALID HTML OR EXCEL", encoding="utf-8")

    xls_valid = msms_dir / "valid.xls"
    _create_sample_maximo_html_xls(xls_valid, [
        ("45001005", "APPR", "CKTN0005AAAA", "PE VALID"),
    ])

    workflow = ConsolidateMsmsWorkflow()
    result = workflow.execute(env, ConsolidateMsmsRequest())

    assert result.rows_appended == 1
    assert len(result.errors) >= 1
    assert "corrupt.xls" in result.errors[0]

    # Valid file moved to completed, corrupt file not moved
    completed_dir = env.storage.get_python_msms_completed_dir()
    assert (completed_dir / "valid.xls").exists()
    assert xls_corrupt.exists()


def test_consolidate_msms_extractor_handles_no_tables_found(tmp_path: Path):
    extractor = ConsolidateMsmsExtractor()
    empty_html = tmp_path / "empty.xls"
    empty_html.write_text("<html><body><p>No table here</p></body></html>", encoding="utf-8")

    rows, error = extractor.read_xls_file(empty_html)
    assert rows == []
    # Should catch ValueError and return error string or empty rows gracefully
    assert error is not None or rows == []


def test_consolidate_msms_extractor_uses_lxml_flavor(tmp_path: Path):
    from unittest.mock import patch
    import pandas as pd

    extractor = ConsolidateMsmsExtractor()
    xls_path = tmp_path / "test.xls"
    _create_sample_maximo_html_xls(xls_path, [("45001001", "APPR", "CKTN0001", "PE TEST")])

    with patch("pandas.read_html", wraps=pd.read_html) as mock_read_html:
        extractor.read_xls_file(xls_path)
        mock_read_html.assert_called_once()
        _, kwargs = mock_read_html.call_args
        assert kwargs.get("flavor") == "lxml"


def test_msms_repository_uses_lxml_flavor(tmp_path: Path):
    from unittest.mock import patch
    import pandas as pd
    from src.msms.repository import LocalExcelMsmsRepository

    repo = LocalExcelMsmsRepository()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [])
    xls_path = tmp_path / "test.xls"
    _create_sample_maximo_html_xls(xls_path, [("45001001", "APPR", "CKTN0001", "PE TEST")])

    with patch("pandas.read_html", wraps=pd.read_html) as mock_read_html:
        res = repo.consolidate_xls_files([xls_path], msms_file)
        mock_read_html.assert_called_once()
        _, kwargs = mock_read_html.call_args
        assert kwargs.get("flavor") == "lxml"
        assert res.rows_appended == 1


def test_consolidate_msms_loader_contiguous_append_with_blank_formatted_rows(tmp_path: Path):
    """Ensure loader overwrites/cleans 100 pre-existing blank formatted rows contiguously."""
    from openpyxl.styles import PatternFill

    loader = ConsolidateMsmsLoader()
    msms_file = tmp_path / "DATA MSMS.xlsx"

    # Create workbook with header and 100 blank formatted rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append(["Work Order", "Location", "Description", "Substation Name ERMS", "FL ERMS", "Cycle Date", "Substation Number"])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in range(2, 102):
        cell = ws.cell(row=r, column=1)
        cell.fill = yellow_fill
        cell.value = ""
    wb.save(msms_file)
    wb.close()

    plan = ConsolidateMsmsPlan(
        target_data_msms=msms_file,
        completed_dir=tmp_path / "COMPLETED",
        rows_to_append=(
            ConsolidateMsmsRow(wo="45001001", location="CKTN0001AAAA", description="PE ONE", fl_erms="CKTN0001/AAAA"),
            ConsolidateMsmsRow(wo="45001002", location="CKTN0002BBBB", description="PE TWO", fl_erms="CKTN0002/BBBB"),
        ),
        files_to_move=(),
        duplicates_skipped=0,
        files_processed=1,
        errors=(),
    )

    loader.load(plan)

    # Verify rows written contiguously without 100 blank rows before or after
    wb_res = openpyxl.load_workbook(msms_file)
    ws_res = wb_res.active
    assert ws_res.max_row == 3  # Header + 2 rows

    rows = list(ws_res.iter_rows(values_only=True))
    assert len(rows) == 3
    assert rows[1][0] == "45001001"
    assert rows[1][1] == "CKTN0001AAAA"
    assert rows[1][4] == "CKTN0001/AAAA"
    assert rows[2][0] == "45001002"
    assert rows[2][1] == "CKTN0002BBBB"
    assert rows[2][4] == "CKTN0002/BBBB"
    wb_res.close()


def test_consolidate_msms_loader_compacts_disconnected_rows(tmp_path: Path):
    """Ensure loader compacts disconnected rows separated by blank blocks contiguously from row 2."""
    from openpyxl.styles import PatternFill

    loader = ConsolidateMsmsLoader()
    msms_file = tmp_path / "DATA MSMS.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append(["Work Order", "Location", "Description", "Substation Name ERMS", "FL ERMS", "Cycle Date", "Substation Number"])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in range(2, 102):
        ws.cell(row=r, column=1).fill = yellow_fill

    # Insert disconnected rows
    ws.cell(row=11, column=1).value = "45001000"
    ws.cell(row=11, column=2).value = "CKTN0000AAAA"
    ws.cell(row=11, column=3).value = "PE EXISTING 1"
    ws.cell(row=11, column=5).value = "CKTN0000/AAAA"

    ws.cell(row=51, column=1).value = "45001001"
    ws.cell(row=51, column=2).value = "CKTN0001BBBB"
    ws.cell(row=51, column=3).value = "PE EXISTING 2"
    ws.cell(row=51, column=5).value = "CKTN0001/BBBB"

    wb.save(msms_file)
    wb.close()

    plan = ConsolidateMsmsPlan(
        target_data_msms=msms_file,
        completed_dir=tmp_path / "COMPLETED",
        rows_to_append=(
            ConsolidateMsmsRow(wo="45001002", location="CKTN0002CCCC", description="PE NEW", fl_erms="CKTN0002/CCCC"),
        ),
        files_to_move=(),
        duplicates_skipped=0,
        files_processed=1,
        errors=(),
    )

    loader.load(plan)

    wb_res = openpyxl.load_workbook(msms_file)
    ws_res = wb_res.active
    assert ws_res.max_row == 4  # Header + 2 existing + 1 new

    rows = list(ws_res.iter_rows(values_only=True))
    assert len(rows) == 4
    assert rows[1][0] == "45001000"
    assert rows[1][1] == "CKTN0000AAAA"
    assert rows[2][0] == "45001001"
    assert rows[2][1] == "CKTN0001BBBB"
    assert rows[3][0] == "45001002"
    assert rows[3][1] == "CKTN0002CCCC"
    wb_res.close()


def test_consolidate_msms_workflow_with_blank_formatted_rows_e2e(tmp_path: Path):
    """End-to-end integration test of ConsolidateMsmsWorkflow with 100 blank formatted rows."""
    from openpyxl.styles import PatternFill

    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append(["Work Order", "Location", "Description", "Substation Name ERMS", "FL ERMS", "Cycle Date", "Substation Number"])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in range(2, 102):
        ws.cell(row=r, column=1).fill = yellow_fill

    # Row 20 has 1 existing record
    ws.cell(row=20, column=1).value = "45001000"
    ws.cell(row=20, column=2).value = "CKTN0000AAAA"
    ws.cell(row=20, column=3).value = "PE EXISTING"
    ws.cell(row=20, column=5).value = "CKTN0000/AAAA"
    wb.save(msms_file)
    wb.close()

    msms_dir = env.storage.get_python_msms_dir()
    xls1 = msms_dir / "45501001.xls"
    _create_sample_maximo_html_xls(xls1, [
        ("45001000", "APPR", "CKTN0000AAAA", "PE EXISTING"),  # Duplicate
        ("45001001", "APPR", "CKTN/PCEJ01565", "PE TELUK SISIK"),  # New
        ("45001002", "APPR", "CKTN0002BBBB", "PE TANJUNG"),  # New
    ])

    workflow = ConsolidateMsmsWorkflow()
    result = workflow.execute(env, ConsolidateMsmsRequest())

    assert result.files_processed == 1
    assert result.rows_appended == 2
    assert result.duplicates_skipped == 1

    wb_res = openpyxl.load_workbook(msms_file)
    ws_res = wb_res.active
    assert ws_res.max_row == 4  # Header + 1 existing + 2 new

    rows = list(ws_res.iter_rows(values_only=True))
    assert len(rows) == 4
    assert rows[1][0] == "45001000"
    assert rows[2][0] == "45001001"
    assert rows[2][4] == "CKTN/PCE/J01565"
    assert rows[3][0] == "45001002"
    assert rows[3][4] == "CKTN0002/BBBB"
    wb_res.close()


