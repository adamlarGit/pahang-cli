"""Unit tests for the 1-Click Substation Post-Processing Orchestrator Service."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch
import openpyxl
import pytest

from src.postprocessing.converters import FakeDocumentConverter
from src.project.environment import ProjectEnvironment
from src.project.models import ProjectMetadata
from src.project.storage import LocalWorkspaceStorage
from src.testsheet.models import SubstationPackage
from src.testsheet.repository import LocalTestsheetPackageRepository
from src.workflows.models import (
    PostProcessingFailure,
    PostProcessingMode,
    PostProcessingRequest,
    PostProcessingSummary,
)
from src.workflows.postprocessing_pipeline import (
    PostProcessingPipelineWorkflow,
    run_postprocessing_pipeline,
)
from src.workflows.postprocessing_preflight import PreFlightValidationError
from src.workflows.service import WorkflowService


@pytest.fixture
def mock_env(tmp_path: Path) -> ProjectEnvironment:
    """Provide an isolated ProjectEnvironment backed by a temporary directory."""
    meta = ProjectMetadata(
        key="test_pahang",
        name="Test Pahang Project",
        base_path=str(tmp_path),
        state="pahang",
        po_number="PO-998877",
        voltage_type="11kV",
        year="2026",
        cycle="Cycle 1",
        technologies=("IR", "US", "TEV"),
    )
    storage = LocalWorkspaceStorage(tmp_path)
    return ProjectEnvironment(metadata=meta, storage=storage)


def _create_sample_testsheet(path: Path, fl: str = "FL-001", station: str = "PE TEST") -> None:
    """Create a valid openpyxl workbook with PCE Testsheet and PCE VI sheets containing placeholders."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "PCE Testsheet"
    ws1["A1"] = "SUBSTATION"
    ws1["B1"] = station
    ws1["A2"] = "FL"
    ws1["B2"] = fl
    ws1["D10"] = "{{signvendor}}"
    ws1["T10"] = "{{signtnb}}"
    ws1["A5"] = None  # blank cell in range A3:Y55
    
    ws2 = wb.create_sheet(title="PCE VI")
    ws2["A1"] = "VISUAL INSPECTION"
    ws2["C6"] = None  # blank cell in range C6:H6
    ws2["C10"] = "{{signvendor}}"
    ws2["K10"] = "{{signtnb}}"
    
    wb.save(path)
    wb.close()


def _setup_substation_files(
    env: ProjectEnvironment,
    date_folder: str = "01-05-2026",
    count: int = 2,
) -> tuple[list[Path], list[Path], list[Path]]:
    """Helper to scaffold matching QUICK REPORT, TESTSHEET, and RAW MATERIAL items."""
    qr_dir = env.get_quick_report_dir() / date_folder
    ts_dir = env.get_testsheet_dir() / date_folder
    raw_dir = env.get_raw_material_dir() / date_folder

    qr_dir.mkdir(parents=True, exist_ok=True)
    ts_dir.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)

    qr_files: list[Path] = []
    ts_files: list[Path] = []
    raw_folders: list[Path] = []

    for i in range(1, count + 1):
        stem = f"0{i}. PE STATION_{i}"
        qr_file = qr_dir / f"{stem}.docx"
        qr_file.write_text(f"dummy docx content {i}")
        qr_files.append(qr_file)

        ts_file = ts_dir / f"{stem}.xlsx"
        _create_sample_testsheet(ts_file, fl=f"FL-00{i}", station=f"PE STATION_{i}")
        ts_files.append(ts_file)

        raw_folder = raw_dir / stem
        raw_folder.mkdir(exist_ok=True)
        raw_folders.append(raw_folder)

    return qr_files, ts_files, raw_folders


def test_happy_path_by_date_with_all_stages(mock_env: ProjectEnvironment, tmp_path: Path) -> None:
    """Happy path BY_DATE mode:

    - Pre-flight check executes.
    - Renaming sync runs.
    - WhatsApp report runs when requested.
    - Signatures applied when enabled.
    - Diagonal borders applied.
    - Testsheet PDF and docx PDF converted and merged into QUICK REPORT/<DATE>/<STEM>.pdf.
    - Original testsheets remain strictly untouched (immutability).
    """
    date_str = "01-05-2026"
    qr_files, ts_files, _ = _setup_substation_files(mock_env, date_str, count=2)

    # Prepare signature image
    sign_dir = tmp_path / "signatures"
    sign_dir.mkdir()
    vendor_sign = sign_dir / "vendor.png"
    import io
    from PIL import Image as PILImage
    img = PILImage.new("RGB", (50, 50), color="blue")
    img.save(vendor_sign)
    tnb_sign = sign_dir / "tnb.png"
    img.save(tnb_sign)

    converter = FakeDocumentConverter()
    progress_messages: list[str] = []

    request = PostProcessingRequest(
        mode=PostProcessingMode.BY_DATE,
        target_dates=(date_str,),
        apply_signatures=True,
        vendor_signature_path=vendor_sign,
        tnb_signature_path=tnb_sign,
        generate_whatsapp=True,
        converter=converter,
        progress_sink=progress_messages.append,
    )

    workflow = PostProcessingPipelineWorkflow()

    with patch("src.workflows.postprocessing_pipeline.validate_postprocessing_preflight") as mock_preflight, \
         patch("src.workflows.postprocessing_pipeline.rename_files_match") as mock_rename, \
         patch("src.workflows.postprocessing_pipeline.run_generate_whatsapp_report") as mock_whatsapp:

        mock_preflight.return_value = MagicMock(is_valid=True)
        mock_rename.return_value = MagicMock(renamed=())
        mock_whatsapp.return_value = MagicMock(output_path=Path("mock_wa.docx"))

        summary = workflow.execute(mock_env, request)

        # 1. Pre-flight check called for the date folder
        assert mock_preflight.call_count == 1
        call_kwargs = mock_preflight.call_args.kwargs
        assert call_kwargs.get("date_folder") == date_str

        # 2. Renaming sync called for quick_report -> testsheet & quick_report -> raw_material
        assert mock_rename.call_count == 2

        # 3. WhatsApp report called
        mock_whatsapp.assert_called_once()

        # 4. Summary results
        assert isinstance(summary, PostProcessingSummary)
        assert len(summary.processed_packages) == 2
        assert len(summary.final_deliverables) == 2
        assert len(summary.failed_packages) == 0
        assert summary.is_successful is True

        # 5. Check deliverables exist in QUICK REPORT/<DATE>/<STEM>.pdf
        for deliv in summary.final_deliverables:
            assert deliv.exists()
            assert deliv.parent == mock_env.get_quick_report_dir() / date_str
            assert deliv.suffix.lower() == ".pdf"

        # 6. Check testsheet PDF in processed_testsheet/pdf/
        for ts in ts_files:
            proc_pdf = ts.parent / "processed_testsheet" / "pdf" / f"{ts.stem}.pdf"
            assert proc_pdf.exists()

        # 7. Check processed testsheet copy exists in processed_testsheet/<STEM>.xlsx
        for ts in ts_files:
            proc_xlsx = ts.parent / "processed_testsheet" / ts.name
            assert proc_xlsx.exists()

        # 8. Testsheet Immutability: Original file still has placeholders in disk content
        for ts in ts_files:
            wb_orig = openpyxl.load_workbook(ts)
            ws_orig = wb_orig["PCE Testsheet"]
            assert ws_orig["D10"].value == "{{signvendor}}"
            assert ws_orig["T10"].value == "{{signtnb}}"
            wb_orig.close()


def test_happy_path_by_date_signatures_disabled_strips_placeholders(
    mock_env: ProjectEnvironment,
) -> None:
    """When signatures are disabled (apply_signatures=False), mode='none' placeholder stripping is executed.

    Working copy has placeholders removed; original testsheet remains untouched.
    """
    date_str = "02-05-2026"
    qr_files, ts_files, _ = _setup_substation_files(mock_env, date_str, count=1)
    converter = FakeDocumentConverter()

    request = PostProcessingRequest(
        mode=PostProcessingMode.BY_DATE,
        target_dates=(date_str,),
        apply_signatures=False,
        generate_whatsapp=False,
        converter=converter,
    )

    workflow = PostProcessingPipelineWorkflow()

    with patch("src.workflows.postprocessing_pipeline.validate_postprocessing_preflight") as mock_preflight, \
         patch("src.workflows.postprocessing_pipeline.rename_files_match") as mock_rename, \
         patch("src.workflows.postprocessing_pipeline.run_generate_whatsapp_report") as mock_whatsapp:

        mock_preflight.return_value = MagicMock(is_valid=True)
        mock_rename.return_value = MagicMock(renamed=())

        summary = workflow.execute(mock_env, request)

        assert summary.is_successful is True
        mock_whatsapp.assert_not_called()

        ts = ts_files[0]
        proc_xlsx = ts.parent / "processed_testsheet" / ts.name
        assert proc_xlsx.exists()

        # Check working copy has placeholders removed
        wb_proc = openpyxl.load_workbook(proc_xlsx)
        ws_proc = wb_proc["PCE Testsheet"]
        assert ws_proc["D10"].value is None or ws_proc["D10"].value == ""
        assert ws_proc["T10"].value is None or ws_proc["T10"].value == ""
        wb_proc.close()

        # Check original source testsheet is untouched
        wb_orig = openpyxl.load_workbook(ts)
        ws_orig = wb_orig["PCE Testsheet"]
        assert ws_orig["D10"].value == "{{signvendor}}"
        assert ws_orig["T10"].value == "{{signtnb}}"
        wb_orig.close()


def test_happy_path_by_fl_mode(mock_env: ProjectEnvironment) -> None:
    """Happy path BY_FL mode:

    - Targets specific substation packages by FL/station.
    - WhatsApp report is automatically skipped even if requested.
    - Renaming sync runs for the parent date folders of the selected FLs.
    """
    date_str = "03-05-2026"
    qr_files, ts_files, _ = _setup_substation_files(mock_env, date_str, count=3)
    converter = FakeDocumentConverter()

    # Request only 1st substation by FL
    request = PostProcessingRequest(
        mode=PostProcessingMode.BY_FL,
        target_fls=("FL-001", "PE STATION_1"),
        generate_whatsapp=True,  # Should be ignored in BY_FL mode
        apply_signatures=False,
        converter=converter,
    )

    workflow = PostProcessingPipelineWorkflow()

    with patch("src.workflows.postprocessing_pipeline.validate_postprocessing_preflight") as mock_preflight, \
         patch("src.workflows.postprocessing_pipeline.rename_files_match") as mock_rename, \
         patch("src.workflows.postprocessing_pipeline.run_generate_whatsapp_report") as mock_whatsapp:

        mock_preflight.return_value = MagicMock(is_valid=True)
        mock_rename.return_value = MagicMock(renamed=())

        summary = workflow.execute(mock_env, request)

        # Preflight & Renaming ran for date_str
        assert mock_preflight.call_count == 1
        call_kwargs = mock_preflight.call_args.kwargs
        assert call_kwargs.get("date_folder") == date_str
        assert mock_rename.call_count == 2

        # WhatsApp skipped in BY_FL mode
        mock_whatsapp.assert_not_called()

        # Only 1 package processed
        assert len(summary.processed_packages) == 1
        assert summary.processed_packages[0].testsheet_xlsx.name == ts_files[0].name


def test_fail_fast_on_preflight_validation_failure(mock_env: ProjectEnvironment) -> None:
    """Fail-fast: PreFlightValidationError immediately aborts pipeline before modifying files or starting conversions."""
    date_str = "04-05-2026"
    _setup_substation_files(mock_env, date_str, count=2)
    converter = FakeDocumentConverter()

    request = PostProcessingRequest(
        mode=PostProcessingMode.BY_DATE,
        target_dates=(date_str,),
        converter=converter,
    )

    workflow = PostProcessingPipelineWorkflow()

    with patch("src.workflows.postprocessing_pipeline.validate_postprocessing_preflight") as mock_preflight, \
         patch("src.workflows.postprocessing_pipeline.rename_files_match") as mock_rename, \
         patch("src.postprocessing.converters.batch_com_session") as mock_session:

        mock_preflight.side_effect = PreFlightValidationError(
            "Count mismatch",
            date_folder=date_str,
            quick_report_count=2,
            testsheet_count=1,
        )

        with pytest.raises(PreFlightValidationError):
            workflow.execute(mock_env, request)

        # Ensure no renaming or COM session occurred
        mock_rename.assert_not_called()
        mock_session.assert_not_called()


def test_per_substation_error_isolation(mock_env: ProjectEnvironment) -> None:
    """Per-substation error isolation:

    When 1 substation fails during COM conversion, remaining substations process successfully,
    and PostProcessingSummary records both succeeded packages and failed items with error details.
    """
    date_str = "05-05-2026"
    qr_files, ts_files, _ = _setup_substation_files(mock_env, date_str, count=3)

    class FailingConverter(FakeDocumentConverter):
        def convert_testsheet_to_pdf(self, xlsx_path: Path, pdf_path: Path, *args, **kwargs) -> Path:
            if "STATION_2" in xlsx_path.name:
                raise RuntimeError("Excel COM RPC Server Unavailable for STATION_2")
            return super().convert_testsheet_to_pdf(xlsx_path, pdf_path, *args, **kwargs)

    converter = FailingConverter()

    request = PostProcessingRequest(
        mode=PostProcessingMode.BY_DATE,
        target_dates=(date_str,),
        apply_signatures=False,
        converter=converter,
    )

    workflow = PostProcessingPipelineWorkflow()

    with patch("src.workflows.postprocessing_pipeline.validate_postprocessing_preflight") as mock_preflight, \
         patch("src.workflows.postprocessing_pipeline.rename_files_match") as mock_rename:

        mock_preflight.return_value = MagicMock(is_valid=True)
        mock_rename.return_value = MagicMock(renamed=())

        summary = workflow.execute(mock_env, request)

        # 2 succeeded (STATION_1 and STATION_3), 1 failed (STATION_2)
        assert len(summary.processed_packages) == 2
        assert len(summary.final_deliverables) == 2
        assert len(summary.failed_packages) == 1

        failed_entry = summary.failed_packages[0]
        assert isinstance(failed_entry, PostProcessingFailure)
        assert "STATION_2" in failed_entry.package.testsheet_xlsx.name
        assert "Excel COM RPC Server Unavailable" in failed_entry.error
        assert summary.is_successful is False


def test_workflow_service_integration(mock_env: ProjectEnvironment) -> None:
    """WorkflowService.run_postprocessing_pipeline delegates correctly to orchestrator."""
    date_str = "06-05-2026"
    _setup_substation_files(mock_env, date_str, count=1)
    converter = FakeDocumentConverter()

    request = PostProcessingRequest(
        mode=PostProcessingMode.BY_DATE,
        target_dates=(date_str,),
        apply_signatures=False,
        converter=converter,
    )

    service = WorkflowService()

    with patch("src.workflows.postprocessing_pipeline.validate_postprocessing_preflight") as mock_preflight, \
         patch("src.workflows.postprocessing_pipeline.rename_files_match") as mock_rename:

        mock_preflight.return_value = MagicMock(is_valid=True)
        mock_rename.return_value = MagicMock(renamed=())

        summary = service.run_postprocessing_pipeline(mock_env, request)

        assert isinstance(summary, PostProcessingSummary)
        assert summary.is_successful is True
        assert len(summary.processed_packages) == 1
