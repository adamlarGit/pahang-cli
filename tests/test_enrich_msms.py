"""Unit and integration tests for EnrichMsmsWorkflow."""
from pathlib import Path
import openpyxl
import pytest

from src.project.environment import ProjectEnvironment
from src.project.models import ProjectMetadata
from src.project.storage import LocalWorkspaceStorage
from src.workflows.models import EnrichMsmsRequest, EnrichMsmsResult
from src.workflows.enrich_msms import (
    EnrichMsmsPreflightGuard,
    EnrichMsmsExtractor,
    EnrichMsmsFilter,
    EnrichMsmsTransformer,
    EnrichMsmsLoader,
    EnrichMsmsAuditor,
    EnrichMsmsWorkflow,
    EnrichCellUpdate,
    EnrichMsmsPlan,
)


def _create_sample_data_msms(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA MSMS"
    ws.append([
        "Work Order",
        "Location",
        "Description",
        "Substation Name ERMS",
        "FL ERMS",
        "Cycle Date",
        "Substation Number",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _create_sample_total_pe(path: Path, rows: list[list], sheet_name: str = "DataCycle1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([
        "PE NO",
        "FL NUMBER",
        "SUBSTATION NAME",
        "DATE",
        "TYPE",
        "WO",
        "SCOPE",
    ])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _setup_project_environment(tmp_path: Path) -> ProjectEnvironment:
    storage = LocalWorkspaceStorage(tmp_path)
    storage.get_python_dir().mkdir(parents=True, exist_ok=True)
    metadata = ProjectMetadata(
        key="pahang_2026",
        name="Pahang 2026 Test",
        po_number="PO42289580",
        state="Pahang",
        voltage_type="11kV",
        year="2026",
        cycle="2",
        technologies=("IR", "DG", "US", "TEV", "VI"),
        base_path=str(tmp_path),
    )
    return ProjectEnvironment(metadata=metadata, storage=storage)


# --- PreflightGuard Tests ---

def test_enrich_msms_preflight_missing_data_msms(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    _create_sample_total_pe(env.storage.get_total_pe_path(), [])

    guard = EnrichMsmsPreflightGuard()
    with pytest.raises(FileNotFoundError, match="DATA MSMS workbook not found"):
        guard.validate(env)


def test_enrich_msms_preflight_missing_total_pe(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    _create_sample_data_msms(env.storage.get_data_msms_path(), [])

    guard = EnrichMsmsPreflightGuard()
    with pytest.raises(FileNotFoundError, match="TOTAL PE workbook not found"):
        guard.validate(env)


def test_enrich_msms_preflight_missing_sheet(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    _create_sample_data_msms(env.storage.get_data_msms_path(), [])
    _create_sample_total_pe(env.storage.get_total_pe_path(), [], sheet_name="WrongSheet")

    guard = EnrichMsmsPreflightGuard()
    with pytest.raises(RuntimeError, match="'DataCycle1' sheet missing"):
        guard.validate(env)


def test_enrich_msms_preflight_success(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    _create_sample_data_msms(env.storage.get_data_msms_path(), [])
    _create_sample_total_pe(env.storage.get_total_pe_path(), [])

    guard = EnrichMsmsPreflightGuard()
    guard.validate(env)  # Should not raise


# --- Extractor Tests ---

def test_enrich_msms_extractor(tmp_path: Path):
    extractor = EnrichMsmsExtractor()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [
        ["45001001", "LOC1", "DESC1", None, None, None, None],
    ])

    pe_file = tmp_path / "TOTAL PE.xlsx"
    _create_sample_total_pe(pe_file, [
        [10, "CKTN0001/BBBB", "PE SUB 1", "2026-01-15", "ATTACHED", "45001001", "FULL"],
    ])

    msms_rows = extractor.read_data_msms_rows(msms_file)
    assert len(msms_rows) == 1
    assert msms_rows[0]["row_idx"] == 2
    assert msms_rows[0]["wo"] == "45001001"

    pe_lookups = extractor.read_total_pe_lookups(pe_file)
    assert "45001001" in pe_lookups.by_wo
    assert pe_lookups.by_wo["45001001"]["substation_name_erms"] == "PE SUB 1"
    assert pe_lookups.by_wo["45001001"]["substation_number"] == 10


# --- Filter & Transformer Tests ---

def test_enrich_msms_filter_and_transformer():
    filter_stage = EnrichMsmsFilter()
    transformer = EnrichMsmsTransformer()

    msms_rows = [
        # Row 2: Unpopulated D-G -> matched and needs update
        {
            "row_idx": 2,
            "wo": "45001001",
            "fl_erms": None,
            "substation_name_erms": None,
            "cycle_date": None,
            "substation_number": None,
        },
        # Row 3: Already populated D-G -> matched but no cell updates needed
        {
            "row_idx": 3,
            "wo": "45001002",
            "fl_erms": "MANUAL/FL",
            "substation_name_erms": "MANUAL NAME",
            "cycle_date": "2026-01-01",
            "substation_number": 99,
        },
        # Row 4: Unmatched WO
        {
            "row_idx": 4,
            "wo": "45009999",
            "fl_erms": None,
            "substation_name_erms": None,
            "cycle_date": None,
            "substation_number": None,
        },
    ]

    total_pe_by_wo = {
        "45001001": {
            "substation_number": 1,
            "fl_erms": "CKTN0001/BBBB",
            "substation_name_erms": "PE ALPHA",
            "cycle_date": "2026-01-10",
        },
        "45001002": {
            "substation_number": 2,
            "fl_erms": "CKTN0002/CCCC",
            "substation_name_erms": "PE BETA",
            "cycle_date": "2026-01-11",
        },
    }

    filtered_updates, matched_count, unmatched_count, unmatched_wos, updated_cells_count = filter_stage.filter_rows(
        msms_rows=msms_rows,
        total_pe_by_wo=total_pe_by_wo,
        total_pe_by_fl={},
    )

    assert matched_count == 2
    assert unmatched_count == 1
    assert unmatched_wos == ("45009999",)
    assert updated_cells_count == 4
    assert len(filtered_updates) == 1
    assert filtered_updates[0].row_index == 2
    assert filtered_updates[0].substation_name_erms == "PE ALPHA"
    assert filtered_updates[0].fl_erms == "CKTN0001/BBBB"
    assert filtered_updates[0].cycle_date == "2026-01-10"
    assert filtered_updates[0].substation_number == 1

    plan = transformer.build_plan(
        data_msms_path=Path("DATA MSMS.xlsx"),
        updates=filtered_updates,
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        unmatched_wos=unmatched_wos,
        updated_cells_count=updated_cells_count,
    )
    assert isinstance(plan, EnrichMsmsPlan)
    assert plan.matched_count == 2


# --- Loader Tests ---

def test_enrich_msms_loader(tmp_path: Path):
    loader = EnrichMsmsLoader()
    msms_file = tmp_path / "DATA MSMS.xlsx"
    _create_sample_data_msms(msms_file, [
        ["45001001", "LOC1", "DESC1", None, None, None, None],
        ["45001002", "LOC2", "DESC2", "EXISTING SUB", "EXISTING/FL", "2026-01-01", 99],
    ])

    plan = EnrichMsmsPlan(
        data_msms_path=msms_file,
        updates=(
            EnrichCellUpdate(
                row_index=2,
                substation_name_erms="PE TELUK SISIK",
                fl_erms="CKTN0001/BBBB",
                cycle_date="2026-01-15",
                substation_number=10,
            ),
        ),
        matched_count=1,
        unmatched_count=0,
        unmatched_wos=(),
        updated_cells_count=4,
    )

    loader.load(plan)

    # Verify openpyxl write
    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Row 2 (updated)
    assert rows[1][0] == "45001001"
    assert rows[1][3] == "PE TELUK SISIK"
    assert rows[1][4] == "CKTN0001/BBBB"
    assert str(rows[1][5]) == "2026-01-15"
    assert rows[1][6] == 10

    # Row 3 (untouched)
    assert rows[2][3] == "EXISTING SUB"
    assert rows[2][4] == "EXISTING/FL"
    assert str(rows[2][5]) == "2026-01-01"
    assert rows[2][6] == 99


# --- End-to-End Integration Workflow Tests ---

def test_enrich_msms_workflow_e2e(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    _create_sample_data_msms(msms_file, [
        ["45001001", "LOC1", "DESC1", None, None, None, None],
        ["45001002", "LOC2", "DESC2", "MANUAL SUB", "MANUAL/FL", "2026-01-05", 99],
        ["45009999", "LOC3", "DESC3", None, None, None, None],
    ])

    pe_file = env.storage.get_total_pe_path()
    _create_sample_total_pe(pe_file, [
        [12, "CKTN0001/BBBB", "PE BANDAR RAYA", "2026-01-15", "ATTACHED", "45001001", "FULL"],
        [15, "CKTN0002/CCCC", "PE BUKIT MEWAH", "2026-01-16", "COMPACT", "45001002", "FULL"],
    ])

    workflow = EnrichMsmsWorkflow()
    result = workflow.execute(env, EnrichMsmsRequest())

    assert isinstance(result, EnrichMsmsResult)
    assert result.matched_count == 2
    assert result.unmatched_count == 1
    assert "45009999" in result.unmatched_wos
    assert result.updated_cells_count == 4

    # Verify openpyxl update
    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Row 2 (45001001): was all None in cols D-G, enriched
    assert rows[1][0] == "45001001"
    assert rows[1][3] == "PE BANDAR RAYA"
    assert rows[1][4] == "CKTN0001/BBBB"
    assert str(rows[1][5]) == "2026-01-15"
    assert rows[1][6] == 12

    # Row 3 (45001002): had manual entries, not overwritten
    assert rows[2][3] == "MANUAL SUB"
    assert rows[2][4] == "MANUAL/FL"
    assert str(rows[2][5]) == "2026-01-05"
    assert rows[2][6] == 99


def test_enrich_msms_filter_matches_via_location_when_fl_erms_blank():
    filter_stage = EnrichMsmsFilter()
    msms_rows = [
        {
            "row_idx": 2,
            "wo": "45001001",
            "location": "CRAU/PCEJ00232",
            "fl_erms": None,
            "substation_name_erms": None,
            "cycle_date": None,
            "substation_number": None,
        }
    ]
    total_pe_by_fl = {
        "CRAU/PCE/J00232": {
            "substation_number": 25,
            "fl_erms": "CRAU/PCE/J00232",
            "substation_name_erms": "PE CRAU JAYA",
            "cycle_date": "2026-02-20",
        },
        "CRAUPCEJ00232": {
            "substation_number": 25,
            "fl_erms": "CRAU/PCE/J00232",
            "substation_name_erms": "PE CRAU JAYA",
            "cycle_date": "2026-02-20",
        },
    }

    filtered_updates, matched_count, unmatched_count, unmatched_wos, updated_cells_count = filter_stage.filter_rows(
        msms_rows=msms_rows,
        total_pe_by_wo={},
        total_pe_by_fl=total_pe_by_fl,
    )

    assert matched_count == 1
    assert unmatched_count == 0
    assert updated_cells_count == 4
    assert len(filtered_updates) == 1
    assert filtered_updates[0].row_index == 2
    assert filtered_updates[0].substation_name_erms == "PE CRAU JAYA"
    assert filtered_updates[0].fl_erms == "CRAU/PCE/J00232"
    assert filtered_updates[0].cycle_date == "2026-02-20"
    assert filtered_updates[0].substation_number == 25


def test_enrich_msms_workflow_e2e_matches_via_location_when_fl_erms_and_wo_blank(tmp_path: Path):
    env = _setup_project_environment(tmp_path)
    msms_file = env.storage.get_data_msms_path()
    # Row 2 has location CRAU/PCEJ00232 but FL ERMS (Col E) is None
    _create_sample_data_msms(msms_file, [
        ["45001001", "CRAU/PCEJ00232", "PE CRAU JAYA DESC", None, None, None, None],
    ])

    pe_file = env.storage.get_total_pe_path()
    # Total PE has WO as None (Col F) but FL NUMBER as CRAU/PCE/J00232
    _create_sample_total_pe(pe_file, [
        [35, "CRAU/PCE/J00232", "PE CRAU JAYA", "2026-03-01", "ATTACHED", None, "FULL"],
    ])

    workflow = EnrichMsmsWorkflow()
    result = workflow.execute(env, EnrichMsmsRequest())

    assert result.matched_count == 1
    assert result.unmatched_count == 0
    assert result.updated_cells_count == 4

    wb = openpyxl.load_workbook(msms_file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    assert rows[1][0] == "45001001"
    assert rows[1][1] == "CRAU/PCEJ00232"
    assert rows[1][3] == "PE CRAU JAYA"
    assert rows[1][4] == "CRAU/PCE/J00232"
    assert str(rows[1][5]) == "2026-03-01"
    assert rows[1][6] == 35

