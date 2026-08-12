import zipfile
import openpyxl
from openpyxl.worksheet.table import Table
from pathlib import Path

from src.workflows.diagonal_borders import process_workbook

def test_diagonal_borders_applies_borders_and_clears_tables(tmp_path: Path):
    # Create a test Excel workbook with a table and blank cells
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCE Testsheet"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Val1"
    ws["B2"] = "Val2"
    ws["A3"] = "" # blank cell in A3:Y55 range
    
    tab = Table(displayName="Table1", ref="A1:B2")
    ws.add_table(tab)
    
    input_file = tmp_path / "001.xlsx"
    wb.save(input_file)
    
    # Process workbook
    out_file = Path(process_workbook(input_file))
    assert out_file.exists()
    
    # Verify openpyxl loaded output has diagonal border on A3
    wb_out = openpyxl.load_workbook(out_file)
    ws_out = wb_out["PCE Testsheet"]
    assert ws_out["A3"].border.diagonalUp is True
    
    # Check that zip relationships do not contain absolute Target="/xl/tables/"
    with zipfile.ZipFile(out_file, 'r') as z:
        for name in z.namelist():
            if name.startswith("xl/worksheets/_rels/"):
                rel_xml = z.read(name).decode("utf-8")
                assert 'Target="/xl/tables/' not in rel_xml, f"Found invalid absolute target path in {name}: {rel_xml}"
                assert 'Target="/xl/drawings/' not in rel_xml, f"Found invalid absolute target path in {name}: {rel_xml}"
